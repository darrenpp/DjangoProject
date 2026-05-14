import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


WORKBOOK_CANDIDATES = [
    Path(
        r"c:\Users\timhi\OneDrive\Desktop\ParotOs\NDOH_Database\Database Template\UTS\Current_DATA\Provisional_and_Full_Registration2024_2026.xlsx"
    ),
    Path("notebooks") / "Provisional2017_2026and_full_license2009_2026.xlsx",
]
DEFAULT_SHEET = "FULL REGO 2009 - current"
OUTPUT_CLEAN = Path("notebooks") / "full_registrations_cleaned.csv"
OUTPUT_ISSUES = Path("notebooks") / "full_registrations_quality_issues.csv"
OUTPUT_SUMMARY = Path("notebooks") / "full_registrations_summary.txt"

FOREIGN_KEYWORDS = [
    "america",
    "american",
    "philippines",
    "philippine",
    "philipine",
    "india",
    "italy",
    "fiji",
    "china",
    "australia",
    "auckland",
    "new zealand",
    "new zealnd",
    "new zeland",
    "nz",
    "uk",
    "united kingdom",
    "usa",
    "united states",
    "kenya",
    "uganda",
    "japan",
    "malaysia",
    "singapore",
    "canada",
    "indonesia",
    "thailand",
]

MONTH_FIXES = {
    "Agu": "Aug",
    "Ago": "Aug",
    "Ap-": "Apr-",
    "Ap ": "Apr ",
    "Jin": "Jun",
    "Jui": "Jul",
    "Mac": "Mar",
    "Ocf": "Oct",
    "Sept": "Sep",
    "Set": "Sep",
}

NULL_LIKE_VALUES = {"", "NAN", "NAT", "NONE", "NULL", "TBA", "N/A", "NA"}
PRACTITIONER_TEXT_KEYWORDS = [
    "P O BOX",
    "PO BOX",
    "PRIVATE MAIL BAG",
    "PROVINCE",
    "HOSPITAL",
    "CLINIC",
    "AUTHORITY",
    "HEALTH",
]


@dataclass
class CleanseResult:
    cleaned: pd.DataFrame
    issues: pd.DataFrame


def resolve_default_workbook():
    for candidate in WORKBOOK_CANDIDATES:
        if candidate.exists():
            return candidate
    return WORKBOOK_CANDIDATES[0]


WORKBOOK_PATH = resolve_default_workbook()


def normalize_text(value):
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_name(value):
    text = normalize_text(value)
    if not text:
        return ""

    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"(?<=[A-Za-z])\d{2,}$", "", text).strip()

    if "," in text:
        left, right = [part.strip() for part in text.split(",", 1)]
        if left and right:
            text = f"{right} {left}"

    text = re.sub(r"\s+", " ", text).strip(" -")
    return text.title()


def split_name(full_name):
    parts = normalize_text(full_name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def normalize_code(value):
    text = normalize_text(value).upper()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def normalize_registration_no(p_code, license_no):
    code = normalize_code(p_code)
    reg = normalize_code(license_no)
    reg = reg.replace("-", " ")
    reg = re.sub(r"\s*/\s*", "/", reg)
    reg = re.sub(r"\s+", " ", reg).strip()

    if reg and code and not reg.startswith(code):
        reg = f"{code} {reg}"

    return re.sub(r"\s+", " ", reg).strip()


def normalize_practitioner_no(value):
    text = normalize_text(value)
    if not text:
        return ""
    text = re.sub(r"\.0$", "", text)
    return text


def is_valid_practitioner_no(value):
    text = normalize_practitioner_no(value)
    if not text:
        return False
    upper_text = text.upper()
    if len(text) > 50:
        return False
    if "," in text:
        return False
    if sum(1 for keyword in PRACTITIONER_TEXT_KEYWORDS if keyword in upper_text):
        return False
    return True


def infer_applicant_type(institution_name):
    institution_lower = normalize_text(institution_name).lower()
    if any(keyword in institution_lower for keyword in FOREIGN_KEYWORDS):
        return "overseas"
    return "national"


def infer_profession_track(qualification_name):
    qualification_lower = normalize_text(qualification_name).lower()
    if "midw" in qualification_lower:
        return "midwifery"
    return "nursing"


def infer_pathway(applicant_type, profession_track):
    if applicant_type == "overseas" and profession_track == "midwifery":
        return "overseas_midwife"
    if applicant_type == "overseas":
        return "overseas_nurse"
    if profession_track == "midwifery":
        return "local_midwifery_graduate"
    return "local_nursing_graduate"


def clean_date_text(raw_value):
    text = normalize_text(raw_value)
    if not text:
        return ""

    if text.upper() in NULL_LIKE_VALUES:
        return ""

    for bad, good in MONTH_FIXES.items():
        text = text.replace(bad, good)

    text = text.replace("_", "-").replace("=", "-")
    text = text.replace(" .", ".").replace(". ", ".")
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"-{2,}", "-", text)
    text = re.sub(r"\.{2,}", ".", text)
    text = re.sub(r"\s+", " ", text).strip()

    if re.match(r"^\d{1,2}-[A-Za-z]{2}-\d{2}$", text):
        month_map = {
            "JA": "Jan",
            "FE": "Feb",
            "MR": "Mar",
            "AP": "Apr",
            "MY": "May",
            "JN": "Jun",
            "JL": "Jul",
            "AU": "Aug",
            "SE": "Sep",
            "OC": "Oct",
            "NO": "Nov",
            "DE": "Dec",
        }
        day_value, month_value, year_value = text.split("-")
        month_key = month_value.upper()
        if month_key in month_map:
            text = f"{day_value}-{month_map[month_key]}-{year_value}"

    return text


def parse_issued_date(raw_value):
    if raw_value is None or pd.isna(raw_value):
        return pd.NaT, "missing_issued_date"

    parsed = pd.NaT
    if isinstance(raw_value, pd.Timestamp):
        parsed = raw_value
    elif hasattr(raw_value, "year") and hasattr(raw_value, "month") and hasattr(raw_value, "day"):
        parsed = pd.Timestamp(raw_value)
    else:
        text = clean_date_text(raw_value)
        if not text:
            return pd.NaT, "missing_issued_date"

        for dayfirst in (True, False):
            attempt = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
            if not pd.isna(attempt):
                parsed = attempt
                break

    if pd.isna(parsed):
        return pd.NaT, "unparsed_issued_date"

    parsed = pd.Timestamp(parsed).normalize()
    current_year = date.today().year

    if parsed.year < 2000:
        return pd.NaT, "invalid_pre_2000_date"

    if parsed.year > current_year + 1 and parsed.year - 20 >= 2000:
        return parsed.replace(year=parsed.year - 20), "repaired_future_year_minus_20"

    return parsed, "valid"


def row_score(record):
    return sum(
        1
        for value in [
            record.get("issued_date"),
            record.get("institution_name"),
            record.get("graduation_year"),
            record.get("qualification_name"),
            record.get("practitioner_no"),
        ]
        if value not in ("", None) and not pd.isna(value)
    )


def load_sheet_rows(workbook_path, sheet_name):
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
        source_id, name, license_type, p_code, license_no, issued_date, institution_name, graduation_year, qualification_name, practitioner_no = row
        if not any(value is not None and str(value).strip() for value in row[:10]):
            continue
        rows.append(
            {
                "source_id": source_id,
                "name_raw": name,
                "license_type": license_type,
                "p_code": p_code,
                "license_no_raw": license_no,
                "issued_date_raw": issued_date,
                "institution_name_raw": institution_name,
                "graduation_year_raw": graduation_year,
                "qualification_name_raw": qualification_name,
                "practitioner_no_raw": practitioner_no,
            }
        )

    return pd.DataFrame(rows)


def cleanse_workbook(workbook_path=WORKBOOK_PATH, sheet_name=DEFAULT_SHEET):
    df = load_sheet_rows(workbook_path=workbook_path, sheet_name=sheet_name)

    if df.empty:
        return CleanseResult(cleaned=pd.DataFrame(), issues=pd.DataFrame())

    df["full_name"] = df["name_raw"].apply(normalize_name)
    df["license_type_clean"] = df["license_type"].apply(normalize_text)
    df["p_code_clean"] = df["p_code"].apply(normalize_code)
    df["license_no_clean"] = df["license_no_raw"].apply(normalize_code)
    df["registration_no"] = df.apply(
        lambda row: normalize_registration_no(row["p_code_clean"], row["license_no_clean"]),
        axis=1,
    )
    df["institution_name"] = df["institution_name_raw"].apply(normalize_text)
    df["qualification_name"] = df["qualification_name_raw"].apply(normalize_text)
    df["practitioner_no"] = df["practitioner_no_raw"].apply(normalize_practitioner_no)
    df["graduation_year"] = pd.to_numeric(df["graduation_year_raw"], errors="coerce")

    issued_values = df["issued_date_raw"].apply(parse_issued_date)
    df["issued_date"] = [item[0] for item in issued_values]
    df["issued_date_status"] = [item[1] for item in issued_values]

    name_parts = df["full_name"].apply(split_name)
    df["first_name"] = [part[0] for part in name_parts]
    df["last_name"] = [part[1] for part in name_parts]

    df["applicant_type"] = df["institution_name"].apply(infer_applicant_type)
    df["profession_track"] = df["qualification_name"].apply(infer_profession_track)
    df["pathway"] = df.apply(
        lambda row: infer_pathway(row["applicant_type"], row["profession_track"]),
        axis=1,
    )
    df["target_model"] = df["profession_track"].map(
        {"midwifery": "midwife", "nursing": "nursingprofessional"}
    )

    df = df[
        (df["license_type_clean"].str.contains("full", case=False, na=False))
        & (df["full_name"] != "")
        & (df["registration_no"] != "")
    ].copy()

    issues = []
    cleaned_records = []

    for registration_no, group in df.groupby("registration_no", dropna=False):
        name_set = {
            normalize_text(name).upper()
            for name in group["full_name"].tolist()
            if normalize_text(name)
        }

        if len(group) > 1 and len(name_set) > 1:
            for _, row in group.iterrows():
                issues.append(
                    {
                        "registration_no": registration_no,
                        "source_id": row["source_id"],
                        "full_name": row["full_name"],
                        "license_no_raw": row["license_no_raw"],
                        "issued_date_raw": row["issued_date_raw"],
                        "qualification_name": row["qualification_name"],
                        "issue_type": "duplicate_registration_conflict",
                        "issue_detail": f"Conflicting names share registration number {registration_no}",
                    }
                )
            continue

        chosen = max(group.to_dict("records"), key=row_score)
        cleaned_records.append(chosen)

        if len(group) > 1:
            for _, row in group.iterrows():
                if row["source_id"] == chosen["source_id"] and row["full_name"] == chosen["full_name"]:
                    continue
                issues.append(
                    {
                        "registration_no": registration_no,
                        "source_id": row["source_id"],
                        "full_name": row["full_name"],
                        "license_no_raw": row["license_no_raw"],
                        "issued_date_raw": row["issued_date_raw"],
                        "qualification_name": row["qualification_name"],
                        "issue_type": "duplicate_registration_dropped",
                        "issue_detail": f"Duplicate row dropped in favor of the most complete record for {registration_no}",
                    }
                )

    cleaned = pd.DataFrame(cleaned_records)
    if cleaned.empty:
        return CleanseResult(cleaned=cleaned, issues=pd.DataFrame(issues))

    cleaned["practitioner_no_is_valid"] = cleaned["practitioner_no"].apply(is_valid_practitioner_no)
    practitioner_counts = cleaned.loc[
        cleaned["practitioner_no_is_valid"], "practitioner_no"
    ].replace("", pd.NA).dropna().value_counts()
    duplicate_practitioners = set(practitioner_counts[practitioner_counts > 1].index.tolist())
    cleaned["practitioner_no_unique"] = cleaned["practitioner_no"].apply(
        lambda value: (
            value
            if value
            and is_valid_practitioner_no(value)
            and value not in duplicate_practitioners
            else ""
        )
    )

    for _, row in cleaned.iterrows():
        row_issues = []
        if pd.isna(row["issued_date"]):
            row_issues.append("Missing or invalid issued date")
        if row["issued_date_status"] != "valid":
            row_issues.append(f"Issued date status: {row['issued_date_status']}")
        if not row["institution_name"] or row["institution_name"].upper() in {"TBA"}:
            row_issues.append("Missing institution")
        if not row["qualification_name"] or row["qualification_name"].upper() in {"TBA"}:
            row_issues.append("Missing qualification")
        if pd.isna(row["graduation_year"]):
            row_issues.append("Missing graduation year")
        if row["practitioner_no"] and row["practitioner_no"] in duplicate_practitioners:
            row_issues.append("Duplicate practitioner number not imported")
        if row["practitioner_no"] and not row["practitioner_no_is_valid"]:
            row_issues.append("Invalid practitioner number format not imported")

        for issue in row_issues:
            issues.append(
                {
                    "registration_no": row["registration_no"],
                    "source_id": row["source_id"],
                    "full_name": row["full_name"],
                    "license_no_raw": row["license_no_raw"],
                    "issued_date_raw": row["issued_date_raw"],
                    "qualification_name": row["qualification_name"],
                    "issue_type": "quality_warning",
                    "issue_detail": issue,
                }
            )

    cleaned = cleaned[
        [
            "registration_no",
            "source_id",
            "full_name",
            "first_name",
            "last_name",
            "p_code_clean",
            "license_no_clean",
            "issued_date",
            "issued_date_status",
            "institution_name",
            "graduation_year",
            "qualification_name",
            "practitioner_no_unique",
            "applicant_type",
            "profession_track",
            "pathway",
            "target_model",
        ]
    ].copy()
    cleaned = cleaned.rename(
        columns={
            "p_code_clean": "p_code",
            "license_no_clean": "license_no",
            "practitioner_no_unique": "practitioner_no",
        }
    )
    cleaned = cleaned.sort_values(["target_model", "registration_no"]).reset_index(drop=True)

    issues_df = pd.DataFrame(issues).sort_values(
        ["issue_type", "registration_no", "source_id"],
        na_position="last",
    )
    return CleanseResult(cleaned=cleaned, issues=issues_df)


def export_outputs(result: CleanseResult):
    result.cleaned.to_csv(OUTPUT_CLEAN, index=False)
    result.issues.to_csv(OUTPUT_ISSUES, index=False)

    summary_lines = [
        "Full Registrations Cleansing Summary",
        f"Cleaned records: {len(result.cleaned)}",
        f"Quality issue rows: {len(result.issues)}",
        f"Nursing professionals: {(result.cleaned['target_model'] == 'nursingprofessional').sum() if not result.cleaned.empty else 0}",
        f"Midwives: {(result.cleaned['target_model'] == 'midwife').sum() if not result.cleaned.empty else 0}",
        f"National records: {(result.cleaned['applicant_type'] == 'national').sum() if not result.cleaned.empty else 0}",
        f"Overseas records: {(result.cleaned['applicant_type'] == 'overseas').sum() if not result.cleaned.empty else 0}",
        f"Missing issued dates: {int(result.cleaned['issued_date'].isna().sum()) if not result.cleaned.empty else 0}",
        f"Unique institutions: {result.cleaned['institution_name'].replace('', pd.NA).dropna().nunique() if not result.cleaned.empty else 0}",
    ]
    OUTPUT_SUMMARY.write_text("\n".join(summary_lines), encoding="utf-8")


if __name__ == "__main__":
    result = cleanse_workbook()
    export_outputs(result)
    print(f"Cleaned records written to: {OUTPUT_CLEAN}")
    print(f"Quality issues written to: {OUTPUT_ISSUES}")
    print(f"Summary written to: {OUTPUT_SUMMARY}")
