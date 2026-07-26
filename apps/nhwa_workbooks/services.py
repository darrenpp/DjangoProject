import re
import ast
import json
import zipfile
from io import BytesIO
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.utils import timezone

from .models import (
    NHWACellEntry,
    NHWACellTemplate,
    NHWAWebSheet,
    NHWAWebWorkbook,
    NHWAWorkbookAuditEvent,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError:  # pragma: no cover - dependency is validated by management command/checks.
    Workbook = None
    load_workbook = None
    column_index_from_string = None
    get_column_letter = None


SOURCE_WORKBOOK = (
    Path(settings.BASE_DIR)
    / "docs"
    / "nhwa_toolkit"
    / "source"
    / "PNG_NHWA_Data_Collection_Toolkit_v2_May2026.xlsx"
)
SOURCE_PRESENTATION = (
    Path(settings.BASE_DIR)
    / "docs"
    / "nhwa_toolkit"
    / "source"
    / "PNG_NHWA_Output2_Presentation_May2026.pptx"
)

OFFICE_WORKBOOKS = {
    "nursing": "Nursing Council NHWA Web Workbook",
    "medical": "Medical Board NHWA Web Workbook",
}

EDITABLE_SOURCE_SHEETS = {
    "T1_PHA_ESTABLISHMENT",
    "T2_TRAINING_SCHOOL",
    "T3_COUNCIL_REGISTER",
    "T4_FINANCE",
    "DATA_QUALITY_CHECKLIST",
}

REFERENCE_SOURCE_SHEETS = {
    "GUIDE",
    "CADRE_MAPPING",
}

WORKBOOK_HEADER_FIELDS = [
    {
        "key": "province_institution",
        "label": "Province / Institution",
        "label_coordinate": "A5",
        "value_coordinate": "A6",
    },
    {
        "key": "reporting_period",
        "label": "Reporting Period",
        "label_coordinate": "D5",
        "value_coordinate": "D6",
    },
    {
        "key": "data_source_system",
        "label": "Data Source System",
        "label_coordinate": "G5",
        "value_coordinate": "G6",
    },
    {
        "key": "completed_by",
        "label": "Completed By",
        "label_coordinate": "K5",
        "value_coordinate": "K6",
    },
    {
        "key": "date",
        "label": "Date",
        "label_coordinate": "O5",
        "value_coordinate": "O6",
    },
]

# The source toolkit is a national master workbook. The web workbook keeps the
# workbook sheets and column layout intact, but hides rows that belong to other
# regulatory bodies so each office only captures its own data.
WORKBOOK_SCOPE_VISIBLE_RANGES = {
    "nursing": {
        "T1_PHA_ESTABLISHMENT": [(1, 11), (26, 35), (90, 92)],
        "T2_TRAINING_SCHOOL": [(1, 11), (16, 22), (46, 46), (48, 50), (53, 55), (66, 93)],
        "T3_COUNCIL_REGISTER": [(1, 14), (21, 25), (28, 28), (38, 40), (43, 44)],
        "T4_FINANCE": [(1, 13), (15, 15), (20, 35), (37, 39), (44, 44), (47, 49), (51, 57), (59, 59)],
    },
    "medical": {
        "T1_PHA_ESTABLISHMENT": [(1, 25), (36, 42), (90, 92)],
        "T2_TRAINING_SCHOOL": [(1, 15), (23, 26), (37, 42), (46, 46), (48, 52), (56, 57), (61, 64), (66, 93)],
        "T3_COUNCIL_REGISTER": [(1, 20), (38, 42), (45, 46)],
        "T4_FINANCE": [(1, 14), (17, 17), (20, 35), (37, 43), (45, 49), (51, 58), (60, 61)],
    },
}

WORKBOOK_SCOPE_NOTES = {
    "nursing": "Nursing Council capture only: nursing personnel, midwifery, nurse training and Nursing Council register/finance rows are shown. Other regulatory rows are hidden and ignored on save.",
    "medical": "Medical Board capture only: medical officers, specialists, CHW/HEO/EHO and approved Medical Board training/finance rows are shown. Other regulatory rows are hidden and ignored on save.",
}

YELLOW_FILL_SUFFIXES = {"FFFDE7", "FFF2CC", "FFFF00"}
GREY_FILL_SUFFIXES = {"EEF2F6", "F8F9FA", "E8EEF4"}


def clean_cell_value(value):
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def rgb_from_cell(cell):
    fill = getattr(cell, "fill", None)
    if not fill or not getattr(fill, "fgColor", None):
        return ""
    if getattr(fill, "fill_type", None) in (None, "none"):
        return ""
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:].upper()
    return ""


def is_dark_hex(rgb):
    if not rgb or len(rgb) != 6:
        return False
    red = int(rgb[0:2], 16)
    green = int(rgb[2:4], 16)
    blue = int(rgb[4:6], 16)
    return (red * Decimal("0.299") + green * Decimal("0.587") + blue * Decimal("0.114")) < 135


def cell_style_json(cell, value, rgb):
    style = {
        "bold": bool(getattr(cell.font, "bold", False)),
        "italic": bool(getattr(cell.font, "italic", False)),
        "fill_rgb": rgb,
        "text_color": "#ffffff" if is_dark_hex(rgb) else "#0f172a",
    }
    return style


def _cell_is_editable(sheet_name, cell, value, rgb):
    if sheet_name not in EDITABLE_SOURCE_SHEETS:
        return False
    if isinstance(value, str) and value.startswith("="):
        return False
    header_value_coordinates = {field["value_coordinate"] for field in WORKBOOK_HEADER_FIELDS}
    if cell.row <= 8 and cell.coordinate not in header_value_coordinates and sheet_name != "DATA_QUALITY_CHECKLIST":
        return False
    if rgb in YELLOW_FILL_SUFFIXES:
        return True
    if cell.coordinate in header_value_coordinates:
        return True
    return False


def _cell_is_heading(cell, value, rgb):
    if not value:
        return False
    if cell.row <= 8:
        return True
    if getattr(cell.font, "bold", False) and rgb not in YELLOW_FILL_SUFFIXES:
        return True
    return rgb and rgb not in YELLOW_FILL_SUFFIXES and rgb not in GREY_FILL_SUFFIXES


def _sheet_purpose(worksheet):
    for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row or 1, 10), values_only=True):
        values = [clean_cell_value(value) for value in row]
        text = " ".join(value for value in values if value)
        if len(text) > 40 and not text.upper().startswith("WHO "):
            return text[:600]
    return ""


@transaction.atomic
def bootstrap_web_workbooks(actor=None, source_path=None):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to bootstrap NHWA web workbooks.")

    source_path = Path(source_path or SOURCE_WORKBOOK)
    if not source_path.exists():
        raise FileNotFoundError(f"NHWA source workbook was not found: {source_path}")

    workbook = load_workbook(source_path, data_only=False)
    created = {"workbooks": 0, "sheets": 0, "cells": 0}

    for office_scope, title in OFFICE_WORKBOOKS.items():
        web_workbook, workbook_created = NHWAWebWorkbook.objects.update_or_create(
            office_scope=office_scope,
            defaults={
                "title": title,
                "slug": f"{office_scope}-nhwa-web-workbook",
                "source_title": "PNG NHWA Data Collection Toolkit v2",
                "source_version": "May 2026",
                "reporting_year": 2025,
                "status": "active",
            },
        )
        if workbook_created:
            web_workbook.created_by = actor
            web_workbook.save(update_fields=["created_by"])
            created["workbooks"] += 1

        for sort_order, worksheet in enumerate(workbook.worksheets, start=1):
            sheet_name = worksheet.title
            web_sheet, sheet_created = NHWAWebSheet.objects.update_or_create(
                workbook=web_workbook,
                source_sheet_name=sheet_name,
                defaults={
                    "title": sheet_name.replace("_", " ").title(),
                    "sort_order": sort_order,
                    "max_row": worksheet.max_row or 0,
                    "max_column": worksheet.max_column or 0,
                    "editable": sheet_name in EDITABLE_SOURCE_SHEETS,
                    "purpose": _sheet_purpose(worksheet),
                    "metadata": {
                        "source_sheet": sheet_name,
                        "reference_only": sheet_name in REFERENCE_SOURCE_SHEETS,
                    },
                },
            )
            if sheet_created:
                created["sheets"] += 1

            max_row = worksheet.max_row or 0
            max_column = worksheet.max_column or 0
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column):
                for cell in row:
                    raw_value = cell.value
                    value = clean_cell_value(raw_value)
                    formula = value if isinstance(raw_value, str) and value.startswith("=") else ""
                    rgb = rgb_from_cell(cell)
                    cell_template, cell_created = NHWACellTemplate.objects.update_or_create(
                        sheet=web_sheet,
                        coordinate=cell.coordinate,
                        defaults={
                            "row_index": cell.row,
                            "column_index": cell.column,
                            "column_letter": get_column_letter(cell.column),
                            "initial_value": "" if formula else value,
                            "formula": formula,
                            "fill_rgb": rgb,
                            "is_editable": _cell_is_editable(sheet_name, cell, raw_value, rgb),
                            "is_formula": bool(formula),
                            "is_heading": _cell_is_heading(cell, value, rgb),
                            "is_required": False,
                            "number_format": clean_cell_value(cell.number_format),
                            "style_json": cell_style_json(cell, raw_value, rgb),
                        },
                    )
                    if cell_created:
                        created["cells"] += 1
                    if cell_template.is_editable:
                        NHWACellEntry.objects.get_or_create(template=cell_template, defaults={"value": value})

        NHWAWorkbookAuditEvent.objects.create(
            workbook=web_workbook,
            actor=actor,
            action="BOOTSTRAPPED",
            details={
                "source": str(source_path),
                "result": created,
                "bootstrapped_at": timezone.now().isoformat(),
            },
        )
    return created


def _decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    try:
        text = str(value).replace(",", "").strip()
        if text.endswith("%"):
            return Decimal(text[:-1].strip()) / Decimal("100")
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _format_decimal(value, decimal_places=None):
    if value is None:
        return ""
    if not isinstance(value, Decimal):
        value = _decimal_or_none(value)
        if value is None:
            return ""
    if decimal_places is not None:
        quantizer = Decimal("1") if decimal_places == 0 else Decimal("1").scaleb(-decimal_places)
        value = value.quantize(quantizer, rounding=ROUND_HALF_UP)
    if value == value.to_integral_value():
        return str(int(value))
    return format(value.normalize(), "f").rstrip("0").rstrip(".")


def _format_cell_display(value, number_format=""):
    number = _decimal_or_none(value)
    if number is None:
        return clean_cell_value(value)

    number_format = number_format or ""
    if "%" in number_format:
        decimal_places = 0
        percent_part = number_format.split("%", 1)[0]
        if "." in percent_part:
            decimal_places = len(percent_part.split(".", 1)[1])
        return f"{_format_decimal(number * Decimal('100'), decimal_places)}%"

    decimal_places = None
    if "." in number_format and not any(token in number_format.lower() for token in ("general", "@")):
        decimal_places = len(number_format.split(".", 1)[1].split(";", 1)[0])
    return _format_decimal(number, decimal_places)


def _coordinate_values(sheet):
    visible_rows = _visible_row_indexes(sheet)
    values = {}
    for cell in sheet.cell_templates.select_related("entry"):
        if visible_rows is not None and cell.row_index not in visible_rows:
            values[cell.coordinate] = ""
            continue
        entry = _entry_for(cell)
        if entry is not None:
            values[cell.coordinate] = entry.value
        elif cell.is_formula:
            values[cell.coordinate] = ""
        else:
            values[cell.coordinate] = cell.initial_value
    return values


def _visible_row_ranges(sheet):
    return WORKBOOK_SCOPE_VISIBLE_RANGES.get(sheet.workbook.office_scope, {}).get(sheet.source_sheet_name)


def _visible_row_indexes(sheet):
    ranges = _visible_row_ranges(sheet)
    if not ranges:
        return None
    rows = set()
    for start, end in ranges:
        rows.update(range(start, end + 1))
    return rows


def is_sheet_row_visible(sheet, row_index):
    visible_rows = _visible_row_indexes(sheet)
    return visible_rows is None or row_index in visible_rows


def sheet_scope_note(sheet):
    return WORKBOOK_SCOPE_NOTES.get(sheet.workbook.office_scope, "")


def _entry_for(cell):
    try:
        return cell.entry
    except ObjectDoesNotExist:
        return None


def _column_range(start, end):
    start_match = re.fullmatch(r"([A-Z]+)(\d+)", start)
    end_match = re.fullmatch(r"([A-Z]+)(\d+)", end)
    if not start_match or not end_match or column_index_from_string is None or get_column_letter is None:
        return []

    start_col = column_index_from_string(start_match.group(1))
    end_col = column_index_from_string(end_match.group(1))
    start_row = int(start_match.group(2))
    end_row = int(end_match.group(2))
    if start_col > end_col:
        start_col, end_col = end_col, start_col
    if start_row > end_row:
        start_row, end_row = end_row, start_row

    return [
        f"{get_column_letter(column_index)}{row_index}"
        for row_index in range(start_row, end_row + 1)
        for column_index in range(start_col, end_col + 1)
    ]


def _split_top_level(value):
    parts = []
    current = []
    depth = 0
    in_string = False
    for character in value:
        if character == '"':
            in_string = not in_string
        elif not in_string:
            if character == "(":
                depth += 1
            elif character == ")":
                depth -= 1
            elif character == "," and depth == 0:
                parts.append("".join(current).strip())
                current = []
                continue
        current.append(character)
    parts.append("".join(current).strip())
    return parts


def _function_args(expression, function_name):
    prefix = f"{function_name}("
    expression = expression.strip()
    if not expression.upper().startswith(prefix):
        return None
    if not expression.endswith(")"):
        return None
    return _split_top_level(expression[len(prefix):-1])


def _cell_is_nonblank(reference, values):
    return clean_cell_value(values.get(reference, "")) != ""


def _cell_is_positive(reference, values):
    number = _decimal_or_none(values.get(reference, ""))
    return number is not None and number > 0


def _cell_is_number(reference, values):
    return _decimal_or_none(values.get(reference, "")) is not None


def _counta_references(argument, values):
    references = _references_from_argument(argument)
    return sum(1 for reference in references if _cell_is_nonblank(reference, values))


def _references_from_argument(argument):
    argument = argument.strip()
    range_match = re.fullmatch(r"([A-Z]{1,3}\d+):([A-Z]{1,3}\d+)", argument)
    if range_match:
        return _column_range(range_match.group(1), range_match.group(2))
    return re.findall(r"\b[A-Z]{1,3}\d+\b", argument)


def _references_from_args(args):
    references = []
    for argument in args:
        for reference in _references_from_argument(argument):
            references.append(reference)
    return references


def _sum_numeric_references(references, values):
    numbers = [_decimal_or_none(values.get(reference, "")) for reference in references]
    numbers = [number for number in numbers if number is not None]
    return sum(numbers, Decimal("0")) if numbers else None


def _blank_if_none(value):
    return "" if value is None else value


def _normalize_formula_expression(expression):
    expression = expression.strip()
    if expression.startswith("="):
        expression = expression[1:]
    return expression.strip()


def _formula_is_numeric_if_series(expression):
    terms = _split_top_level_by_operator(expression, "+")
    if not terms:
        return False
    return all(re.fullmatch(r"IF\(ISNUMBER\([A-Z]{1,3}\d+\),[A-Z]{1,3}\d+,0\)", term, re.IGNORECASE) for term in terms)


def _split_top_level_by_operator(value, operator):
    parts = []
    current = []
    depth = 0
    in_string = False
    for character in value:
        if character == '"':
            in_string = not in_string
        elif not in_string:
            if character == "(":
                depth += 1
            elif character == ")":
                depth -= 1
            elif character == operator and depth == 0:
                parts.append("".join(current).strip())
                current = []
                continue
        current.append(character)
    parts.append("".join(current).strip())
    return [part for part in parts if part]


def _evaluate_condition(condition, values):
    condition = condition.strip()

    nonblank_match = re.fullmatch(r"([A-Z]{1,3}\d+)<>\"\"", condition)
    if nonblank_match:
        return _cell_is_nonblank(nonblank_match.group(1), values)

    positive_match = re.fullmatch(r"([A-Z]{1,3}\d+)>0", condition)
    if positive_match:
        return _cell_is_positive(positive_match.group(1), values)

    sum_positive_match = re.fullmatch(r"([A-Z]{1,3}\d+)\+([A-Z]{1,3}\d+)>0", condition)
    if sum_positive_match:
        left = _decimal_or_none(values.get(sum_positive_match.group(1), ""))
        right = _decimal_or_none(values.get(sum_positive_match.group(2), ""))
        if left is None or right is None:
            return False
        return (left + right) > 0

    return False


def _evaluate_logical_condition(condition, values):
    condition = condition.strip()
    upper = condition.upper()
    if upper.startswith("AND("):
        conditions = _function_args(condition, "AND") or []
        return all(_evaluate_logical_condition(item, values) for item in conditions)
    if upper.startswith("OR("):
        conditions = _function_args(condition, "OR") or []
        return any(_evaluate_logical_condition(item, values) for item in conditions)
    return _evaluate_condition(condition, values)


def _safe_arithmetic(expression, values):
    expression = expression.strip()
    references = set(re.findall(r"\b[A-Z]{1,3}\d+\b", expression))
    environment = {}
    for reference in references:
        number = _decimal_or_none(values.get(reference, ""))
        if number is None:
            return None
        environment[reference] = number

    allowed_nodes = (
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.UAdd,
        ast.USub,
        ast.Load,
        ast.Name,
        ast.Constant,
    )
    try:
        tree = ast.parse(expression, mode="eval")
    except SyntaxError:
        return None
    if any(not isinstance(node, allowed_nodes) for node in ast.walk(tree)):
        return None
    if any(isinstance(node, ast.Name) and node.id not in environment for node in ast.walk(tree)):
        return None
    try:
        result = eval(compile(tree, "<nhwa_formula>", "eval"), {"__builtins__": {}}, environment)
    except Exception:
        return None
    return result if isinstance(result, Decimal) else _decimal_or_none(result)


def _numeric_if_sum(formula, values):
    refs = re.findall(r"IF\(ISNUMBER\(([A-Z]{1,3}\d+)\),\1,0\)", formula)
    if not refs:
        return None
    return sum((_decimal_or_none(values.get(reference, "")) or Decimal("0")) for reference in refs)


def _round_average_from_expression(expression, values):
    total = _numeric_if_sum(expression, values) or Decimal("0")
    average_refs = re.findall(
        r"IF\(ISNUMBER\(([A-Z]{1,3}\d+)\),\1,0\)",
        expression,
        flags=re.IGNORECASE,
    )
    numeric_count = sum(
        1
        for reference in average_refs
        if _cell_is_number(reference, values)
    )
    return total / Decimal(max(1, numeric_count))


def _evaluate_true_expression(true_value, values):
    true_value = true_value.strip()
    upper = true_value.upper()

    if upper.startswith("IFERROR("):
        inner_args = _function_args(true_value, "IFERROR") or []
        if not inner_args:
            return ""
        return _evaluate_true_expression(inner_args[0], values)

    sum_match = re.fullmatch(r"SUM\((.+)\)", true_value, flags=re.IGNORECASE)
    if sum_match:
        references = _references_from_args([sum_match.group(1)])
        return _blank_if_none(_sum_numeric_references(references, values))

    round_match = re.fullmatch(r"ROUND\((.+),(\d+)\)", true_value, flags=re.IGNORECASE)
    if round_match:
        average = _round_average_from_expression(round_match.group(1), values)
        decimals = int(round_match.group(2))
        quantizer = Decimal("1") if decimals == 0 else Decimal("1").scaleb(-decimals)
        return average.quantize(quantizer, rounding=ROUND_HALF_UP)

    if _formula_is_numeric_if_series(true_value):
        return _blank_if_none(_numeric_if_sum(true_value, values))

    return _blank_if_none(_safe_arithmetic(true_value, values))


def evaluate_formula(formula, values):
    if not formula:
        return ""

    expression = _normalize_formula_expression(formula)

    if _formula_is_numeric_if_series(expression):
        return _numeric_if_sum(expression, values) or Decimal("0")

    if expression.upper().startswith("IF("):
        args = _function_args(expression, "IF")
        if args and len(args) == 3:
            condition, true_value, false_value = args
            if false_value != '""':
                return ""

            counta_match = re.fullmatch(r"COUNTA\((.+)\)>0", condition, flags=re.IGNORECASE)
            if counta_match:
                counta_args = _split_top_level(counta_match.group(1))
                if sum(_counta_references(argument, values) for argument in counta_args) <= 0:
                    return ""
                return _evaluate_true_expression(true_value, values)

            if not _evaluate_logical_condition(condition, values):
                return ""
            return _evaluate_true_expression(true_value, values)

    return _blank_if_none(_safe_arithmetic(expression, values))


def ensure_sheet_entry_state(sheet):
    """Keep existing bootstrapped sheets aligned with current entry-sheet policy."""
    should_be_editable = sheet.source_sheet_name in EDITABLE_SOURCE_SHEETS
    if sheet.editable != should_be_editable:
        sheet.editable = should_be_editable
        metadata = dict(sheet.metadata or {})
        metadata["reference_only"] = not should_be_editable
        sheet.metadata = metadata
        sheet.save(update_fields=["editable", "metadata"])
    return sheet


def build_sheet_header_fields(sheet):
    if sheet.source_sheet_name not in EDITABLE_SOURCE_SHEETS:
        return []

    coordinates = {
        field["label_coordinate"] for field in WORKBOOK_HEADER_FIELDS
    } | {
        field["value_coordinate"] for field in WORKBOOK_HEADER_FIELDS
    }
    cells = {
        cell.coordinate: cell
        for cell in sheet.cell_templates.filter(coordinate__in=coordinates).select_related("entry")
    }

    fields = []
    for field in WORKBOOK_HEADER_FIELDS:
        label_cell = cells.get(field["label_coordinate"])
        value_cell = cells.get(field["value_coordinate"])
        if value_cell is None or value_cell.is_formula:
            continue

        changed_fields = []
        if not value_cell.is_editable:
            value_cell.is_editable = True
            changed_fields.append("is_editable")
        if not value_cell.fill_rgb:
            value_cell.fill_rgb = "FFFDE7"
            changed_fields.append("fill_rgb")
        if changed_fields:
            value_cell.save(update_fields=changed_fields)

        entry, _created = NHWACellEntry.objects.get_or_create(
            template=value_cell,
            defaults={"value": clean_cell_value(value_cell.initial_value)},
        )
        label = clean_cell_value(label_cell.initial_value).rstrip(":") if label_cell else field["label"]
        fields.append(
            {
                "key": field["key"],
                "label": label or field["label"],
                "coordinate": value_cell.coordinate,
                "template_id": value_cell.id,
                "input_name": f"cell_{value_cell.id}",
                "value": clean_cell_value(entry.value if entry is not None else value_cell.initial_value),
                "is_editable": value_cell.is_editable,
            }
        )
    return fields


def build_sheet_grid(sheet):
    values = _coordinate_values(sheet)
    visible_rows = _visible_row_indexes(sheet)
    formula_results = {}
    cells_by_row = {}
    for cell in sheet.cell_templates.all():
        if cell.is_formula:
            formula_results[cell.coordinate] = evaluate_formula(cell.formula, {**values, **formula_results})
        cells_by_row.setdefault(cell.row_index, []).append(cell)

    rows = []
    for row_index in range(1, sheet.max_row + 1):
        if visible_rows is not None and row_index not in visible_rows:
            continue
        row_cells = []
        row_has_visible_content = False
        for cell in sorted(cells_by_row.get(row_index, []), key=lambda item: item.column_index):
            entry = _entry_for(cell)
            value = formula_results.get(cell.coordinate)
            if value is None:
                value = entry.value if entry is not None else cell.initial_value
            display_value = _format_cell_display(value, cell.number_format) if cell.is_formula else clean_cell_value(value)
            display_value = _scoped_display_value(sheet, cell, display_value)
            row_has_visible_content = row_has_visible_content or bool(display_value or cell.initial_value or cell.formula or cell.fill_rgb)
            is_dark_fill = is_dark_hex(cell.fill_rgb)
            is_dark_empty = is_dark_fill and not bool(display_value or cell.initial_value or cell.formula or cell.is_editable)
            row_cells.append(
                {
                    "template": cell,
                    "value": display_value,
                    "is_formula": cell.is_formula,
                    "is_editable": cell.is_editable,
                    "is_dark_fill": is_dark_fill,
                    "is_dark_empty": is_dark_empty,
                    "is_blank_locked": not bool(display_value or cell.formula or cell.is_editable),
                    "css": _template_css(cell),
                    "input_name": f"cell_{cell.id}",
                }
            )
        if row_has_visible_content:
            rows.append({"index": row_index, "cells": row_cells})
    return rows


def _scoped_display_value(sheet, cell, value):
    scope = sheet.workbook.office_scope
    sheet_name = sheet.source_sheet_name
    if sheet_name == "T3_COUNCIL_REGISTER" and cell.coordinate == "A2":
        office = "Nursing Council" if scope == "nursing" else "Medical Board"
        return f"Module 1 (Practising vs. Non-Practising Distinction) | For completion by {office}"
    if sheet_name == "T2_TRAINING_SCHOOL" and cell.coordinate == "A1":
        if scope == "nursing":
            return "TOOL 2 - NURSING AND MIDWIFERY TRAINING SCHOOL TEMPLATE"
        if scope == "medical":
            return "TOOL 2 - MEDICAL BOARD TRAINING SCHOOL / MEDICAL EDUCATION UNIT TEMPLATE"
    if sheet_name == "T2_TRAINING_SCHOOL" and cell.coordinate == "A2":
        if scope == "nursing":
            return "Module 3 (Education and Training) | For completion by Nursing Council training institution administrators"
        if scope == "medical":
            return "Module 3 (Education and Training) | For completion by Medical Board training institution administrators and the Medical Education Unit"
    return value


def _template_css(cell):
    style = []
    rgb = cell.fill_rgb
    if rgb:
        if cell.is_editable:
            style.append("background-color: #fffdf2;")
            style.append("color: #102b45;")
        elif cell.is_formula:
            style.append("background-color: #eef2f7;")
            style.append("color: #102b45;")
        elif is_dark_hex(rgb):
            has_content = bool(clean_cell_value(cell.initial_value) or cell.formula or cell.is_heading)
            if has_content:
                style.append("background-color: #12304a;")
                style.append("color: #ffffff;")
            else:
                style.append("background-color: #f8fafc;")
                style.append("color: #64748b;")
        else:
            style.append(f"background-color: #{rgb};")
            style.append(f"color: {cell.style_json.get('text_color', '#0f172a')};")
    if cell.style_json.get("bold"):
        style.append("font-weight: 800;")
    return " ".join(style)


def workbook_completion(workbook):
    editable_cell_ids = []
    formulas = 0
    for sheet in workbook.sheets.all():
        visible_rows = _visible_row_indexes(sheet)
        cells = NHWACellTemplate.objects.filter(sheet=sheet)
        if visible_rows is not None:
            cells = cells.filter(row_index__in=visible_rows)
        editable_cell_ids.extend(cells.filter(is_editable=True).values_list("id", flat=True))
        formulas += cells.filter(is_formula=True).count()

    editable = len(editable_cell_ids)
    filled = NHWACellEntry.objects.filter(template_id__in=editable_cell_ids, value__gt="").count()
    return {
        "editable": editable,
        "filled": filled,
        "formulas": formulas,
        "percent": round((filled / editable) * 100, 1) if editable else 0,
    }


def source_document_statuses():
    documents = [
        ("workbook", "PNG NHWA Data Collection Toolkit v2", SOURCE_WORKBOOK),
        ("presentation", "PNG NHWA Output 2 guidance", SOURCE_PRESENTATION),
    ]
    return [
        {
            "key": key,
            "title": title,
            "path": str(path),
            "exists": path.exists(),
            "size_kb": round(path.stat().st_size / 1024, 1) if path.exists() else None,
            "policy": "Controlled source artefact; rendered through platform workflow, not imported as registry data.",
        }
        for key, title, path in documents
    ]


def _editable_cells_for_sheet(sheet):
    cells = NHWACellTemplate.objects.filter(sheet=sheet, is_editable=True)
    visible_rows = _visible_row_indexes(sheet)
    if visible_rows is not None:
        cells = cells.filter(row_index__in=visible_rows)
    return cells


def sheet_completion(sheet):
    editable_cell_ids = list(_editable_cells_for_sheet(sheet).values_list("id", flat=True))
    editable = len(editable_cell_ids)
    filled_ids = set(
        NHWACellEntry.objects
        .filter(template_id__in=editable_cell_ids)
        .exclude(value="")
        .values_list("template_id", flat=True)
    )
    missing_cells = (
        NHWACellTemplate.objects
        .filter(id__in=[cell_id for cell_id in editable_cell_ids if cell_id not in filled_ids])
        .order_by("row_index", "column_index")[:12]
    )
    return {
        "editable": editable,
        "filled": len(filled_ids),
        "missing": max(editable - len(filled_ids), 0),
        "percent": round((len(filled_ids) / editable) * 100, 1) if editable else 0,
        "missing_examples": [
            f"{cell.sheet.source_sheet_name}!{cell.coordinate}"
            for cell in missing_cells
        ],
    }


def workbook_readiness(workbook):
    sheet_rows = []
    checklist = None
    for sheet in workbook.sheets.all():
        completion = sheet_completion(sheet)
        row = {
            "sheet": sheet,
            "completion": completion,
            "is_checklist": sheet.source_sheet_name == "DATA_QUALITY_CHECKLIST",
        }
        if row["is_checklist"]:
            checklist = row
        sheet_rows.append(row)

    workbook_totals = workbook_completion(workbook)
    checklist_complete = bool(
        checklist
        and checklist["completion"]["editable"] > 0
        and checklist["completion"]["missing"] == 0
    )
    return {
        "completion": workbook_totals,
        "sheet_rows": sheet_rows,
        "checklist": checklist,
        "checklist_complete": checklist_complete,
        "ready_for_signoff": checklist_complete and workbook.status == "active",
        "export_ready": workbook.status == "locked",
    }


def lock_workbooks_for_signoff(actor=None, scopes=None):
    scopes = scopes or ("nursing", "medical")
    locked = []
    blocked = []
    for workbook in NHWAWebWorkbook.objects.filter(office_scope__in=scopes).prefetch_related("sheets"):
        readiness = workbook_readiness(workbook)
        if not readiness["checklist_complete"]:
            blocked.append({
                "workbook": workbook,
                "reason": "Data quality checklist is incomplete.",
            })
            continue
        if workbook.status != "locked":
            workbook.status = "locked"
            workbook.save(update_fields=["status", "updated_at"])
            NHWAWorkbookAuditEvent.objects.create(
                workbook=workbook,
                actor=actor,
                action="LOCKED",
                details={
                    "signoff": True,
                    "checklist_complete": True,
                    "locked_at": timezone.now().isoformat(),
                    "direction": "Registry -> NHWA workbook -> sign-off -> export",
                },
            )
        locked.append(workbook)
    return {"locked": locked, "blocked": blocked}


def unlock_workbooks(actor=None, scopes=None):
    scopes = scopes or ("nursing", "medical")
    unlocked = []
    for workbook in NHWAWebWorkbook.objects.filter(office_scope__in=scopes, status="locked"):
        workbook.status = "active"
        workbook.save(update_fields=["status", "updated_at"])
        NHWAWorkbookAuditEvent.objects.create(
            workbook=workbook,
            actor=actor,
            action="UNLOCKED",
            details={
                "reason": "System Admin reopened workbook for correction.",
                "unlocked_at": timezone.now().isoformat(),
            },
        )
        unlocked.append(workbook)
    return unlocked


def _safe_sheet_title(value, used_titles):
    base = re.sub(r"[:\\/?*\[\]]", "_", value or "Sheet")[:31] or "Sheet"
    title = base
    counter = 2
    while title in used_titles:
        suffix = f" {counter}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_titles.add(title)
    return title


def _cell_export_value(sheet, cell, values, formula_results):
    if cell.is_formula:
        result = evaluate_formula(cell.formula, {**values, **formula_results})
        formula_results[cell.coordinate] = result
        return _format_cell_display(result, cell.number_format)
    entry = _entry_for(cell)
    return clean_cell_value(entry.value if entry is not None else cell.initial_value)


def build_web_workbook_xlsx(workbook):
    if Workbook is None:
        raise RuntimeError("openpyxl is required to export NHWA workbook packages.")

    export_workbook = Workbook()
    default_sheet = export_workbook.active
    export_workbook.remove(default_sheet)
    used_titles = set()

    for sheet in workbook.sheets.all().prefetch_related("cell_templates__entry"):
        export_sheet = export_workbook.create_sheet(_safe_sheet_title(sheet.source_sheet_name, used_titles))
        values = _coordinate_values(sheet)
        formula_results = {}
        visible_rows = _visible_row_indexes(sheet)
        for cell in sheet.cell_templates.all():
            if visible_rows is not None and cell.row_index not in visible_rows:
                continue
            export_sheet.cell(
                row=cell.row_index,
                column=cell.column_index,
                value=_cell_export_value(sheet, cell, values, formula_results),
            )

    buffer = BytesIO()
    export_workbook.save(buffer)
    return buffer.getvalue()


def build_submission_manifest(workbooks):
    return {
        "generated_at": timezone.now().isoformat(),
        "direction": "Registry / Analytics / Finance / Facilities -> NHWA workbook cells -> review and sign-off -> NHWA export/report",
        "registry_writeback": "disabled",
        "source_documents": source_document_statuses(),
        "workbooks": [
            {
                "id": workbook.id,
                "title": workbook.title,
                "office_scope": workbook.office_scope,
                "reporting_year": workbook.reporting_year,
                "status": workbook.status,
                "completion": workbook_completion(workbook),
                "readiness": {
                    "checklist_complete": workbook_readiness(workbook)["checklist_complete"],
                    "export_ready": workbook.status == "locked",
                },
            }
            for workbook in workbooks
        ],
    }


def build_submission_pack(workbooks, actor=None):
    workbooks = list(workbooks)
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("NHWA_Submission_Manifest.json", json.dumps(build_submission_manifest(workbooks), indent=2))
        for workbook in workbooks:
            filename = f"{workbook.office_scope}_NHWA_{workbook.reporting_year}.xlsx"
            archive.writestr(filename, build_web_workbook_xlsx(workbook))
            NHWAWorkbookAuditEvent.objects.create(
                workbook=workbook,
                actor=actor,
                action="VIEWED",
                details={
                    "exported": True,
                    "filename": filename,
                    "exported_at": timezone.now().isoformat(),
                },
            )
    return buffer.getvalue()


def save_sheet_entries(sheet, post_data, actor, request=None):
    if not sheet.editable:
        return 0
    changed = 0
    visible_rows = _visible_row_indexes(sheet)
    for cell in sheet.cell_templates.filter(is_editable=True):
        if visible_rows is not None and cell.row_index not in visible_rows:
            continue
        key = f"cell_{cell.id}"
        if key not in post_data:
            continue
        new_value = clean_cell_value(post_data.get(key))
        entry, _created = NHWACellEntry.objects.get_or_create(template=cell)
        if entry.value != new_value:
            entry.value = new_value
            entry.updated_by = actor
            entry.save(update_fields=["value", "updated_by", "updated_at"])
            changed += 1
    if changed:
        NHWAWorkbookAuditEvent.objects.create(
            workbook=sheet.workbook,
            sheet=sheet,
            actor=actor,
            action="SHEET_SAVED",
            details={"changed_cells": changed},
            ip_address=request.META.get("REMOTE_ADDR") if request else None,
            user_agent=request.META.get("HTTP_USER_AGENT", "") if request else "",
        )
    return changed
