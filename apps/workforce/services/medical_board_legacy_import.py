from collections import Counter
from datetime import date
from decimal import Decimal
import os
from pathlib import Path
import re
from types import SimpleNamespace

from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.workforce.models import (
    DataImportBatch,
    ImportedWorkbookSheet,
    MedicalDoctor,
    PracticingLicenseRecord,
    Qualification,
)
from apps.workforce.services.medical_board_workbook_import import (
    clean_text,
    extract_amount,
    normalize_registration,
    parse_date,
    split_name,
    title_name,
)


DEFAULT_MEDICAL_BOARD_LEGACY_WORKBOOKS = [
    Path(
        r"c:\Users\darre\OneDrive\Documents\ProjectApps\databasedocuments"
        r"\medicalboarddata\Current database on Medical Board"
        r"\All ANAESTHESTICS 2026\All ANAESTHESTICS 2026.xlsx"
    ),
    Path(
        r"c:\Users\darre\OneDrive\Documents\ProjectApps\databasedocuments"
        r"\medicalboarddata\Current database on Medical Board"
        r"\All EHO'S  2026\All EHO'S  2026.xlsx"
    ),
    Path(
        r"c:\Users\darre\OneDrive\Documents\ProjectApps\databasedocuments"
        r"\medicalboarddata\Current database on Medical Board"
        r"\All MLT-MLA 2026\All MLT-MLA 2026.xlsx"
    ),
    Path(
        r"c:\Users\darre\OneDrive\Documents\ProjectApps\databasedocuments"
        r"\medicalboarddata\EXCEL DATA"
        r"\LIST OF EHO'S MIX  2026 UPDATES\LIST OF EHO'S MIX  2026 UPDATES.xlsx"
    ),
    Path(
        r"c:\Users\darre\OneDrive\Documents\ProjectApps\databasedocuments"
        r"\medicalboarddata\EXCEL DATA"
        r"\OVERSEAS DATABASE REGISTRATION 2026\OVERSEAS DATABASE REGISTRATION 2026.xlsx"
    ),
]

HEADER_ALIASES = {
    "app for": "application_for",
    "arc#": "registration_card",
    "atp / frc": "license_status",
    "atp / full r c": "license_status",
    "atp date": "atp_date",
    "authority to practice / full r c": "license_status",
    "amount": "amount",
    "card #": "registration_card",
    "cert #": "registration_card",
    "cert#": "registration_card",
    "date of the receipt": "payment_date",
    "doctors name": "full_name",
    "employer": "employer_name",
    "full registration": "full_registration_no",
    "full rego": "full_registration_no",
    "full reg": "full_registration_no",
    "last atp": "last_atp_year",
    "last atp year": "last_atp_year",
    "mb rego card#": "registration_card",
    "mb rego no.#": "registration_card",
    "medical practioners#": "medical_practitioner_no",
    "mp#": "medical_practitioner_no",
    "name": "name",
    "names": "full_name",
    "nationality": "nationality",
    "officer name": "full_name",
    "app/status": "license_status",
    "app status": "license_status",
    "certificate no": "registration_card",
    "full": "full_registration_no",
    "arch#": "practitioner_no",
    "arch #": "practitioner_no",
    "practitioner's no": "practitioner_no",
    "pn.#": "practitioner_no",
    "pn.#": "practitioner_no",
    "pn#": "practitioner_no",
    "practisioners #": "practitioner_no",
    "practioners #/nos": "practitioner_no",
    "practitioner no.": "practitioner_no",
    "practitioner no": "practitioner_no",
    "prov": "provisional_no",
    "prov mlt#/nos": "provisional_no",
    "province": "province",
    "place of employment": "employer_name",
    "professional": "professional_category",
    "professional-cadres": "professional_category",
    "professional-cahres": "professional_category",
    "qualification": "qualification",
    "qualification/designation": "qualification",
    "reciept date": "payment_date",
    "reciept no": "receipt_number",
    "reciept no#": "receipt_number",
    "reciept#": "receipt_number",
    "receipt no": "receipt_number",
    "receipt no#": "receipt_number",
    "receipt#": "receipt_number",
    "rego date": "registration_date",
    "registration date": "registration_date",
    "remarks": "remarks",
    "spec": "specialist_no",
    "spec#": "specialist_no",
    "speciality": "specialty",
    "specialty": "specialty",
    "surname": "surname",
    "year": "record_year",
    "reciept #": "receipt_number",
}

SHEET_CATEGORY_HINTS = {
    "EHO": "Environmental Health Officer",
    "HEO": "Health Extension Officer",
    "NAT-ANAESTHESTICS": "Anaesthetics/ATO",
    "NAT-MLT-MLA": "Medical Laboratory Technician / Assistant",
    "NAT-PHYSIOTHERAPIST": "Physiotherapist",
    "ALL OVERSESA MEMBERS MB": "Overseas Medical Board Practitioner",
}

DOCTOR_KEYWORDS = (
    "BDS",
    "DENTAL",
    "DOCTOR",
    "MBBS",
    "MEDICAL PRACTITIONER",
    "MEDICAL PRACTIONER",
    "MD ",
    "PAEDIATRIC",
    "PEADIATRIC",
    "SPECIALIST",
    "SURGEON",
)

GENERIC_REGISTRATION_CODES = {
    "ATO01",
    "MB.01",
    "MLT/MLA01",
    "PHYSIO-01",
}


def normalize_header(value):
    text = clean_text(value).lower()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def safe_sheet_label(workbook_path, sheet_name):
    label = f"{Path(workbook_path).stem}::{sheet_name}"
    return label[:255]


def normalise_source_key(value):
    text = clean_text(value).upper()
    return re.sub(r"[^A-Z0-9]+", "-", text).strip("-")


def compact_source_key(workbook_path, source_sheet, row_number):
    workbook_key = normalise_source_key(Path(workbook_path).stem)[:14]
    sheet_key = normalise_source_key(source_sheet)[:14]
    return f"{row_number}-{workbook_key}-{sheet_key}"


def is_generic_registration_code(value):
    text = clean_text(value).upper()
    if not text:
        return True
    if text in GENERIC_REGISTRATION_CODES:
        return True
    return bool(re.fullmatch(r"[A-Z]{2,6}\.?\d{1,2}", text))


def latest_year(value):
    parsed = parse_date(value)
    if parsed:
        return parsed.year
    years = [
        int(match.group(0))
        for match in re.finditer(r"(?:19|20)\d{2}", clean_text(value))
        if 1901 <= int(match.group(0)) <= date.today().year + 1
    ]
    if years:
        return max(years)
    text = clean_text(value)
    if re.fullmatch(r"\d{2}", text):
        year = int(text)
        return 2000 + year if year <= date.today().year % 100 + 1 else 1900 + year
    return None


def infer_category(sheet_name, qualification, professional="", specialty=""):
    professional = clean_text(professional)
    specialty = clean_text(specialty)
    if specialty:
        return specialty
    if professional:
        return professional
    sheet_upper = clean_text(sheet_name).upper()
    for hint, label in SHEET_CATEGORY_HINTS.items():
        if hint in sheet_upper:
            return label
    return clean_text(qualification) or "Medical Board Practitioner"


def infer_target_model(category, qualification, sheet_name):
    text = f"{category} {qualification} {sheet_name}".upper()
    if "CHW" in text or "COMMUNITY HEALTH" in text:
        return "communityhealthworker"
    if any(keyword in text for keyword in DOCTOR_KEYWORDS):
        return "medicaldoctor"
    return "other"


def infer_specialty_value(category, qualification):
    text = f"{category} {qualification}".upper()
    if "PUBLIC HEALTH" in text:
        return "public_health"
    if "PAEDIATR" in text or "PEADIATR" in text:
        return "paediatric_child_health"
    if "RADIOLOG" in text:
        return "radiology"
    if "PATHOLOG" in text:
        return "pathology"
    if "MICROBIOLOG" in text:
        return "microbiology"
    if "DERMATOLOG" in text:
        return "dermatology"
    if "PSYCHIAT" in text:
        return "psychiatry_mental_health"
    if "CARDIO" in text:
        return "cardiology"
    if "SPECIALIST" in text:
        return "other"
    return ""


def record_type_from_status(status, provisional_no="", full_registration_no=""):
    text = clean_text(status).upper()
    if "FULL" in text or "FRC" in text or "R C" in text or clean_text(full_registration_no):
        return "full"
    if "PROV" in text or clean_text(provisional_no):
        return "provisional"
    return "workforce_listing"


def applicant_type_from_nationality(nationality):
    text = clean_text(nationality).upper()
    return "national" if text in {"PNG", "PAPUA NEW GUINEA", "P.N.G."} else "overseas"


def year_columns(headers):
    columns = []
    for index, header in enumerate(headers):
        text = clean_text(header)
        if re.fullmatch(r"(?:19|20)\d{2}", text):
            year = int(text)
            if 1901 <= year <= date.today().year + 1:
                columns.append((index, year))
    return columns


def normalize_sheet_name(value):
    return re.sub(r"\s+", " ", clean_text(value).upper()).strip()


def expand_combined_registration_fields(record):
    text = clean_text(record.get("registration_card"))
    if not text:
        return record
    patterns = {
        "provisional_no": r"\bPROV\.?\s*([A-Z0-9/-]+)",
        "full_registration_no": r"\bFULL\.?\s*([A-Z0-9/-]+)",
        "specialist_no": r"\bSPEC\.?\s*([A-Z0-9/-]+)",
    }
    for field_name, pattern in patterns.items():
        if record.get(field_name):
            continue
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            record[field_name] = match.group(1)
    return record


class MedicalBoardLegacyWorkbookImporter:
    def __init__(self, workbook_paths=None, initiated_by=None, include_sheets=None, dry_run=False, preview=False):
        self.workbook_paths = [Path(path) for path in (workbook_paths or DEFAULT_MEDICAL_BOARD_LEGACY_WORKBOOKS)]
        self.initiated_by = initiated_by
        self.include_sheets = {normalize_sheet_name(value) for value in (include_sheets or []) if clean_text(value)}
        self.dry_run = bool(dry_run or preview)
        self.summary = Counter()
        self._row_dedupes = set()
        self._licence_dedupes = set()
        self._atp_dedupes = set()
        self._sheet_reports = []

    def import_workbooks(self):
        existing_paths = [path for path in self.workbook_paths if path.exists()]
        missing_paths = [str(path) for path in self.workbook_paths if not path.exists()]
        if not existing_paths:
            raise FileNotFoundError("No Medical Board legacy workbooks were found.")

        workbook_metadata = []
        for path in existing_paths:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                workbook_metadata.append({
                    "path": path,
                    "sheet_count": len(workbook.sheetnames),
                    "row_count": sum(workbook[name].max_row or 0 for name in workbook.sheetnames),
                })
            finally:
                workbook.close()

        if self.dry_run:
            for item in workbook_metadata:
                self._import_single_workbook(None, item["path"], preview=True)
            return SimpleNamespace(
                id="preview",
                status="preview",
                summary={
                    "sheets_imported": self.summary.get("sheets_imported", 0),
                    "sheets_skipped": self.summary.get("sheets_skipped", 0),
                    "sheet_reports": self._sheet_reports,
                    "missing_files": missing_paths,
                    **dict(self.summary),
                },
            )

        batch = DataImportBatch.objects.create(
            source_file_name="Medical Board legacy workbooks",
            source_file_path=os.linesep.join(str(item["path"]) for item in workbook_metadata),
            source_kind="medical_board_workbook",
            status="running",
            total_sheets=sum(item["sheet_count"] for item in workbook_metadata),
            total_rows=sum(item["row_count"] for item in workbook_metadata),
            initiated_by=self.initiated_by,
            summary={"missing_files": missing_paths},
        )

        try:
            with transaction.atomic():
                for item in workbook_metadata:
                    self._import_single_workbook(batch, item["path"])
            batch.status = "completed"
            batch.processed_rows = self.summary.get("records_created", 0)
            batch.summary = {**dict(self.summary), "missing_files": missing_paths}
            batch.completed_at = timezone.now()
            batch.save(update_fields=["status", "processed_rows", "summary", "completed_at"])
            return batch
        except Exception as exc:
            batch.status = "failed"
            batch.summary = {"error": str(exc), **dict(self.summary), "missing_files": missing_paths}
            batch.completed_at = timezone.now()
            batch.save(update_fields=["status", "summary", "completed_at"])
            raise

    def _should_process_sheet(self, sheet_name):
        if not self.include_sheets:
            return True
        normalized = normalize_sheet_name(sheet_name)
        return normalized in self.include_sheets or any(
            requested in normalized for requested in self.include_sheets
        )

    def _import_single_workbook(self, batch, workbook_path, preview=False):
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        should_update_batch = batch is not None and not preview and not self.dry_run
        try:
            for worksheet in workbook.worksheets:
                if not self._should_process_sheet(worksheet.title):
                    self.summary["sheets_skipped"] += 1
                    self._sheet_reports.append(
                        {
                            "source_file": str(workbook_path),
                            "source_sheet": worksheet.title,
                            "status": "skipped",
                            "note": "Not included by --sheets filter.",
                        }
                    )
                    continue
                self._import_worksheet(batch, workbook_path, worksheet)
                if should_update_batch:
                    batch.processed_sheets += 1
                    batch.save(update_fields=["processed_sheets"])
        finally:
            workbook.close()

    def _find_header(self, worksheet):
        fallback_header = None
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row or 0, 10), values_only=True),
            start=1,
        ):
            headers = [clean_text(value) for value in row]
            mapped = {}
            for index, header in enumerate(headers):
                canonical = HEADER_ALIASES.get(normalize_header(header))
                if canonical and canonical not in mapped:
                    mapped[canonical] = index
            if "full_name" in mapped or ("name" in mapped and "surname" in mapped):
                return row_number, headers, mapped, "standard"
            if "name" in mapped:
                return row_number, headers, mapped, "standard"
            if year_columns(headers):
                candidate = (
                    "practitioner_no" in mapped
                    or "registration_card" in mapped
                    or "license_status" in mapped
                    or "receipt_number" in mapped
                    or "payment_date" in mapped
                    or "amount" in mapped
                )
                if candidate and fallback_header is None:
                    fallback_header = (row_number, headers, mapped, "legacy_atp")
            if not fallback_header and year_columns(headers):
                fallback_header = (row_number, headers, mapped, "legacy_atp")
        return fallback_header or (None, [], {}, None)

    def _import_worksheet(self, batch, workbook_path, worksheet):
        sheet_name = safe_sheet_label(workbook_path, worksheet.title)
        header_row, headers, header_map, sheet_layout = self._find_header(worksheet)
        raw_rows = max((worksheet.max_row or 0) - (header_row or 0), 0)
        sheet = self._new_sheet(
            batch=batch,
            sheet_name=sheet_name,
            source_file=str(workbook_path),
            source_sheet=worksheet.title,
            sheet_type="legacy_atp" if sheet_layout == "legacy_atp" else "medical_board_legacy",
            header_row=header_row,
            raw_rows=raw_rows,
        )
        if not header_row:
            self.summary["sheets_skipped"] += 1
            self.summary["total_rows_seen"] += 0
            self._sheet_reports.append({
                "source_file": str(workbook_path),
                "source_sheet": worksheet.title,
                "status": "skipped",
                "reason": "No recognized header row found.",
                "layout": "unrecognized",
            })
            self._finalize_sheet(sheet, 0, 0)
            return

        imported = skipped = 0
        if sheet_layout == "legacy_atp":
            imported, skipped = self._import_legacy_atp_sheet(
                batch,
                sheet,
                workbook_path,
                worksheet,
                header_row,
                headers,
                header_map,
            )
        else:
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                record = self._row_to_record(headers, header_map, row)
                if not record.get("full_name"):
                    skipped += 1
                    continue
                created = self._record_legacy_row(
                    batch,
                    sheet,
                    workbook_path,
                    worksheet.title,
                    row_number,
                    record,
                )
                self._record_year_columns(
                    batch,
                    sheet,
                    worksheet.title,
                    row_number,
                    record,
                    headers,
                    row,
                )
                imported += created
                if not created:
                    skipped += 1
        self._finalize_sheet(sheet, imported, skipped)
        self._sheet_reports.append({
            "source_file": str(workbook_path),
            "source_sheet": worksheet.title,
            "status": "processed",
            "layout": sheet_layout,
            "imported_rows": imported,
            "skipped_rows": skipped,
        })

    def _new_sheet(self, batch, sheet_name, source_file, source_sheet, sheet_type, header_row, raw_rows):
        if self.dry_run:
            return {
                "sheet_name": sheet_name,
                "sheet_type": sheet_type,
                "status": "processed" if header_row else "skipped",
                "raw_rows": raw_rows,
                "notes": "" if header_row else "No Medical Board row header found.",
                "metadata": {"source_file": source_file, "source_sheet": source_sheet},
            }
        return ImportedWorkbookSheet.objects.create(
            batch=batch,
            sheet_name=sheet_name,
            sheet_type=sheet_type,
            status="processed" if header_row else "skipped",
            raw_rows=raw_rows,
            notes="" if header_row else "No Medical Board row header found.",
            metadata={"source_file": source_file, "source_sheet": source_sheet},
        )

    def _finalize_sheet(self, sheet, imported_rows, skipped_rows):
        if self.dry_run:
            sheet["imported_rows"] = imported_rows
            sheet["skipped_rows"] = skipped_rows
            if sheet.get("status") != "skipped":
                self.summary["sheets_imported"] += 1
            return
        sheet.imported_rows = imported_rows
        sheet.skipped_rows = skipped_rows
        sheet.status = "processed"
        sheet.save(update_fields=["imported_rows", "skipped_rows", "status"])
        self.summary["sheets_imported"] += 1

    def _row_to_record(self, headers, header_map, row):
        record = {}
        for field_name, index in header_map.items():
            if index < len(row):
                record[field_name] = clean_text(row[index])
        record["raw_headers"] = headers
        if not record.get("full_name"):
            first = clean_text(record.get("name"))
            surname = clean_text(record.get("surname"))
            if first or surname:
                record["full_name"] = title_name(f"{first} {surname}".strip())
        expand_combined_registration_fields(record)
        return record

    def _source_registration(self, record, source_key, target_model):
        candidates = [
            record.get("practitioner_no"),
            record.get("medical_practitioner_no"),
            record.get("full_registration_no"),
            record.get("specialist_no"),
            record.get("provisional_no"),
            record.get("registration_card"),
        ]
        for candidate in candidates:
            text = clean_text(candidate).upper()
            if text and not is_generic_registration_code(text):
                prefix = "MD" if target_model == "medicaldoctor" else ("CHW" if target_model == "communityhealthworker" else "MB")
                return normalize_registration(text, prefix=prefix)
        prefix = "MD" if target_model == "medicaldoctor" else ("CHW" if target_model == "communityhealthworker" else "MB-AH")
        return normalize_registration(source_key, prefix=prefix)

    def _sheet_name(self, sheet):
        if isinstance(sheet, dict):
            return sheet.get("sheet_name", "")
        return sheet.sheet_name

    def _source_sheet_key(self, workbook_path, source_sheet, row_number):
        return compact_source_key(workbook_path, source_sheet, row_number)

    def _existing_record(
        self,
        source_sheet_name,
        source_row,
        record_type,
        *,
        record_year=None,
        reference_number=None,
        target_model=None,
        registration_no="",
        full_name="",
        practitioner_number="",
        payment_date=None,
    ):
        if self.dry_run:
            return False
        query = PracticingLicenseRecord.objects.filter(
            source_sheet_name=source_sheet_name,
            source_row=source_row,
            record_type=record_type,
        )
        if record_year is not None:
            query = query.filter(record_year=record_year)
        if clean_text(reference_number):
            query = query.filter(reference_number=reference_number)
        if target_model:
            query = query.filter(target_model=target_model)
        if query.exists():
            return True

        business_query = PracticingLicenseRecord.objects.filter(record_type=record_type)
        if target_model:
            business_query = business_query.filter(target_model=target_model)
        if clean_text(registration_no):
            business_query = business_query.filter(registration_no=registration_no)
        elif clean_text(practitioner_number):
            business_query = business_query.filter(practitioner_number=practitioner_number)
        elif clean_text(full_name):
            business_query = business_query.filter(full_name__iexact=full_name)
        else:
            return False
        if record_year is not None:
            business_query = business_query.filter(record_year=record_year)
        if clean_text(reference_number):
            business_query = business_query.filter(reference_number=reference_number)
        if payment_date:
            business_query = business_query.filter(payment_date=payment_date)
        return business_query.exists()

    def _has_direct_payment_hint(self, record):
        return bool(
            clean_text(record.get("receipt_number"))
            or clean_text(record.get("payment_date"))
            or clean_text(record.get("amount"))
        )

    def _import_legacy_atp_sheet(self, batch, sheet, workbook_path, worksheet, header_row, headers, header_map):
        source_sheet = worksheet.title
        imported = skipped = 0
        source_sheet_name = self._sheet_name(sheet)
        fallback_name_index = 6
        name_index = header_map.get("full_name") or header_map.get("name") or fallback_name_index
        if name_index is None:
            name_index = fallback_name_index

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            record = self._row_to_record(headers, header_map, row)
            if not record.get("full_name"):
                fallback_name = clean_text(row[name_index] if name_index < len(row) else None)
                if fallback_name:
                    record["full_name"] = title_name(fallback_name)
                    record.setdefault("practitioner_no", clean_text(row[name_index + 1]) if name_index + 1 < len(row) else "")
                else:
                    skipped += 1
                    continue

            category = infer_category(
                source_sheet,
                record.get("qualification"),
                record.get("professional_category"),
                record.get("specialty"),
            )
            target_model = infer_target_model(category, record.get("qualification"), source_sheet)
            full_name = title_name(record.get("full_name"))
            first_name, last_name = split_name(full_name)
            source_key = self._source_sheet_key(workbook_path, source_sheet, row_number)
            registration_no = self._source_registration(record, source_key, target_model)
            practitioner_number = clean_text(
                record.get("practitioner_no")
                or record.get("medical_practitioner_no")
                or record.get("practitioner_number")
            )
            applicant_type = applicant_type_from_nationality(record.get("nationality"))
            record_year = (
                latest_year(record.get("record_year"))
                or latest_year(record.get("last_atp_year"))
                or latest_year(record.get("atp_date"))
                or latest_year(record.get("registration_date"))
            )
            year_hits = 0

            for position, (index, year) in enumerate(year_columns(headers), start=1):
                payment_text = clean_text(row[index] if index < len(row) else None)
                receipt_text = clean_text(row[index + 1] if index + 1 < len(row) else None)
                if not payment_text and not receipt_text and not clean_text(record.get("payment_date")) and not clean_text(record.get("amount")):
                    continue
                dedupe_key = (
                    full_name.upper(),
                    source_sheet_name.upper(),
                    str(row_number),
                    str(year),
                    clean_text(record.get("practitioner_no")),
                )
                if dedupe_key in self._atp_dedupes:
                    self.summary["duplicate_licence_rows_skipped"] += 1
                    continue
                self._atp_dedupes.add(dedupe_key)
                amount = extract_amount(payment_text or record.get("amount"))
                payment_dt = parse_date(payment_text) or parse_date(record.get("payment_date"))
                source_row = (row_number * 100) + position
                if self._existing_record(
                    source_sheet_name=source_sheet_name,
                    source_row=source_row,
                    record_type="practicing_license",
                    record_year=year,
                    target_model=target_model,
                    registration_no=registration_no,
                    full_name=full_name,
                    practitioner_number=practitioner_number,
                    reference_number=receipt_text or payment_text,
                    payment_date=payment_dt,
                ):
                    self.summary["duplicate_licence_rows_skipped"] += 1
                    year_hits += 1
                    continue
                if self.dry_run:
                    self.summary["records_created"] += 1
                    self.summary["practicing_license_rows"] += 1
                    self.summary["record_type_practicing_license"] += 1
                    imported += 1
                    year_hits += 1
                    continue

                PracticingLicenseRecord.objects.create(
                    batch=batch,
                    sheet=sheet,
                    record_type="practicing_license",
                    target_model=target_model,
                    source_sheet_name=source_sheet_name,
                    source_row=source_row,
                    record_year=year,
                    full_name=full_name,
                    first_name=first_name,
                    last_name=last_name,
                    registration_no=registration_no,
                    practitioner_number=practitioner_number,
                    applicant_type=applicant_type,
                    qualification_name=record.get("qualification", ""),
                    category=category,
                    workplace_address=record.get("employer_name", ""),
                    province=record.get("province", ""),
                    payment_date=payment_dt,
                    amount=amount,
                    reference_number=receipt_text or payment_text,
                    raw_payload=self._payload(record, workbook_path, source_sheet, category, target_model, source_key),
                )
                self.summary["records_created"] += 1
                self.summary["practicing_license_rows"] += 1
                self.summary["record_type_practicing_license"] += 1
                self.summary["category_{}".format(category)] += 1
                imported += 1
                year_hits += 1

            payment_created = 0
            if self._has_direct_payment_hint(record):
                payment_created = self._record_payment(
                    batch,
                    sheet,
                    row_number,
                    record,
                    full_name,
                    first_name,
                    last_name,
                    registration_no,
                    practitioner_number,
                    applicant_type,
                    category,
                    target_model,
                    record_year,
                    parse_date(record.get("payment_date")) or parse_date(record.get("atp_date")),
                    extract_amount(record.get("amount")),
                    record.get("receipt_number"),
                )
            imported += payment_created
            if not year_hits and not payment_created:
                skipped += 1
        return imported, skipped

    def _record_legacy_row(self, batch, sheet, workbook_path, source_sheet, row_number, record):
        source_key = self._source_sheet_key(workbook_path, source_sheet, row_number)
        category = infer_category(
            source_sheet,
            record.get("qualification"),
            record.get("professional_category"),
            record.get("specialty"),
        )
        target_model = infer_target_model(category, record.get("qualification"), source_sheet)
        record_type = record_type_from_status(
            record.get("license_status"),
            provisional_no=record.get("provisional_no"),
            full_registration_no=record.get("full_registration_no"),
        )
        full_name = title_name(record.get("full_name"))
        first_name, last_name = split_name(full_name)
        record_year = (
            latest_year(record.get("record_year"))
            or latest_year(record.get("last_atp_year"))
            or latest_year(record.get("atp_date"))
            or latest_year(record.get("registration_date"))
        )
        issued_date = parse_date(record.get("atp_date")) or parse_date(record.get("registration_date"))
        registration_no = self._source_registration(record, source_key, target_model)
        practitioner_number = clean_text(
            record.get("practitioner_no")
            or record.get("medical_practitioner_no")
            or record.get("practitioner_number")
        )
        receipt_number = clean_text(record.get("receipt_number"))
        applicant_type = applicant_type_from_nationality(record.get("nationality"))
        source_sheet_name = self._sheet_name(sheet)
        if self._existing_record(
            source_sheet_name=source_sheet_name,
            source_row=row_number,
            record_type=record_type,
            record_year=record_year,
            reference_number=receipt_number,
            target_model=target_model,
            registration_no=registration_no,
            full_name=full_name,
            practitioner_number=practitioner_number,
        ):
            self.summary["duplicate_rows_skipped"] += 1
            return 0
        identity_key = (
            full_name.upper(),
            category.upper(),
            registration_no.upper(),
            practitioner_number.upper(),
            str(record_year or ""),
            receipt_number.upper(),
        )
        if identity_key in self._row_dedupes:
            self.summary["duplicate_rows_skipped"] += 1
            return 0
        self._row_dedupes.add(identity_key)

        medical_doctor = None
        if target_model == "medicaldoctor":
            if not self.dry_run:
                medical_doctor = self._upsert_medical_doctor(
                    registration_no,
                    first_name,
                    last_name,
                    record,
                    category,
                    applicant_type,
                    issued_date,
                )

        if self.dry_run:
            self.summary["records_created"] += 1
            self.summary[f"{target_model}_rows"] += 1
            self.summary[f"category_{category}"] += 1
            self.summary[f"record_type_{record_type}"] += 1
        else:
            PracticingLicenseRecord.objects.create(
                batch=batch,
                sheet=sheet,
                record_type=record_type,
                target_model=target_model,
                source_sheet_name=source_sheet_name,
                source_row=row_number,
                record_year=record_year,
                full_name=full_name,
                first_name=first_name,
                last_name=last_name,
                registration_no=registration_no,
                practitioner_number=practitioner_number,
                applicant_type=applicant_type,
                nationality=record.get("nationality", ""),
                qualification_name=record.get("qualification", ""),
                category=category,
                workplace_address=record.get("employer_name", ""),
                province=record.get("province", ""),
                issued_date=issued_date,
                reference_number=receipt_number,
                amount=extract_amount(record.get("remarks") or record.get("amount") or record.get("receipt_number")),
                raw_payload=self._payload(record, workbook_path, source_sheet, category, target_model, source_key),
            )
            self.summary["records_created"] += 1
            self.summary[f"{target_model}_rows"] += 1
            self.summary[f"category_{category}"] += 1
            self.summary[f"record_type_{record_type}"] += 1
        if target_model == "medicaldoctor" and medical_doctor:
            self.summary["medical_doctors_upserted"] += 1

        if self._has_direct_payment_hint(record):
            self._record_payment(
                batch,
                sheet,
                row_number,
                record,
                full_name,
                first_name,
                last_name,
                registration_no,
                practitioner_number,
                applicant_type,
                category,
                target_model,
                record_year,
                parse_date(record.get("payment_date")) or issued_date,
                extract_amount(record.get("amount") or record.get("remarks")),
                record.get("receipt_number"),
            )
        return 1

    def _record_year_columns(self, batch, sheet, source_sheet, row_number, record, headers, row):
        full_name = title_name(record.get("full_name"))
        if not full_name:
            return
        first_name, last_name = split_name(full_name)
        category = infer_category(
            source_sheet,
            record.get("qualification"),
            record.get("professional_category"),
            record.get("specialty"),
        )
        target_model = infer_target_model(category, record.get("qualification"), source_sheet)
        source_sheet_name = self._sheet_name(sheet)
        source_key = f"{row_number}-{normalise_source_key(source_sheet_name)[:24]}"
        registration_no = self._source_registration(record, source_key, target_model)
        practitioner_number = clean_text(record.get("practitioner_no") or record.get("medical_practitioner_no"))
        applicant_type = applicant_type_from_nationality(record.get("nationality"))

        for position, (index, year) in enumerate(year_columns(headers), start=1):
            payment_text = clean_text(row[index] if index < len(row) else None)
            receipt_text = clean_text(row[index + 1] if index + 1 < len(row) else None)
            if not payment_text and not receipt_text:
                continue
            key = (
                full_name.upper(),
                category.upper(),
                source_sheet_name.upper(),
                str(year),
                payment_text.upper(),
                receipt_text.upper(),
            )
            if key in self._licence_dedupes:
                self.summary["duplicate_licence_rows_skipped"] += 1
                continue
            self._licence_dedupes.add(key)
            source_row = (row_number * 100) + position
            if self._existing_record(
                source_sheet_name=source_sheet_name,
                source_row=source_row,
                record_type="practicing_license",
                record_year=year,
                target_model=target_model,
                registration_no=registration_no,
                full_name=full_name,
                practitioner_number=practitioner_number,
                reference_number=receipt_text or payment_text,
                payment_date=parse_date(payment_text),
            ):
                self.summary["duplicate_licence_rows_skipped"] += 1
                continue
            if self.dry_run:
                self.summary["records_created"] += 1
                self.summary["practicing_license_rows"] += 1
                self.summary["record_type_practicing_license"] += 1
                continue
            PracticingLicenseRecord.objects.create(
                batch=batch,
                sheet=sheet,
                record_type="practicing_license",
                target_model=target_model,
                source_sheet_name=source_sheet_name,
                source_row=source_row,
                record_year=year,
                full_name=full_name,
                first_name=first_name,
                last_name=last_name,
                registration_no=registration_no,
                practitioner_number=practitioner_number,
                applicant_type=applicant_type,
                qualification_name=record.get("qualification", ""),
                category=category,
                workplace_address=record.get("employer_name", ""),
                province=record.get("province", ""),
                payment_date=parse_date(payment_text),
                amount=extract_amount(payment_text),
                reference_number=receipt_text or payment_text,
                raw_payload={
                    "source_sheet": source_sheet,
                    "payment": payment_text,
                    "receipt": receipt_text,
                    "year": year,
                    "form_code": "MBRN",
                    "profession_track": "medical_doctor" if target_model == "medicaldoctor" else "allied_health",
                },
            )
            self.summary["records_created"] += 1
            self.summary["practicing_license_rows"] += 1
            self.summary["record_type_practicing_license"] += 1

    def _record_payment(
        self,
        batch,
        sheet,
        row_number,
        record,
        full_name,
        first_name,
        last_name,
        registration_no,
        practitioner_number,
        applicant_type,
        category,
        target_model,
        record_year,
        payment_date,
        payment_amount,
        reference_number,
    ):
        source_row = (row_number * 1000) + 1
        reference_number = clean_text(reference_number)
        if self._existing_record(
            source_sheet_name=self._sheet_name(sheet),
            source_row=source_row,
            record_type="payment",
            record_year=record_year,
            reference_number=reference_number,
            target_model=target_model,
            registration_no=registration_no,
            full_name=full_name,
            practitioner_number=practitioner_number,
            payment_date=payment_date,
        ):
            self.summary["duplicate_rows_skipped"] += 1
            return 0
        if self.dry_run:
            if reference_number or payment_amount is not None:
                self.summary["records_created"] += 1
                self.summary["payment_rows"] += 1
                self.summary["record_type_payment"] += 1
                return 1
            return 0
        if not reference_number and payment_amount is None:
            return 0
        PracticingLicenseRecord.objects.create(
            batch=batch,
            sheet=sheet,
            record_type="payment",
            target_model=target_model,
            source_sheet_name=self._sheet_name(sheet),
            source_row=source_row,
            record_year=record_year,
            full_name=full_name,
            first_name=first_name,
            last_name=last_name,
            registration_no=registration_no,
            practitioner_number=practitioner_number,
            applicant_type=applicant_type,
            qualification_name=record.get("qualification", ""),
            category=category,
            workplace_address=record.get("employer_name", ""),
            province=record.get("province", ""),
            payment_date=payment_date,
            amount=payment_amount,
            reference_number=reference_number,
            raw_payload={**record, "form_code": "MBRN", "payment_source": "legacy_medical_board_workbook"},
        )
        self.summary["records_created"] += 1
        self.summary["payment_rows"] += 1
        self.summary["record_type_payment"] += 1
        return 1

    def _payload(self, record, workbook_path, source_sheet, category, target_model, source_key):
        return {
            "source_file": str(workbook_path),
            "source_sheet": source_sheet,
            "source_key": source_key,
            "form_code": "MBSP" if "SPECIALIST" in category.upper() else "MBRN",
            "profession_track": "medical_doctor" if target_model == "medicaldoctor" else "allied_health",
            "category": category,
            "professional_category": record.get("professional_category", ""),
            "specialty": record.get("specialty", ""),
            "full_name": record.get("full_name", ""),
            "nationality": record.get("nationality", ""),
            "registration_no": record.get("registration_card", ""),
            "practitioner_no": record.get("practitioner_no") or record.get("medical_practitioner_no", ""),
            "qualification": record.get("qualification", ""),
            "application_for": record.get("application_for", ""),
            "year": record.get("record_year", ""),
            "employer_name": record.get("employer_name", ""),
            "license_status": record.get("license_status", ""),
            "atp_date": record.get("atp_date", ""),
            "receipt_number": record.get("receipt_number", ""),
            "province": record.get("province", ""),
            "remarks": record.get("remarks", ""),
        }

    def _upsert_medical_doctor(self, registration_no, first_name, last_name, record, category, applicant_type, issued_date):
        specialty = infer_specialty_value(category, record.get("qualification", ""))
        existing = MedicalDoctor.objects.filter(registration_no=registration_no).only("pk").first()
        defaults = {
            "first_name": first_name or "Medical",
            "last_name": last_name or "Practitioner",
            "applicant_type": applicant_type,
            "nationality": record.get("nationality", ""),
            "full_address": record.get("employer_name", ""),
            "province": record.get("province", ""),
            "date_issued": issued_date,
            "is_active": True,
        }
        if specialty or not existing:
            defaults["specialty"] = specialty
        doctor, _created = MedicalDoctor.objects.update_or_create(
            registration_no=registration_no,
            defaults=defaults,
        )
        qualification = record.get("qualification", "")
        if qualification:
            Qualification.objects.update_or_create(
                content_type=ContentType.objects.get_for_model(doctor),
                object_id=doctor.pk,
                qualification_name=qualification[:200],
                defaults={
                    "institution_name": record.get("employer_name", "")[:255],
                    "program_completed": category[:255],
                    "qualification_type": "Medical Board legacy import",
                    "country": record.get("nationality", "") or "Papua New Guinea",
                },
            )
        return doctor
