from collections import Counter
from datetime import date, datetime, timedelta
from decimal import Decimal
import os
from pathlib import Path
import re

import pandas as pd
from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.workforce.models import (
    CommunityHealthWorker,
    DataImportBatch,
    ImportedWorkbookSheet,
    PracticingLicenseRecord,
    Qualification,
    TrainingInstitution,
)


_DEFAULT_WORKBOOK_CANDIDATES = [
    Path(os.environ["MEDICAL_BOARD_CHW_WORKBOOK_PATH"])
    for _ in [0]
    if os.environ.get("MEDICAL_BOARD_CHW_WORKBOOK_PATH")
] + [
    Path(
        r"c:\Users\darre\OneDrive\Documents\ProjectApps\databasedocuments"
        r"\medicalboarddata\Current database on Medical Board"
        r"\CHW 1985-2026 DATABASE CURRENTLY UPDATING"
        r"\CHW 1985-2026 DATABASE CURRENTLY UPDATING.xlsx"
    ),
    Path(r"d:\Database Template\Medical Board\CHW 1985-2026 DATABASE CURRENTLY UPDATING.xlsx"),
]
DEFAULT_MEDICAL_BOARD_WORKBOOK = next(
    (path for path in _DEFAULT_WORKBOOK_CANDIDATES if path.exists()),
    _DEFAULT_WORKBOOK_CANDIDATES[0],
)

MEDICAL_BOARD_CHW_MARKER_SHEETS = {
    "CHW",
    "ATP DATABASE ONLY",
    "NOT ON DATABASE",
    "NOT ON DATABASE CHW",
    "SCHOOL ADDRESS",
    "RECIEPT TRACKER",
    "RECEIPT TRACKER",
}

PROVINCES = [
    "Autonomous Region of Bougainville",
    "Central",
    "Chimbu",
    "East New Britain",
    "East Sepik",
    "Eastern Highlands",
    "Enga",
    "Gulf",
    "Hela",
    "Jiwaka",
    "Madang",
    "Manus",
    "Milne Bay",
    "Morobe",
    "National Capital District",
    "New Ireland",
    "Northern",
    "Oro",
    "Sandaun",
    "Simbu",
    "Southern Highlands",
    "Western",
    "Western Highlands",
    "West New Britain",
]

PROVINCE_ALIASES = {
    "NCD": "National Capital District",
    "W.H.P": "Western Highlands",
    "WHP": "Western Highlands",
    "S.H.P": "Southern Highlands",
    "SHP": "Southern Highlands",
    "E.H.P": "Eastern Highlands",
    "EHP": "Eastern Highlands",
    "W.S.P": "Sandaun",
    "ORO": "Northern",
    "NORTHERN PROVINCE": "Northern",
    "MOROBE PROVINCE": "Morobe",
    "MADANG PROVINCE": "Madang",
}

CHW_HEADER_ALIASES = {
    "address": "address",
    "amount": "amount",
    "app/status": "license_status",
    "apply for": "application_for",
    "arch#": "practitioner_number",
    "cert #": "registry",
    "certificate no": "registry",
    "date": "date",
    "date of the receipt": "payment_date",
    "dob": "date_of_birth",
    "email": "email",
    "expiry due date": "expiry_due_date",
    "gender": "gender",
    "license#": "registry",
    "marital status": "marital_status",
    "name": "name",
    "nationality": "nationality",
    "phone": "primary_phone",
    "place of employment": "employer_name",
    "place of orgin": "place_of_origin",
    "place of origin": "place_of_origin",
    "practitioner's no": "practitioner_number",
    "professional": "professional_cadre",
    "professional-cadres": "professional_cadre",
    "professional-cahres": "professional_cadre",
    "qualification": "qualification",
    "qualification2": "qualification",
    "qualifications": "qualification",
    "qualifications2": "qualification",
    "qualifications\\designation": "qualification",
    "receipt": "receipt",
    "receipt no": "receipt",
    "receipt no.": "receipt",
    "receipt no#": "receipt",
    "reciept": "receipt",
    "reciept no": "receipt",
    "remarks": "remarks",
    "registry #": "registry",
    "school address": "address",
    "surname": "surname",
}


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    text = text.replace("\u25cf", "").replace("\u26ab", "")
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value):
    return re.sub(r"\s+", " ", clean_text(value).lower()).strip()


def normalize_sheet_name(value):
    return re.sub(r"\s+", " ", clean_text(value).upper()).strip()


def is_chw_register_sheet(sheet_name):
    return normalize_sheet_name(sheet_name) == "CHW"


def is_pending_chw_sheet(sheet_name):
    normalized = normalize_sheet_name(sheet_name)
    return normalized == "NOT ON DATABASE" or (
        "NOT ON DATABASE" in normalized and "CHW" in normalized
    )


def is_school_address_sheet(sheet_name):
    return normalize_sheet_name(sheet_name) == "SCHOOL ADDRESS"


def is_atp_database_sheet(sheet_name):
    return normalize_sheet_name(sheet_name) == "ATP DATABASE ONLY"


def is_medical_board_chw_workbook(workbook_path):
    path = Path(workbook_path)
    if not path.exists():
        return False
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        normalized_sheet_names = {normalize_sheet_name(name) for name in workbook.sheetnames}
        return (
            "CHW" in normalized_sheet_names
            and bool(normalized_sheet_names.intersection(MEDICAL_BOARD_CHW_MARKER_SHEETS))
        )
    finally:
        workbook.close()


def title_name(value):
    text = clean_text(value)
    if "," in text:
        family, given = [part.strip() for part in text.split(",", 1)]
        text = f"{given} {family}".strip()
    return text.title()


def split_name(value):
    text = title_name(value)
    parts = text.split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def parse_date(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        return value.date() if 1901 <= value.year <= date.today().year + 1 else None
    if isinstance(value, date):
        return value if 1901 <= value.year <= date.today().year + 1 else None
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        serial_value = float(value)
        if 20000 <= serial_value <= 60000:
            parsed = date(1899, 12, 30) + timedelta(days=int(serial_value))
            return parsed if 1901 <= parsed.year <= date.today().year + 1 else None

    text = clean_text(value)
    if not text:
        return None
    text = text.replace("_", ".").replace("-", ".")
    text = re.sub(r"\s+", "", text)
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed if 1901 <= parsed.year <= date.today().year + 1 else None
        except ValueError:
            continue
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    parsed_date = pd.Timestamp(parsed).date()
    return parsed_date if 1901 <= parsed_date.year <= date.today().year + 1 else None


def parse_birth_date(value):
    parsed = parse_date(value)
    if parsed:
        return parsed
    text = clean_text(value)
    if not text:
        return None
    text = text.replace("_", ".").replace("-", ".")
    text = re.sub(r"\s+", "", text)
    for fmt in ("%d.%m.%y", "%d/%m/%y"):
        try:
            candidate = datetime.strptime(text, fmt).date()
        except ValueError:
            continue
        if candidate > date.today():
            candidate = date(candidate.year - 100, candidate.month, candidate.day)
        return candidate if 1901 <= candidate.year <= date.today().year else None
    return None


def extract_year(value, fallback_date=None):
    parsed = parse_date(value)
    if parsed:
        return parsed.year
    text = clean_text(value)
    match = re.search(r"(19|20)\d{2}", text)
    if match:
        year = int(match.group(0))
        if 1980 <= year <= date.today().year + 1:
            return year
    return fallback_date.year if fallback_date else None


def normalize_registration(value, prefix="CHW"):
    text = clean_text(value).upper()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    text = re.sub(r"\s+", "", text)
    return f"{prefix}-{text}"[:50]


def extract_practitioner_number(remarks):
    text = clean_text(remarks).upper()
    match = re.search(r"P\s*#\s*:?\s*([A-Z0-9/-]+)", text)
    return match.group(1)[:100] if match else ""


def extract_student_id(remarks):
    text = clean_text(remarks).upper()
    match = re.search(r"STUDENT\s*ID\s*[-:]?\s*([A-Z0-9/-]+)", text)
    return match.group(1)[:100] if match else ""


def extract_province(address):
    text = clean_text(address).upper()
    for alias, province in PROVINCE_ALIASES.items():
        if alias in text:
            return province
    for province in PROVINCES:
        if province.upper() in text:
            return province
    return ""


def extract_school_name(qualification):
    text = clean_text(qualification)
    match = re.search(r"\(([^)]+)\)", text)
    if match:
        return match.group(1).strip().title()
    match = re.search(r"CERT(?:IFICATE)? IN CHW,\s*([^,]+)", text, flags=re.IGNORECASE)
    return match.group(1).strip().title() if match else ""


def normalize_gender(value):
    text = clean_text(value).lower()
    if text in {"m", "male"}:
        return "Male"
    if text in {"f", "female"}:
        return "Female"
    return ""


def extract_amount(value):
    text = clean_text(value).upper()
    match = re.search(r"\bK\s*([0-9]+(?:\.[0-9]{1,2})?)", text)
    return Decimal(match.group(1)) if match else None


def find_chw_header(worksheet, *, max_rows=10):
    best = (None, {})
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row or 0, max_rows), values_only=True),
        start=1,
    ):
        mapped = {}
        for index, value in enumerate(row):
            key = CHW_HEADER_ALIASES.get(normalize_header(value))
            if key and key not in mapped:
                mapped[key] = index
        if "name" in mapped and ("registry" in mapped or "qualification" in mapped or "professional_cadre" in mapped):
            return row_number, mapped
        if len(mapped) > len(best[1]):
            best = (row_number, mapped)
    return best if "name" in best[1] else (None, {})


def row_value(cells, header_map, *keys):
    for key in keys:
        index = header_map.get(key)
        if index is not None and index < len(cells):
            value = clean_text(cells[index])
            if value:
                return value
    return ""


def row_date(cells, header_map, *keys):
    for key in keys:
        index = header_map.get(key)
        if index is not None and index < len(cells):
            parsed = parse_date(cells[index])
            if parsed:
                return parsed
    return None


def row_birth_date(cells, header_map):
    index = header_map.get("date_of_birth")
    if index is not None and index < len(cells):
        return parse_birth_date(cells[index])
    return None


def chw_full_name(record):
    name = clean_text(record.get("name"))
    surname = clean_text(record.get("surname"))
    if surname and surname.upper() not in name.upper().split():
        return title_name(f"{name} {surname}".strip())
    return title_name(name)


def is_header_like_chw_record(record):
    name = clean_text(record.get("name")).lower()
    registry = clean_text(record.get("registry")).lower()
    return name in {"name", "names", "full name"} or registry in {"cert #", "license#", "registry #"}


def atp_year_columns(worksheet):
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    columns = []
    for index, value in enumerate(header):
        text = clean_text(value)
        if re.fullmatch(r"(19|20)\d{2}", text):
            year = int(text)
            if 1980 <= year <= date.today().year + 1:
                columns.append((index, year))
    return columns


def atp_header(worksheet):
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row or 0, 5), values_only=True),
        start=1,
    ):
        columns = []
        for index, value in enumerate(row):
            text = clean_text(value)
            if re.fullmatch(r"(19|20)\d{2}", text):
                year = int(text)
                if 1980 <= year <= date.today().year + 1:
                    columns.append((index, year))
        mapped = {}
        for index, value in enumerate(row):
            key = CHW_HEADER_ALIASES.get(normalize_header(value))
            if key and key not in mapped:
                mapped[key] = index
        if columns and "name" in mapped:
            return row_number, mapped, columns
    return 1, {"registry": 0, "date": 1, "name": 2, "practitioner_number": 3}, atp_year_columns(worksheet)


class MedicalBoardWorkbookImporter:
    def __init__(self, workbook_path=DEFAULT_MEDICAL_BOARD_WORKBOOK, initiated_by=None, dry_run=False):
        self.workbook_path = Path(workbook_path)
        self.initiated_by = initiated_by
        self.dry_run = dry_run
        self.summary = Counter()

    def _existing_record(
        self,
        source_sheet_name,
        source_row,
        record_type,
        *,
        target_model="communityhealthworker",
        registration_no="",
        full_name="",
        record_year=None,
        reference_number="",
        practitioner_number="",
        payment_date=None,
    ):
        if self.dry_run:
            return False
        if PracticingLicenseRecord.objects.filter(
            source_sheet_name=source_sheet_name,
            source_row=source_row,
            record_type=record_type,
            target_model=target_model,
        ).exists():
            return True

        query = PracticingLicenseRecord.objects.filter(
            record_type=record_type,
            target_model=target_model,
        )
        if registration_no:
            query = query.filter(registration_no=registration_no)
        elif practitioner_number:
            query = query.filter(practitioner_number=practitioner_number)
        elif full_name:
            query = query.filter(full_name__iexact=full_name)
        else:
            return False

        if record_type == "workforce_listing":
            return query.exists()
        if record_year:
            query = query.filter(record_year=record_year)
        if reference_number:
            query = query.filter(reference_number=reference_number)
        if payment_date:
            query = query.filter(payment_date=payment_date)
        return query.exists()

    def import_workbook(self):
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Medical Board workbook not found: {self.workbook_path}")

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        batch = DataImportBatch.objects.create(
            source_file_name=self.workbook_path.name,
            source_file_path=str(self.workbook_path),
            source_kind="medical_board_workbook",
            status="running",
            total_sheets=len(workbook.sheetnames),
            total_rows=sum(workbook[name].max_row or 0 for name in workbook.sheetnames),
            initiated_by=self.initiated_by,
        )

        try:
            with transaction.atomic():
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    if is_chw_register_sheet(sheet_name):
                        self._import_chw_sheet(batch, worksheet)
                    elif "CHWS" in normalize_sheet_name(sheet_name) and "FILE" in normalize_sheet_name(sheet_name):
                        self._import_chw_sheet(batch, worksheet)
                    elif is_pending_chw_sheet(sheet_name):
                        self._import_pending_chw_sheet(batch, worksheet)
                    elif is_school_address_sheet(sheet_name):
                        self._import_school_sheet(batch, worksheet)
                    elif is_atp_database_sheet(sheet_name):
                        self._import_atp_sheet(batch, worksheet)
                    else:
                        ImportedWorkbookSheet.objects.create(
                            batch=batch,
                            sheet_name=sheet_name,
                            sheet_type="reference",
                            status="skipped",
                            raw_rows=worksheet.max_row or 0,
                            notes="Reference/chart sheet retained in workbook but not row-imported.",
                        )
                    batch.processed_sheets += 1
                    batch.save(update_fields=["processed_sheets"])

            batch.status = "completed"
            batch.summary = dict(self.summary)
            batch.processed_rows = self.summary.get("processed_rows", 0)
            batch.completed_at = timezone.now()
            batch.save(update_fields=["status", "summary", "processed_rows", "completed_at"])
            return batch
        except Exception as exc:
            batch.status = "failed"
            batch.summary = {"error": str(exc), **dict(self.summary)}
            batch.completed_at = timezone.now()
            batch.save(update_fields=["status", "summary", "completed_at"])
            raise
        finally:
            workbook.close()

    def _record_import(self, batch, sheet, row_number, record, *, pending=False):
        if is_header_like_chw_record(record):
            self.summary["skipped_header_rows"] += 1
            return None
        registration_no = normalize_registration(record.get("registry"))
        full_name = chw_full_name(record)
        if not full_name:
            self.summary["skipped_missing_name"] += 1
            return None

        issued_date = parse_date(record.get("date"))
        year = extract_year(record.get("date"), issued_date)
        first_name, last_name = split_name(full_name)
        qualification = clean_text(record.get("qualification"))
        address = clean_text(record.get("address"))
        remarks = clean_text(record.get("remarks"))
        receipt = clean_text(record.get("receipt"))
        practitioner_number = (
            clean_text(record.get("practitioner_number"))
            or extract_practitioner_number(remarks)
            or extract_student_id(remarks)
        )
        province = extract_province(address)

        if not registration_no:
            safe_key = practitioner_number or f"{full_name}-{qualification}" or f"{sheet.sheet_name}-{row_number}"
            registration_no = normalize_registration(safe_key, prefix="CHW-PENDING")

        chw = self._upsert_chw_profile(
            registration_no=registration_no,
            first_name=first_name,
            last_name=last_name,
            record=record,
            address=address,
            province=province,
            practitioner_number=practitioner_number,
            qualification=qualification,
        )
        self._save_qualification(chw, qualification)

        if self._existing_record(
            sheet.sheet_name,
            row_number,
            "workforce_listing",
            registration_no=registration_no,
            full_name=full_name,
            record_year=year,
            practitioner_number=practitioner_number,
        ):
            self.summary["duplicate_rows_skipped"] += 1
            return chw

        PracticingLicenseRecord.objects.create(
            batch=batch,
            sheet=sheet,
            record_type="workforce_listing",
            target_model="communityhealthworker",
            source_sheet_name=sheet.sheet_name,
            source_row=row_number,
            record_year=year,
            full_name=full_name,
            first_name=first_name,
            last_name=last_name,
            registration_no=registration_no,
            practitioner_number=practitioner_number,
            applicant_type="national",
            qualification_name=qualification,
            category="Community Health Worker - Pending" if pending else "Community Health Worker",
            institution_name=extract_school_name(qualification),
            workplace_address=address,
            province=province,
            issued_date=issued_date,
            payment_date=record.get("payment_date"),
            amount=extract_amount(record.get("amount")),
            reference_number=receipt,
            raw_payload={
                "registry": clean_text(record.get("registry")),
                "date": clean_text(record.get("date")),
                "address": address,
                "receipt": receipt,
                "remarks": remarks,
                "pending_import": pending,
                "date_of_birth": record.get("date_of_birth").isoformat() if record.get("date_of_birth") else "",
                "marital_status": clean_text(record.get("marital_status")),
                "gender": normalize_gender(record.get("gender")),
                "nationality": clean_text(record.get("nationality")),
                "phone": clean_text(record.get("primary_phone")),
                "email": clean_text(record.get("email")),
                "application_for": clean_text(record.get("application_for")),
                "professional_cadre": clean_text(record.get("professional_cadre")),
                "license_status": clean_text(record.get("license_status")),
                "expiry_due_date": record.get("expiry_due_date").isoformat() if record.get("expiry_due_date") else "",
            },
        )
        self.summary["chw_imported"] += 1
        if pending:
            self.summary["pending_chw_imported"] += 1
        self.summary["processed_rows"] += 1
        if year:
            self.summary[f"year_{year}"] += 1
        return chw

    def _upsert_chw_profile(
        self,
        *,
        registration_no,
        first_name,
        last_name,
        record,
        address,
        province,
        practitioner_number,
        qualification,
    ):
        defaults = {
            "first_name": first_name or "CHW",
            "last_name": last_name or "Record",
            "applicant_type": "national",
            "date_of_birth": record.get("date_of_birth"),
            "gender": normalize_gender(record.get("gender")),
            "marital_status": clean_text(record.get("marital_status"))[:30],
            "nationality": clean_text(record.get("nationality"))[:100] or "PNG",
            "primary_phone": clean_text(record.get("primary_phone"))[:20] or None,
            "email": clean_text(record.get("email"))[:254],
            "full_address": address,
            "province": province,
            "community_id": practitioner_number[:50],
            "training_level": qualification[:100],
            "is_active": True,
        }
        existing = CommunityHealthWorker.objects.filter(registration_no=registration_no).first()
        if not existing:
            return CommunityHealthWorker.objects.create(registration_no=registration_no, **defaults)
        changed = []
        for field, value in defaults.items():
            current = getattr(existing, field)
            if value not in ("", None) or current in ("", None):
                if current != value:
                    setattr(existing, field, value)
                    changed.append(field)
        if changed:
            existing.save(update_fields=changed + ["updated_at"])
        return existing

    def _record_chw_payment(self, batch, sheet, row_number, record, chw):
        reference_number = clean_text(record.get("receipt"))
        amount = extract_amount(record.get("amount"))
        payment_date = record.get("payment_date")
        if not reference_number and amount is None and not payment_date:
            return None
        full_name = chw_full_name(record)
        first_name, last_name = split_name(full_name)
        record_year = extract_year(payment_date, payment_date) or extract_year(record.get("date"))
        registration_no = chw.registration_no
        practitioner_number = clean_text(record.get("practitioner_number")) or chw.community_id
        source_row = (row_number * 1000) + 1
        if self._existing_record(
            sheet.sheet_name,
            source_row,
            "payment",
            registration_no=registration_no,
            full_name=full_name,
            record_year=record_year,
            reference_number=reference_number,
            practitioner_number=practitioner_number,
            payment_date=payment_date,
        ):
            self.summary["duplicate_rows_skipped"] += 1
            return None
        raw_payload = {
            key: (value.isoformat() if hasattr(value, "isoformat") else value)
            for key, value in record.items()
        }
        PracticingLicenseRecord.objects.create(
            batch=batch,
            sheet=sheet,
            record_type="payment",
            target_model="communityhealthworker",
            source_sheet_name=sheet.sheet_name,
            source_row=source_row,
            record_year=record_year,
            full_name=full_name,
            first_name=first_name,
            last_name=last_name,
            registration_no=registration_no,
            practitioner_number=practitioner_number,
            applicant_type="national",
            qualification_name=clean_text(record.get("qualification")),
            category="Community Health Worker Payment",
            workplace_address=clean_text(record.get("address")),
            province=extract_province(record.get("address")),
            payment_date=payment_date,
            amount=amount,
            reference_number=reference_number,
            raw_payload={**raw_payload, "payment_source": "medical_board_chw_workbook"},
        )
        self.summary["payments_imported"] += 1
        self.summary["processed_rows"] += 1
        return True

    def _import_chw_sheet(self, batch, worksheet):
        header_row, header_map = find_chw_header(worksheet)
        sheet = ImportedWorkbookSheet.objects.create(
            batch=batch,
            sheet_name=worksheet.title,
            sheet_type="medical_board_chw",
            status="processed",
            raw_rows=max((worksheet.max_row or 1) - (header_row or 1), 0),
        )
        if not header_row:
            sheet.status = "skipped"
            sheet.notes = "No CHW person header found."
            sheet.save(update_fields=["status", "notes"])
            self.summary["sheets_skipped"] += 1
            return

        imported = skipped = 0
        for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            cells = list(row)
            record = {
                "registry": row_value(cells, header_map, "registry"),
                "date": row_value(cells, header_map, "date"),
                "name": row_value(cells, header_map, "name"),
                "surname": row_value(cells, header_map, "surname"),
                "date_of_birth": row_birth_date(cells, header_map),
                "marital_status": row_value(cells, header_map, "marital_status"),
                "nationality": row_value(cells, header_map, "nationality", "place_of_origin"),
                "gender": row_value(cells, header_map, "gender"),
                "primary_phone": row_value(cells, header_map, "primary_phone"),
                "email": row_value(cells, header_map, "email"),
                "address": row_value(cells, header_map, "address", "employer_name"),
                "qualification": row_value(cells, header_map, "qualification"),
                "application_for": row_value(cells, header_map, "application_for"),
                "professional_cadre": row_value(cells, header_map, "professional_cadre"),
                "license_status": row_value(cells, header_map, "license_status"),
                "receipt": row_value(cells, header_map, "receipt"),
                "payment_date": row_date(cells, header_map, "payment_date"),
                "amount": row_value(cells, header_map, "amount"),
                "expiry_due_date": row_date(cells, header_map, "expiry_due_date"),
                "practitioner_number": row_value(cells, header_map, "practitioner_number"),
                "remarks": row_value(cells, header_map, "remarks"),
            }
            if not chw_full_name(record):
                skipped += 1
                continue
            chw = self._record_import(
                batch,
                sheet,
                row_number,
                record,
            )
            if chw:
                imported += 1
                self._record_chw_payment(batch, sheet, row_number, record, chw)
            else:
                skipped += 1
        sheet.imported_rows = imported
        sheet.skipped_rows = skipped
        sheet.metadata = {"header_row": header_row, "mapped_columns": sorted(header_map)}
        sheet.save(update_fields=["imported_rows", "skipped_rows", "metadata"])

    def _import_atp_sheet(self, batch, worksheet):
        header_row, header_map, year_columns = atp_header(worksheet)
        sheet = ImportedWorkbookSheet.objects.create(
            batch=batch,
            sheet_name=worksheet.title,
            sheet_type="medical_board_chw_practicing_licences",
            status="processed",
            raw_rows=max((worksheet.max_row or 1) - header_row, 0),
        )
        imported = skipped = 0
        for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            cells = list(row)
            record = {
                "registry": row_value(cells, header_map, "registry"),
                "date": row_value(cells, header_map, "date"),
                "name": row_value(cells, header_map, "name"),
                "surname": row_value(cells, header_map, "surname"),
                "practitioner_number": row_value(cells, header_map, "practitioner_number"),
            }
            if is_header_like_chw_record(record):
                skipped += 1
                continue
            full_name = chw_full_name(record)
            if not full_name:
                skipped += 1
                continue

            first_name, last_name = split_name(full_name)
            registration_no = normalize_registration(record.get("registry"))
            if not registration_no:
                registration_no = normalize_registration(f"ATP-{row_number}", prefix="CHW")
            practitioner_number = clean_text(record.get("practitioner_number"))
            base_issued_date = parse_date(record.get("date"))
            row_imported = 0

            for position, (column_index, year) in enumerate(year_columns, start=1):
                payment_text = clean_text(cells[column_index] if column_index < len(cells) else None)
                receipt_text = clean_text(cells[column_index + 1] if column_index + 1 < len(cells) else None)
                if not payment_text and not receipt_text:
                    continue
                source_row = (row_number * 100) + position
                if self._existing_record(
                    source_sheet_name=sheet.sheet_name,
                    source_row=source_row,
                    record_type="practicing_license",
                    registration_no=registration_no,
                    full_name=full_name,
                    record_year=year,
                    reference_number=receipt_text or payment_text,
                    practitioner_number=practitioner_number,
                    payment_date=parse_date(payment_text),
                ):
                    self.summary["duplicate_licence_rows_skipped"] += 1
                    continue

                PracticingLicenseRecord.objects.create(
                    batch=batch,
                    sheet=sheet,
                    record_type="practicing_license",
                    target_model="communityhealthworker",
                    source_sheet_name=sheet.sheet_name,
                    source_row=source_row,
                    record_year=year,
                    full_name=full_name,
                    first_name=first_name,
                    last_name=last_name,
                    registration_no=registration_no,
                    practitioner_number=practitioner_number,
                    applicant_type="national",
                    category="Community Health Worker Practising Licence",
                    issued_date=base_issued_date,
                    payment_date=parse_date(payment_text),
                    amount=extract_amount(payment_text),
                    reference_number=receipt_text or payment_text,
                    raw_payload={
                        "registry": clean_text(record.get("registry")),
                        "date": clean_text(record.get("date")),
                        "arch": practitioner_number,
                        "year": year,
                        "payment": payment_text,
                        "receipt": receipt_text,
                        "source_row": row_number,
                    },
                )
                imported += 1
                row_imported += 1
                self.summary["chw_practicing_licences_imported"] += 1
                self.summary["processed_rows"] += 1
                self.summary[f"licence_year_{year}"] += 1

            if not row_imported:
                skipped += 1

        sheet.imported_rows = imported
        sheet.skipped_rows = skipped
        sheet.metadata = {"header_row": header_row, "mapped_columns": sorted(header_map), "year_columns": [year for _, year in year_columns]}
        sheet.save(update_fields=["imported_rows", "skipped_rows", "metadata"])

    def _import_pending_chw_sheet(self, batch, worksheet):
        header_row, header_map = find_chw_header(worksheet)
        sheet = ImportedWorkbookSheet.objects.create(
            batch=batch,
            sheet_name=worksheet.title,
            sheet_type="medical_board_chw_pending",
            status="processed",
            raw_rows=max((worksheet.max_row or 1) - (header_row or 1), 0),
        )
        if not header_row:
            sheet.status = "skipped"
            sheet.notes = "No pending CHW person header found."
            sheet.save(update_fields=["status", "notes"])
            self.summary["sheets_skipped"] += 1
            return

        imported = skipped = 0
        for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            cells = list(row)
            record = {
                "registry": row_value(cells, header_map, "registry"),
                "date": row_value(cells, header_map, "date"),
                "name": row_value(cells, header_map, "name"),
                "surname": row_value(cells, header_map, "surname"),
                "date_of_birth": row_birth_date(cells, header_map),
                "marital_status": row_value(cells, header_map, "marital_status"),
                "nationality": row_value(cells, header_map, "nationality", "place_of_origin"),
                "gender": row_value(cells, header_map, "gender"),
                "primary_phone": row_value(cells, header_map, "primary_phone"),
                "email": row_value(cells, header_map, "email"),
                "address": row_value(cells, header_map, "address", "employer_name"),
                "qualification": row_value(cells, header_map, "qualification"),
                "application_for": row_value(cells, header_map, "application_for"),
                "professional_cadre": row_value(cells, header_map, "professional_cadre"),
                "license_status": row_value(cells, header_map, "license_status"),
                "receipt": row_value(cells, header_map, "receipt"),
                "payment_date": row_date(cells, header_map, "payment_date"),
                "amount": row_value(cells, header_map, "amount"),
                "expiry_due_date": row_date(cells, header_map, "expiry_due_date"),
                "practitioner_number": row_value(cells, header_map, "practitioner_number"),
                "remarks": row_value(cells, header_map, "remarks"),
            }
            if not chw_full_name(record):
                skipped += 1
                continue
            chw = self._record_import(
                batch,
                sheet,
                row_number,
                record,
                pending=True,
            )
            if chw:
                imported += 1
                self._record_chw_payment(batch, sheet, row_number, record, chw)
            else:
                skipped += 1
        sheet.imported_rows = imported
        sheet.skipped_rows = skipped
        sheet.metadata = {"header_row": header_row, "mapped_columns": sorted(header_map)}
        sheet.save(update_fields=["imported_rows", "skipped_rows", "metadata"])

    def _import_school_sheet(self, batch, worksheet):
        sheet = ImportedWorkbookSheet.objects.create(
            batch=batch,
            sheet_name=worksheet.title,
            sheet_type="medical_board_training_institutions",
            status="processed",
            raw_rows=worksheet.max_row or 0,
        )
        imported = 0
        stats = []
        for row in worksheet.iter_rows(min_row=4, values_only=True):
            number, school, province, address = (list(row) + [None] * 4)[:4]
            school_name = clean_text(school)
            if not school_name or not clean_text(number):
                continue
            institution, _ = TrainingInstitution.objects.update_or_create(
                name=school_name.title(),
                defaults={"type": "CHW Training School", "is_active": True},
            )
            stats.append({
                "institution_id": institution.pk,
                "school": institution.name,
                "province": clean_text(province),
                "address": clean_text(address),
            })
            imported += 1
        sheet.imported_rows = imported
        sheet.metadata = {"training_institutions": stats[:40]}
        sheet.save(update_fields=["imported_rows", "metadata"])
        self.summary["training_institutions_imported"] += imported

    def _save_qualification(self, chw, qualification):
        if not qualification:
            return
        content_type = ContentType.objects.get_for_model(chw)
        institution_name = extract_school_name(qualification)
        institution = None
        if institution_name:
            institution, _ = TrainingInstitution.objects.get_or_create(
                name=institution_name,
                defaults={"type": "CHW Training School", "is_active": True},
            )
        Qualification.objects.update_or_create(
            content_type=content_type,
            object_id=chw.pk,
            qualification_name=qualification[:200],
            defaults={
                "institution": institution,
                "institution_name": institution_name,
                "program_completed": "Community Health Worker",
                "qualification_type": "Certificate",
                "country": "Papua New Guinea",
            },
        )
