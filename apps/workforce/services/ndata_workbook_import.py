from collections import Counter
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re

import pandas as pd
from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.chartsheet import Chartsheet

from apps.dashboard.models import Receipt
from apps.workforce.models import (
    Application,
    CommunityHealthWorker,
    DataImportBatch,
    HealthStudent,
    ImportedWorkbookSheet,
    MedicalDoctor,
    Midwife,
    NurseAide,
    NursingProfessional,
    PracticingLicenseRecord,
    Qualification,
    TrainingInstitution,
    WorkforceSnapshot,
)
from apps.workforce.services.institution_classification import (
    applicant_type_for_institution,
    classify_training_institution,
)


DEFAULT_WORKBOOK = Path(r"d:\2026 Current N-DATA Statistics & Tracking - SECTIONS (Autosaved).xlsx")

def normalize_text(value):
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_name(value):
    text = normalize_text(value)
    if not text:
        return ""
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\s+", " ", text).strip(" -")
    if "," in text:
        left, right = [part.strip() for part in text.split(",", 1)]
        if left and right:
            text = f"{right} {left}"
    return text.title()


def split_name(full_name):
    parts = normalize_text(full_name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def normalize_identifier(value):
    text = normalize_text(value).upper()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    if re.fullmatch(r"\d+\.\d+", text):
        try:
            text = str(int(Decimal(text)))
        except (InvalidOperation, ValueError):
            pass
    text = text.replace("_", " ").replace("-", " - ")
    text = re.sub(r"\s*/\s*", "/", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def safe_secondary_identifier(value):
    text = normalize_identifier(value)
    if not text or len(text) > 50:
        return ""
    if text in {"TBA", "N/A", "NA", "NIL", "NONE", "UNKNOWN", "-"}:
        return ""
    return text


def normalize_registration_no(value, prefix=""):
    base = normalize_identifier(value)
    prefix = normalize_identifier(prefix)
    if not base and prefix:
        return prefix
    if prefix and base and not base.startswith(prefix):
        base = f"{prefix} {base}"
    base = re.sub(r"\s+", " ", base).strip()
    if len(base) > 100:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}", base):
        return ""
    return base


def normalize_provisional_registration(value, prefix=""):
    base = normalize_identifier(value)
    prefix = normalize_identifier(prefix)
    if not base:
        return ""
    if prefix in {"PRO", "PROV", "PROVISIONAL", ""}:
        base = re.sub(r"^PROV?[-\s]*", "", base, flags=re.IGNORECASE)
        base = base.replace(" - ", "-").replace(" ", "")
        return f"PROV-{base}"[:100]
    return normalize_registration_no(value, prefix)


def registration_for_model(value):
    value = normalize_registration_no(value)
    if not value or len(value) > 50:
        return ""
    if any(token in value for token in [" MW ", " NRS ", " MID ", " AND "]):
        return ""
    return value


def parse_date(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        serial_value = float(value)
        if 20000 <= serial_value <= 60000:
            excel_origin = date(1899, 12, 30)
            parsed_date = excel_origin + timedelta(days=int(serial_value))
            if 1901 <= parsed_date.year <= date.today().year + 1:
                return parsed_date
    if isinstance(value, datetime):
        if value.year < 1901:
            return None
        return value.date()
    if isinstance(value, date):
        if value.year < 1901:
            return None
        return value

    text = normalize_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{4}", text) and int(text) > date.today().year + 1:
        return None

    fixes = {
        "Sept.": "Sep",
        "Sept": "Sep",
        "Janurary": "January",
        "Konedoubu": "Konedobu",
        "Jui": "Jul",
        "Jin": "Jun",
    }
    for bad, good in fixes.items():
        text = text.replace(bad, good)
    text = text.replace("_", "-").replace("=", "-")
    text = re.sub(r"\s*([./])\s*", r"\1", text)
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"([A-Za-z]+)-(\d{2,4})", r"\1 \2", text)
    text = re.sub(r"\.{2,}", ".", text)
    text = re.sub(r"-{2,}", "-", text)
    text = re.sub(r"\s+", " ", text).strip()

    if re.fullmatch(r"\d{1,2} [A-Za-z]{3}\. - \d{2}", text):
        text = text.replace(". -", "-")

    parsed = pd.NaT
    for dayfirst in (True, False):
        attempt = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
        if not pd.isna(attempt):
            parsed = attempt
            break

    if pd.isna(parsed):
        return None

    parsed_date = pd.Timestamp(parsed).date()
    if parsed_date.year < 1901:
        return None
    if parsed_date.year > date.today().year + 1 and parsed_date.year - 20 >= 2000:
        parsed_date = parsed_date.replace(year=parsed_date.year - 20)
    return parsed_date


def parse_year(value, fallback_date=None):
    if value is None or pd.isna(value):
        return fallback_date.year if fallback_date else None
    text = normalize_text(value)
    if not text:
        return fallback_date.year if fallback_date else None
    if re.fullmatch(r"\d{4}", text):
        return int(text)
    if re.fullmatch(r"\d{5}(?:\.\d+)?", text):
        parsed = parse_date(float(text))
        if parsed:
            return parsed.year
    parsed = parse_date(value)
    if parsed:
        return parsed.year
    if fallback_date:
        return fallback_date.year
    return None


def parse_decimal(value):
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def infer_applicant_type(*values):
    combined = " ".join(normalize_text(value) for value in values if normalize_text(value))
    return applicant_type_for_institution(combined)


def infer_target_model(category="", qualification="", registration_no=""):
    text = " ".join(
        normalize_text(value).lower()
        for value in [category, qualification, registration_no]
        if normalize_text(value)
    )
    if "doctor" in text or text.startswith("md"):
        return "medicaldoctor"
    if "nurse aide" in text or "n/aide" in text or "nurse-aide" in text:
        return "nurseaide"
    if "community health" in text or "chw" in text:
        return "communityhealthworker"
    if "midw" in text:
        return "midwife"
    return "nursingprofessional"


def qualification_for_target(target_model, category, qualification):
    base = normalize_text(qualification) or normalize_text(category)
    if not base:
        if target_model == "midwife":
            return "Midwifery"
        if target_model == "nurseaide":
            return "Nurse Aide"
        return "General Nursing"
    return base[:100]


def safe_institution(name):
    clean = normalize_text(name)
    if not clean or clean.upper() in {"TBA", "N/A", "NA"}:
        return None
    institution_type = classify_training_institution(clean)
    institution, created = TrainingInstitution.objects.get_or_create(
        name=clean[:255],
        defaults={"type": institution_type},
    )
    if not created and institution.type != institution_type:
        institution.type = institution_type
        institution.save(update_fields=["type"])
    return institution


def unique_preserve_order(values):
    seen = set()
    result = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


class NDataWorkbookImporter:
    def __init__(
        self,
        workbook_path=DEFAULT_WORKBOOK,
        initiated_by=None,
        sync_live_profiles=True,
        deduplicate_rows=True,
        blank_row_limit=200,
        sheet_names=None,
    ):
        self.workbook_path = Path(workbook_path)
        self.initiated_by = initiated_by
        self.sync_live_profiles = sync_live_profiles
        self.deduplicate_rows = deduplicate_rows
        self.blank_row_limit = blank_row_limit
        self.sheet_names = unique_preserve_order(
            normalize_text(sheet_name)
            for sheet_name in (sheet_names or [])
            if normalize_text(sheet_name)
        )
        self.batch = None
        self._model_map = {
            "nursingprofessional": NursingProfessional,
            "midwife": Midwife,
            "medicaldoctor": MedicalDoctor,
            "communityhealthworker": CommunityHealthWorker,
            "nurseaide": NurseAide,
            "healthstudent": HealthStudent,
        }
        self._professional_cache = {
            key: {obj.registration_no: obj for obj in model.objects.exclude(registration_no__isnull=True)}
            for key, model in self._model_map.items()
        }
        self._application_ct_cache = {}
        self.summary = {
            "records_created": Counter(),
            "records_updated": Counter(),
            "practice_records": 0,
            "duplicate_rows_skipped": 0,
            "current_year_field_updates": 0,
            "receipts": 0,
            "sheets_processed": 0,
            "sheets_skipped": 0,
        }
        self._practice_record_keys = set()

    def import_workbook(self):
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            non_chart_sheets = [
                name for name in workbook.sheetnames
                if not isinstance(workbook[name], Chartsheet)
            ]
            selected_sheet_names = self._selected_workbook_sheet_names(non_chart_sheets)
            self.batch = DataImportBatch.objects.create(
                source_file_name=self.workbook_path.name,
                source_file_path=str(self.workbook_path),
                source_kind="ndata_workbook",
                status="running",
                total_sheets=len(selected_sheet_names),
                initiated_by=self.initiated_by,
            )

            try:
                with transaction.atomic():
                    for name in selected_sheet_names:
                        ws = workbook[name]
                        if isinstance(ws, Chartsheet):
                            continue
                        self._process_sheet(ws)
                    self._hydrate_from_current_year_rows()
                    self._sync_snapshots()
                    self.batch.status = "completed"
                    self.batch.completed_at = timezone.now()
                    self.batch.summary = self._build_summary()
                    self.batch.save(update_fields=["status", "completed_at", "summary", "processed_sheets", "processed_rows", "total_rows"])
            except Exception as exc:
                self.batch.status = "failed"
                self.batch.completed_at = timezone.now()
                self.batch.summary = self._build_summary(error=str(exc))
                self.batch.save(update_fields=["status", "completed_at", "summary"])
                raise
        finally:
            workbook.close()

        return self.batch

    def _selected_workbook_sheet_names(self, non_chart_sheets):
        if not self.sheet_names:
            return non_chart_sheets

        available_by_key = {name.strip().lower(): name for name in non_chart_sheets}
        missing = [
            sheet_name for sheet_name in self.sheet_names
            if sheet_name.strip().lower() not in available_by_key
        ]
        if missing:
            available = ", ".join(non_chart_sheets)
            requested = ", ".join(missing)
            raise ValueError(
                f"Selected sheet not found: {requested}. Available sheets: {available}"
            )

        return [
            available_by_key[sheet_name.strip().lower()]
            for sheet_name in self.sheet_names
        ]

    def _build_summary(self, error=""):
        return {
            "error": error,
            "records_created": dict(self.summary["records_created"]),
            "records_updated": dict(self.summary["records_updated"]),
            "practice_records": self.summary["practice_records"],
            "duplicate_rows_skipped": self.summary["duplicate_rows_skipped"],
            "current_year_field_updates": self.summary["current_year_field_updates"],
            "sync_live_profiles": self.sync_live_profiles,
            "receipts": self.summary["receipts"],
            "sheets_processed": self.summary["sheets_processed"],
            "sheets_skipped": self.summary["sheets_skipped"],
            "selected_sheet_names": self.sheet_names,
        }

    def _process_sheet(self, ws):
        sheet_name = ws.title
        sheet_key = sheet_name.lower()
        sheet = ImportedWorkbookSheet.objects.create(
            batch=self.batch,
            sheet_name=sheet_name,
            raw_rows=max(ws.max_row - 1, 0),
            status="pending",
        )

        if "prov rego" in sheet_key:
            handler = self._import_provisional_sheet
            sheet_type = "provisional"
        elif "full rego" in sheet_key:
            handler = self._import_full_sheet
            sheet_type = "full"
        elif "temp cert" in sheet_key or "temporary cert" in sheet_key:
            handler = self._import_temporary_sheet
            sheet_type = "temporary"
        elif "atp record" in sheet_key:
            handler = self._import_atp_record_sheet
            sheet_type = "practicing_license"
        elif "atp" in sheet_key and ("payment" in sheet_key or "updated spreadsheet" in sheet_key or re.search(r"\batp 20", sheet_key) or "2014" in sheet_key):
            handler = self._import_atp_payment_sheet
            sheet_type = "payment"
        elif any(token in sheet_key for token in ["active", "list", "sheet2", "print 2022", "pom gen", "update list"]):
            handler = self._import_listing_sheet
            sheet_type = "workforce_listing"
        else:
            sheet.sheet_type = "unsupported"
            sheet.status = "skipped"
            sheet.notes = "No importer configured for this sheet."
            sheet.save(update_fields=["sheet_type", "status", "notes"])
            self.summary["sheets_skipped"] += 1
            self.batch.processed_sheets += 1
            self.batch.save(update_fields=["processed_sheets"])
            return

        sheet.sheet_type = sheet_type
        sheet.save(update_fields=["sheet_type"])

        imported_rows, skipped_rows, notes = handler(ws, sheet)
        sheet.imported_rows = imported_rows
        sheet.skipped_rows = skipped_rows
        sheet.status = "processed"
        sheet.notes = notes
        sheet.save(update_fields=["imported_rows", "skipped_rows", "status", "notes"])

        self.summary["sheets_processed"] += 1
        self.batch.processed_sheets += 1
        self.batch.processed_rows += imported_rows
        self.batch.total_rows += imported_rows + skipped_rows
        self.batch.save(update_fields=["processed_sheets", "processed_rows", "total_rows"])

    def _practice_record_defaults(self, **kwargs):
        payload = kwargs.pop("raw_payload", {})
        defaults = {
            "batch": self.batch,
            "raw_payload": payload,
        }
        defaults.update(kwargs)
        return defaults

    def _iter_value_rows(self, ws, min_row, max_col):
        blank_count = 0
        for idx, row in enumerate(ws.iter_rows(min_row=min_row, max_col=max_col, values_only=True), start=min_row):
            if all(not normalize_text(value) for value in row):
                blank_count += 1
                if blank_count >= self.blank_row_limit:
                    break
                continue
            blank_count = 0
            yield idx, row

    def _practice_record_dedupe_key(self, record_type, kwargs):
        registration_no = normalize_registration_no(kwargs.get("registration_no"))
        practitioner_number = normalize_identifier(kwargs.get("practitioner_number"))
        full_name = normalize_name(kwargs.get("full_name"))
        dob = kwargs.get("date_of_birth") or ""
        identity = registration_no or practitioner_number or f"{full_name}|{dob}"
        return (
            record_type,
            kwargs.get("target_model", ""),
            kwargs.get("record_year"),
            identity,
            kwargs.get("issued_date"),
            kwargs.get("payment_date"),
            normalize_identifier(kwargs.get("reference_number")),
            normalize_text(kwargs.get("category")).lower(),
        )

    def _create_practice_record(self, sheet, source_row, record_type, **kwargs):
        if self.deduplicate_rows:
            dedupe_key = self._practice_record_dedupe_key(record_type, kwargs)
            if dedupe_key in self._practice_record_keys:
                self.summary["duplicate_rows_skipped"] += 1
                return None
            self._practice_record_keys.add(dedupe_key)
            if self.batch and PracticingLicenseRecord.objects.filter(
                batch__source_file_name=self.batch.source_file_name,
                source_sheet_name=sheet.sheet_name,
                source_row=source_row,
                record_type=record_type,
            ).exists():
                self.summary["duplicate_rows_skipped"] += 1
                return None
        record = PracticingLicenseRecord.objects.create(
            sheet=sheet,
            source_sheet_name=sheet.sheet_name,
            source_row=source_row,
            record_type=record_type,
            **self._practice_record_defaults(**kwargs),
        )
        self.summary["practice_records"] += 1
        return record

    def _identity_key_for_record(self, record):
        registration_no = normalize_registration_no(record.registration_no)
        practitioner_number = normalize_identifier(record.practitioner_number)
        if registration_no:
            return ("registration", registration_no)
        if practitioner_number:
            return ("practitioner", practitioner_number)
        name = normalize_name(record.full_name)
        dob = record.date_of_birth.isoformat() if record.date_of_birth else ""
        return ("name_dob", f"{name}|{dob}") if name and dob else ("name", name)

    def _hydrate_from_current_year_rows(self):
        batch_records = PracticingLicenseRecord.objects.filter(batch=self.batch)
        current_year = (
            batch_records.filter(record_year__isnull=False)
            .order_by("-record_year")
            .values_list("record_year", flat=True)
            .first()
        )
        if not current_year:
            return
        field_names = [
            "gender",
            "date_of_birth",
            "practitioner_number",
            "applicant_type",
            "nationality",
            "qualification_name",
            "category",
            "workplace_address",
            "province",
        ]
        current_profiles = {}
        for record in batch_records.filter(record_year=current_year, record_type="practicing_license"):
            key = self._identity_key_for_record(record)
            if not key[1]:
                continue
            profile = current_profiles.setdefault(key, {})
            for field_name in field_names:
                value = getattr(record, field_name)
                if value not in (None, "") and field_name not in profile:
                    profile[field_name] = value

        updates = 0
        for record in batch_records.exclude(record_year=current_year):
            profile = current_profiles.get(self._identity_key_for_record(record))
            if not profile:
                continue
            update_fields = []
            for field_name, value in profile.items():
                if getattr(record, field_name) in (None, "") and value not in (None, ""):
                    setattr(record, field_name, value)
                    update_fields.append(field_name)
            if update_fields:
                record.save(update_fields=update_fields + ["updated_at"])
                updates += 1
        self.summary["current_year_field_updates"] = updates

    def _model_for_target(self, target_model):
        return self._model_map.get(target_model, NursingProfessional)

    def _content_type_for_model(self, model):
        if model not in self._application_ct_cache:
            self._application_ct_cache[model] = ContentType.objects.get_for_model(model)
        return self._application_ct_cache[model]

    def _update_professional(self, target_model, registration_no, full_name, **kwargs):
        if not registration_no:
            return None, False

        model = self._model_for_target(target_model)
        cache = self._professional_cache[target_model]
        professional = cache.get(registration_no)
        created = False
        if professional is None:
            first_name, last_name = split_name(full_name)
            defaults = {
                "registration_no": registration_no,
                "first_name": first_name[:100],
                "last_name": last_name[:100],
            }
            professional = model(**defaults)
            created = True

        qualification_name = kwargs.pop("qualification_name", "")
        institution_name = kwargs.pop("institution_name", "")
        payment_date = kwargs.pop("payment_date", None)
        record_year = kwargs.pop("record_year", None)
        create_qualification = kwargs.pop("create_qualification", True)

        for field_name, value in kwargs.items():
            if value in (None, ""):
                continue
            if field_name == "registration_number":
                value = safe_secondary_identifier(value)
                if not value:
                    continue
                if model.objects.exclude(pk=professional.pk).filter(registration_number=value).exists():
                    continue
            current = getattr(professional, field_name, None)
            if current in (None, "") or (
                isinstance(value, date) and isinstance(current, date) and value > current
            ):
                setattr(professional, field_name, value)

        if hasattr(professional, "qualification_level") and qualification_name:
            if not getattr(professional, "qualification_level", ""):
                professional.qualification_level = qualification_name[:100]

        if payment_date and hasattr(professional, "license_expiry_date"):
            expiry_date = date(record_year or payment_date.year, 12, 31)
            current_expiry = getattr(professional, "license_expiry_date", None)
            if current_expiry is None or expiry_date > current_expiry:
                professional.license_expiry_date = expiry_date

        if institution_name and not getattr(professional, "full_address", ""):
            professional.full_address = institution_name[:1000]

        professional.save()
        cache[registration_no] = professional

        counter_key = target_model
        if created:
            self.summary["records_created"][counter_key] += 1
        else:
            self.summary["records_updated"][counter_key] += 1

        if create_qualification and qualification_name and target_model in {"nursingprofessional", "midwife"}:
            institution = safe_institution(institution_name)
            Qualification.objects.update_or_create(
                content_type=self._content_type_for_model(model),
                object_id=professional.id,
                qualification_name=qualification_name[:200],
                defaults={
                    "institution": institution,
                    "institution_name": normalize_text(institution_name)[:255],
                    "program_completed": qualification_name[:255],
                    "completion_year": record_year,
                    "qualification_type": "Imported Workbook",
                },
            )

        return professional, created

    def _upsert_application(self, professional, form_code, form_title, issued_date, profession_track, payload):
        if not professional:
            return
        model = professional.__class__
        expiry_date = None
        if form_code in {"NC2", "NC3", "NC5", "NC8"} and issued_date:
            expiry_date = date(issued_date.year, 12, 31)
        elif form_code == "NC1" and issued_date:
            expiry_date = issued_date.replace(year=issued_date.year)  # keep same date if unknown duration

        Application.objects.update_or_create(
            content_type=self._content_type_for_model(model),
            object_id=professional.id,
            form_code=form_code,
            defaults={
                "form_title": form_title,
                "pathway": "overseas_nurse" if professional.applicant_type == "overseas" else "local_nursing_graduate",
                "profession_track": profession_track,
                "status": "approved",
                "submitted_date": issued_date or date.today(),
                "approved_date": issued_date,
                "expiry_date": expiry_date,
                "payload": payload,
                "reviewer_notes": f"Imported from {self.workbook_path.name} ({payload.get('source_sheet', '')}).",
            },
        )

    def _upsert_receipt(self, full_name, practitioner_number, payment_date, amount, reference_number, description, payment_method=""):
        if not payment_date or amount is None:
            return
        receipt_datetime = timezone.make_aware(datetime.combine(payment_date, datetime.min.time()))
        if reference_number:
            receipt, created = Receipt.objects.update_or_create(
                official_receipt_no=reference_number[:50],
                defaults={
                    "amount": amount,
                    "description": description[:1000],
                    "status": "completed",
                    "receipt_date": receipt_datetime,
                    "payment_method": (payment_method or "imported")[:50],
                    "officer_receiving": full_name[:255],
                    "practitioner_number": practitioner_number[:100],
                },
            )
        else:
            existing = Receipt.objects.filter(
                practitioner_number=practitioner_number[:100],
                amount=amount,
                receipt_date=receipt_datetime,
            ).first()
            if existing:
                receipt = existing
                created = False
            else:
                receipt = Receipt.objects.create(
                    amount=amount,
                    description=description[:1000],
                    status="completed",
                    receipt_date=receipt_datetime,
                    payment_method=(payment_method or "imported")[:50],
                    officer_receiving=full_name[:255],
                    practitioner_number=practitioner_number[:100],
                )
                created = True
        if created:
            self.summary["receipts"] += 1

    def _import_provisional_sheet(self, ws, sheet):
        imported = 0
        skipped = 0
        for idx, row in self._iter_value_rows(ws, min_row=5, max_col=9):
            source_id, name, registration_label, prefix, provisional_no, issued_raw, institution, year_raw, qualification = row
            full_name = normalize_name(name)
            if not full_name or "NAME" == full_name.upper():
                skipped += 1
                continue
            issued_date = parse_date(issued_raw)
            record_year = parse_year(year_raw, issued_date)
            registration_no = normalize_provisional_registration(provisional_no, prefix or "PRO")
            qualification_name = normalize_text(qualification)
            target_model = infer_target_model(
                category=normalize_text(registration_label) or qualification_name,
                qualification=qualification_name,
                registration_no=registration_no,
            )
            record = self._create_practice_record(
                sheet,
                idx,
                "provisional",
                target_model=target_model,
                record_year=record_year,
                full_name=full_name,
                first_name=split_name(full_name)[0],
                last_name=split_name(full_name)[1],
                registration_no=registration_no,
                qualification_name=qualification_name,
                institution_name=normalize_text(institution),
                issued_date=issued_date,
                category=normalize_text(registration_label) or "Provisional",
                applicant_type=infer_applicant_type(institution),
                raw_payload={
                    "source_id": source_id,
                    "registration_label": normalize_text(registration_label),
                    "prefix": normalize_text(prefix),
                    "provisional_no": normalize_text(provisional_no),
                },
            )
            if record is None:
                skipped += 1
                continue
            imported += 1
        return imported, skipped, "Imported provisional records into practice-history store."

    def _import_full_sheet(self, ws, sheet):
        imported = 0
        skipped = 0
        seen = set()
        for idx, row in self._iter_value_rows(ws, min_row=5, max_col=10):
            source_id, name, license_label, prefix, license_no, issued_raw, institution, year_raw, qualification, practitioner_no = row
            full_name = normalize_name(name)
            if not full_name:
                skipped += 1
                continue
            registration_no = normalize_registration_no(license_no, prefix)
            if not registration_no:
                skipped += 1
                continue
            dedupe_key = (registration_no, full_name)
            if dedupe_key in seen:
                skipped += 1
                continue
            seen.add(dedupe_key)

            issued_date = parse_date(issued_raw)
            record_year = parse_year(year_raw, issued_date)
            category = normalize_text(qualification)
            target_model = infer_target_model(category=category, qualification=category, registration_no=registration_no)
            practitioner_number = normalize_identifier(practitioner_no)
            applicant_type = infer_applicant_type(institution, qualification)

            record = self._create_practice_record(
                sheet,
                idx,
                "full",
                target_model=target_model,
                record_year=record_year,
                full_name=full_name,
                first_name=split_name(full_name)[0],
                last_name=split_name(full_name)[1],
                registration_no=registration_no,
                practitioner_number=practitioner_number,
                qualification_name=category,
                institution_name=normalize_text(institution),
                issued_date=issued_date,
                category=normalize_text(license_label) or category,
                applicant_type=applicant_type,
                raw_payload={"source_id": source_id},
            )
            if record is None:
                skipped += 1
                continue

            model_reg = registration_for_model(registration_no)
            if self.sync_live_profiles and model_reg:
                professional, _ = self._update_professional(
                    target_model,
                    model_reg,
                    full_name,
                    qualification_name=qualification_for_target(target_model, category, category),
                    institution_name=normalize_text(institution),
                    applicant_type=applicant_type,
                    registration_number=practitioner_number[:50] if len(practitioner_number) <= 50 else "",
                    date_issued=issued_date,
                    record_year=record_year,
                )
                self._upsert_application(
                    professional,
                    "NC2",
                    "Application for Full Licence",
                    issued_date,
                    "midwifery" if target_model == "midwife" else "nursing",
                    {"source_sheet": sheet.sheet_name, "source_id": source_id, "registration_no": registration_no},
                )
            imported += 1
        return imported, skipped, "Imported full-registration rows and synchronized professionals."

    def _import_temporary_sheet(self, ws, sheet):
        start_row = 2 if ws.title == "TEMP CERT 2020" else 5
        imported = 0
        skipped = 0
        for idx, row in self._iter_value_rows(ws, min_row=start_row, max_col=8):
            name, registration_label, prefix, license_no, issued_raw, institution, year_raw, qualification = row[:8]
            full_name = normalize_name(name)
            if not full_name or full_name.upper() == "NAME":
                skipped += 1
                continue
            registration_no = normalize_registration_no(license_no, prefix or "TEM")
            if not registration_no:
                skipped += 1
                continue
            issued_date = parse_date(issued_raw)
            record_year = parse_year(year_raw, issued_date)
            qualification_name = normalize_text(qualification)
            applicant_type = infer_applicant_type(institution, qualification, registration_label)
            target_model = infer_target_model(category="Temporary Nurse", qualification=qualification_name, registration_no=registration_no)

            record = self._create_practice_record(
                sheet,
                idx,
                "temporary",
                target_model=target_model,
                record_year=record_year,
                full_name=full_name,
                first_name=split_name(full_name)[0],
                last_name=split_name(full_name)[1],
                registration_no=registration_no,
                qualification_name=qualification_name,
                institution_name=normalize_text(institution),
                issued_date=issued_date,
                category=normalize_text(registration_label) or "Temporary Certificate",
                applicant_type=applicant_type,
                raw_payload={"prefix": prefix},
            )
            if record is None:
                skipped += 1
                continue

            model_reg = registration_for_model(registration_no)
            if self.sync_live_profiles and model_reg:
                professional, _ = self._update_professional(
                    target_model,
                    model_reg,
                    full_name,
                    qualification_name=qualification_for_target(target_model, qualification_name, qualification_name),
                    institution_name=normalize_text(institution),
                    applicant_type=applicant_type,
                    date_issued=issued_date,
                    record_year=record_year,
                )
                self._upsert_application(
                    professional,
                    "NC8",
                    "Application for Temporary Licence",
                    issued_date,
                    "temporary",
                    {"source_sheet": sheet.sheet_name, "registration_no": registration_no},
                )
            imported += 1
        return imported, skipped, "Imported temporary certificates and synchronized professionals."

    def _import_atp_record_sheet(self, ws, sheet):
        start_row = 4 if "2022" in ws.title else 3
        imported = 0
        skipped = 0
        payment_year = re.search(r"(20\d{2})", ws.title)
        sheet_year = int(payment_year.group(1)) if payment_year else None

        for idx, row in self._iter_value_rows(ws, min_row=start_row, max_col=16):
            values = list(row) + [None] * (16 - len(row))
            name = values[1]
            full_name = normalize_name(name)
            if not full_name:
                skipped += 1
                continue

            gender = normalize_text(values[2])
            dob = parse_date(values[3])
            registration_no = normalize_registration_no(values[4])
            practitioner_number = normalize_identifier(values[5])
            category = normalize_text(values[6])
            qualification_name = normalize_text(values[7]) if "2026" in ws.title else ""
            country = normalize_text(values[8] if "2026" in ws.title else values[7])
            workplace = normalize_text(values[9] if "2026" in ws.title else values[8])
            province = normalize_text(values[10] if "2026" in ws.title else values[9])
            reference_number = ""
            if "2026" in ws.title:
                payment_date = parse_date(values[11])
                renewal_fee = parse_decimal(values[12])
                overseas_fee = parse_decimal(values[13])
                late_fee = parse_decimal(values[14])
                payment_method = normalize_text(values[15])
                if payment_method.upper().startswith("R"):
                    reference_number = payment_method
            else:
                payment_date = parse_date(values[10])
                renewal_fee = parse_decimal(values[11])
                late_fee = parse_decimal(values[12])
                overseas_fee = parse_decimal(values[13])
                payment_method = normalize_text(values[14] if "2022" in ws.title else values[15] if "2024" in ws.title else "")
                if payment_method.upper().startswith("R"):
                    reference_number = payment_method
            record_year = parse_year(sheet_year, payment_date) or sheet_year
            applicant_type = infer_applicant_type(country, workplace)
            target_model = infer_target_model(category=category, qualification=qualification_name, registration_no=registration_no)

            record = self._create_practice_record(
                sheet,
                idx,
                "practicing_license",
                target_model=target_model,
                record_year=record_year,
                full_name=full_name,
                first_name=split_name(full_name)[0],
                last_name=split_name(full_name)[1],
                gender=gender,
                date_of_birth=dob,
                registration_no=registration_no,
                practitioner_number=practitioner_number,
                applicant_type=applicant_type,
                nationality=country,
                qualification_name=qualification_name,
                category=category,
                workplace_address=workplace,
                province=province,
                payment_date=payment_date,
                renewal_fee=renewal_fee,
                overseas_fee=overseas_fee,
                late_fee=late_fee,
                reference_number=reference_number,
                payment_method=payment_method,
                raw_payload={"sheet_year": sheet_year},
            )
            if record is None:
                skipped += 1
                continue

            model_reg = registration_for_model(registration_no)
            if self.sync_live_profiles and model_reg:
                professional, _ = self._update_professional(
                    target_model,
                    model_reg,
                    full_name,
                    qualification_name=qualification_for_target(target_model, category, qualification_name or category),
                    institution_name=workplace,
                    applicant_type=applicant_type,
                    gender=gender if gender in {"Male", "Female"} else "",
                    date_of_birth=dob,
                    province=province[:100],
                    full_address=workplace,
                    registration_number=practitioner_number[:50] if len(practitioner_number) <= 50 else "",
                    payment_date=payment_date,
                    record_year=record_year,
                    create_qualification=False,
                )
                description = f"Imported ATP record for {record_year} from {sheet.sheet_name}"
                reference = ""
                if payment_method.startswith("R"):
                    reference = payment_method
                self._upsert_receipt(
                    full_name,
                    practitioner_number,
                    payment_date,
                    renewal_fee or overseas_fee,
                    reference,
                    description,
                    payment_method=payment_method or "ATP record",
                )
            imported += 1
        return imported, skipped, "Imported ATP practicing-license records."

    def _import_atp_payment_sheet(self, ws, sheet):
        if "2020-2021 payment" in ws.title:
            start_row = 3
        else:
            start_row = 2

        imported = 0
        skipped = 0
        sheet_year_match = re.search(r"(20\d{2})", ws.title)
        sheet_year = int(sheet_year_match.group(1)) if sheet_year_match else None

        for idx, row in self._iter_value_rows(ws, min_row=start_row, max_col=12):
            values = list(row) + [None] * (12 - len(row))
            if "2020-2021 payment" in ws.title:
                full_name = normalize_name(values[1])
                workplace = normalize_text(values[2])
                province = normalize_text(values[3])
                payment_date = parse_date(values[4])
                amount = parse_decimal(values[5])
                overseas_fee = parse_decimal(values[6])
                penalty_fee = parse_decimal(values[7])
                renewal_fee = parse_decimal(values[8])
                reference = normalize_text(values[10])
                category = "ATP Payment"
                registration_no = ""
                practitioner_number = ""
                nationality = "Overseas" if overseas_fee else "PNG"
            else:
                full_name = normalize_name(values[2])
                workplace = normalize_text(values[3])
                province = normalize_text(values[4])
                payment_date = parse_date(values[9] or values[1])
                amount = parse_decimal(values[8])
                overseas_fee = None
                penalty_fee = None
                renewal_fee = amount
                reference = normalize_text(values[7])
                category = normalize_text(values[5]) or "ATP Payment"
                registration_no = ""
                practitioner_number = ""
                nationality = normalize_text(values[6])

            if not full_name:
                skipped += 1
                continue

            record_year = parse_year(sheet_year, payment_date) or sheet_year
            applicant_type = infer_applicant_type(nationality, workplace)

            record = self._create_practice_record(
                sheet,
                idx,
                "payment",
                target_model="other",
                record_year=record_year,
                full_name=full_name,
                first_name=split_name(full_name)[0],
                last_name=split_name(full_name)[1],
                registration_no=registration_no,
                practitioner_number=practitioner_number,
                applicant_type=applicant_type,
                nationality=nationality,
                category=category,
                workplace_address=workplace,
                province=province,
                payment_date=payment_date,
                amount=amount,
                renewal_fee=renewal_fee,
                overseas_fee=overseas_fee,
                penalty_fee=penalty_fee,
                reference_number=reference,
                payment_method="Imported ATP Payment",
                raw_payload={"sheet_year": sheet_year},
            )
            if record is None:
                skipped += 1
                continue
            self._upsert_receipt(
                full_name,
                practitioner_number,
                payment_date,
                amount or renewal_fee or overseas_fee,
                reference,
                f"Imported payment row from {sheet.sheet_name}",
                payment_method="Imported ATP Payment",
            )
            imported += 1

        return imported, skipped, "Imported ATP payment rows and receipt records."

    def _import_listing_sheet(self, ws, sheet):
        sheet_key = ws.title.lower()
        imported = 0
        skipped = 0

        if ws.title == "Sheet2":
            iterator = self._iter_value_rows(ws, min_row=2, max_col=9)
        elif "active gnurse" in sheet_key:
            iterator = self._iter_value_rows(ws, min_row=4, max_col=12)
        elif "midwifery list" in sheet_key:
            iterator = self._iter_value_rows(ws, min_row=5, max_col=9)
        elif "pom gen. 2020" in sheet_key:
            iterator = self._iter_value_rows(ws, min_row=4, max_col=7)
        elif "update list 2020" in sheet_key:
            iterator = self._iter_value_rows(ws, min_row=3, max_col=5)
        else:
            iterator = self._iter_value_rows(ws, min_row=2, max_col=7)

        for idx, row in iterator:
            values = list(row)
            full_name = ""
            registration_no = ""
            practitioner_number = ""
            category = ""
            workplace = ""
            province = ""
            payment_date = None
            qualification = ""
            gender = ""
            dob = None
            target_model = "nursingprofessional"
            record_year = None

            if ws.title == "Sheet2":
                _, name, gender, dob_raw, reg_no, practitioner_no, category, _, workplace = values[:9]
                full_name = normalize_name(name)
                registration_no = normalize_registration_no(reg_no)
                practitioner_number = normalize_identifier(practitioner_no)
                category = normalize_text(category)
                workplace = normalize_text(workplace)
                dob = parse_date(dob_raw)
                target_model = infer_target_model(category=category, registration_no=registration_no)
                province = "Manus"
            elif "active gnurse" in sheet_key:
                _, surname_like, name_like, _, _, employer, address, province_raw, _, _, marital, dob_raw = values[:12]
                full_name = normalize_name(name_like or surname_like)
                workplace = normalize_text(employer)
                province = normalize_text(province_raw)
                category = "General Nurse"
                gender = ""
                dob = parse_date(dob_raw)
                target_model = "nursingprofessional"
                record_year = 2013
            elif "midwifery list" in sheet_key:
                license_no, name, prefix, number, issued_raw, course, _, _, _ = values[:9]
                full_name = normalize_name(name)
                registration_no = normalize_registration_no(number, prefix)
                category = "Midwife"
                course_text = normalize_text(course)
                qualification = course_text
                issued_text = normalize_text(issued_raw)
                embedded_date_patterns = [
                    r"(\d{1,2}\s*[./-]\s*\d{1,2}\s*[./-]\s*\d{2,4})",
                    r"(\d{1,2}\s+[A-Za-z]+\s*-\s*\d{2,4})",
                ]
                payment_date = None
                for pattern in embedded_date_patterns:
                    embedded_date_match = re.search(pattern, course_text)
                    if not embedded_date_match:
                        continue
                    embedded_date = embedded_date_match.group(1)
                    payment_date = parse_date(embedded_date)
                    if payment_date:
                        qualification = normalize_text(course_text.replace(embedded_date, "")).strip("- ").strip()
                        break
                if not payment_date and " - " in course_text:
                    issued_piece, qualification_piece = course_text.split(" - ", 1)
                    parsed_embedded_date = parse_date(issued_piece)
                    if parsed_embedded_date:
                        payment_date = parsed_embedded_date
                        qualification = qualification_piece
                    else:
                        payment_date = parse_date(issued_text)
                elif not payment_date:
                    payment_date = parse_date(issued_text)
                target_model = "midwife"
                record_year = parse_year(payment_date, payment_date) or 2014
            elif "pom gen. 2020" in sheet_key:
                _, _, name, hospital, province_raw, payment_raw, amount = values[:7]
                full_name = normalize_name(name)
                workplace = normalize_text(hospital)
                province = normalize_text(province_raw)
                payment_date = parse_date(payment_raw)
                category = "General Nurse"
                target_model = "nursingprofessional"
                record_year = parse_year(payment_raw, payment_date) or 2020
                amount = parse_decimal(amount)
            elif "update list 2020" in sheet_key:
                particulars, first_name, practitioner_no, reg_no, _ = values[:5]
                full_name = normalize_name(particulars or first_name)
                practitioner_number = normalize_identifier(practitioner_no)
                registration_no = normalize_identifier(reg_no)
                category = "Estimated Nursing Data"
                target_model = infer_target_model(category=category, registration_no=registration_no)
                record_year = 2020
            else:  # print 2022 and similar
                practitioner_no, name, workplace, province_raw, payment_raw, *_rest = values + [None] * (7 - len(values))
                full_name = normalize_name(name)
                practitioner_number = normalize_identifier(practitioner_no)
                province = normalize_text(province_raw)
                workplace = normalize_text(workplace)
                payment_date = parse_date(payment_raw)
                category = "General Nurse"
                target_model = "nursingprofessional"
                record_year = parse_year(payment_raw, payment_date) or 2022

            if not full_name:
                skipped += 1
                continue

            applicant_type = infer_applicant_type(workplace, qualification)
            if record_year is None:
                record_year = parse_year(payment_date, payment_date)

            record = self._create_practice_record(
                sheet,
                idx,
                "workforce_listing",
                target_model=target_model,
                record_year=record_year,
                full_name=full_name,
                first_name=split_name(full_name)[0],
                last_name=split_name(full_name)[1],
                gender=gender,
                date_of_birth=dob,
                registration_no=registration_no,
                practitioner_number=practitioner_number,
                applicant_type=applicant_type,
                qualification_name=qualification,
                category=category,
                workplace_address=workplace,
                province=province,
                payment_date=payment_date,
                raw_payload={},
            )
            if record is None:
                skipped += 1
                continue

            model_reg = registration_for_model(registration_no)
            if self.sync_live_profiles and model_reg:
                self._update_professional(
                    target_model,
                    model_reg,
                    full_name,
                    qualification_name=qualification_for_target(target_model, category, qualification),
                    institution_name=workplace,
                    applicant_type=applicant_type,
                    gender=gender if gender in {"Male", "Female"} else "",
                    date_of_birth=dob,
                    province=province[:100],
                    full_address=workplace,
                    registration_number=practitioner_number[:50] if len(practitioner_number) <= 50 else "",
                    payment_date=payment_date,
                    record_year=record_year,
                    create_qualification=False,
                )
            imported += 1

        return imported, skipped, "Imported workforce listing rows."

    def _sync_snapshots(self):
        WorkforceSnapshot.objects.filter(year__gt=date.today().year).delete()
        batch_records = PracticingLicenseRecord.objects.filter(
            batch=self.batch,
            record_year__isnull=False,
            record_year__lte=date.today().year,
        )
        years = unique_preserve_order(
            sorted(batch_records.values_list("record_year", flat=True).distinct())
        )
        for year_value in years:
            year_records = list(batch_records.filter(record_year=year_value))
            if not year_records:
                continue

            active_registrations = {
                record.registration_no
                for record in year_records
                if record.registration_no
            }

            def count_target(target_model):
                return len({
                    record.registration_no or record.full_name
                    for record in year_records
                    if record.target_model == target_model
                })

            near_retirement = 0
            for record in year_records:
                if record.date_of_birth and (year_value - record.date_of_birth.year) >= 55:
                    near_retirement += 1

            WorkforceSnapshot.objects.update_or_create(
                year=year_value,
                defaults={
                    "total_active_workers": len(active_registrations),
                    "total_nurses": count_target("nursingprofessional"),
                    "total_doctors": count_target("medicaldoctor"),
                    "total_midwives": count_target("midwife"),
                    "total_chw": count_target("communityhealthworker"),
                    "new_registrations": len({
                        record.registration_no or record.full_name
                        for record in year_records
                        if record.record_type in {"full_approved", "temporary"}
                    }),
                    "renewals": len({
                        record.registration_no or record.full_name
                        for record in year_records
                        if record.record_type in {"practicing_license", "payment"}
                    }),
                    "retirements": 0,
                    "new_graduates_joined": len({
                        record.registration_no or record.full_name
                        for record in year_records
                        if record.record_type == "provisional"
                    }),
                    "nearing_retirement": near_retirement,
                },
            )
