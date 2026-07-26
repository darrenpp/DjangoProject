from datetime import date
from io import BytesIO
import os
from tempfile import NamedTemporaryFile, TemporaryDirectory

from django import forms
from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook
from rest_framework.test import APIClient

from apps.competency.models import CompetencyAssessment
from apps.dashboard.models import Receipt
from apps.documents.models import Document
from apps.workforce.models import (
    Application,
    ApplicationChecklistItem,
    ApplicationFormResponse,
    AuditLog,
    DeceasedNotification,
    EmployerVerificationRequest,
    ApplicationPathway,
    Cadre,
    DeclarationTemplate,
    DocumentRequirement,
    DynamicFormDefinition,
    FeeSchedule,
    CommunityHealthWorker,
    DataImportBatch,
    DocumentType,
    MedicalDoctor,
    Midwife,
    NurseAide,
    NursingProfessional,
    IssuedLicenceDocument,
    PracticingLicenseRecord,
    RegulatoryBody,
    SupervisorAssignment,
    TrainingInstitution,
)
from apps.notifications.models import EnquiryMessageAttachment, EnquiryThread, Notification
from apps.workforce.forms import (
    MedicalBoardAccreditationChecklistForm,
    MedicalBoardChwRegistrationForm,
    MedicalBoardPrivateHealthFacilityChecklistForm,
    MedicalBoardRenewalRegistrationForm,
    MedicalBoardSpecialistApplicationForm,
    MedicalBoardTrainingCollegeFacilityForm,
)
from apps.workforce.services.nursing_council_workflows import (
    NURSING_COUNCIL_CODE,
    NursingCouncilValidationService,
    approve_deceased_notification,
    approve_nursing_application,
    build_nursing_workflow_rows,
    build_public_form_guide,
    complete_supervisor_competency,
    create_deceased_notification,
    create_employer_verification_request,
    create_supervisor_assignment,
    ensure_nursing_council_configuration,
    generate_application_checklist,
    search_public_nursing_register,
)
from apps.workforce.services.licence_issuance import issue_application_licence_document
from apps.workforce.services.ai_import_cleanser import cleanse_import_row
from apps.workforce.services.medical_board_workbook_import import (
    MedicalBoardWorkbookImporter,
    is_medical_board_chw_workbook,
)
from apps.workforce.services.medical_board_legacy_import import MedicalBoardLegacyWorkbookImporter


def _medical_board_chw_workbook_bytes():
    workbook = Workbook()
    chw_sheet = workbook.active
    chw_sheet.title = "CHW"
    chw_sheet.append(["CERT #", "Date", "Name", "Address", "Qualifications", "Receipt No.", "Remarks"])
    chw_sheet.append([
        "1",
        "05.09.85",
        "RURI, YEME",
        "BUNAPAS HEALTH CENTRE BOX 1080, MADANG PROVINCE",
        "APO CERT (GAUBIN) 1976",
        "AR-5901",
        "P#: 016835 APPROVED: 05.09.85",
    ])

    atp_sheet = workbook.create_sheet("ATP DATABASE ONLY")
    atp_sheet.append(["CERT #", "Date", "Name", "ARCH#", "2024", None])
    atp_sheet.append(["1", "05.09.85", "RURI, YEME", "CHW01", "05.03.2024 K15", "R00001699968"])

    pending_sheet = workbook.create_sheet("NOT ON DATABASE CHW")
    pending_sheet.append(["Registry #", "Date", "Name", "Address", "Qualifications\\DESIGNATION", "Receipt No.", "Remarks"])
    pending_sheet.append([None, "13.04.2018", "LUCAS BALPHINA", "KUNDIAWA", "CERTIFICATE", "G548197", "STUDENT ID-160043"])

    school_sheet = workbook.create_sheet("SCHOOL ADDRESS")
    school_sheet.append(["SCHOOL ADDRESS"])
    school_sheet.append([])
    school_sheet.append([None, "SCHOOLS", "PROVINCES", "ADDRESSES"])
    school_sheet.append(["1", "RUMGINAE", "WESTERN", "RUMGINAE CHWTS P O BOX 34, KIUNGA"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _medical_board_legacy_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ALL OVERSESA MEMBERS MB"
    sheet.append([
        "No. ",
        "Doctors Name",
        "Nationality",
        "ARC#",
        "Qualification/DESIGNATION",
        "App for",
        "Year",
        "Employer",
        "ATP / Full R C",
        "ATP DATE",
        "RECIEPT#",
        "PROV MLT#/NOS",
        "SPEC#",
        "Medical Practioners#",
        "REMARKS",
    ])
    sheet.append(["1", "SIOK TAN", "SINGAPORE", "MB.01", "MBBS", None, "2025", "YWAM", "FULL", "24.06.2025", "R00001940233", None, None, "134", None])
    sheet.append(["2", "HEE KYUN KOOK", "KOREA", "MB.01", "PARAMEDIC", None, "2025", "GLOBAL MISSION", None, None, None, None, None, None, "NO ADDITIONAL INFORMATION"])

    eho_sheet = workbook.create_sheet("EHO")
    eho_sheet.append([])
    eho_sheet.append([])
    eho_sheet.append([None, "Officer Name ", "MB Rego CARD# ", "ATP / FRC", "Registration Date", "LAST ATP Year", "Practisioners #", "Qualification", "Province", "Reciept No#", "REMARKS"])
    eho_sheet.append(["1", "Aaron Gwamatae", "301", None, "1990", "2006", "P#022372", "Diploma in Health Science", "MADANG", "RG5181", "Last ATP01.08.2006 Sect 15"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class ProfessionalDetailAccessTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.medical_registrar = user_model.objects.create_user(
            username="medical_registrar",
            password="StrongPass123!",
            role="registrar",
            department="Medical Board",
        )
        self.nursing_registrar = user_model.objects.create_user(
            username="nursing_registrar",
            password="StrongPass123!",
            role="registrar",
            department="Nursing Council",
        )
        self.professional = NursingProfessional.objects.create(
            first_name="Apiasets",
            last_name="Nurse",
            registration_no="NC-23228",
            email="apiasets@example.test",
        )

    def test_records_list_uses_model_aware_detail_link_for_professionals(self):
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse("record_list", args=["nursingprofessional"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            reverse("record_detail", args=["nursingprofessional", self.professional.pk]),
        )
        self.assertNotContains(
            response,
            reverse("professional_detail", args=[self.professional.pk]),
        )

    def test_existing_professional_detail_url_allows_records_hub_staff(self):
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse("professional_detail", args=[self.professional.pk]))

        self.assertEqual(response.status_code, 200)

    def test_medical_board_staff_cannot_access_nursing_professional_records(self):
        self.client.force_login(self.medical_registrar)

        list_response = self.client.get(reverse("record_list", args=["nursingprofessional"]))
        detail_response = self.client.get(reverse("professional_detail", args=[self.professional.pk]))

        self.assertEqual(list_response.status_code, 404)
        self.assertEqual(detail_response.status_code, 404)


class RecordHubOfficeScopeTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.medical_registrar = user_model.objects.create_user(
            username="medical_board_registrar",
            password="StrongPass123!",
            role="registrar",
            department="Medical Board",
        )
        self.nursing_registrar = user_model.objects.create_user(
            username="nursing_council_registrar",
            password="StrongPass123!",
            role="registrar",
            department="Nursing Council",
        )
        self.nurse = NursingProfessional.objects.create(
            first_name="Nora",
            last_name="Nurse",
            registration_no="NC-100",
            email="nora@example.test",
        )
        self.midwife = Midwife.objects.create(
            first_name="Mira",
            last_name="Midwife",
            registration_no="MW-100",
            email="mira@example.test",
        )
        self.nurse_aide = NurseAide.objects.create(
            first_name="Ari",
            last_name="Aide",
            registration_no="NA-100",
            email="ari@example.test",
        )
        self.doctor = MedicalDoctor.objects.create(
            first_name="Dina",
            last_name="Doctor",
            registration_no="MD-100",
            email="dina@example.test",
        )
        self.chw = CommunityHealthWorker.objects.create(
            first_name="Chris",
            last_name="Worker",
            registration_no="CHW-100",
            email="chris@example.test",
        )
        self.nursing_application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(self.nurse),
            object_id=self.nurse.pk,
            form_code="NC3",
            pathway="other",
            form_title="Nursing renewal",
        )
        self.medical_application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(self.doctor),
            object_id=self.doctor.pk,
            form_code="MD2",
            pathway="medical_board",
            form_title="Medical renewal",
        )
        self.nursing_receipt = Receipt.objects.create(
            user=self.nursing_registrar,
            application=self.nursing_application,
            receipt_number="NC-RCT-100",
            amount="50.00",
            status="completed",
        )
        self.medical_receipt = Receipt.objects.create(
            user=self.medical_registrar,
            application=self.medical_application,
            receipt_number="MB-RCT-100",
            amount="75.00",
            status="completed",
        )
        self.imported_nursing_receipt = Receipt.objects.create(
            receipt_number="NC-IMPORT-100",
            amount="25.00",
            description="N-DATA imported receipt",
            status="completed",
        )
        self.system_admin = user_model.objects.create_user(
            username="system_admin",
            password="StrongPass123!",
            role="admin",
            is_superuser=True,
            is_staff=True,
        )

    def test_medical_board_records_hub_excludes_nursing_record_types(self):
        self.client.force_login(self.medical_registrar)

        response = self.client.get(reverse("records_home"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Medical Board Records Hub")
        self.assertContains(response, "Medical Doctors")
        self.assertContains(response, "Community Health Workers")
        self.assertNotContains(response, "Nursing Professionals")
        self.assertNotContains(response, "Nurse Aides")
        self.assertNotContains(response, "Midwifes")
        visible_slugs = {item["slug"] for item in response.context["models"]}
        self.assertNotIn("user", visible_slugs)
        self.assertNotIn("report", visible_slugs)
        self.assertNotIn("ocrdocument", visible_slugs)

    def test_nursing_council_records_hub_excludes_medical_record_types(self):
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse("records_home"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nursing Council Records Hub")
        self.assertContains(response, "Nursing Professionals")
        self.assertContains(response, "Nurse Aides")
        self.assertContains(response, "Midwifes")
        self.assertNotContains(response, "Medical Doctors")
        self.assertNotContains(response, "Community Health Workers")
        visible_slugs = {item["slug"] for item in response.context["models"]}
        self.assertNotIn("user", visible_slugs)
        self.assertNotIn("report", visible_slugs)
        self.assertNotIn("ocrdocument", visible_slugs)

    def test_office_registrar_can_delete_scoped_professional_records(self):
        self.client.force_login(self.medical_registrar)

        response = self.client.post(reverse("record_delete", args=["medicaldoctor", self.doctor.pk]))

        self.assertRedirects(response, reverse("record_list", args=["medicaldoctor"]))
        self.assertFalse(MedicalDoctor.objects.filter(pk=self.doctor.pk).exists())

    def test_nursing_professionals_list_uses_server_side_datatable_and_crud_actions(self):
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse("record_list", args=["nursingprofessional"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="recordsHubTable"')
        self.assertContains(response, 'data-server-side="1"')
        self.assertContains(response, 'id="recordsTableFilterForm"')
        self.assertContains(response, 'id="recordsPageLength"')
        self.assertContains(response, 'id="recordsGlobalSearch"')
        self.assertContains(response, "First")
        self.assertContains(response, "Previous")
        self.assertContains(response, "Next")
        self.assertContains(response, "Last")
        self.assertContains(response, "Showing 1-1 of 1")
        self.assertContains(response, reverse("record_list_data", args=["nursingprofessional"]))
        self.assertContains(response, reverse("record_create", args=["nursingprofessional"]))
        self.assertContains(response, reverse("record_detail", args=["nursingprofessional", self.nurse.pk]))
        self.assertContains(response, reverse("record_update", args=["nursingprofessional", self.nurse.pk]))
        self.assertContains(response, reverse("record_delete", args=["nursingprofessional", self.nurse.pk]))

    def test_nursing_professionals_datatable_endpoint_searches_sorts_and_returns_crud_actions(self):
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse("record_list_data", args=["nursingprofessional"]), {
            "draw": "3",
            "start": "0",
            "length": "25",
            "search[value]": "Nora",
            "order[0][column]": "3",
            "order[0][dir]": "asc",
            "columns[3][name]": "first_name",
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["draw"], 3)
        self.assertEqual(payload["recordsFiltered"], 1)
        self.assertEqual(payload["data"][0]["first_name"], "Nora")
        self.assertIn(reverse("record_detail", args=["nursingprofessional", self.nurse.pk]), payload["data"][0]["actions"])
        self.assertIn(reverse("record_update", args=["nursingprofessional", self.nurse.pk]), payload["data"][0]["actions"])
        self.assertIn(reverse("record_delete", args=["nursingprofessional", self.nurse.pk]), payload["data"][0]["actions"])

    def test_nursing_professionals_list_keeps_page_size_in_pagination_links(self):
        for index in range(30):
            NursingProfessional.objects.create(
                first_name=f"Paged {index}",
                last_name="Nurse",
                registration_no=f"NC-PAGE-{index}",
            )
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse("record_list", args=["nursingprofessional"]), {"per_page": "25"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Showing 1-25 of 31")
        self.assertContains(response, "Page 1 of 2")
        self.assertContains(response, "per_page=25&amp;page=2")

    def test_system_admin_can_use_generic_delete(self):
        doctor = MedicalDoctor.objects.create(
            first_name="Delete",
            last_name="Test",
            registration_no="MD-DELETE",
        )
        self.client.force_login(self.system_admin)

        response = self.client.post(reverse("record_delete", args=["medicaldoctor", doctor.pk]))

        self.assertRedirects(response, reverse("record_list", args=["medicaldoctor"]))
        self.assertFalse(MedicalDoctor.objects.filter(pk=doctor.pk).exists())

    def test_receipt_records_are_separated_by_office(self):
        self.client.force_login(self.medical_registrar)

        medical_response = self.client.get(reverse("record_list", args=["receipt"]))

        self.assertEqual(medical_response.status_code, 200)
        self.assertContains(medical_response, self.medical_receipt.receipt_number)
        self.assertNotContains(medical_response, self.nursing_receipt.receipt_number)
        self.assertNotContains(medical_response, self.imported_nursing_receipt.receipt_number)

        self.client.force_login(self.nursing_registrar)
        nursing_response = self.client.get(reverse("record_list", args=["receipt"]))

        self.assertEqual(nursing_response.status_code, 200)
        self.assertContains(nursing_response, self.nursing_receipt.receipt_number)
        self.assertContains(nursing_response, self.imported_nursing_receipt.receipt_number)
        self.assertNotContains(nursing_response, self.medical_receipt.receipt_number)

    def test_medical_board_can_open_chw_import_record_from_ndata_workbook(self):
        ndata_batch = DataImportBatch.objects.create(
            source_file_name="2026 Current ATP-DATA Statistics & Tracking latest.xlsx",
            source_kind="ndata_workbook",
            status="completed",
        )
        chw_import_record = PracticingLicenseRecord.objects.create(
            batch=ndata_batch,
            source_sheet_name="ATP RECORD 2026",
            source_row=1110,
            record_type="practicing_license",
            target_model="communityhealthworker",
            full_name="Aine Tongamp",
            registration_no="G 7383",
            record_year=2026,
        )

        self.client.force_login(self.medical_registrar)
        detail_response = self.client.get(reverse("record_detail", args=["practicinglicenserecord", chw_import_record.pk]))
        edit_response = self.client.get(reverse("record_update", args=["practicinglicenserecord", chw_import_record.pk]))

        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "Aine Tongamp")
        self.assertEqual(edit_response.status_code, 200)
        self.assertIn(ndata_batch, list(edit_response.context["form"].fields["batch"].queryset))

        self.client.force_login(self.nursing_registrar)
        nursing_response = self.client.get(reverse("record_detail", args=["practicinglicenserecord", chw_import_record.pk]))
        nursing_edit_response = self.client.get(reverse("record_update", args=["practicinglicenserecord", chw_import_record.pk]))

        self.assertEqual(nursing_response.status_code, 200)
        self.assertContains(nursing_response, "Aine Tongamp")
        self.assertNotContains(nursing_response, reverse("record_update", args=["practicinglicenserecord", chw_import_record.pk]))
        self.assertEqual(nursing_edit_response.status_code, 404)

    def test_medical_board_edit_form_uses_medical_cadres_and_specialty_dropdown(self):
        Cadre.objects.create(name="Nursing", category="nursing")
        Cadre.objects.create(name="Midwifery", category="midwifery")
        self.client.force_login(self.medical_registrar)

        response = self.client.get(reverse("record_update", args=["medicaldoctor", self.doctor.pk]))

        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertIsInstance(form.fields["specialty"].widget, forms.Select)
        self.assertIn("public_health", [value for value, _label in form.fields["specialty"].choices])
        cadre_names = set(form.fields["cadre"].queryset.values_list("name", flat=True))
        self.assertIn("Medical Doctor", cadre_names)
        self.assertIn("Medical Specialist", cadre_names)
        self.assertNotIn("Nursing", cadre_names)
        self.assertNotIn("Midwifery", cadre_names)

    def test_nursing_edit_form_does_not_offer_medical_board_cadres(self):
        Cadre.objects.create(name="Medical Doctor", category="medical")
        Cadre.objects.create(name="Community Health Worker", category="chw")
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse("record_update", args=["nursingprofessional", self.nurse.pk]))

        self.assertEqual(response.status_code, 200)
        cadre_names = set(response.context["form"].fields["cadre"].queryset.values_list("name", flat=True))
        self.assertIn("Nursing", cadre_names)
        self.assertNotIn("Medical Doctor", cadre_names)
        self.assertNotIn("Community Health Worker", cadre_names)

    def test_document_type_records_are_separated_by_office(self):
        nursing_doc = DocumentType.objects.create(
            name="Nursing transcript",
            description="Nursing Council document requirement: Transcript",
            is_required=True,
        )
        medical_doc = DocumentType.objects.create(
            name="Medical board qualification",
            description="Medical Board document requirement: Qualification evidence",
            is_required=True,
        )

        self.client.force_login(self.medical_registrar)
        medical_response = self.client.get(reverse("record_list", args=["documenttype"]))

        self.assertEqual(medical_response.status_code, 200)
        self.assertIn(medical_doc, list(medical_response.context["objects"]))
        self.assertNotIn(nursing_doc, list(medical_response.context["objects"]))

        self.client.force_login(self.nursing_registrar)
        nursing_response = self.client.get(reverse("record_list", args=["documenttype"]))

        self.assertEqual(nursing_response.status_code, 200)
        self.assertIn(nursing_doc, list(nursing_response.context["objects"]))
        self.assertNotIn(medical_doc, list(nursing_response.context["objects"]))

    def test_medical_specialist_records_hub_view_is_medical_only(self):
        self.doctor.specialty = "public_health"
        self.doctor.save(update_fields=["specialty"])
        self.client.force_login(self.medical_registrar)

        home_response = self.client.get(reverse("records_home"))
        list_response = self.client.get(reverse("record_list", args=["medicalspecialist"]))

        self.assertEqual(home_response.status_code, 200)
        self.assertContains(home_response, "Medical Specialists")
        self.assertEqual(list_response.status_code, 200)
        self.assertContains(list_response, "Dina")

        self.client.force_login(self.nursing_registrar)
        nursing_response = self.client.get(reverse("record_list", args=["medicalspecialist"]))

        self.assertEqual(nursing_response.status_code, 404)

    def test_import_record_editor_scopes_batches_and_target_models(self):
        medical_batch = DataImportBatch.objects.create(
            source_file_name="medical.xlsx",
            source_kind="medical_board_workbook",
            status="completed",
        )
        nursing_batch = DataImportBatch.objects.create(
            source_file_name="nursing.xlsx",
            source_kind="ndata_workbook",
            status="completed",
        )

        self.client.force_login(self.medical_registrar)
        medical_response = self.client.get(reverse("record_create", args=["practicinglicenserecord"]))

        self.assertEqual(medical_response.status_code, 200)
        self.assertIn(medical_batch, list(medical_response.context["form"].fields["batch"].queryset))
        self.assertNotIn(nursing_batch, list(medical_response.context["form"].fields["batch"].queryset))
        medical_targets = {value for value, _label in medical_response.context["form"].fields["target_model"].choices}
        self.assertIn("medicaldoctor", medical_targets)
        self.assertIn("communityhealthworker", medical_targets)
        self.assertIn("other", medical_targets)
        self.assertNotIn("nursingprofessional", medical_targets)

        self.client.force_login(self.nursing_registrar)
        nursing_response = self.client.get(reverse("record_create", args=["practicinglicenserecord"]))

        self.assertEqual(nursing_response.status_code, 200)
        self.assertIn(nursing_batch, list(nursing_response.context["form"].fields["batch"].queryset))
        self.assertNotIn(medical_batch, list(nursing_response.context["form"].fields["batch"].queryset))
        nursing_targets = {value for value, _label in nursing_response.context["form"].fields["target_model"].choices}
        self.assertIn("nursingprofessional", nursing_targets)
        self.assertNotIn("medicaldoctor", nursing_targets)


class MobileDataCollectionApiTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.medical_registrar = user_model.objects.create_user(
            username="medical_mobile_registrar",
            password="StrongPass123!",
            role="registrar",
            department="Medical Board",
        )
        self.nursing_registrar = user_model.objects.create_user(
            username="nursing_mobile_registrar",
            password="StrongPass123!",
            role="registrar",
            department="Nursing Council",
        )
        self.client = APIClient()

    def test_mobile_bootstrap_is_scoped_to_medical_board_forms(self):
        self.client.force_authenticate(user=self.medical_registrar)

        response = self.client.get(reverse("workforce_api:mobile_bootstrap"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["officer"]["office_scope"], "medical")
        self.assertIn("CHW1", response.data["enabled_forms"])
        self.assertIn("MBSP", response.data["enabled_forms"])
        self.assertNotIn("NC1", response.data["enabled_forms"])

    def test_medical_mobile_user_cannot_sync_nursing_form(self):
        self.client.force_authenticate(user=self.medical_registrar)

        response = self.client.post(
            reverse("workforce_api:mobile_sync_batch"),
            {
                "device_id": "device-1",
                "client_batch_id": "batch-1",
                "records": [{
                    "client_record_id": "nursing-cross-office",
                    "office_scope": "nursing",
                    "form_code": "NC1",
                    "target_model": "healthstudent",
                    "person": {"first_name": "Mary", "last_name": "Kila"},
                }],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["accepted"], [])
        self.assertEqual(len(response.data["rejected"]), 1)
        self.assertFalse(Application.objects.filter(payload__client_record_id="nursing-cross-office").exists())

    def test_mobile_sync_creates_pending_application_and_is_idempotent(self):
        payload = {
            "device_id": "device-2",
            "client_batch_id": "batch-2",
            "app_version": "1.0.0",
            "records": [{
                "client_record_id": "medical-chw-1",
                "office_scope": "medical",
                "form_code": "CHW1",
                "target_model": "communityhealthworker",
                "person": {
                    "first_name": "Ally",
                    "last_name": "Mark",
                    "primary_phone": "70000000",
                },
                "qualification": {"training_level": "CHW"},
                "payload": {"declaration_accepted": True},
            }],
        }
        self.client.force_authenticate(user=self.medical_registrar)

        first_response = self.client.post(reverse("workforce_api:mobile_sync_batch"), payload, format="json")
        second_response = self.client.post(reverse("workforce_api:mobile_sync_batch"), payload, format="json")

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 200)
        application_id = first_response.data["accepted"][0]["server_application_id"]
        self.assertEqual(second_response.data["accepted"][0]["server_application_id"], application_id)
        self.assertEqual(Application.objects.filter(payload__client_record_id="medical-chw-1").count(), 1)
        application = Application.objects.get(pk=application_id)
        self.assertEqual(application.status, "pending")
        self.assertEqual(application.form_code, "CHW1")
        self.assertEqual(application.payload["office_scope"], "medical")
        self.assertTrue(ApplicationFormResponse.objects.filter(application=application, form_code="CHW1").exists())
        self.assertTrue(AuditLog.objects.filter(action="MOBILE_RECORD_SYNCED", entity_id=str(application.pk)).exists())

    def test_mobile_attachment_upload_links_file_to_pending_application(self):
        self.client.force_authenticate(user=self.medical_registrar)
        sync_response = self.client.post(
            reverse("workforce_api:mobile_sync_batch"),
            {
                "records": [{
                    "client_record_id": "medical-attachment-1",
                    "office_scope": "medical",
                    "form_code": "CHW1",
                    "target_model": "communityhealthworker",
                    "person": {"first_name": "Attachment", "last_name": "Test"},
                }]
            },
            format="json",
        )
        application_id = sync_response.data["accepted"][0]["server_application_id"]
        upload = SimpleUploadedFile("id.txt", b"test document", content_type="text/plain")

        response = self.client.post(
            reverse("workforce_api:mobile_attachment_upload"),
            {
                "client_record_id": "medical-attachment-1",
                "server_application_id": application_id,
                "attachment_id": "attachment-1",
                "document_code": "id_document",
                "sha256": "abc123",
                "file": upload,
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 201)
        application = Application.objects.get(pk=application_id)
        self.assertEqual(Document.objects.filter(related_object_id=application_id).count(), 1)
        self.assertEqual(application.payload["mobile_uploaded_attachments"][0]["attachment_id"], "attachment-1")
        self.assertTrue(AuditLog.objects.filter(action="MOBILE_ATTACHMENT_UPLOADED", entity_id=str(application.pk)).exists())


class MedicalBoardWorkbookImportTests(TestCase):
    def test_importer_processes_actual_medical_board_chw_sheet_names(self):
        workbook_bytes = _medical_board_chw_workbook_bytes()
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as workbook_file:
            workbook_file.write(workbook_bytes)
            workbook_path = workbook_file.name

        try:
            self.assertTrue(is_medical_board_chw_workbook(workbook_path))
            batch = MedicalBoardWorkbookImporter(workbook_path=workbook_path).import_workbook()
        finally:
            os.unlink(workbook_path)

        self.assertEqual(batch.source_kind, "medical_board_workbook")
        self.assertEqual(batch.status, "completed")
        self.assertEqual(CommunityHealthWorker.objects.count(), 2)
        self.assertEqual(
            PracticingLicenseRecord.objects.filter(
                batch=batch,
                target_model="communityhealthworker",
                record_type="workforce_listing",
            ).count(),
            2,
        )
        self.assertEqual(
            PracticingLicenseRecord.objects.filter(
                batch=batch,
                target_model="communityhealthworker",
                record_type="practicing_license",
            ).count(),
            1,
        )
        self.assertTrue(TrainingInstitution.objects.filter(name="Rumginae").exists())
        self.assertEqual(batch.summary["pending_chw_imported"], 1)
        self.assertEqual(batch.summary["chw_practicing_licences_imported"], 1)
        self.assertTrue(
            batch.sheets.filter(sheet_name="NOT ON DATABASE CHW", status="processed").exists()
        )

    def test_importer_maps_current_chw_layout_and_skips_business_duplicates(self):
        workbook = Workbook()
        chw_sheet = workbook.active
        chw_sheet.title = "CHW"
        chw_sheet.append([
            "LICENSE#",
            "DATE",
            "NAME",
            "SURNAME",
            "DOB",
            "MARITAL STATUS",
            "NATIONALITY",
            "GENDER",
            "PHONE",
            "EMAIL",
            "SCHOOL ADDRESS",
            "Qualifications",
            "APPLY FOR",
            "QUALIFICATION",
        ])
        chw_sheet.append([
            "1",
            "05.09.85",
            "RURI, YEME",
            "",
            "15.02.60",
            "Married",
            "PNG",
            "M",
            "70000001",
            "ruri@example.test",
            "BUNAPAS HEALTH CENTRE BOX 1080, MADANG PROVINCE",
            "APO CERT (GAUBIN) 1976",
            "FULL",
            "CHW",
        ])
        atp_sheet = workbook.create_sheet("ATP DATABASE ONLY")
        atp_sheet.append(["CERT #", "Date", "Name", "SURNAME", "ARCH#", "DATE2", "2017", "RECIEPT"])
        atp_sheet.append(["CERT #", "Date", "Name", "SURNAME", "ARCH#", "DATE", 2017, "RECIEPT"])
        atp_sheet.append(["1", "05.09.85", "RURI, YEME", "", "P#001", "", "05.03.2017 K15", "R0001"])
        output = BytesIO()
        workbook.save(output)

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as workbook_file:
            workbook_file.write(output.getvalue())
            workbook_path = workbook_file.name

        try:
            MedicalBoardWorkbookImporter(workbook_path=workbook_path).import_workbook()
            MedicalBoardWorkbookImporter(workbook_path=workbook_path).import_workbook()
        finally:
            os.unlink(workbook_path)

        chw = CommunityHealthWorker.objects.get(registration_no="CHW-1")
        self.assertEqual(chw.first_name, "Yeme")
        self.assertEqual(chw.last_name, "Ruri")
        self.assertEqual(chw.date_of_birth.isoformat(), "1960-02-15")
        self.assertEqual(chw.gender, "Male")
        self.assertEqual(chw.email, "ruri@example.test")
        self.assertEqual(chw.province, "Madang")
        self.assertEqual(
            PracticingLicenseRecord.objects.filter(
                target_model="communityhealthworker",
                registration_no="CHW-1",
                record_type="workforce_listing",
            ).count(),
            1,
        )
        self.assertEqual(
            PracticingLicenseRecord.objects.filter(
                target_model="communityhealthworker",
                registration_no="CHW-1",
                record_type="practicing_license",
                record_year=2017,
            ).count(),
            1,
        )

    def test_medical_board_bulk_upload_uses_medical_importer(self):
        user = get_user_model().objects.create_user(
            username="medical_import_registrar",
            password="StrongPass123!",
            role="registrar",
            department="Medical Board",
        )
        upload = SimpleUploadedFile(
            "medical-board-chw.xlsx",
            _medical_board_chw_workbook_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.force_login(user)

        response = self.client.post(reverse("import_data"), {"file": upload})

        self.assertEqual(response.status_code, 302)
        batch = DataImportBatch.objects.get()
        self.assertEqual(batch.source_kind, "medical_board_workbook")
        self.assertEqual(CommunityHealthWorker.objects.count(), 2)
        self.assertFalse(DataImportBatch.objects.filter(source_kind="ndata_workbook").exists())


class MedicalBoardLegacyWorkbookImportTests(TestCase):
    def test_legacy_import_aligns_doctor_and_allied_rows_to_medical_board(self):
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as workbook_file:
            workbook_file.write(_medical_board_legacy_workbook_bytes())
            workbook_path = workbook_file.name

        try:
            batch = MedicalBoardLegacyWorkbookImporter(workbook_paths=[workbook_path]).import_workbooks()
        finally:
            os.unlink(workbook_path)

        self.assertEqual(batch.source_kind, "medical_board_workbook")
        self.assertEqual(batch.status, "completed")
        self.assertEqual(MedicalDoctor.objects.count(), 1)
        self.assertEqual(MedicalDoctor.objects.get().specialty, "")
        self.assertEqual(
            PracticingLicenseRecord.objects.filter(batch=batch, target_model="medicaldoctor").count(),
            2,
        )
        self.assertEqual(
            PracticingLicenseRecord.objects.filter(batch=batch, target_model="other", category__icontains="Environmental Health Officer").count(),
            2,
        )
        self.assertFalse(NursingProfessional.objects.exists())
        self.assertFalse(NurseAide.objects.exists())
        self.assertFalse(Midwife.objects.exists())

    def test_medical_board_bulk_upload_accepts_legacy_workbook_without_nursing_import(self):
        user = get_user_model().objects.create_user(
            username="medical_legacy_import_registrar",
            password="StrongPass123!",
            role="registrar",
            department="Medical Board",
        )
        upload = SimpleUploadedFile(
            "medical-board-legacy.xlsx",
            _medical_board_legacy_workbook_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.force_login(user)

        response = self.client.post(reverse("import_data"), {"file": upload})

        self.assertEqual(response.status_code, 302)
        batch = DataImportBatch.objects.get()
        self.assertEqual(batch.source_kind, "medical_board_workbook")
        self.assertEqual(MedicalDoctor.objects.count(), 1)
        self.assertTrue(PracticingLicenseRecord.objects.filter(target_model="other").exists())
        self.assertFalse(DataImportBatch.objects.filter(source_kind="ndata_workbook").exists())


class MedicalBoardFormTests(TestCase):
    def test_pdf_mapped_medical_board_forms_validate_and_expose_expected_fields(self):
        profile_data = {
            "applicant_type": "national",
            "date_of_birth": "1985-06-12",
            "gender": "Male",
            "nationality": "Papua New Guinea",
            "contact_number": "70000001",
            "email_address": "applicant@example.com",
            "full_address": "Waigani, Port Moresby",
            "province": "National Capital District",
            "applicant_signature": "Dr Maria Kalo",
            "institution_attended": "University of Papua New Guinea",
            "institute_name": "University of Papua New Guinea",
            "program_completed": "Bachelor of Medicine",
            "date_of_completion": "2010-12-10",
            "employment_status": "full_time",
            "employer_name": "Port Moresby General Hospital",
            "place_of_work": "hospital",
            "registration_no": "MB-1001",
            "licence_number": "MB-LIC-1001",
        }
        facility_data = {
            "facility_owner": "National Department of Health",
            "ownership": "public",
            "facility_level": "district",
            "province": "National Capital District",
            "district": "Port Moresby",
            "physical_address": "Waigani Drive",
            "contact_person": "Maria Kalo",
            "contact_number": "70000002",
            "email_address": "facility@example.com",
            "services_offered": "General clinical services",
            "staffing_summary": "Registered clinical staff available",
            "equipment_and_supplies": "Essential equipment available",
            "infection_control_measures": "Standard infection control procedures",
            "emergency_readiness": "Emergency response plan available",
            "applicant_signature": "Maria Kalo",
        }
        cases = [
            (
                MedicalBoardSpecialistApplicationForm,
                {
                    **profile_data,
                    "full_name": "Dr Maria Kalo",
                    "practitioner_stream": "medical_practitioner",
                    "specialty": "public_health",
                },
                ["specialty", "qualifications_summary", "commissioner_of_oaths"],
            ),
            (
                MedicalBoardRenewalRegistrationForm,
                {
                    **profile_data,
                    "first_name": "Maria",
                    "last_name": "Kalo",
                    "practitioner_no": "MB-P-1001",
                },
                ["application_types", "practitioner_categories", "postgrad_qualification_type_3"],
            ),
            (
                MedicalBoardChwRegistrationForm,
                {
                    **profile_data,
                    "full_name": "Samuel Henao",
                    "registration_no": "CHW-1001",
                    "community_id": "CHW-COM-1001",
                    "training_level": "Certificate",
                },
                ["community_id", "training_level"],
            ),
            (
                MedicalBoardAccreditationChecklistForm,
                {
                    **facility_data,
                    "facility_name": "Central Health Training Institution",
                    "declaration": "on",
                },
                ["accreditation_curriculum_status", "accreditation_head_of_training_institution_comments"],
            ),
            (
                MedicalBoardPrivateHealthFacilityChecklistForm,
                {
                    **facility_data,
                    "facility_name": "Boroko Private Clinic",
                    "ownership": "private",
                    "declaration": "on",
                },
                ["private_health_application_form_status", "private_health_floor_plan_comments"],
            ),
            (
                MedicalBoardTrainingCollegeFacilityForm,
                {
                    **facility_data,
                    "facility_name": "Koki Dental Training Clinic",
                    "applicant_full_name": "Maria Kalo",
                    "applicant_address": "Waigani, Port Moresby",
                    "operation_type": "Dental training clinic",
                    "premises_description": "Standalone clinic with training rooms",
                    "declared_at": "Port Moresby",
                    "oath_witness": "Commissioner of Oaths",
                    "declaration": "on",
                },
                ["hsfc_requirement_formal_application_status", "staff_roster", "registrar_signature"],
            ),
        ]

        for form_class, data, expected_fields in cases:
            with self.subTest(form_code=form_class.form_code):
                form = form_class(data=data)

                self.assertTrue(form.is_valid(), form.errors.as_json())
                for field_name in expected_fields:
                    self.assertIn(field_name, form.fields)
                self.assertGreaterEqual(len(form.section_layout), 4)

    def test_medical_board_submission_records_response_without_nursing_checklist(self):
        form = MedicalBoardSpecialistApplicationForm(data={
            "applicant_type": "national",
            "date_of_birth": "1985-06-12",
            "gender": "Male",
            "nationality": "Papua New Guinea",
            "contact_number": "70000001",
            "email_address": "applicant@example.com",
            "full_address": "Waigani, Port Moresby",
            "province": "National Capital District",
            "applicant_signature": "Dr Maria Kalo",
            "institution_attended": "University of Papua New Guinea",
            "institute_name": "University of Papua New Guinea",
            "program_completed": "Bachelor of Medicine",
            "date_of_completion": "2010-12-10",
            "employment_status": "full_time",
            "employer_name": "Port Moresby General Hospital",
            "place_of_work": "hospital",
            "registration_no": "MB-1001",
            "licence_number": "MB-LIC-1001",
            "full_name": "Dr Maria Kalo",
            "practitioner_stream": "medical_practitioner",
            "specialty": "public_health",
        })
        self.assertTrue(form.is_valid(), form.errors.as_json())

        application = form.save()

        self.assertEqual(application.form_code, "MBSP")
        self.assertEqual(application.pathway, "medical_board")
        self.assertEqual(application.payload["specialty"], "public_health")
        self.assertTrue(ApplicationFormResponse.objects.filter(application=application, form_code="MBSP").exists())
        self.assertFalse(ApplicationChecklistItem.objects.filter(application=application).exists())


class MedicalBoardApplicantRegistrationNavigationTests(TestCase):
    def setUp(self):
        self.user_model = get_user_model()

    def test_medical_board_applicants_see_medical_registration_sidebar_link(self):
        cases = {
            "doctor": "doctor_dashboard",
            "chw": "chw_dashboard",
        }

        for role, dashboard_name in cases.items():
            with self.subTest(role=role):
                user = self.user_model.objects.create_user(
                    username=f"{role}_applicant",
                    password="StrongPass123!",
                    role=role,
                    department="Medical Board",
                )
                self.client.force_login(user)

                response = self.client.get(reverse(dashboard_name))

                self.assertEqual(response.status_code, 200)
                self.assertContains(response, f'href="{reverse("medical_board_register")}" class="nav-link"')
                self.client.logout()

    def test_medical_board_applicants_are_redirected_from_nursing_form_index(self):
        for role in ("doctor", "chw"):
            with self.subTest(role=role):
                user = self.user_model.objects.create_user(
                    username=f"{role}_redirect",
                    password="StrongPass123!",
                    role=role,
                    department="Medical Board",
                )
                self.client.force_login(user)

                response = self.client.get(reverse("nursing_forms_portal"))

                self.assertRedirects(response, reverse("medical_board_register"), fetch_redirect_response=False)
                self.client.logout()


class AIImportCleanserTests(TestCase):
    def test_local_cleanser_normalizes_row_and_flags_review_items(self):
        result = cleanse_import_row(
            {
                "Full Name": "  maria   test  ",
                "Province": "Morob",
                "Registration No": " rn 100 ",
                "Payment Date": "2050-03-17",
            },
            row_number=1,
            source_label="sample.xlsx",
            scope="nursing",
        )

        self.assertEqual(result["provider"], "local")
        self.assertEqual(result["normalized_row"]["full_name"], "maria test")
        self.assertEqual(result["normalized_row"]["registration_no"], "RN 100")
        self.assertTrue(result["requires_human_review"])
        issue_types = {issue["issue_type"] for issue in result["issues"]}
        self.assertTrue({"province_fuzzy_match", "unknown_province"} & issue_types)
        self.assertIn("future_date", issue_types)


class NursingCouncilWorkflowConfigurationTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        ensure_nursing_council_configuration()

    def test_configuration_seed_creates_expected_nursing_council_records(self):
        body = RegulatoryBody.objects.get(code=NURSING_COUNCIL_CODE)

        self.assertEqual(ApplicationPathway.objects.filter(regulatory_body=body).count(), 12)
        self.assertEqual(DynamicFormDefinition.objects.filter(regulatory_body=body).count(), 18)
        self.assertEqual(DocumentRequirement.objects.filter(pathway__regulatory_body=body).count(), 60)
        self.assertEqual(FeeSchedule.objects.filter(regulatory_body=body).count(), 10)
        self.assertEqual(DeclarationTemplate.objects.filter(regulatory_body=body).count(), 11)

    def test_public_form_guide_uses_configured_pathways_and_forms(self):
        guide = build_public_form_guide()
        graduate_pathway = "PNG Graduate Nurse Provisional Licence (PNG_NURSE_GRAD_PROV)"

        self.assertIn(graduate_pathway, guide)
        self.assertIn(("NC1", "Application for Provisional Licence to Practice"), guide[graduate_pathway])
        self.assertIn(("G4", "Statement of Competency for Graduate Nurses"), guide[graduate_pathway])
        self.assertIn("Special application forms", guide)
        self.assertIn(
            ("NC9", "Temporary Licence to Practise Criteria for Overseas Nurses Checklist (Revised 2023)"),
            guide["Special application forms"],
        )
        self.assertFalse(any("DECEASED_NOTICE" in pathway for pathway in guide))

    def test_dashboard_workflow_rows_include_full_configured_pathways(self):
        rows = build_nursing_workflow_rows()
        codes = {row["code"] for row in rows}

        self.assertEqual(len(rows), 12)
        self.assertIn("OVERSEAS_TEMP", codes)
        self.assertIn("DECEASED_NOTICE", codes)
        self.assertIn("EMPLOYER_VERIFY", codes)

    def test_renewal_validation_blocks_missing_employment_status(self):
        professional = NursingProfessional.objects.create(
            first_name="Maria",
            last_name="Test",
            gender="Female",
            date_of_birth=date(1990, 1, 1),
            registration_no="RN-TEST-001",
        )
        application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(professional),
            object_id=professional.pk,
            form_code="NC3",
            pathway="local_nursing_graduate",
            payload={"pathway_code": "PNG_RENEWAL"},
        )

        result = NursingCouncilValidationService(application).validate_for_status("submitted")

        self.assertFalse(result["can_proceed"])
        self.assertIn("Renewal applications must capture employment status.", result["errors"])

    def test_checklist_generator_creates_required_items_from_pathway_config(self):
        professional = NursingProfessional.objects.create(
            first_name="Lina",
            last_name="Graduate",
            gender="Female",
            date_of_birth=date(2001, 2, 3),
            registration_no="GRAD-TEST-001",
        )
        application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(professional),
            object_id=professional.pk,
            form_code="NC1",
            pathway="local_nursing_graduate",
            payload={"pathway_code": "PNG_NURSE_GRAD_PROV"},
        )

        items = generate_application_checklist(application)
        labels = set(
            ApplicationChecklistItem.objects.filter(application=application).values_list(
                "document_requirement__label",
                flat=True,
            )
        )

        self.assertEqual(len(items), 5)
        self.assertIn("Academic award", labels)
        self.assertIn("G4 competency statement", labels)

    def test_approval_creates_licence_record_and_audit_log(self):
        professional = NursingProfessional.objects.create(
            first_name="Anna",
            last_name="Renewal",
            gender="Female",
            date_of_birth=date(1988, 5, 1),
            registration_no="RN-RENEW-001",
            registration_number="P-RENEW-001",
        )
        application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(professional),
            object_id=professional.pk,
            form_code="NC3",
            pathway="local_nursing_graduate",
            payload={
                "pathway_code": "PNG_RENEWAL",
                "declaration_acceptance": True,
                "employment_status": "full_time",
                "employer_name": "Public Hospital",
                "facility_name": "Public Hospital",
                "province": "National Capital District",
                "position_title": "Nurse",
                "area_of_employment": "government",
                "start_date": "2026-01-01",
            },
        )
        for item in generate_application_checklist(application):
            item.status = "accepted"
            item.save(update_fields=["status"])
        Receipt.objects.create(
            receipt_number="",
            amount="70.00",
            status="completed",
            application=application,
        )

        result = approve_nursing_application(application)
        application.refresh_from_db()
        professional.refresh_from_db()

        self.assertTrue(result["approved"])
        self.assertEqual(application.status, "approved")
        self.assertEqual(professional.license_expiry_date, date(date.today().year, 12, 31))
        self.assertTrue(PracticingLicenseRecord.objects.filter(source_row=application.pk, record_type="practicing_license").exists())
        self.assertTrue(AuditLog.objects.filter(action="REGISTRAR_APPROVED", entity_id=str(application.pk)).exists())

    def test_full_registration_approval_creates_an_approved_full_licence_record(self):
        """Keep applicant-stage full imports separate from registrar-approved licences."""
        professional = NursingProfessional.objects.create(
            first_name="Lina",
            last_name="Full",
            gender="Female",
            date_of_birth=date(1992, 4, 8),
            registration_no="RN-FULL-APPROVED-001",
            registration_number="P-FULL-APPROVED-001",
        )
        application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(professional),
            object_id=professional.pk,
            form_code="NC2",
            pathway="local_nursing_graduate",
            payload={
                "pathway_code": "PNG_PROV_TO_FULL",
                "declaration_acceptance": True,
                "institution_attended": "Test Nursing College",
            },
        )
        for item in generate_application_checklist(application):
            item.status = "accepted"
            item.save(update_fields=["status"])
        assignment = create_supervisor_assignment(
            application=application,
            supervisor_name="Senior Nurse",
        )
        complete_supervisor_competency(
            assignment=assignment,
            result="competent",
            comments="Ready for full registration.",
        )
        Receipt.objects.create(
            receipt_number="FULL-APPROVED-REC-001",
            amount="50.00",
            status="completed",
            application=application,
        )

        result = approve_nursing_application(application)

        self.assertTrue(result["approved"])
        self.assertTrue(
            PracticingLicenseRecord.objects.filter(
                source_row=application.pk,
                record_type="full_approved",
            ).exists()
        )
        self.assertFalse(
            PracticingLicenseRecord.objects.filter(
                source_row=application.pk,
                record_type="full",
            ).exists()
        )

    def test_approved_application_can_issue_official_document_by_mailbox_and_email(self):
        user_model = get_user_model()
        professional = NursingProfessional.objects.create(
            first_name="Issue",
            last_name="Client",
            gender="Female",
            date_of_birth=date(1991, 6, 1),
            registration_no="RN-ISSUE-001",
            registration_number="P-ISSUE-001",
            email="issue.client@example.com",
        )
        professional_ct = ContentType.objects.get_for_model(professional)
        recipient = user_model.objects.create_user(
            username="issue_client",
            email="issue.client@example.com",
            password="pass",
            role="nurse",
            professional_content_type=professional_ct,
            professional_object_id=professional.pk,
            professional_record_status="linked",
        )
        registrar = user_model.objects.create_user(
            username="issue_registrar",
            email="registrar@example.com",
            password="pass",
            role="registrar",
            role_approved=True,
            department="Nursing Council",
        )
        application = Application.objects.create(
            content_type=professional_ct,
            object_id=professional.pk,
            form_code="NC3",
            pathway="local_nursing_graduate",
            payload={
                "pathway_code": "PNG_RENEWAL",
                "declaration_acceptance": True,
                "employment_status": "full_time",
                "employer_name": "Public Hospital",
                "facility_name": "Public Hospital",
                "province": "National Capital District",
                "position_title": "Nurse",
                "area_of_employment": "government",
                "start_date": "2026-01-01",
            },
        )
        for item in generate_application_checklist(application):
            item.status = "accepted"
            item.save(update_fields=["status"])
        Receipt.objects.create(
            receipt_number="ISSUE-REC-001",
            amount="70.00",
            status="completed",
            application=application,
        )
        approve_nursing_application(application, actor=registrar)
        application.refresh_from_db()

        with TemporaryDirectory() as media_root, override_settings(
            MEDIA_ROOT=media_root,
            EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        ):
            mail.outbox = []
            issued = issue_application_licence_document(
                application,
                issuer=registrar,
                delivery_channel="both",
            )

            self.assertEqual(issued.document_type, "authority_to_practice")
            self.assertEqual(issued.status, "sent")
            self.assertTrue(issued.email_sent)
            self.assertTrue(issued.mailbox_sent)
            self.assertEqual(issued.recipient_user, recipient)
            self.assertEqual(len(mail.outbox), 1)
            self.assertTrue(issued.file.name.endswith(".pdf"))
            self.assertTrue(EnquiryThread.objects.filter(pk=issued.mailbox_thread_id, recipient_user=recipient).exists())
            self.assertTrue(EnquiryMessageAttachment.objects.filter(message__thread=issued.mailbox_thread).exists())
            self.assertTrue(Notification.objects.filter(user=recipient, subject__icontains="Authority to Practice").exists())
            self.assertTrue(IssuedLicenceDocument.objects.filter(pk=issued.pk, practicing_record__record_type="practicing_license").exists())

    def test_public_register_search_returns_safe_fields_only(self):
        NursingProfessional.objects.create(
            first_name="Safe",
            last_name="Search",
            gender="Female",
            date_of_birth=date(1992, 8, 2),
            registration_no="RN-SAFE-001",
            registration_number="PN-SAFE-001",
            email="private@example.com",
            primary_phone="12345",
        )

        rows = search_public_nursing_register(query="Safe")

        self.assertEqual(rows[0]["full_name"], "Safe Search")
        self.assertIn("registration_number", rows[0])
        self.assertNotIn("email", rows[0])
        self.assertNotIn("date_of_birth", rows[0])
        self.assertNotIn("primary_phone", rows[0])

        full_name_rows = search_public_nursing_register(query="Safe Search")
        self.assertEqual(full_name_rows[0]["full_name"], "Safe Search")

    def test_public_register_search_returns_public_safe_imported_records(self):
        nursing_batch = DataImportBatch.objects.create(
            source_file_name="nursing-atp.xlsx",
            source_kind="nursing_license_workbook",
            status="completed",
        )
        medical_batch = DataImportBatch.objects.create(
            source_file_name="medical.xlsx",
            source_kind="medical_board_workbook",
            status="completed",
        )
        PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            source_sheet_name="ATP RECORD 2026",
            source_row=10,
            record_type="practicing_license",
            target_model="nursingprofessional",
            full_name="Imported Safe",
            first_name="Imported",
            last_name="Safe",
            registration_no="PG 2026",
            practitioner_number="PN-IMPORT",
            record_year=date.today().year,
            category="General Nurse",
            reference_number="PRIVATE-RECEIPT",
        )
        PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            source_sheet_name="Payments",
            source_row=11,
            record_type="payment",
            target_model="nursingprofessional",
            full_name="Imported Safe",
            first_name="Imported",
            last_name="Safe",
            registration_no="PG 2026",
            practitioner_number="PN-IMPORT",
            record_year=date.today().year,
            category="General Nurse",
            reference_number="PRIVATE-PAYMENT",
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            source_sheet_name="Medical",
            source_row=12,
            record_type="practicing_license",
            target_model="medicaldoctor",
            full_name="Imported Safe",
            first_name="Imported",
            last_name="Safe",
            registration_no="MD 2026",
            practitioner_number="MP-IMPORT",
            record_year=date.today().year,
            category="Medical Doctor",
        )

        rows = search_public_nursing_register(query="Imported Safe")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["full_name"], "Imported Safe")
        self.assertEqual(rows[0]["registration_number"], "PG 2026")
        self.assertEqual(rows[0]["professional_category"], "Registered Nurse")
        self.assertEqual(rows[0]["licence_status"], "Active")
        self.assertNotIn("reference_number", rows[0])
        self.assertNotIn("payment_method", rows[0])
        self.assertNotIn("date_of_birth", rows[0])

    def test_public_register_search_url_renders_html_page_for_browser(self):
        NursingProfessional.objects.create(
            first_name="Safe",
            last_name="Search",
            registration_no="RN-SAFE-HTML",
            registration_number="PN-SAFE-HTML",
            email="private@example.com",
            primary_phone="12345",
        )

        response = self.client.get(reverse("public_nursing_register_search_root"), {"q": "Safe"})

        self.assertEqual(response.status_code, 200)
        self.assertIn("text/html", response["Content-Type"])
        self.assertContains(response, "PNG Nursing Council Public Register Verification")
        self.assertContains(response, "Safe Search")
        self.assertContains(response, "RN-SAFE-HTML")
        self.assertContains(response, reverse("workforce_map") + "?office=nursing")
        self.assertContains(response, reverse("workforce_map"))
        self.assertNotContains(response, "private@example.com")
        self.assertNotContains(response, "12345")

    def test_public_register_search_url_keeps_json_mode(self):
        NursingProfessional.objects.create(
            first_name="Safe",
            last_name="Json",
            registration_no="RN-SAFE-JSON",
            registration_number="PN-SAFE-JSON",
            email="private-json@example.com",
            primary_phone="67890",
        )

        response = self.client.get(
            reverse("public_nursing_register_search_root"),
            {"q": "Safe", "format": "json"},
            HTTP_ACCEPT="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/json")
        payload = response.json()
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["results"][0]["full_name"], "Safe Json")
        self.assertNotIn("email", payload["results"][0])
        self.assertNotIn("primary_phone", payload["results"][0])

    def test_public_medical_board_register_search_renders_html_page_for_browser(self):
        MedicalDoctor.objects.create(
            first_name="Public",
            last_name="Doctor",
            registration_no="MD-PUBLIC-001",
            registration_number="MP-PUBLIC-001",
            email="private-medical@example.com",
            primary_phone="77777",
            specialty="General Practice",
            license_expiry_date=date(2030, 1, 1),
        )

        response = self.client.get(reverse("public_medical_board_register_search_root"), {"q": "Public"})

        self.assertEqual(response.status_code, 200)
        self.assertIn("text/html", response["Content-Type"])
        self.assertContains(response, "PNG Medical Board Public Register Verification")
        self.assertContains(response, "Public Doctor")
        self.assertContains(response, "MD-PUBLIC-001")
        self.assertContains(response, reverse("workforce_map") + "?office=medical")
        self.assertNotContains(response, "private-medical@example.com")
        self.assertNotContains(response, "77777")

    def test_public_medical_board_register_search_keeps_json_mode(self):
        CommunityHealthWorker.objects.create(
            first_name="Public",
            last_name="Chw",
            registration_no="CHW-PUBLIC-001",
            community_id="CHW-COMM-001",
            email="private-chw@example.com",
            primary_phone="88888",
            training_level="Certificate",
        )

        response = self.client.get(
            reverse("public_medical_board_register_search_root"),
            {"q": "Chw", "format": "json"},
            HTTP_ACCEPT="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/json")
        payload = response.json()
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["results"][0]["full_name"], "Public Chw")
        self.assertNotIn("email", payload["results"][0])
        self.assertNotIn("primary_phone", payload["results"][0])

    def test_medical_board_forms_catalogue_includes_initial_doctor_application(self):
        response = self.client.get(reverse("medical_board_register"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Initial Medical Practitioner Registration")
        self.assertContains(response, reverse("medical_board_form_register", args=["MD1"]))

    def test_deceased_notification_approval_deactivates_practitioner(self):
        professional = NursingProfessional.objects.create(
            first_name="Late",
            last_name="Practitioner",
            gender="Female",
            date_of_birth=date(1975, 1, 1),
            registration_no="RN-LATE-001",
            is_active=True,
        )
        user = get_user_model().objects.create_user(
            username="late.practitioner",
            password="StrongPass123!",
            role="nurse",
            professional_content_type=ContentType.objects.get_for_model(professional),
            professional_object_id=professional.pk,
            professional_record_status="linked",
        )
        notification = create_deceased_notification(
            actor=None,
            name_at_report="Late Practitioner",
            date_of_death=date(2026, 5, 1),
            registration_number="RN-LATE-001",
        )

        approve_deceased_notification(notification)
        professional.refresh_from_db()
        user.refresh_from_db()

        self.assertFalse(professional.is_active)
        self.assertEqual(professional.license_expiry_date, date(2026, 5, 1))
        self.assertEqual(user.professional_record_status, "deceased")
        self.assertEqual(DeceasedNotification.objects.get(pk=notification.pk).verification_status, "approved")

    def test_employer_verification_returns_safe_snapshot(self):
        NursingProfessional.objects.create(
            first_name="Employer",
            last_name="Visible",
            gender="Female",
            date_of_birth=date(1990, 1, 1),
            registration_no="RN-EMP-001",
            email="hidden@example.com",
        )

        verification = create_employer_verification_request(
            actor=None,
            employer_name="Test Employer",
            registration_number="RN-EMP-001",
        )

        self.assertEqual(verification.status, "verified")
        self.assertEqual(verification.safe_result_json["full_name"], "Employer Visible")
        self.assertNotIn("email", verification.safe_result_json)
        self.assertTrue(EmployerVerificationRequest.objects.filter(pk=verification.pk).exists())

    def test_supervisor_assignment_completion_creates_competency_assessment(self):
        professional = NursingProfessional.objects.create(
            first_name="Competent",
            last_name="Applicant",
            gender="Female",
            date_of_birth=date(1995, 4, 4),
            registration_no="RN-COMP-001",
        )
        application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(professional),
            object_id=professional.pk,
            form_code="NC2",
            pathway="local_nursing_graduate",
            payload={"pathway_code": "PNG_FULL_REG"},
        )

        assignment = create_supervisor_assignment(application=application, supervisor_name="Senior Nurse")
        assessment = complete_supervisor_competency(assignment=assignment, result="competent", comments="Ready for full practice.")
        assignment.refresh_from_db()

        self.assertEqual(assignment.status, "completed")
        self.assertTrue(assessment.is_passed)
        self.assertTrue(SupervisorAssignment.objects.filter(pk=assignment.pk).exists())
        self.assertTrue(CompetencyAssessment.objects.filter(pk=assessment.pk).exists())
