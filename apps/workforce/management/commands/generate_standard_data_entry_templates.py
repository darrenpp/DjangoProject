from pathlib import Path

from django import forms
from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from apps.workforce.forms import (
    ChwPublicRegistrationForm,
    GraduateMidwifeBatchListForm,
    GraduateMidwivesChecklistForm,
    GraduateNurseBatchListForm,
    GraduateNursesChecklistForm,
    GraduateVitaeForm,
    HealthStudentPublicRegistrationForm,
    MedicalBoardAccreditationChecklistForm,
    MedicalBoardChwRegistrationForm,
    MedicalBoardPrivateHealthFacilityChecklistForm,
    MedicalBoardRenewalRegistrationForm,
    MedicalBoardSpecialistApplicationForm,
    MedicalBoardTrainingCollegeFacilityForm,
    MedicalDoctorPublicRegistrationForm,
    MidwifeCompetencyStatementForm,
    NC10ChildNursingCompetencyForm,
    NC11DoubleMajorChecklistForm,
    NC1ProvisionalLicenceForm,
    NC2FullLicenceForm,
    NC3RenewalLicenceForm,
    NC4ProvisionalChecklistForm,
    NC5OverseasFullRegistrationForm,
    NC6NursingCompetencyForm,
    NC7MidwiferyCompetencyForm,
    NC8TemporaryLicenceForm,
    NC9TemporaryChecklistForm,
    NurseAidePublicRegistrationForm,
    NurseCompetencyStatementForm,
    NursingFullLicenseForm,
    NursingPublicRegistrationForm,
    NursingRenewalForm,
)


DATA_ROWS = 250
HEADER_FILL = PatternFill("solid", fgColor="12324A")
REQUIRED_FILL = PatternFill("solid", fgColor="D9EAD3")
OPTIONAL_FILL = PatternFill("solid", fgColor="EAF2F8")
LOCKED_FILL = PatternFill("solid", fgColor="F4F7F6")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)

COMMON_IMPORT_COLUMNS = [
    ("source_reference", "Source file, paper file, or officer reference", False),
    ("record_action", "create, update, or review", False),
    ("record_status", "pending, approved, rejected, or imported", False),
    ("notes", "Internal import or cleansing notes", False),
]

MEDICAL_SHEETS = [
    ("MedicalDoctor", "MD1", MedicalDoctorPublicRegistrationForm),
    ("CommunityHealthWorker", "CHW1", ChwPublicRegistrationForm),
    ("MBSP_Specialist", "MBSP", MedicalBoardSpecialistApplicationForm),
    ("MBRN_Renewal", "MBRN", MedicalBoardRenewalRegistrationForm),
    ("CHW1_Form", "CHW1", MedicalBoardChwRegistrationForm),
    ("MBAC_Accreditation", "MBAC", MedicalBoardAccreditationChecklistForm),
    ("MBPF_PrivateFacility", "MBPF", MedicalBoardPrivateHealthFacilityChecklistForm),
    ("MBTC_TrainingFacility", "MBTC", MedicalBoardTrainingCollegeFacilityForm),
    ("AlliedHealth_Intake", "MBRN", MedicalBoardRenewalRegistrationForm),
]

NURSING_SHEETS = [
    ("NursingProfessional", "NC1", NursingPublicRegistrationForm),
    ("NursingFullLicense", "NC2", NursingFullLicenseForm),
    ("NursingRenewal", "NC3", NursingRenewalForm),
    ("NurseAide", "NC2", NurseAidePublicRegistrationForm),
    ("HealthStudent", "G3", HealthStudentPublicRegistrationForm),
    ("G1_GradChecklist", "G1", GraduateNursesChecklistForm),
    ("G2_GradBatch", "G2", GraduateNurseBatchListForm),
    ("G3_GraduateVitae", "G3", GraduateVitaeForm),
    ("G4_NurseCompetency", "G4", NurseCompetencyStatementForm),
    ("G5_MidwifeCompetency", "G5", MidwifeCompetencyStatementForm),
    ("G6_MidwifeChecklist", "G6", GraduateMidwivesChecklistForm),
    ("G7_MidwifeBatch", "G7", GraduateMidwifeBatchListForm),
    ("NC1_Provisional", "NC1", NC1ProvisionalLicenceForm),
    ("NC2_FullLicense", "NC2", NC2FullLicenceForm),
    ("NC3_Renewal", "NC3", NC3RenewalLicenceForm),
    ("NC4_ProvChecklist", "NC4", NC4ProvisionalChecklistForm),
    ("NC5_OverseasFull", "NC5", NC5OverseasFullRegistrationForm),
    ("NC6_NurseCompetency", "NC6", NC6NursingCompetencyForm),
    ("NC7_MidwifeCompetency", "NC7", NC7MidwiferyCompetencyForm),
    ("NC8_Temporary", "NC8", NC8TemporaryLicenceForm),
    ("NC9_TempChecklist", "NC9", NC9TemporaryChecklistForm),
    ("NC10_ChildNursing", "NC10", NC10ChildNursingCompetencyForm),
    ("NC11_DoubleMajor", "NC11", NC11DoubleMajorChecklistForm),
]


def field_type(field):
    if isinstance(field, forms.DateField):
        return "date"
    if isinstance(field, forms.DecimalField):
        return "decimal"
    if isinstance(field, forms.IntegerField):
        return "integer"
    if isinstance(field, forms.BooleanField):
        return "yes_no"
    if isinstance(field, forms.FileField):
        return "file_path"
    if getattr(field, "choices", None):
        return "choice"
    return "text"


def field_choices(field):
    choices = []
    for value, label in getattr(field, "choices", []) or []:
        if value in {None, ""}:
            continue
        choices.append(str(value))
    return choices


def form_fields(form_class):
    form = form_class()
    fields = []
    for name, field in form.fields.items():
        if name == "captcha":
            continue
        fields.append({
            "name": name,
            "label": str(field.label or name),
            "required": bool(field.required),
            "type": field_type(field),
            "choices": field_choices(field),
        })
    return fields


def safe_title(value, used_titles):
    base = value[:31]
    title = base
    suffix = 1
    while title in used_titles:
        suffix += 1
        title = f"{base[:28]}_{suffix}"
    used_titles.add(title)
    return title


def add_guide_sheet(workbook, office_label):
    sheet = workbook.active
    sheet.title = "Import Guide"
    sheet.append([f"{office_label} Standard Data Entry Workbook"])
    sheet.append(["Purpose", "Use these table sheets for data entry so future imports match the Django system fields."])
    sheet.append(["Header rule", "Do not rename columns. Add records from row 2 down."])
    sheet.append(["Dates", "Use YYYY-MM-DD where possible."])
    sheet.append(["Files", "For document/photo fields, enter the saved file path or document reference."])
    sheet.append(["Scope", f"Rows in this workbook are for {office_label} only. Do not mix regulatory bodies."])
    sheet["A1"].font = TITLE_FONT
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows(min_row=2, max_row=6, min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def add_dictionary_sheet(workbook):
    sheet = workbook.create_sheet("Field Dictionary")
    sheet.append(["office", "sheet_name", "form_code", "field_name", "label", "required", "type", "choices"])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    sheet.freeze_panes = "A2"
    widths = [18, 28, 12, 36, 48, 12, 16, 80]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    return sheet


def add_data_sheet(workbook, dictionary_sheet, office_label, used_titles, display_title, form_code, form_class):
    title = safe_title(display_title, used_titles)
    sheet = workbook.create_sheet(title)
    fields = form_fields(form_class)
    headers = [item[0] for item in COMMON_IMPORT_COLUMNS] + [field["name"] for field in fields]
    sheet.append(headers)

    for index, (name, _label, _required) in enumerate(COMMON_IMPORT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = LOCKED_FILL
        cell.font = Font(bold=True)
        cell.comment = Comment("Standard import metadata column.", "Codex")
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(name) + 3, 18), 36)

    for offset, field in enumerate(fields, start=len(COMMON_IMPORT_COLUMNS) + 1):
        cell = sheet.cell(row=1, column=offset)
        cell.fill = REQUIRED_FILL if field["required"] else OPTIONAL_FILL
        cell.font = Font(bold=True)
        cell.comment = Comment(
            f"{field['label']}\nRequired: {'Yes' if field['required'] else 'No'}\nType: {field['type']}",
            "Codex",
        )
        sheet.column_dimensions[get_column_letter(offset)].width = min(max(len(field["name"]) + 3, 18), 42)
        dictionary_sheet.append([
            office_label,
            title,
            form_code,
            field["name"],
            field["label"],
            "Yes" if field["required"] else "No",
            field["type"],
            ", ".join(field["choices"]),
        ])

    for row in range(2, DATA_ROWS + 2):
        sheet.cell(row=row, column=2).value = "create"
        sheet.cell(row=row, column=3).value = "pending"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{DATA_ROWS + 1}"
    sheet.protection.sheet = False

    action_validation = DataValidation(type="list", formula1='"create,update,review"', allow_blank=True)
    status_validation = DataValidation(type="list", formula1='"pending,approved,rejected,imported"', allow_blank=True)
    sheet.add_data_validation(action_validation)
    sheet.add_data_validation(status_validation)
    action_validation.add(f"B2:B{DATA_ROWS + 1}")
    status_validation.add(f"C2:C{DATA_ROWS + 1}")

    for offset, field in enumerate(fields, start=len(COMMON_IMPORT_COLUMNS) + 1):
        column_letter = get_column_letter(offset)
        if field["type"] == "yes_no":
            validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        elif field["choices"] and len(",".join(field["choices"])) < 240:
            validation = DataValidation(type="list", formula1=f'"{",".join(field["choices"])}"', allow_blank=True)
        else:
            continue
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{DATA_ROWS + 1}")

def build_workbook(office_label, sheet_specs):
    workbook = Workbook()
    add_guide_sheet(workbook, office_label)
    dictionary_sheet = add_dictionary_sheet(workbook)
    used_titles = {"Import Guide", "Field Dictionary"}
    for display_title, form_code, form_class in sheet_specs:
        add_data_sheet(workbook, dictionary_sheet, office_label, used_titles, display_title, form_code, form_class)
    return workbook


class Command(BaseCommand):
    help = "Generate standard Excel data-entry templates from the current Django form fields."

    def add_arguments(self, parser):
        parser.add_argument(
            "--output-dir",
            default=str(settings.BASE_DIR / "docs" / "import_templates"),
            help="Directory where the Medical Board and Nursing Council templates will be written.",
        )

    def handle(self, *args, **options):
        output_dir = Path(options["output_dir"])
        output_dir.mkdir(parents=True, exist_ok=True)

        medical_path = output_dir / "NDOH_Medical_Board_Standard_Data_Entry_Template.xlsx"
        nursing_path = output_dir / "NDOH_Nursing_Council_Standard_Data_Entry_Template.xlsx"

        medical_workbook = build_workbook("Medical Board", MEDICAL_SHEETS)
        medical_workbook.save(medical_path)

        nursing_workbook = build_workbook("Nursing Council", NURSING_SHEETS)
        nursing_workbook.save(nursing_path)

        self.stdout.write(self.style.SUCCESS(f"Created {medical_path}"))
        self.stdout.write(self.style.SUCCESS(f"Created {nursing_path}"))
