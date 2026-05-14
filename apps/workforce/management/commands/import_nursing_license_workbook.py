from datetime import date, datetime, timedelta
from pathlib import Path
import re

import pandas as pd
from django.contrib.contenttypes.models import ContentType
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.workforce.models import (
    Application,
    Cadre,
    DataImportBatch,
    ImportedWorkbookSheet,
    Midwife,
    NursingProfessional,
    PracticingLicenseRecord,
    Qualification,
    TrainingInstitution,
)
from apps.workforce.services.institution_classification import (
    applicant_type_for_institution,
    classify_training_institution,
)


NULL_LIKE_VALUES = {"", "NAN", "NAT", "NONE", "NULL", "TBA", "N/A", "NA", "-"}
MONTH_FIXES = {
    "Agu": "Aug",
    "Ago": "Aug",
    "Ap-": "Apr-",
    "Jin": "Jun",
    "Jui": "Jul",
    "Mac": "Mar",
    "Ocf": "Oct",
    "Sept": "Sep",
    "Set": "Sep",
}
ALIGNMENT_SUMMARY_KEYS = (
    "provisional_sheet",
    "full_sheet",
    "source_rows_seen",
    "blank_rows_skipped",
    "missing_name_rows_skipped",
    "missing_registration_rows_skipped",
    "provisional_rows_ready",
    "full_rows_ready",
    "duplicate_rows_dropped",
    "duplicate_registration_repairs",
    "cleaned_rows",
    "provisional_cleaned",
    "full_cleaned",
    "national_institution_rows",
    "overseas_institution_rows",
)


def normalize_text(value):
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(value):
    text = normalize_text(value).lower()
    text = text.replace("licence", "license").replace("lisence", "license").replace("lisenc", "license")
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def normalize_name(value):
    text = normalize_text(value)
    if not text:
        return ""
    text = re.sub(r"\([^)]*\)", "", text)
    if "," in text:
        left, right = [part.strip() for part in text.split(",", 1)]
        if left and right:
            text = f"{right} {left}"
    text = re.sub(r"\s+", " ", text).strip(" -")
    return text.title()


def split_name(full_name):
    parts = normalize_text(full_name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def clean_identifier(value):
    text = normalize_text(value).upper()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    text = text.replace("_", " ").replace("-", " ")
    text = re.sub(r"\s*/\s*", "/", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_registration_no(prefix, number):
    prefix = clean_identifier(prefix)
    number = clean_identifier(number)
    if not number:
        return ""
    if prefix and not number.startswith(prefix):
        return f"{prefix} {number}".strip()
    return number


def parse_date(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        return value.date() if value.year >= 1901 else None
    if isinstance(value, date):
        return value if value.year >= 1901 else None

    text = normalize_text(value)
    if not text or text.upper() in NULL_LIKE_VALUES:
        return None
    for bad, good in MONTH_FIXES.items():
        text = text.replace(bad, good)
    text = text.replace("_", "-").replace("=", "-")
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"\s+", " ", text).strip()
    for dayfirst in (True, False):
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
        if not pd.isna(parsed):
            parsed_date = pd.Timestamp(parsed).date()
            if parsed_date.year < 1901:
                return None
            if parsed_date.year > date.today().year + 1 and parsed_date.year - 20 >= 2000:
                parsed_date = parsed_date.replace(year=parsed_date.year - 20)
            return parsed_date
    return None


def parse_year(value, fallback=None):
    if value is None or pd.isna(value):
        return fallback.year if fallback else None
    text = normalize_text(value)
    if re.fullmatch(r"\d{4}", text):
        return int(text)
    parsed = parse_date(value)
    if parsed:
        return parsed.year
    return fallback.year if fallback else None


def infer_applicant_type(*values):
    combined = " ".join(normalize_text(value) for value in values if normalize_text(value))
    return applicant_type_for_institution(combined)


def infer_profession_track(qualification):
    return "midwifery" if "midw" in normalize_text(qualification).lower() else "nursing"


def model_key_for_qualification(qualification):
    return "midwife" if infer_profession_track(qualification) == "midwifery" else "nursingprofessional"


def row_score(row):
    return sum(
        1
        for key in (
            "issued_date",
            "institution_name",
            "graduation_year",
            "qualification_name",
            "practitioner_number",
            "source_id",
        )
        if row.get(key) not in ("", None)
    )


def json_safe(value):
    if value is None or pd.isna(value):
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


class Command(BaseCommand):
    help = "Cleanse and import the combined provisional and full licence workbook."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the workbook containing provisional and full licence sheets.")
        parser.add_argument("--dry-run", action="store_true", help="Read and cleanse the workbook without saving.")
        parser.add_argument(
            "--append",
            action="store_true",
            help="Append a new import batch instead of replacing an earlier batch for the same source path.",
        )
        parser.add_argument(
            "--delete-recent-imports",
            action="store_true",
            help=(
                "Before importing, delete recent nursing licence import batches whose source path "
                "or cleansing totals do not align with this workbook. In --dry-run mode this only previews deletions."
            ),
        )
        parser.add_argument(
            "--recent-import-hours",
            type=int,
            default=24,
            help="How far back --delete-recent-imports should look for misaligned batches. Default: 24 hours.",
        )

    def handle(self, *args, **options):
        workbook = Path(options["file"])
        if not workbook.exists():
            raise CommandError(f"Workbook not found: {workbook}")

        dry_run = options["dry_run"]
        append = options["append"]
        delete_recent_imports = options["delete_recent_imports"]
        recent_import_hours = max(options["recent_import_hours"], 1)

        provisional_rows, full_rows, workbook_stats = self._load_rows(workbook)
        cleaned_rows, cleanse_stats = self._dedupe_rows(provisional_rows + full_rows)

        summary = {
            **workbook_stats,
            **cleanse_stats,
            "cleaned_rows": len(cleaned_rows),
            "provisional_cleaned": sum(1 for row in cleaned_rows if row["record_type"] == "provisional"),
            "full_cleaned": sum(1 for row in cleaned_rows if row["record_type"] == "full"),
            "national_institution_rows": sum(
                1 for row in cleaned_rows if row["institution_name"] and row["applicant_type"] == "national"
            ),
            "overseas_institution_rows": sum(
                1 for row in cleaned_rows if row["institution_name"] and row["applicant_type"] == "overseas"
            ),
        }
        misaligned_recent_imports = []
        if delete_recent_imports:
            misaligned_recent_imports = self._misaligned_recent_imports(
                workbook=workbook,
                summary=summary,
                recent_import_hours=recent_import_hours,
            )
            self._write_recent_import_delete_plan(misaligned_recent_imports, recent_import_hours)

        if dry_run:
            self.stdout.write(self.style.WARNING("Dry run selected. No database records were changed."))
            self._write_summary(summary)
            return

        with transaction.atomic():
            deleted_recent_imports = 0
            deleted_recent_import_ids = []
            if delete_recent_imports and misaligned_recent_imports:
                deleted_recent_import_ids = [item["batch"].id for item in misaligned_recent_imports]
                deleted_recent_imports = len(deleted_recent_import_ids)
                DataImportBatch.objects.filter(id__in=deleted_recent_import_ids).delete()

            replaced_batches = 0
            if not append:
                old_batches = DataImportBatch.objects.filter(
                    source_kind="nursing_license_workbook",
                    source_file_path=str(workbook),
                )
                replaced_batches = old_batches.count()
                old_batches.delete()

            batch = DataImportBatch.objects.create(
                source_file_name=workbook.name,
                source_file_path=str(workbook),
                source_kind="nursing_license_workbook",
                status="running",
                total_sheets=2,
            )
            import_stats = self._import_cleaned_rows(batch, cleaned_rows)
            batch.status = "completed"
            batch.completed_at = timezone.now()
            batch.processed_sheets = batch.sheets.filter(status="processed").count()
            batch.total_rows = summary["source_rows_seen"]
            batch.processed_rows = len(cleaned_rows)
            batch.summary = {
                **summary,
                **import_stats,
                "replaced_batches": replaced_batches,
                "deleted_recent_misaligned_imports": deleted_recent_imports,
                "deleted_recent_misaligned_import_ids": deleted_recent_import_ids,
            }
            batch.save(update_fields=["status", "completed_at", "processed_sheets", "total_rows", "processed_rows", "summary"])

        self.stdout.write(self.style.SUCCESS(f"Imported nursing licence workbook batch #{batch.id}"))
        self._write_summary(batch.summary)

    def _load_rows(self, workbook_path):
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        provisional_sheet = self._find_sheet(workbook.sheetnames, "prov")
        full_sheet = self._find_sheet(workbook.sheetnames, "full")
        if not provisional_sheet or not full_sheet:
            raise CommandError("Workbook must contain one provisional sheet and one full licence sheet.")

        provisional_rows, provisional_stats = self._read_sheet(workbook[provisional_sheet], "provisional")
        full_rows, full_stats = self._read_sheet(workbook[full_sheet], "full")
        return provisional_rows, full_rows, {
            "provisional_sheet": provisional_sheet,
            "full_sheet": full_sheet,
            "source_rows_seen": provisional_stats["source_rows_seen"] + full_stats["source_rows_seen"],
            "blank_rows_skipped": provisional_stats["blank_rows_skipped"] + full_stats["blank_rows_skipped"],
            "missing_name_rows_skipped": provisional_stats["missing_name_rows_skipped"] + full_stats["missing_name_rows_skipped"],
            "missing_registration_rows_skipped": (
                provisional_stats["missing_registration_rows_skipped"] + full_stats["missing_registration_rows_skipped"]
            ),
            "provisional_rows_ready": len(provisional_rows),
            "full_rows_ready": len(full_rows),
        }

    def _find_sheet(self, sheet_names, token):
        for sheet_name in sheet_names:
            if token in sheet_name.lower():
                return sheet_name
        return ""

    def _canonical_header(self, header, sheet_type):
        key = normalize_key(header)
        if key in {"id", "no"}:
            return "source_id"
        if key in {"name", "full_name"}:
            return "name"
        if key == "registration_no":
            return "license_type" if sheet_type == "provisional" else "registration_no"
        if key in {"license_type", "license_type_"}:
            return "license_prefix" if sheet_type == "provisional" else "license_type"
        if key in {"license_no", "no"}:
            return "license_no"
        if key in {"provisional_no", "provisional_number"}:
            return "provisional_no"
        if key in {"issued_date", "issue_date"}:
            return "issued_date"
        if key in {"institution_attended", "instititution_attended", "institution_attend", "name_of_school_institute"}:
            return "institution_name"
        if key in {"year", "graduation_year", "year_graduate"}:
            return "graduation_year"
        if key in {"qualification", "qualification_name"}:
            return "qualification_name"
        if key in {"practitioners_no", "practitioner_no", "practitioners_number"}:
            return "practitioner_number"
        return key

    def _read_sheet(self, worksheet, sheet_type):
        headers = [self._canonical_header(value, sheet_type) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows = []
        stats = {
            "source_rows_seen": 0,
            "blank_rows_skipped": 0,
            "missing_name_rows_skipped": 0,
            "missing_registration_rows_skipped": 0,
        }
        blank_streak = 0

        for row_index, values in enumerate(worksheet.iter_rows(min_row=2, max_col=len(headers), values_only=True), start=2):
            raw_values = dict(zip(headers, values))
            if not any(value is not None and normalize_text(value) for value in values):
                stats["blank_rows_skipped"] += 1
                blank_streak += 1
                if blank_streak >= 500:
                    break
                continue

            blank_streak = 0
            stats["source_rows_seen"] += 1

            full_name = normalize_name(raw_values.get("name"))
            if not full_name:
                stats["missing_name_rows_skipped"] += 1
                continue

            if sheet_type == "provisional":
                registration_no = normalize_registration_no(raw_values.get("license_prefix") or "PRO", raw_values.get("provisional_no"))
                form_code = "NC1"
                form_title = "Application for Provisional Licence"
                qualification_type = "Provisional Licence"
            else:
                registration_no = normalize_registration_no("", raw_values.get("license_no") or raw_values.get("registration_no"))
                form_code = "NC2"
                form_title = "Application for Full Licence"
                qualification_type = "Full Licence"

            if not registration_no:
                source_id = clean_identifier(raw_values.get("source_id"))
                registration_no = normalize_registration_no("PROV" if sheet_type == "provisional" else "FULL", source_id)

            if not registration_no:
                stats["missing_registration_rows_skipped"] += 1
                continue

            issued_date = parse_date(raw_values.get("issued_date"))
            graduation_year = parse_year(raw_values.get("graduation_year"), issued_date)
            institution_name = normalize_text(raw_values.get("institution_name"))
            if institution_name.upper() in NULL_LIKE_VALUES:
                institution_name = ""
            qualification_name = normalize_text(raw_values.get("qualification_name"))
            applicant_type = infer_applicant_type(institution_name)
            profession_track = infer_profession_track(qualification_name)
            target_model = model_key_for_qualification(qualification_name)
            first_name, last_name = split_name(full_name)

            rows.append(
                {
                    "source_sheet": worksheet.title,
                    "source_row": row_index,
                    "source_id": clean_identifier(raw_values.get("source_id")),
                    "record_type": sheet_type,
                    "target_model": target_model,
                    "form_code": form_code,
                    "form_title": form_title,
                    "qualification_type": qualification_type,
                    "full_name": full_name,
                    "first_name": first_name,
                    "last_name": last_name,
                    "registration_no": registration_no[:100],
                    "practitioner_number": clean_identifier(raw_values.get("practitioner_number")),
                    "applicant_type": applicant_type,
                    "nationality": "Papua New Guinea" if applicant_type == "national" else "",
                    "institution_name": institution_name[:255],
                    "qualification_name": qualification_name[:255],
                    "graduation_year": graduation_year,
                    "issued_date": issued_date,
                    "record_year": issued_date.year if issued_date else graduation_year,
                    "profession_track": profession_track,
                    "raw_payload": {str(key): json_safe(value) for key, value in raw_values.items()},
                }
            )

        return rows, stats

    def _dedupe_rows(self, rows):
        exact_rows = {}
        duplicate_rows_dropped = 0

        for row in rows:
            key = (row["record_type"], row["registration_no"].upper(), row["full_name"].upper())
            existing = exact_rows.get(key)
            if existing is None:
                exact_rows[key] = row
                continue

            duplicate_rows_dropped += 1
            if row_score(row) > row_score(existing):
                exact_rows[key] = row

        repaired_registration_conflicts = 0
        rows_by_registration = {}
        for row in exact_rows.values():
            key = (row["record_type"], row["registration_no"].upper())
            rows_by_registration.setdefault(key, []).append(row)

        cleaned_rows = []
        for group in rows_by_registration.values():
            group = sorted(group, key=lambda item: (item["source_row"], item["full_name"]))
            names = {row["full_name"].upper() for row in group}
            if len(names) == 1:
                cleaned_rows.extend(group)
                continue

            for index, row in enumerate(group):
                if index == 0:
                    cleaned_rows.append(row)
                    continue
                repaired = row.copy()
                repaired["raw_payload"] = {
                    **repaired.get("raw_payload", {}),
                    "source_registration_no": repaired["registration_no"],
                    "registration_no_repair": "duplicate_registration_name_conflict",
                }
                suffix = repaired["source_id"] or f"ROW{repaired['source_row']}"
                repaired["registration_no"] = f"{repaired['registration_no']} ID{suffix}"[:100]
                cleaned_rows.append(repaired)
                repaired_registration_conflicts += 1

        cleaned = sorted(
            cleaned_rows,
            key=lambda item: (item["record_type"], item["registration_no"], item["source_row"]),
        )
        return cleaned, {
            "duplicate_rows_dropped": duplicate_rows_dropped,
            "duplicate_registration_repairs": repaired_registration_conflicts,
        }

    def _misaligned_recent_imports(self, workbook, summary, recent_import_hours):
        since = timezone.now() - timedelta(hours=recent_import_hours)
        recent_batches = (
            DataImportBatch.objects.filter(
                source_kind="nursing_license_workbook",
                started_at__gte=since,
            )
            .exclude(status="running")
            .order_by("-started_at")
        )

        misaligned = []
        for batch in recent_batches:
            reasons = self._batch_alignment_reasons(batch, workbook, summary)
            if reasons:
                misaligned.append({"batch": batch, "reasons": reasons})
        return misaligned

    def _batch_alignment_reasons(self, batch, workbook, summary):
        reasons = []
        batch_path = normalize_text(batch.source_file_path).casefold()
        current_path = str(workbook).casefold()
        if batch_path != current_path:
            reasons.append("source_file_path")

        stored_summary = batch.summary or {}
        for key in ALIGNMENT_SUMMARY_KEYS:
            if stored_summary.get(key) != summary.get(key):
                reasons.append(key)

        if batch.processed_rows and batch.processed_rows != summary["cleaned_rows"]:
            reasons.append("processed_rows")
        return reasons

    def _import_cleaned_rows(self, batch, rows):
        nursing_cadre, _ = Cadre.objects.get_or_create(name="Nursing", defaults={"category": "nursing"})
        midwifery_cadre, _ = Cadre.objects.get_or_create(name="Midwifery", defaults={"category": "midwifery"})
        content_types = {
            "nursingprofessional": ContentType.objects.get_for_model(NursingProfessional),
            "midwife": ContentType.objects.get_for_model(Midwife),
        }
        model_map = {
            "nursingprofessional": NursingProfessional,
            "midwife": Midwife,
        }
        sheet_map = {}
        for sheet_name, sheet_rows in self._rows_by_sheet(rows).items():
            sheet_map[sheet_name] = ImportedWorkbookSheet.objects.create(
                batch=batch,
                sheet_name=sheet_name,
                sheet_type="provisional" if any(row["record_type"] == "provisional" for row in sheet_rows) else "full",
                status="processed",
                raw_rows=len(sheet_rows),
                imported_rows=len(sheet_rows),
            )

        stats = {
            "license_records_created": 0,
            "professionals_created": 0,
            "professionals_updated": 0,
            "applications_created": 0,
            "applications_updated": 0,
            "qualifications_created": 0,
            "qualifications_updated": 0,
            "institutions_created": 0,
            "institutions_updated": 0,
        }

        for row in rows:
            sheet = sheet_map[row["source_sheet"]]
            PracticingLicenseRecord.objects.create(
                batch=batch,
                sheet=sheet,
                source_sheet_name=row["source_sheet"],
                source_row=row["source_row"],
                record_type=row["record_type"],
                target_model=row["target_model"],
                record_year=row["record_year"],
                full_name=row["full_name"][:255],
                first_name=row["first_name"][:100],
                last_name=row["last_name"][:100],
                registration_no=row["registration_no"][:100],
                practitioner_number=row["practitioner_number"][:100],
                applicant_type=row["applicant_type"],
                nationality=row["nationality"],
                qualification_name=row["qualification_name"][:255],
                category=row["qualification_type"],
                institution_name=row["institution_name"][:255],
                issued_date=row["issued_date"],
                raw_payload=row["raw_payload"],
            )
            stats["license_records_created"] += 1

            model = model_map[row["target_model"]]
            content_type = content_types[row["target_model"]]
            cadre = midwifery_cadre if row["target_model"] == "midwife" else nursing_cadre
            professional, created = self._upsert_professional(model, cadre, row)
            stats["professionals_created" if created else "professionals_updated"] += 1

            institution, inst_created, inst_updated = self._upsert_institution(row)
            stats["institutions_created"] += int(inst_created)
            stats["institutions_updated"] += int(inst_updated)

            qualification_created = self._upsert_qualification(content_type, professional, institution, row)
            stats["qualifications_created" if qualification_created else "qualifications_updated"] += 1

            application_created = self._upsert_application(content_type, professional, row)
            stats["applications_created" if application_created else "applications_updated"] += 1

        return stats

    def _rows_by_sheet(self, rows):
        grouped = {}
        for row in rows:
            grouped.setdefault(row["source_sheet"], []).append(row)
        return grouped

    def _upsert_professional(self, model, cadre, row):
        registration_no = row["registration_no"][:50]
        professional = model.objects.filter(registration_no=registration_no).first()
        created = False
        if professional is None:
            professional = model(registration_no=registration_no)
            created = True

        professional.first_name = row["first_name"][:100] or professional.first_name
        professional.last_name = row["last_name"][:100] or professional.last_name
        professional.applicant_type = row["applicant_type"] or professional.applicant_type
        professional.date_issued = row["issued_date"] or professional.date_issued
        if row["issued_date"]:
            professional.license_expiry_date = (
                row["issued_date"] + timedelta(days=180)
                if row["record_type"] == "provisional"
                else row["issued_date"] + timedelta(days=1095)
            )
        if row["qualification_name"]:
            professional.qualification_level = row["qualification_name"][:100]
        if (
            row["practitioner_number"]
            and len(row["practitioner_number"]) <= 50
            and not model.objects.exclude(pk=professional.pk).filter(registration_number=row["practitioner_number"]).exists()
        ):
            professional.registration_number = row["practitioner_number"]
        professional.cadre = cadre
        professional.is_active = True
        professional.save()
        return professional, created

    def _upsert_institution(self, row):
        institution_name = row["institution_name"]
        if not institution_name:
            return None, False, False
        institution_type = classify_training_institution(institution_name)
        institution, created = TrainingInstitution.objects.get_or_create(
            name=institution_name,
            defaults={"type": institution_type, "is_active": True},
        )
        updated = False
        if not created and institution.type != institution_type:
            institution.type = institution_type
            institution.save(update_fields=["type"])
            updated = True
        return institution, created, updated

    def _upsert_qualification(self, content_type, professional, institution, row):
        qualification_name = row["qualification_name"] or row["qualification_type"]
        qualification, created = Qualification.objects.update_or_create(
            content_type=content_type,
            object_id=professional.id,
            qualification_name=qualification_name[:200],
            defaults={
                "institution": institution,
                "institution_name": row["institution_name"][:255],
                "program_completed": qualification_name[:255],
                "completion_year": row["graduation_year"],
                "qualification_type": row["qualification_type"],
                "country": "Papua New Guinea" if row["applicant_type"] == "national" else "",
            },
        )
        return created

    def _upsert_application(self, content_type, professional, row):
        expiry_date = None
        if row["issued_date"]:
            expiry_date = (
                row["issued_date"] + timedelta(days=180)
                if row["record_type"] == "provisional"
                else row["issued_date"] + timedelta(days=1095)
            )
        application, created = Application.objects.update_or_create(
            content_type=content_type,
            object_id=professional.id,
            form_code=row["form_code"],
            defaults={
                "pathway": "overseas_midwife"
                if row["applicant_type"] == "overseas" and row["profession_track"] == "midwifery"
                else "overseas_nurse"
                if row["applicant_type"] == "overseas"
                else "local_midwifery_graduate"
                if row["profession_track"] == "midwifery"
                else "local_nursing_graduate",
                "form_title": row["form_title"],
                "profession_track": row["profession_track"],
                "status": "approved",
                "approved_date": row["issued_date"],
                "expiry_date": expiry_date,
                "payload": {
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                    "source_id": row["source_id"],
                    "registration_no": row["registration_no"],
                    "institution_name": row["institution_name"],
                },
                "reviewer_notes": (
                    f"Imported from {row['source_sheet']} row {row['source_row']} "
                    f"as {row['qualification_type']}."
                ),
            },
        )
        if row["issued_date"]:
            application.submitted_date = row["issued_date"]
            application.save(update_fields=["submitted_date"])
        return created

    def _write_summary(self, summary):
        for key, value in summary.items():
            self.stdout.write(f"{key}: {value}")

    def _write_recent_import_delete_plan(self, misaligned_recent_imports, recent_import_hours):
        if not misaligned_recent_imports:
            self.stdout.write(
                self.style.SUCCESS(
                    f"No misaligned nursing licence import batches found in the last {recent_import_hours} hour(s)."
                )
            )
            return

        self.stdout.write(
            self.style.WARNING(
                f"Misaligned nursing licence import batches found in the last {recent_import_hours} hour(s):"
            )
        )
        for item in misaligned_recent_imports:
            batch = item["batch"]
            reasons = ", ".join(item["reasons"])
            self.stdout.write(f"  batch #{batch.id}: {batch.source_file_name} ({reasons})")
