from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.contrib.contenttypes.models import ContentType
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

from apps.dashboard.models import Receipt
from apps.workforce.models import (
    AuditLog,
    Application,
    CommunityHealthWorker,
    HealthStudent,
    MedicalDoctor,
    Midwife,
    NurseAide,
    NursingProfessional,
    PracticingLicenseRecord,
)
from apps.workforce.services.ndata_workbook_import import (
    infer_target_model,
    normalize_identifier,
    normalize_name,
    normalize_registration_no,
    normalize_text,
    parse_date,
    parse_decimal,
    qualification_for_target,
)


DEFAULT_WORKBOOK = Path(
    r"C:\Users\darre\Documents\ProjectApps\databasedocuments\spreadsheets"
    r"\ATP_LATEST\ATP_LATEST_FROM_JOYCE\2026 Current ATP-DATA Statistics & Tracking latest.xlsx"
)
DEFAULT_SHEET = "ATP RECORD 2026"
PLACEHOLDER_VALUES = {"", "-", "--", "N/A", "NA", "NIL", "NONE", "NULL", "UNKNOWN", "TBA", "TBD"}
LIVE_MODELS = (
    ("nursing", "Nursing Professional", NursingProfessional),
    ("nursing", "Midwife", Midwife),
    ("nursing", "Nurse Aide", NurseAide),
    ("nursing", "Graduand / Health Student", HealthStudent),
    ("medical", "Medical Doctor", MedicalDoctor),
    ("medical", "Community Health Worker", CommunityHealthWorker),
)
PROFESSIONAL_MODEL_MAP = {
    "nursingprofessional": NursingProfessional,
    "midwife": Midwife,
    "nurseaide": NurseAide,
    "healthstudent": HealthStudent,
    "medicaldoctor": MedicalDoctor,
    "communityhealthworker": CommunityHealthWorker,
}
G_FORM_CODES = {"G1", "G2", "G3", "G4", "G5", "G6", "G7"}


@dataclass
class AtpWorkbookRow:
    row_number: int
    full_name: str
    normalized_name: str
    gender: str
    date_of_birth: object
    registration_no: str
    compact_registration: str
    registration_suffix: str
    practitioner_number: str
    category: str
    qualification_name: str
    nationality: str
    workplace_address: str
    province: str
    payment_date: object
    renewal_fee: Decimal | None
    overseas_fee: Decimal | None
    late_fee: Decimal | None
    receipt_number: str
    target_model: str


def clean_value(value) -> str:
    text = normalize_text(value)
    return "" if text.upper() in PLACEHOLDER_VALUES else text


def compact_identifier(value) -> str:
    text = clean_value(value).upper()
    return re.sub(r"[^A-Z0-9]+", "", text)


def registration_suffix(value) -> str:
    compact = compact_identifier(value)
    if not compact:
        return ""
    match = re.search(r"(\d{2,})$", compact)
    return match.group(1) if match else compact


def is_provisional_registration(value) -> bool:
    compact = compact_identifier(value)
    return compact.startswith("PROV") or compact.startswith("PRO")


def clean_receipt(value) -> str:
    text = clean_value(value).upper().replace(" ", "")
    if text in PLACEHOLDER_VALUES:
        return ""
    return text[:120]


def amount_from_row(row: AtpWorkbookRow) -> Decimal | None:
    parts = [row.renewal_fee, row.overseas_fee, row.late_fee]
    values = [value for value in parts if value is not None]
    if not values:
        return None
    return sum(values, Decimal("0"))


def identity_key_for_object(obj, label: str) -> tuple[str, str]:
    reg = compact_identifier(getattr(obj, "registration_no", ""))
    reg2 = compact_identifier(getattr(obj, "registration_number", ""))
    name = normalize_name(f"{getattr(obj, 'first_name', '')} {getattr(obj, 'last_name', '')}").upper()
    dob = getattr(obj, "date_of_birth", None)
    if reg:
        return "strong_registration", f"{label}:registration:{reg}"
    if reg2:
        return "strong_registration_number", f"{label}:registration_number:{reg2}"
    if name and dob:
        return "strong_name_dob", f"{label}:name_dob:{name}:{dob.isoformat()}"
    if name:
        return "weak_name_only", f"{label}:name:{name}"
    return "missing_identity", f"{label}:missing:{obj.pk}"


class Command(BaseCommand):
    help = (
        "Backfill missing 2026 ATP receipt/DOB details from the source workbook and "
        "produce separated full, ATP, provisional, and graduand/student statistics."
    )

    def add_arguments(self, parser):
        parser.add_argument("--file", default=str(DEFAULT_WORKBOOK), help="Path to the 2026 ATP workbook.")
        parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet containing ATP 2026 rows.")
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Apply high-confidence missing-field updates. Without this, only a dry-run report is written.",
        )
        parser.add_argument("--report-dir", default="docs/reports", help="Directory for reconciliation reports.")

    def handle(self, *args, **options):
        workbook_path = Path(options["file"])
        sheet_name = options["sheet"]
        apply_changes = options["apply"]
        report_dir = Path(options["report_dir"])
        report_dir.mkdir(parents=True, exist_ok=True)

        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")

        self.stdout.write(f"Reading ATP source workbook: {workbook_path}")
        source_rows = self._read_source_rows(workbook_path, sheet_name)
        if not source_rows:
            raise CommandError(f"No ATP source rows found in sheet '{sheet_name}'.")

        timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
        with transaction.atomic():
            result = self._reconcile(source_rows, apply_changes=apply_changes)
            if not apply_changes:
                transaction.set_rollback(True)

        result["mode"] = "apply" if apply_changes else "dry_run"
        result["source_workbook"] = str(workbook_path)
        result["source_sheet"] = sheet_name
        result["generated_at"] = timezone.localtime().isoformat()
        result["statistics"] = self._statistics()

        paths = self._write_reports(report_dir, timestamp, result)
        if apply_changes:
            AuditLog.objects.create(
                action="reconcile_2026_atp_receipts",
                entity_type="workforce.practicing_license_record",
                entity_id="ATP RECORD 2026",
                new_values_json={
                    "workbook": str(workbook_path),
                    "updated_practice_records": result["updated_practice_records"],
                    "updated_professionals": result["updated_professionals"],
                    "receipts_created": result["receipts_created"],
                    "receipts_updated": result["receipts_updated"],
                    "report": str(paths["markdown"]),
                },
            )

        self.stdout.write(self.style.SUCCESS("ATP 2026 reconciliation complete."))
        self.stdout.write(f"Mode: {result['mode']}")
        self.stdout.write(f"Source rows read: {result['source_rows_read']}")
        self.stdout.write(f"Matched source rows: {result['matched_source_rows']}")
        self.stdout.write(f"Unmatched source rows: {result['unmatched_source_rows']}")
        self.stdout.write(f"Practice records updated: {result['updated_practice_records']}")
        self.stdout.write(f"Professional records updated: {result['updated_professionals']}")
        self.stdout.write(f"Receipts created/updated: {result['receipts_created']}/{result['receipts_updated']}")
        self.stdout.write(f"Conflicts requiring review: {len(result['conflicts'])}")
        self.stdout.write(f"Report: {paths['markdown']}")
        self.stdout.write(f"JSON: {paths['json']}")

    def _read_source_rows(self, workbook_path: Path, sheet_name: str) -> list[AtpWorkbookRow]:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise CommandError(f"Worksheet '{sheet_name}' not found. Available: {', '.join(workbook.sheetnames)}")

        worksheet = workbook[sheet_name]
        rows: list[AtpWorkbookRow] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=3, max_col=16, values_only=True),
            start=3,
        ):
            values = list(row) + [None] * (16 - len(row))
            full_name = normalize_name(values[1])
            if not full_name:
                continue
            registration_no = normalize_registration_no(values[4])
            category = clean_value(values[6])
            qualification_name = clean_value(values[7])
            receipt_number = clean_receipt(values[15])
            target_model = infer_target_model(
                category=category,
                qualification=qualification_name,
                registration_no=registration_no,
            )
            rows.append(
                AtpWorkbookRow(
                    row_number=row_number,
                    full_name=full_name,
                    normalized_name=normalize_name(full_name).upper(),
                    gender=clean_value(values[2]),
                    date_of_birth=parse_date(values[3]),
                    registration_no=registration_no,
                    compact_registration=compact_identifier(registration_no),
                    registration_suffix=registration_suffix(registration_no),
                    practitioner_number=clean_value(normalize_identifier(values[5])),
                    category=category,
                    qualification_name=qualification_name,
                    nationality=clean_value(values[8]),
                    workplace_address=clean_value(values[9]),
                    province=clean_value(values[10]),
                    payment_date=parse_date(values[11]),
                    renewal_fee=parse_decimal(values[12]),
                    overseas_fee=parse_decimal(values[13]),
                    late_fee=parse_decimal(values[14]),
                    receipt_number=receipt_number,
                    target_model=target_model,
                )
            )
        return rows

    def _reconcile(self, source_rows: list[AtpWorkbookRow], apply_changes: bool) -> dict:
        practice_by_source_row, practice_by_reg_name, practice_by_suffix_name = self._build_practice_indexes()
        professionals_by_reg_name, professionals_by_suffix_name = self._build_professional_indexes()

        result = {
            "source_rows_read": len(source_rows),
            "matched_source_rows": 0,
            "unmatched_source_rows": 0,
            "updated_practice_records": 0,
            "updated_professionals": 0,
            "practice_field_updates": Counter(),
            "professional_field_updates": Counter(),
            "receipts_created": 0,
            "receipts_updated": 0,
            "receipts_skipped_missing_amount_or_date": 0,
            "conflicts": [],
            "unmatched_rows": [],
            "sample_rows": [],
        }

        for row in source_rows:
            practice_records = self._match_practice_records(
                row,
                practice_by_source_row,
                practice_by_reg_name,
                practice_by_suffix_name,
            )
            if practice_records:
                result["matched_source_rows"] += 1
            else:
                result["unmatched_source_rows"] += 1
                result["unmatched_rows"].append(self._source_row_summary(row))

            for record in practice_records:
                changed_fields = self._update_practice_record(record, row, result)
                if changed_fields:
                    result["updated_practice_records"] += 1
                    if apply_changes:
                        record.save(update_fields=changed_fields + ["updated_at"])

            professional_records = self._match_professionals(
                row,
                professionals_by_reg_name,
                professionals_by_suffix_name,
            )
            for professional in professional_records:
                changed_fields = self._update_professional(professional, row, result)
                if changed_fields:
                    result["updated_professionals"] += 1
                    if apply_changes:
                        professional.save(update_fields=changed_fields)

            receipt_action = self._upsert_receipt(row, apply_changes)
            if receipt_action == "created":
                result["receipts_created"] += 1
            elif receipt_action == "updated":
                result["receipts_updated"] += 1
            elif receipt_action == "skipped":
                result["receipts_skipped_missing_amount_or_date"] += 1

            if row.full_name.lower() in {"everestar aitsi", "everester aitsi"}:
                result["sample_rows"].append(
                    {
                        **self._source_row_summary(row),
                        "matched_practice_record_ids": [record.id for record in practice_records],
                        "matched_professional_record_ids": [
                            f"{obj.__class__.__name__}:{obj.id}" for obj in professional_records
                        ],
                    }
                )

        result["practice_field_updates"] = dict(result["practice_field_updates"])
        result["professional_field_updates"] = dict(result["professional_field_updates"])
        return result

    def _build_practice_indexes(self):
        by_source_row = defaultdict(list)
        by_reg_name = defaultdict(list)
        by_suffix_name = defaultdict(list)
        queryset = PracticingLicenseRecord.objects.filter(record_type="practicing_license", record_year=2026)
        for record in queryset.iterator(chunk_size=2000):
            if record.source_sheet_name == DEFAULT_SHEET:
                by_source_row[record.source_row].append(record)
            name = normalize_name(record.full_name).upper()
            reg = compact_identifier(record.registration_no)
            suffix = registration_suffix(record.registration_no)
            if reg and name:
                by_reg_name[(reg, name)].append(record)
            if suffix and name:
                by_suffix_name[(suffix, name)].append(record)
        return by_source_row, by_reg_name, by_suffix_name

    def _build_professional_indexes(self):
        by_reg_name = defaultdict(list)
        by_suffix_name = defaultdict(list)
        for _scope, label, model in LIVE_MODELS:
            for obj in model.objects.filter(is_active=True).iterator(chunk_size=2000):
                name = normalize_name(f"{getattr(obj, 'first_name', '')} {getattr(obj, 'last_name', '')}").upper()
                registrations = {
                    compact_identifier(getattr(obj, "registration_no", "")),
                    compact_identifier(getattr(obj, "registration_number", "")),
                }
                suffixes = {
                    registration_suffix(getattr(obj, "registration_no", "")),
                    registration_suffix(getattr(obj, "registration_number", "")),
                }
                for reg in registrations:
                    if reg and name:
                        by_reg_name[(reg, name)].append(obj)
                for suffix in suffixes:
                    if suffix and name:
                        by_suffix_name[(suffix, name)].append(obj)
        return by_reg_name, by_suffix_name

    def _match_practice_records(self, row, by_source_row, by_reg_name, by_suffix_name):
        matches = list(by_source_row.get(row.row_number, []))
        if not matches and row.compact_registration:
            matches = list(by_reg_name.get((row.compact_registration, row.normalized_name), []))
        if not matches and row.registration_suffix:
            matches = list(by_suffix_name.get((row.registration_suffix, row.normalized_name), []))
        return self._unique_objects(matches)

    def _match_professionals(self, row, by_reg_name, by_suffix_name):
        matches = []
        if row.compact_registration:
            matches.extend(by_reg_name.get((row.compact_registration, row.normalized_name), []))
        if row.registration_suffix:
            matches.extend(by_suffix_name.get((row.registration_suffix, row.normalized_name), []))
        return self._unique_objects(matches)

    def _unique_objects(self, objects):
        seen = set()
        unique = []
        for obj in objects:
            key = (obj.__class__.__name__, obj.pk)
            if key in seen:
                continue
            seen.add(key)
            unique.append(obj)
        return unique

    def _update_practice_record(self, record, row, result) -> list[str]:
        changed = []
        field_values = {
            "date_of_birth": row.date_of_birth,
            "reference_number": row.receipt_number,
            "payment_method": row.receipt_number if row.receipt_number else "",
            "payment_date": row.payment_date,
            "renewal_fee": row.renewal_fee,
            "overseas_fee": row.overseas_fee,
            "late_fee": row.late_fee,
            "practitioner_number": row.practitioner_number,
            "qualification_name": row.qualification_name,
            "workplace_address": row.workplace_address,
            "province": row.province,
            "gender": row.gender,
            "nationality": row.nationality,
            "category": row.category,
        }
        for field, value in field_values.items():
            if self._set_missing_or_log_conflict(
                record,
                field,
                value,
                row,
                result,
                conflict_group="practice_record",
            ):
                changed.append(field)
                result["practice_field_updates"][field] += 1
        return changed

    def _update_professional(self, professional, row, result) -> list[str]:
        changed = []
        field_values = {
            "date_of_birth": row.date_of_birth,
            "gender": row.gender if row.gender in {"Male", "Female"} else "",
            "province": row.province[:100] if row.province else "",
            "full_address": row.workplace_address,
            "registration_number": row.practitioner_number[:50]
            if row.practitioner_number and row.practitioner_number.upper() not in PLACEHOLDER_VALUES
            else "",
        }
        if hasattr(professional, "qualification_level"):
            field_values["qualification_level"] = qualification_for_target(
                row.target_model,
                row.category,
                row.qualification_name or row.category,
            )
        for field, value in field_values.items():
            if not hasattr(professional, field):
                continue
            if self._set_missing_or_log_conflict(
                professional,
                field,
                value,
                row,
                result,
                conflict_group="professional",
            ):
                changed.append(field)
                result["professional_field_updates"][field] += 1
        return changed

    def _set_missing_or_log_conflict(self, obj, field, value, row, result, conflict_group) -> bool:
        if value in {None, ""}:
            return False
        current = getattr(obj, field)
        current_blank = current in {None, ""}
        if isinstance(current, str):
            current_blank = clean_value(current) == ""
        if current_blank:
            try:
                model_field = obj._meta.get_field(field)
            except Exception:
                model_field = None
            if (
                model_field is not None
                and getattr(model_field, "unique", False)
                and obj.__class__.objects.exclude(pk=obj.pk).filter(**{field: value}).exists()
            ):
                result["conflicts"].append(
                    {
                        "group": conflict_group,
                        "model": obj.__class__.__name__,
                        "record_id": obj.pk,
                        "source_row": row.row_number,
                        "full_name": row.full_name,
                        "registration_no": row.registration_no,
                        "field": field,
                        "existing_value": "",
                        "source_value": str(value),
                        "action": "skipped_unique_conflict",
                    }
                )
                return False
            setattr(obj, field, value)
            return True
        if str(current).strip() != str(value).strip():
            result["conflicts"].append(
                {
                    "group": conflict_group,
                    "model": obj.__class__.__name__,
                    "record_id": obj.pk,
                    "source_row": row.row_number,
                    "full_name": row.full_name,
                    "registration_no": row.registration_no,
                    "field": field,
                    "existing_value": str(current),
                    "source_value": str(value),
                    "action": "not_overwritten",
                }
            )
        return False

    def _upsert_receipt(self, row: AtpWorkbookRow, apply_changes: bool) -> str:
        amount = amount_from_row(row)
        if not row.receipt_number or not row.payment_date or amount is None:
            return "skipped" if row.receipt_number else "none"
        if not apply_changes:
            return "created" if not Receipt.objects.filter(official_receipt_no=row.receipt_number[:50]).exists() else "updated"

        receipt_datetime = timezone.make_aware(datetime.combine(row.payment_date, datetime.min.time()))
        receipt, created = Receipt.objects.update_or_create(
            official_receipt_no=row.receipt_number[:50],
            defaults={
                "amount": amount,
                "description": f"Imported ATP 2026 receipt for {row.full_name}",
                "status": "completed",
                "receipt_date": receipt_datetime,
                "payment_method": "Imported ATP receipt",
                "practitioner_number": row.practitioner_number[:100],
            },
        )
        return "created" if created else "updated"

    def _statistics(self) -> dict:
        live_rows = []
        totals = Counter()
        provisional_live = self._live_provisional_summary()
        for scope, label, model in LIVE_MODELS:
            queryset = model.objects.filter(is_active=True)
            if model in {NursingProfessional, Midwife, NurseAide}:
                queryset = queryset.exclude(Q(registration_no__istartswith="PROV") | Q(registration_no__istartswith="PRO "))
            distinct, weak, missing, duplicates = self._trusted_distinct_queryset(queryset, label)
            live_rows.append(
                {
                    "scope": scope,
                    "category": label,
                    "raw_active_rows": queryset.count(),
                    "trusted_distinct_workers": distinct,
                    "weak_name_only_identities": weak,
                    "missing_identity_rows": missing,
                    "duplicate_identity_groups": duplicates,
                }
            )
            totals["raw_active_rows"] += queryset.count()
            totals["trusted_distinct_workers"] += distinct
            totals["weak_name_only_identities"] += weak
            totals["missing_identity_rows"] += missing
            totals["duplicate_identity_groups"] += duplicates

        imported = {
            "atp_2026": self._imported_distinct(PracticingLicenseRecord.objects.filter(record_type="practicing_license", record_year=2026)),
            "atp_all_years": self._imported_distinct(PracticingLicenseRecord.objects.filter(record_type="practicing_license")),
            "full_license_all_years": self._imported_distinct(PracticingLicenseRecord.objects.filter(record_type="full")),
            "provisional_license_all_years": self._imported_distinct(PracticingLicenseRecord.objects.filter(record_type="provisional")),
            "provisional_rows_mapped_to_healthstudent": PracticingLicenseRecord.objects.filter(
                record_type="provisional",
                target_model="healthstudent",
            ).count(),
            "g_form_graduand_applications": Application.objects.filter(form_code__in=G_FORM_CODES).count(),
        }
        return {
            "live_registered_workers_excluding_provisional_prefix": {
                "rows": live_rows,
                "totals": dict(totals),
            },
            "live_provisional_licence_holders": provisional_live,
            "imported_license_categories": imported,
            "counting_note": (
                "PRO/PROV registration prefixes are separated as provisional licence holders. "
                "Graduands/students are counted only from HealthStudent records and G1-G7 form applications. "
                "Imported provisional rows mapped to healthstudent are flagged for reclassification review, "
                "not treated as G-form graduands."
            ),
        }

    def _live_provisional_summary(self) -> dict:
        rows = []
        total = Counter()
        for model in (NursingProfessional, Midwife, NurseAide):
            queryset = model.objects.filter(is_active=True).filter(
                Q(registration_no__istartswith="PROV") | Q(registration_no__istartswith="PRO ")
            )
            distinct, weak, missing, duplicates = self._trusted_distinct_queryset(queryset, model.__name__)
            row = {
                "model": model.__name__,
                "raw_active_rows": queryset.count(),
                "trusted_distinct_workers": distinct,
                "weak_name_only_identities": weak,
                "missing_identity_rows": missing,
                "duplicate_identity_groups": duplicates,
            }
            rows.append(row)
            for key, value in row.items():
                if isinstance(value, int):
                    total[key] += value
        return {"rows": rows, "totals": dict(total)}

    def _trusted_distinct_queryset(self, queryset, label: str):
        identities = defaultdict(int)
        for obj in queryset.iterator(chunk_size=2000):
            identities[identity_key_for_object(obj, label)] += 1
        distinct = sum(1 for (strength, _key) in identities if strength.startswith("strong"))
        weak = sum(1 for (strength, _key) in identities if strength == "weak_name_only")
        missing = sum(count for (strength, _key), count in identities.items() if strength == "missing_identity")
        duplicates = sum(1 for count in identities.values() if count > 1)
        return distinct, weak, missing, duplicates

    def _imported_distinct(self, queryset) -> dict:
        identities = defaultdict(int)
        target_counter = Counter()
        for record in queryset.iterator(chunk_size=4000):
            target = record.target_model or "other"
            reg = compact_identifier(record.registration_no)
            practitioner = compact_identifier(record.practitioner_number)
            name = normalize_name(record.full_name).upper()
            dob = record.date_of_birth.isoformat() if record.date_of_birth else ""
            if reg:
                key = ("strong_registration", f"{target}:registration:{reg}")
            elif practitioner:
                key = ("strong_practitioner", f"{target}:practitioner:{practitioner}")
            elif name and dob:
                key = ("strong_name_dob", f"{target}:name_dob:{name}:{dob}")
            elif name:
                key = ("weak_name_only", f"{target}:name:{name}")
            else:
                key = ("missing_identity", f"{target}:missing:{record.id}")
            identities[key] += 1
            target_counter[target] += 1
        return {
            "raw_rows": queryset.count(),
            "trusted_distinct_workers": sum(1 for (strength, _key) in identities if strength.startswith("strong")),
            "estimated_distinct_workers_including_name_only": sum(
                1 for (strength, _key) in identities if strength.startswith("strong") or strength == "weak_name_only"
            ),
            "duplicate_identity_groups": sum(1 for count in identities.values() if count > 1),
            "duplicate_rows_collapsed": sum(max(count - 1, 0) for count in identities.values()),
            "by_target_model": dict(sorted(target_counter.items())),
        }

    def _source_row_summary(self, row: AtpWorkbookRow) -> dict:
        return {
            "source_row": row.row_number,
            "full_name": row.full_name,
            "registration_no": row.registration_no,
            "practitioner_number": row.practitioner_number,
            "date_of_birth": row.date_of_birth.isoformat() if row.date_of_birth else "",
            "receipt_number": row.receipt_number,
            "payment_date": row.payment_date.isoformat() if row.payment_date else "",
            "category": row.category,
            "target_model": row.target_model,
        }

    def _write_reports(self, report_dir: Path, timestamp: str, result: dict) -> dict:
        json_path = report_dir / f"atp_2026_reconciliation_{timestamp}.json"
        md_path = report_dir / f"atp_2026_reconciliation_{timestamp}.md"
        conflicts_path = report_dir / f"atp_2026_reconciliation_conflicts_{timestamp}.csv"
        unmatched_path = report_dir / f"atp_2026_reconciliation_unmatched_{timestamp}.csv"
        json_path.write_text(json.dumps(result, indent=2, default=str), encoding="utf-8")
        md_path.write_text(self._markdown(result), encoding="utf-8")
        self._write_csv(conflicts_path, result["conflicts"])
        self._write_csv(unmatched_path, result["unmatched_rows"])
        return {
            "json": json_path,
            "markdown": md_path,
            "conflicts_csv": conflicts_path,
            "unmatched_csv": unmatched_path,
        }

    def _markdown(self, result: dict) -> str:
        stats = result["statistics"]
        lines = [
            "# ATP 2026 Reconciliation And Registry Segregation Report",
            "",
            f"Generated at: {result['generated_at']}",
            f"Mode: {result['mode']}",
            f"Source workbook: {result['source_workbook']}",
            f"Source sheet: {result['source_sheet']}",
            "",
            "## Reconciliation Results",
            "",
            f"- Source rows read: {result['source_rows_read']}",
            f"- Matched source rows: {result['matched_source_rows']}",
            f"- Unmatched source rows: {result['unmatched_source_rows']}",
            f"- Existing ATP practice records updated: {result['updated_practice_records']}",
            f"- Existing professional records updated: {result['updated_professionals']}",
            f"- Receipt records created: {result['receipts_created']}",
            f"- Receipt records updated: {result['receipts_updated']}",
            f"- Conflicts requiring registrar review: {len(result['conflicts'])}",
            "",
            "## Live Registered Workers Excluding PRO/PROV Provisional Prefixes",
            "",
            "| Category | Raw active rows | Trusted distinct workers | Weak identities | Missing identity rows | Duplicate groups |",
            "|---|---:|---:|---:|---:|---:|",
        ]
        for row in stats["live_registered_workers_excluding_provisional_prefix"]["rows"]:
            lines.append(
                f"| {row['category']} | {row['raw_active_rows']} | {row['trusted_distinct_workers']} | "
                f"{row['weak_name_only_identities']} | {row['missing_identity_rows']} | {row['duplicate_identity_groups']} |"
            )
        totals = stats["live_registered_workers_excluding_provisional_prefix"]["totals"]
        lines.append(
            f"| Total | {totals.get('raw_active_rows', 0)} | {totals.get('trusted_distinct_workers', 0)} | "
            f"{totals.get('weak_name_only_identities', 0)} | {totals.get('missing_identity_rows', 0)} | "
            f"{totals.get('duplicate_identity_groups', 0)} |"
        )
        lines.extend(
            [
                "",
                "## Live Provisional Licence Holders",
                "",
                "| Model | Raw active rows | Trusted distinct workers | Weak identities | Missing identity rows | Duplicate groups |",
                "|---|---:|---:|---:|---:|---:|",
            ]
        )
        for row in stats["live_provisional_licence_holders"]["rows"]:
            lines.append(
                f"| {row['model']} | {row['raw_active_rows']} | {row['trusted_distinct_workers']} | "
                f"{row['weak_name_only_identities']} | {row['missing_identity_rows']} | {row['duplicate_identity_groups']} |"
            )
        total = stats["live_provisional_licence_holders"]["totals"]
        lines.append(
            f"| Total | {total.get('raw_active_rows', 0)} | {total.get('trusted_distinct_workers', 0)} | "
            f"{total.get('weak_name_only_identities', 0)} | {total.get('missing_identity_rows', 0)} | "
            f"{total.get('duplicate_identity_groups', 0)} |"
        )
        lines.extend(["", "## Imported Licence Categories", ""])
        for key, value in stats["imported_license_categories"].items():
            if isinstance(value, dict):
                lines.append(
                    f"- {key}: {value['trusted_distinct_workers']} trusted distinct workers "
                    f"from {value['raw_rows']} raw rows; duplicate rows collapsed: {value['duplicate_rows_collapsed']}."
                )
            else:
                lines.append(f"- {key}: {value}")
        lines.extend(
            [
                "",
                "## Counting Rule",
                "",
                stats["counting_note"],
                "",
                "No conflicting non-empty fields were overwritten. Conflicts and unmatched rows are written to the CSV outputs for registrar review.",
            ]
        )
        if result["sample_rows"]:
            lines.extend(["", "## Sample Check - Everestar Aitsi", ""])
            for row in result["sample_rows"]:
                lines.append(
                    f"- Source row {row['source_row']}: DOB {row['date_of_birth'] or '-'}, "
                    f"receipt {row['receipt_number'] or '-'}, matched ATP record IDs {row['matched_practice_record_ids']}, "
                    f"matched professional IDs {row['matched_professional_record_ids']}."
                )
        return "\n".join(lines)

    def _write_csv(self, path: Path, rows: list[dict]):
        if not rows:
            path.write_text("", encoding="utf-8")
            return
        fieldnames = list(rows[0].keys())
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
