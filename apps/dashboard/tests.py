from datetime import date
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest import mock

from django.core.management import call_command
from django.core.management.base import CommandError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.core.cache import cache
from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.accounts.models import SecurityAuditEvent
from apps.common.models import DuplicateReviewQueue
from apps.complaints.models import ComplaintCase, DisciplinaryCase, RegulatoryDecisionRecord
from apps.dashboard.ai_provider import ai_provider_status
from apps.dashboard.assistant_rag import (
    build_vector_index,
    knowledge_index_is_stale,
    rag_status,
    retrieve_vector_sources,
)
from apps.dashboard.models import (
    AssistantConversation,
    AssistantFeedback,
    AssistantMessage,
    FAQCategory,
    FAQEntry,
    ForumCategory,
    ForumPost,
    ForumTopic,
    MappedEntity,
    NursingCouncilBoardActionItem,
    NursingCouncilBoardAgendaItem,
    NursingCouncilBoardAttendance,
    NursingCouncilBoardMeeting,
    NursingCouncilBoardPaper,
    NursingAnalyticsSnapshot,
    NursingLifecycleFact,
    NursingPractitionerIndex,
    NursingStageYearMetric,
    RegistryArchiveRecord,
    Receipt,
)
from apps.dashboard.nhwa_toolkit import _cell_style
from apps.dashboard.reference_breakdown import build_reference_breakdown
from apps.dashboard.registry_archive import active_professional_count, sync_registry_archives
from apps.dashboard.staff_ai import build_staff_ai_chat_response, staff_ai_question_needs_knowledge_search
from apps.dashboard.views import _registrar_worker_origin_context
from apps.documents.models import Document
from apps.workforce.models import (
    Application,
    AuditLog,
    Cadre,
    CommunityHealthWorker,
    CPDRecord,
    DataImportBatch,
    EmploymentRecord,
    Facility,
    HealthStudent,
    Location,
    MedicalDoctor,
    MissingDataReview,
    Midwife,
    NurseAide,
    NursingProfessional,
    PostingHistory,
    PracticingLicenseRecord,
    Qualification,
    TrainingInstitution,
    WorkforceSnapshot,
)


def _relative_luminance(hex_color):
    value = hex_color.strip().lstrip('#')
    channels = [int(value[index:index + 2], 16) / 255 for index in (0, 2, 4)]
    converted = [
        channel / 12.92 if channel <= 0.03928 else ((channel + 0.055) / 1.055) ** 2.4
        for channel in channels
    ]
    return 0.2126 * converted[0] + 0.7152 * converted[1] + 0.0722 * converted[2]


def contrast_ratio(foreground_hex, background_hex):
    foreground = _relative_luminance(foreground_hex)
    background = _relative_luminance(background_hex)
    lighter = max(foreground, background)
    darker = min(foreground, background)
    return (lighter + 0.05) / (darker + 0.05)


class DashboardHeaderTitleTests(TestCase):
    def test_clinical_dashboards_show_board_specific_welcome_title(self):
        user_model = get_user_model()
        cases = [
            ('nurse.user', 'nurse', 'nurse_dashboard', 'PNG Nursing Council Online Platform Dashboard'),
            ('doctor.user', 'doctor', 'doctor_dashboard', 'Medical Board Online Platform Dashboard'),
            ('chw.user', 'chw', 'chw_dashboard', 'Medical Board Online Platform Dashboard'),
            ('graduand.user', 'graduand', 'graduand_dashboard', 'PNG Nursing Council Online Platform Dashboard'),
        ]

        for username, role, url_name, expected_title in cases:
            with self.subTest(role=role):
                user = user_model.objects.create_user(
                    username=username,
                    password='StrongPass123!',
                    role=role,
                )
                self.client.force_login(user)

                response = self.client.get(reverse(url_name))

                self.assertEqual(response.status_code, 200)
                expected_display_name = username.replace('.', ' ').title()
                self.assertContains(response, f'Welcome, {expected_display_name} To Your {expected_title}')
                self.assertContains(response, '<section class="content-header ndoh-platform-header')
                self.assertNotContains(response, 'data-platform-header-toggle')
                self.assertNotContains(response, 'Hide Header')
                self.assertNotContains(response, 'Show Header')
                self.assertNotContains(response, 'ndoh-header-collapsible')
                self.assertNotContains(response, 'is-collapsed')
                self.assertContains(response, 'National_emblem_of_Papua_New_Guinea_(variant).svg.png')
                self.assertContains(response, 'width: min(520px, 42vw);')
                self.assertContains(response, '.content-wrapper::before')
                self.assertContains(response, '.content-header.ndoh-platform-header::after')
                if 'PNG Nursing Council' in expected_title:
                    self.assertContains(response, 'ndoh-nursing-council-header')
                else:
                    self.assertContains(response, 'ndoh-medical-board-header')
                    self.assertNotContains(response, 'ndoh-nursing-council-header')
                self.assertNotContains(
                    response,
                    'The National Department Of Health Regulatory Bodies Nursing Council &amp; The Medical Board Online Workforce System',
                )
                self.client.logout()


class NursingCouncilBoardPortalTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.board_user = user_model.objects.create_user(
            username='board.member',
            password='StrongPass123!',
            role='board_member',
            first_name='Board',
            last_name='Member',
            department='Nursing Council Board',
        )
        self.registrar = user_model.objects.create_user(
            username='board.secretary',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        self.meeting = NursingCouncilBoardMeeting.objects.create(
            title='July Nursing Council Board Meeting',
            scheduled_for=timezone.now() + timedelta(days=7),
            meeting_mode='hybrid',
            location='NDoH Boardroom / Teams',
            status='papers_issued',
            quorum_required=1,
            chair=self.board_user,
            secretary=self.registrar,
            created_by=self.registrar,
            public_summary='Registration, education, standards, and conduct assurance.',
        )
        self.agenda_item = NursingCouncilBoardAgendaItem.objects.create(
            meeting=self.meeting,
            order=1,
            title='Registration Committee report',
            purpose='decision',
            category='registration',
            status='ready',
            summary='Applicant exceptions and provisional to full recommendations.',
            presenter=self.registrar,
        )
        document = Document.objects.create(
            title='July Board Pack',
            office_scope='nursing',
            status='active',
            is_record=True,
            created_by=self.registrar,
        )
        NursingCouncilBoardPaper.objects.create(
            meeting=self.meeting,
            agenda_item=self.agenda_item,
            title='Registration Committee Board Paper',
            document=document,
            classification='private',
            status='issued',
            version_label='v1',
            prepared_by=self.registrar,
        )
        Application.objects.create(
            form_code='NC1',
            status='pending',
            payload={'full_name': 'Board Queue Applicant'},
        )
        ComplaintCase.objects.create(
            office_scope='nursing',
            title='Conduct matter for board noting',
            description='High risk conduct concern.',
            risk_level='high',
            priority='high',
            subject_name='Nursing Practitioner',
            subject_identifier='NC-BOARD-1',
            created_by=self.registrar,
        )
        RegulatoryDecisionRecord.objects.create(
            office_scope='nursing',
            decision_type='registration',
            status='draft',
            title='Draft registration decision',
            subject_name='Draft Decision Applicant',
            decision_text='Decision pending board review.',
            rationale='Committee recommendation.',
            created_by=self.registrar,
        )

    def test_board_member_main_dashboard_redirects_to_board_portal(self):
        self.client.force_login(self.board_user)

        response = self.client.get(reverse('main_dashboard'))

        self.assertRedirects(response, reverse('board_nursing_dashboard'))

    def test_home_page_includes_nursing_council_board_portal_entry(self):
        nurse_user = get_user_model().objects.create_user(
            username='normal.nurse.home',
            password='StrongPass123!',
            role='nurse',
            first_name='Normal',
            last_name='Nurse',
        )
        response = self.client.get(reverse('portal_home'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nursing Council Board Sign In')
        self.assertContains(response, reverse('board_login'))

        self.client.force_login(nurse_user)
        response = self.client.get(reverse('portal_home'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Nursing Council Board Portal')
        self.assertNotContains(response, reverse('board_login'))

        self.client.force_login(self.board_user)
        response = self.client.get(reverse('portal_home'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nursing Council Board Portal')
        self.assertContains(response, reverse('board_nursing_dashboard'))

    def test_sidebar_hides_board_portal_from_normal_staff_account(self):
        normal_staff = get_user_model().objects.create_user(
            username='normal.nurse.portal',
            password='StrongPass123!',
            role='nurse',
            first_name='Normal',
            last_name='Nurse',
        )

        self.client.force_login(normal_staff)
        response = self.client.get(reverse('nurse_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'NC Board Portal')
        self.assertNotContains(response, reverse('board_nursing_dashboard'))

    def test_anonymous_board_portal_redirects_to_board_sign_in(self):
        response = self.client.get(reverse('nursing_council_board_portal'))

        self.assertRedirects(
            response,
            f"{reverse('board_login')}?next={reverse('nursing_council_board_portal')}",
            fetch_redirect_response=False,
        )

    def test_board_portal_renders_meeting_pack_and_decision_queue(self):
        self.client.force_login(self.board_user)

        response = self.client.get(reverse('nursing_council_board_portal'))

        self.assertRedirects(response, reverse('board_nursing_dashboard'), fetch_redirect_response=False)

        response = self.client.get(reverse('board_nursing_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PNG Nursing Council Board Governance Portal')
        self.assertContains(response, 'July Nursing Council Board Meeting')
        self.assertContains(response, 'Registration Committee report')
        self.assertContains(response, 'Registration Committee Board Paper')
        self.assertContains(response, 'Attendance, Quorum, and Conflict Declarations')
        self.assertContains(response, 'Nursing Council Board')
        self.assertNotContains(response, 'NC1 application #')
        self.assertNotContains(response, 'Conduct matter for board noting')
        self.assertNotContains(response, 'Draft registration decision')
        self.assertNotContains(response, 'Workbooks')
        self.assertNotContains(response, 'NC Operations')
        self.assertNotContains(response, reverse('complaint_case_list'))
        self.assertNotContains(response, reverse('regulatory_decision_list'))

    def test_board_member_cannot_open_nursing_portal_data(self):
        self.client.force_login(self.board_user)
        application = Application.objects.get(form_code='NC1')
        decision = RegulatoryDecisionRecord.objects.get(title='Draft registration decision')

        portal_response = self.client.get(reverse('nursing_council_portal'))
        application_response = self.client.get(reverse('application_detail', args=[application.pk]))
        complaints_response = self.client.get(reverse('complaint_case_list') + '?office=nursing')
        decisions_response = self.client.get(reverse('regulatory_decision_list') + '?office=nursing')
        decision_detail_response = self.client.get(reverse('regulatory_decision_detail', args=[decision.decision_uuid]))
        repository_response = self.client.get(reverse('repository_search'))

        self.assertRedirects(portal_response, reverse('main_dashboard'), fetch_redirect_response=False)
        self.assertEqual(application_response.status_code, 404)
        self.assertEqual(complaints_response.status_code, 403)
        self.assertEqual(decisions_response.status_code, 403)
        self.assertEqual(decision_detail_response.status_code, 404)
        self.assertEqual(repository_response.status_code, 403)

    def test_board_member_can_record_attendance_conflict_and_action(self):
        self.client.force_login(self.board_user)

        attendance_response = self.client.post(reverse('nursing_council_board_portal'), {
            'board_action': 'record_attendance',
            'meeting_id': self.meeting.pk,
            'attendance_status': 'present',
            'conflict_declared': 'on',
            'recusal_required': 'on',
            'conflict_note': 'Recused from one registration item.',
        })

        self.assertRedirects(
            attendance_response,
            reverse('nursing_council_board_portal') + '#board-attendance',
            fetch_redirect_response=False,
        )
        attendance = NursingCouncilBoardAttendance.objects.get(meeting=self.meeting, member=self.board_user)
        self.assertEqual(attendance.attendance_status, 'present')
        self.assertTrue(attendance.conflict_declared)
        self.assertTrue(attendance.recusal_required)
        self.assertEqual(attendance.role_on_board, 'chair')

        action_response = self.client.post(reverse('nursing_council_board_portal'), {
            'board_action': 'add_action_item',
            'meeting_id': self.meeting.pk,
            'title': 'Follow up Registration Committee conditions',
            'description': 'Return conditions register to the next board meeting.',
            'due_date': (timezone.localdate() + timedelta(days=14)).isoformat(),
            'priority': 'high',
        })

        self.assertRedirects(
            action_response,
            reverse('nursing_council_board_portal') + '#board-actions',
            fetch_redirect_response=False,
        )
        action = NursingCouncilBoardActionItem.objects.get(meeting=self.meeting, owner=self.board_user)
        self.assertEqual(action.title, 'Follow up Registration Committee conditions')
        self.assertEqual(action.priority, 'high')
        self.assertEqual(action.status, 'open')


class SystemAdminCommandCentreTests(TestCase):
    def setUp(self):
        cache.clear()
        self.admin_user = get_user_model().objects.create_user(
            username='system.command.admin',
            password='StrongPass123!',
            role='admin',
            is_staff=True,
            is_superuser=True,
            role_approved=True,
            operations_approved=True,
        )

    def test_admin_dashboard_renders_command_centre_and_priority_metrics(self):
        current_year = timezone.localdate().year
        DataImportBatch.objects.create(
            source_file_name='nursing-source.xlsx',
            source_kind='ndata_workbook',
            status='completed',
            total_rows=100,
            processed_rows=100,
        )
        DataImportBatch.objects.create(
            source_file_name='medical-source.xlsx',
            source_kind='medical_board_workbook',
            status='failed',
            total_rows=50,
            processed_rows=10,
        )
        WorkforceSnapshot.objects.create(
            year=current_year,
            total_active_workers=15,
            total_nurses=8,
            total_doctors=3,
            total_midwives=2,
            total_chw=2,
            nearing_retirement=1,
        )
        nurse = NursingProfessional.objects.create(
            first_name='Admin',
            last_name='Review',
            registration_no='ADMIN-NURSE-1',
            license_expiry_date=timezone.localdate() + timedelta(days=20),
        )
        nurse_ct = ContentType.objects.get_for_model(NursingProfessional)
        MissingDataReview.objects.create(
            content_type=nurse_ct,
            object_id=nurse.pk,
            full_name='Admin Review',
            professional_type='Nursing Professional',
            missing_fields=['Email address', 'Phone number'],
            missing_count=2,
            severity='high',
        )
        DuplicateReviewQueue.objects.create(
            content_type=nurse_ct,
            object_id=nurse.pk,
            suspected_duplicate={'registration_no': 'ADMIN-NURSE-1'},
            similarity_score=0.96,
        )
        Application.objects.create(form_code='NC1', status='pending')
        Application.objects.filter(form_code='NC1').update(
            submitted_date=timezone.localdate() - timedelta(days=20)
        )
        Receipt.objects.create(
            user=self.admin_user,
            receipt_number='ADMIN-RCT-1',
            amount='50.00',
            status='pending',
            payer_match_confidence='ambiguous',
        )
        get_user_model().objects.create_user(
            username='unlinked.normal.user',
            password='StrongPass123!',
            role='nurse',
            professional_record_status='unmatched',
        )
        SecurityAuditEvent.objects.create(
            username='failed.login',
            action='LOGIN_FAILED',
            path='/accounts/login/',
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('admin_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'System Administration Command Centre')
        self.assertContains(response, 'Priority Action Queue')
        self.assertContains(response, 'System Control Indicators')
        self.assertContains(response, 'Regulatory Body Service Status')
        self.assertContains(response, 'International Standards Readiness')
        self.assertContains(response, 'Operations Console')
        self.assertContains(response, 'Import And Snapshot Health')
        self.assertContains(response, 'Imported Source Rows')
        self.assertContains(response, 'Open Data Quality Reviews')
        self.assertContains(response, 'Receipted Payments')
        self.assertContains(response, 'Select ATP Workbook')
        self.assertContains(response, 'Select Full-Licence Workbook')
        self.assertContains(response, 'Select Provisional Workbook')
        self.assertContains(response, reverse('nursing_council_portal') + '?import=atp#nursing-public-protection')
        self.assertContains(response, reverse('nursing_council_portal') + '?import=full_licence#nursing-public-protection')
        self.assertContains(response, reverse('nursing_council_portal') + '?import=provisional#nursing-public-protection')
        self.assertNotContains(response, 'data-admin-command="import_current_atp_workbook"')
        self.assertNotContains(response, 'data-admin-command="import_provisional_licenses"')
        self.assertContains(response, 'data-admin-command="audit_missing_data"')
        self.assertContains(response, reverse('public_nursing_register_search_root'))
        self.assertContains(response, reverse('public_medical_board_register_search_root'))
        self.assertEqual(response.context['admin_duplicate_review_count'], 1)
        self.assertEqual(response.context['admin_overdue_application_count'], 1)
        self.assertEqual(response.context['admin_receipt_mismatch_count'], 1)
        self.assertEqual(response.context['high_priority_missing_data_count'], 1)
        self.assertGreaterEqual(response.context['admin_unlinked_practitioner_count'], 1)
        self.assertGreaterEqual(response.context['admin_import_attention_count'], 1)


class DashboardSearchNursingAnalyticsTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.nursing_registrar = user_model.objects.create_user(
            username='nursing.analytics.search',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

    def test_long_excel_row_search_finds_active_nursing_analytics_fact(self):
        snapshot = NursingAnalyticsSnapshot.objects.create(
            source_file_name='PNG_Nursing_Council_Integrated_Dashboard_Model.xlsx',
            source_file_hash='snapshot-search-hash',
            is_active=True,
            kpi_summary={
                'total_lifecycle_records': 34851,
                'clean_full_licence_records': 6695,
            },
        )
        fact = NursingLifecycleFact.objects.create(
            snapshot=snapshot,
            record_id='FULL-005873',
            lifecycle_stage='Full Licence',
            licence_status='Full',
            lifecycle_order=2,
            cycle_year=2024,
            full_name='Henny Sakatias',
            person_group_key='HENNY SAKATIAS',
            registration_no='GD 5360',
            cadre='Registered Nurse',
            institution='Tuna Bay School of Nursing',
            facility='East Sepik Provincial Health Authority',
            province='East Sepik Province',
            source_sheet='Clean_Full_Licence',
            source_row=5887,
            record_quality='High',
            raw_payload={'Source_Row': 5887, 'Original_Name': 'Henny Sakatias'},
        )

        self.client.force_login(self.nursing_registrar)
        response = self.client.get(reverse('dashboard_search'), {
            'q': (
                '5887 1396 510 Henny Sakatias Henny Sakatias Full GD 5360 '
                '03.10.2025 Tuna Bay School of Nursing 2024 Diploma in General Nursing '
                'East Sepik Provincial Health Authority Private Mail Bag WEWAK East Sepik Province'
            ),
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nursing Analytics Snapshot Records')
        self.assertContains(response, 'Henny Sakatias')
        self.assertContains(response, 'GD 5360')
        self.assertContains(response, 'Full Licence')
        self.assertContains(response, 'Analytics snapshot only; not an operational registry row.')
        detail_url = reverse('nursing_analytics_fact_detail', args=[fact.pk])
        self.assertContains(response, detail_url)
        self.assertContains(response, 'Open')

        detail_response = self.client.get(detail_url)
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, 'Nursing Analytics Record')
        self.assertContains(detail_response, 'Henny Sakatias')
        self.assertContains(detail_response, 'GD 5360')
        self.assertContains(detail_response, 'Clean_Full_Licence')
        self.assertContains(detail_response, 'This is a read-only Nursing Council analytics snapshot record.')

    def test_search_links_complete_nursing_analytics_pathway(self):
        snapshot = NursingAnalyticsSnapshot.objects.create(
            source_file_name='PNG_Nursing_Council_Integrated_Dashboard_Model.xlsx',
            source_file_hash='snapshot-pathway-search-hash',
            is_active=True,
            kpi_summary={
                'total_lifecycle_records': 3,
                'clean_provisional_records': 1,
                'clean_full_licence_records': 1,
                'clean_atp_records': 1,
            },
        )
        provisional = NursingLifecycleFact.objects.create(
            snapshot=snapshot,
            record_id='PROV-ALEX-000001',
            lifecycle_stage='Provisional Licence',
            licence_status='Provisional',
            lifecycle_order=1,
            cycle_year=2018,
            full_name='Alex Paito',
            person_group_key='ALEX PAITO',
            identity_confidence='High',
            registration_no='PRO 3100',
            cadre='Nursing Graduand',
            institution='Lutheran School of Nursing',
            source_sheet='Clean_Provisional',
            source_row=2707,
            record_quality='High',
        )
        NursingLifecycleFact.objects.create(
            snapshot=snapshot,
            record_id='FULL-ALEX-000001',
            lifecycle_stage='Full Licence',
            licence_status='Full Licence',
            lifecycle_order=2,
            cycle_year=2018,
            full_name='Alex Paito',
            person_group_key='ALEX PAITO',
            identity_confidence='High',
            registration_no='GD 2293',
            cadre='Registered Nurse',
            institution='Lutheran School of Nursing',
            source_sheet='Clean_Full_Licence',
            source_row=2803,
            record_quality='High',
        )
        NursingLifecycleFact.objects.create(
            snapshot=snapshot,
            record_id='ATP-ALEX-000001',
            lifecycle_stage='Authority to Practice',
            licence_status='ATP / Renewal',
            lifecycle_order=3,
            cycle_year=2024,
            full_name='Alex Paito',
            person_group_key='ALEX PAITO',
            identity_confidence='High',
            registration_no='GD 2229',
            cadre='Registered Nurse',
            facility='University Of Goroka',
            source_sheet='ATP RECORD 2024',
            source_row=3778,
            record_quality='Medium',
        )
        NursingPractitionerIndex.objects.create(
            snapshot=snapshot,
            practitioner_group_id='PG-ALEX-PAITO',
            person_group_key='ALEX PAITO',
            representative_name='Alex Paito',
            identity_confidence='High',
            record_count=3,
            stages_present='Provisional Licence; Full Licence; Authority to Practice',
            has_provisional=True,
            has_full_licence=True,
            has_atp=True,
            first_year=2018,
            latest_year=2024,
            latest_atp_year=2024,
            latest_cadre='Registered Nurse',
            latest_facility='University Of Goroka',
            registration_nos='PRO 3100; GD 2293; GD 2229',
            practitioner_nos='',
        )

        self.client.force_login(self.nursing_registrar)
        response = self.client.get(reverse('dashboard_search'), {'q': 'Alex Paito'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Alex Paito')
        self.assertContains(response, 'Linked pathway')
        self.assertContains(response, 'Provisional Licence / Full Licence / Authority to Practice')
        self.assertContains(response, 'Open pathway')
        pathway_url = reverse('nursing_analytics_pathway_detail', args=[provisional.pk])
        self.assertContains(response, pathway_url)

        detail_response = self.client.get(pathway_url)
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, 'Nursing Analytics Pathway')
        self.assertContains(detail_response, 'Complete provisional to ATP pathway')
        self.assertContains(detail_response, 'PRO 3100')
        self.assertContains(detail_response, 'GD 2293')
        self.assertContains(detail_response, 'GD 2229')


class ReceiptOwnershipLinkingTests(TestCase):
    def test_application_receipt_links_to_professional_owner(self):
        nurse = NursingProfessional.objects.create(
            first_name='Receipt',
            last_name='Owner',
            registration_no='RN-REC-100',
        )
        application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(nurse),
            object_id=nurse.pk,
            form_code='NC3',
            pathway='other',
            status='approved',
        )
        receipt = Receipt.objects.create(
            application=application,
            receipt_number='APP-RCT-100',
            official_receipt_no='APP-OR-100',
            amount='100.00',
            status='completed',
        )

        call_command('link_receipts_to_individual_records', '--apply')

        receipt.refresh_from_db()
        self.assertEqual(receipt.payer_content_type, ContentType.objects.get_for_model(nurse))
        self.assertEqual(receipt.payer_object_id, nurse.pk)
        self.assertEqual(receipt.payer_match_confidence, 'application')

    def test_imported_receipt_number_links_to_practicing_license_record(self):
        batch = DataImportBatch.objects.create(
            source_file_name='atp-link-test.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        record = PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=10,
            record_year=2026,
            full_name='Receipt Match',
            registration_no='RN-REC-200',
            reference_number='OR-REC-200',
            payment_method='OR-REC-200',
            payment_date=timezone.localdate(),
            amount='150.00',
        )
        receipt = Receipt.objects.create(
            receipt_number='LOCAL-RCT-200',
            official_receipt_no='OR-REC-200',
            amount='150.00',
            status='completed',
            receipt_date=timezone.now(),
        )

        call_command('link_receipts_to_individual_records', '--apply')

        receipt.refresh_from_db()
        self.assertEqual(receipt.payer_content_type, ContentType.objects.get_for_model(record))
        self.assertEqual(receipt.payer_object_id, record.pk)
        self.assertEqual(receipt.payer_match_confidence, 'receipt_number')
        self.assertFalse(MissingDataReview.objects.filter(
            content_type=ContentType.objects.get_for_model(Receipt),
            object_id=receipt.pk,
        ).exclude(status='resolved').exists())

    def test_unmatched_receipt_creates_high_value_review(self):
        receipt = Receipt.objects.create(
            receipt_number='UNMATCHED-RCT-100',
            official_receipt_no='UNMATCHED-OR-100',
            amount='250.00',
            status='completed',
        )

        call_command('link_receipts_to_individual_records', '--apply')

        receipt.refresh_from_db()
        self.assertEqual(receipt.payer_match_confidence, 'unlinked')
        review = MissingDataReview.objects.get(
            content_type=ContentType.objects.get_for_model(Receipt),
            object_id=receipt.pk,
        )
        self.assertEqual(review.severity, 'high')
        self.assertEqual(review.status, 'under_review')
        self.assertIn('High value review', review.missing_fields[0])

    def test_linked_user_profile_sees_original_application_and_imported_receipts(self):
        from apps.accounts.professional_linking import attach_professional_record
        from apps.dashboard.views import _receipt_queryset_for_user

        user = get_user_model().objects.create_user(
            username='receipt.profile.user',
            password='StrongPass123!',
            role='nurse',
            registration_number='RN-PROFILE-100',
        )
        nurse = NursingProfessional.objects.create(
            first_name='Profile',
            last_name='Owner',
            registration_no='RN-PROFILE-100',
        )
        attach_professional_record(user, nurse, status='linked')
        application = Application.objects.create(
            content_type=ContentType.objects.get_for_model(nurse),
            object_id=nurse.pk,
            form_code='NC3',
            pathway='other',
            status='approved',
        )
        application_receipt = Receipt.objects.create(
            application=application,
            receipt_number='PROFILE-APP-RCT',
            official_receipt_no='PROFILE-APP-OR',
            amount='75.00',
            status='completed',
        )

        batch = DataImportBatch.objects.create(
            source_file_name='profile-atp.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        record = PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=88,
            record_year=2026,
            full_name='Profile Owner',
            registration_no='RN-PROFILE-100',
            payment_date=timezone.localdate(),
        )
        imported_receipt = Receipt.objects.create(
            receipt_number='PROFILE-IMP-RCT',
            official_receipt_no='PROFILE-IMP-OR',
            amount='90.00',
            status='completed',
            payer_content_type=ContentType.objects.get_for_model(record),
            payer_object_id=record.pk,
            payer_match_confidence='receipt_number',
        )
        unrelated_receipt = Receipt.objects.create(
            receipt_number='PROFILE-OTHER-RCT',
            official_receipt_no='PROFILE-OTHER-OR',
            amount='120.00',
            status='completed',
        )

        receipt_ids = set(_receipt_queryset_for_user(user).values_list('id', flat=True))

        self.assertIn(application_receipt.pk, receipt_ids)
        self.assertIn(imported_receipt.pk, receipt_ids)
        self.assertNotIn(unrelated_receipt.pk, receipt_ids)


class EngagementPlatformTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.staff_user = user_model.objects.create_user(
            username='nursing.engagement.staff',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        location = Location.objects.create(province='National Capital District', district='Port Moresby')
        Facility.objects.create(
            name='Port Moresby General Hospital',
            type='Hospital',
            ownership='public',
            location=location,
        )
        TrainingInstitution.objects.create(
            name='Tuna Bay School of Nursing',
            type='Nursing School',
            location_name='East Sepik Province',
            regulatory_body_name='Nursing Council',
            source_reference='test fixture',
        )

    def test_seed_engagement_platform_creates_faq_forum_and_map_references(self):
        call_command('seed_engagement_platform')

        self.assertGreaterEqual(FAQCategory.objects.count(), 3)
        self.assertGreaterEqual(FAQEntry.objects.count(), 4)
        self.assertTrue(ForumCategory.objects.filter(slug='public-questions').exists())
        self.assertTrue(MappedEntity.objects.filter(name='Tuna Bay School of Nursing', office_scope='nursing').exists())
        self.assertTrue(MappedEntity.objects.filter(name='Port Moresby General Hospital').exists())

        faq_response = self.client.get(reverse('public_faqs'))
        forum_response = self.client.get(reverse('forum_index'))
        map_response = self.client.get(reverse('workforce_map') + '?office=nursing')

        self.assertEqual(faq_response.status_code, 200)
        self.assertContains(faq_response, 'Public FAQs')
        self.assertEqual(forum_response.status_code, 200)
        self.assertContains(forum_response, 'Public Questions')
        self.assertEqual(map_response.status_code, 200)
        self.assertContains(map_response, 'Mapped Schools, Institutions and Facilities')
        self.assertContains(map_response, reverse('public_nursing_register_search_root'))
        self.assertContains(map_response, 'Tuna Bay School of Nursing')
        self.assertContains(map_response, 'API key missing')
        self.assertContains(map_response, 'manage.py geocode_mapped_entities --limit 100')

    @override_settings(GOOGLE_MAPS_API_KEY='test-maps-key')
    def test_workforce_map_renders_google_maps_embed_iframe(self):
        response = self.client.get(reverse('workforce_map'), {'q': 'Space Needle, Seattle WA'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Google Maps Embed Layer')
        self.assertContains(response, '<iframe', html=False)
        self.assertContains(response, 'width="600"', html=False)
        self.assertContains(response, 'height="450"', html=False)
        self.assertContains(response, 'loading="lazy"', html=False)
        self.assertContains(response, 'allowfullscreen', html=False)
        self.assertContains(response, 'referrerpolicy="no-referrer-when-downgrade"', html=False)
        self.assertContains(
            response,
            'https://www.google.com/maps/embed/v1/place?key=test-maps-key&amp;q=Space+Needle%2C+Seattle+WA',
            html=False,
        )

    @override_settings(GOOGLE_MAPS_API_KEY='')
    def test_geocode_command_requires_google_maps_api_key(self):
        call_command('seed_engagement_platform')

        with self.assertRaises(CommandError):
            call_command('geocode_mapped_entities', limit=1)

    def test_public_forum_posts_are_moderated_and_staff_can_approve(self):
        call_command('seed_engagement_platform', skip_map_refresh=True)
        category = ForumCategory.objects.get(slug='public-questions')

        response = self.client.post(reverse('forum_category_detail', args=[category.slug]), {
            'public_author_name': 'Public User',
            'public_author_email': 'public@example.com',
            'title': 'How do I verify ATP?',
            'body': 'Please confirm the ATP verification process.',
        })

        self.assertEqual(response.status_code, 302)
        topic = ForumTopic.objects.get(title='How do I verify ATP?')
        post = ForumPost.objects.get(topic=topic)
        self.assertEqual(topic.status, 'pending')
        self.assertEqual(post.status, 'pending')

        public_topic_response = self.client.get(reverse('forum_topic_detail', args=[category.slug, topic.slug]))
        self.assertEqual(public_topic_response.status_code, 403)

        self.client.force_login(self.staff_user)
        staff_topic_response = self.client.get(reverse('forum_topic_detail', args=[category.slug, topic.slug]))
        self.assertEqual(staff_topic_response.status_code, 200)
        approve_response = self.client.post(reverse('forum_moderate_post', args=[post.pk]), {'action': 'approved'})
        self.assertEqual(approve_response.status_code, 302)
        post.refresh_from_db()
        topic.refresh_from_db()
        self.assertEqual(post.status, 'approved')
        self.assertEqual(topic.status, 'approved')


class RegisteredNursesExcelExportTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.registrar = user_model.objects.create_user(
            username='nursing.registrar.export',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        NursingProfessional.objects.create(
            first_name='Mary',
            last_name='Kila',
            registration_no='RN-100',
            registration_number='REG-100',
            gender='Female',
            applicant_type='national',
            province='National Capital District',
            email='mary@example.com',
            primary_phone='70000001',
            license_expiry_date=timezone.localdate() + timedelta(days=365),
            date_issued=timezone.localdate(),
        )
        NursingProfessional.objects.create(
            first_name='Ruth',
            last_name='Wari',
            registration_no='RN-101',
            gender='Female',
            applicant_type='overseas',
            province='Morobe',
            license_expiry_date=timezone.localdate() - timedelta(days=10),
        )

    def test_export_nurses_button_returns_excel_analytics_workbook(self):
        self.client.force_login(self.registrar)

        response = self.client.get(reverse('registered_nurses_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('registered_nurses_analytics.xlsx', response['Content-Disposition'])
        workbook = load_workbook(BytesIO(response.content), read_only=False)
        self.assertIn('Read Me First', workbook.sheetnames)
        self.assertIn('Executive Summary', workbook.sheetnames)
        self.assertIn('Charts', workbook.sheetnames)
        self.assertIn('Registered Nurses', workbook.sheetnames)
        self.assertIn('Data Quality', workbook.sheetnames)
        self.assertEqual(workbook['Registered Nurses']['A2'].value, 'Full Name')
        self.assertEqual(workbook['Registered Nurses']['A3'].value, 'Mary Kila')
        self.assertGreaterEqual(len(workbook['Charts']._charts), 4)

    def test_legacy_registered_nurses_csv_url_returns_excel_workbook(self):
        self.client.force_login(self.registrar)

        response = self.client.get(reverse('report_csv', args=['registered_nurses']))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('registered_nurses_analytics.xlsx', response['Content-Disposition'])


class DuplicateReviewWorkflowAccessTests(TestCase):
    def test_data_quality_reviewer_can_open_duplicate_review_workflow(self):
        user_model = get_user_model()
        data_quality_reviewer = user_model.objects.create_user(
            username='data.quality.reviewer',
            password='StrongPass123!',
            role='reviewer',
            department='Data Quality',
        )
        self.client.force_login(data_quality_reviewer)

        response = self.client.get(reverse('duplicate_review_workflow'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Duplicate Review Workflow')
        self.assertContains(response, 'All Regulatory Offices')


class FinancialForecastAccessTests(TestCase):
    def test_operations_approved_data_quality_reviewer_can_open_financial_forecast(self):
        user_model = get_user_model()
        data_quality_reviewer = user_model.objects.create_user(
            username='data.quality.elevated',
            password='StrongPass123!',
            role='reviewer',
            department='Data Quality',
            operations_approved=True,
        )
        self.client.force_login(data_quality_reviewer)

        response = self.client.get(reverse('financial_forecast_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Financial Forecast and Receipt Tracking')
        self.assertContains(response, 'Current scope:')
        self.assertContains(response, 'All Regulatory Offices')
        self.assertContains(response, 'Nursing Council Financial Forecast')
        self.assertContains(response, 'Medical Board Financial Forecast')


class ReferenceBreakdownTests(TestCase):
    def setUp(self):
        cache.clear()
        TrainingInstitution.objects.create(name='Pacific Adventist University', type='')
        TrainingInstitution.objects.create(name='PAU', type='')
        TrainingInstitution.objects.create(name='Lae School of Nursing', type='')
        TrainingInstitution.objects.create(name='APIASETS School of Nursing', type='')
        TrainingInstitution.objects.create(name='Rumginae CHW Training School', type='CHW Training School')
        TrainingInstitution.objects.create(name='Auckland University, New Zealand', type='')
        TrainingInstitution.objects.create(name='Ackland University, New Zealnd', type='National Institution')
        TrainingInstitution.objects.create(name='Adventist U/Philippine', type='National Institution')
        TrainingInstitution.objects.create(name='America', type='National Institution')
        TrainingInstitution.objects.create(name="St Mary's SON Chitradurga, India", type='National Institution')
        TrainingInstitution.objects.create(name='2016', type='CHW Training School')

        batch = DataImportBatch.objects.create(source_file_name='test.xlsx', status='completed')
        PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='workforce_listing',
            target_model='nursingprofessional',
            source_sheet_name='April 2026',
            source_row=1,
            full_name='Jane Doe',
            workplace_address='Port Moresby General Hospital, P O Box 1, Boroko',
        )
        PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='workforce_listing',
            target_model='nursingprofessional',
            source_sheet_name='April 2026',
            source_row=2,
            full_name='John Doe',
            workplace_address='Port Moresby General Hospital P O Box 1, Boroko',
        )

    def test_reference_breakdown_separates_png_nursing_school_categories(self):
        breakdown = build_reference_breakdown()

        self.assertEqual(breakdown['png_nursing_school_count'], 20)
        self.assertEqual(breakdown['mapped_nursing_reference_count'], 4)
        self.assertEqual(breakdown['chw_training_reference_count'], 1)
        self.assertEqual(breakdown['medical_board_chw_training_reference_count'], 1)
        self.assertEqual(breakdown['medical_board_chw_training_examples'], ['Rumginae CHW Training School'])
        self.assertEqual(breakdown['overseas_institution_reference_count'], 5)
        self.assertEqual(breakdown['national_institution_reference_count'], 4)
        self.assertEqual(breakdown['legacy_institution_reference_count'], 1)
        self.assertEqual(breakdown['facility_grouped_reference_count'], 1)
        self.assertEqual(breakdown['facility_raw_reference_count'], 2)

    def test_overall_dashboard_does_not_count_workplace_refs_as_facilities(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='facility_admin',
            password='StrongPass123!',
            role='admin',
            is_staff=True,
        )
        Application.objects.create(form_code='NC1', status='pending')
        self.client.force_login(user)

        response = self.client.get(reverse('advanced_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['facility_count'], 0)
        self.assertEqual(response.context['provisional_applicant_count'], 1)
        self.assertEqual(response.context['graduand_count'], 1)
        self.assertEqual(response.context['reference_breakdown']['facility_grouped_reference_count'], 1)
        self.assertContains(response, 'Graduands / Provisional Applicants')
        self.assertContains(response, 'Verified Facilities')
        self.assertContains(response, 'Imported workplace references, cleaned names')


class OverallDashboardLicenseRecordTablesTests(TestCase):
    def setUp(self):
        cache.clear()
        self.user = get_user_model().objects.create_user(
            username='overall_nursing_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        nursing_batch = DataImportBatch.objects.create(
            source_file_name='nursing-licences.xlsx',
            source_kind='nursing_license_workbook',
            status='completed',
        )
        self.atp_record = PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=1,
            record_year=2026,
            full_name='Dashboard ATP Nurse',
            registration_no='NC-ATP-DASH',
            practitioner_number='PN-ATP-DASH',
            category='General Nursing',
        )
        self.full_record = PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            record_type='full',
            target_model='midwife',
            source_sheet_name='FULL LICENSE',
            source_row=2,
            record_year=2026,
            full_name='Dashboard Full Midwife',
            registration_no='NC-FULL-DASH',
        )
        self.provisional_record = PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            record_type='provisional',
            target_model='healthstudent',
            source_sheet_name='PROV REGO',
            source_row=3,
            record_year=2026,
            full_name='Dashboard Provisional Graduand',
            registration_no='NC-PROV-DASH',
        )
        medical_batch = DataImportBatch.objects.create(
            source_file_name='medical-licences.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            record_type='practicing_license',
            target_model='medicaldoctor',
            source_sheet_name='MEDICAL ATP',
            source_row=1,
            record_year=2026,
            full_name='Hidden Medical ATP',
            registration_no='MB-ATP-DASH',
        )

    def test_overall_dashboard_renders_three_server_side_license_tables(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse('advanced_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ATP, Full-License and Provisional Records')
        self.assertContains(response, 'dashboard-atp-records-table')
        self.assertContains(response, 'dashboard-full-license-records-table')
        self.assertContains(response, 'dashboard-provisional-records-table')
        self.assertContains(response, reverse('dashboard_license_record_table', args=['atp']))
        self.assertContains(response, reverse('record_create', args=['practicinglicenserecord']) + '?record_type=practicing_license')
        tables = {table['key']: table for table in response.context['license_record_tables']}
        self.assertEqual(tables['atp']['count'], 1)
        self.assertEqual(tables['full-license']['count'], 1)
        self.assertEqual(tables['provisional']['count'], 1)

    def test_license_table_endpoint_filters_by_type_scope_and_returns_crud_actions(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse('dashboard_license_record_table', args=['atp']),
            {
                'draw': '7',
                'start': '0',
                'length': '10',
                'search[value]': 'NC-ATP-DASH',
                'order[0][column]': '0',
                'order[0][dir]': 'asc',
                'columns[0][name]': 'full_name',
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['draw'], 7)
        self.assertEqual(payload['recordsTotal'], 1)
        self.assertEqual(payload['recordsFiltered'], 1)
        self.assertEqual(len(payload['data']), 1)
        row = payload['data'][0]
        self.assertEqual(row['full_name'], 'Dashboard ATP Nurse')
        self.assertEqual(row['registration_no'], 'NC-ATP-DASH')
        self.assertIn(reverse('record_detail', args=['practicinglicenserecord', self.atp_record.pk]), row['actions'])
        self.assertIn(reverse('record_update', args=['practicinglicenserecord', self.atp_record.pk]), row['actions'])
        self.assertIn(reverse('record_delete', args=['practicinglicenserecord', self.atp_record.pk]), row['actions'])


def _append_sheet(workbook, title, rows):
    sheet = workbook.create_sheet(title)
    for row in rows:
        sheet.append(row)
    return sheet


def _build_nursing_analytics_workbook(path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _append_sheet(workbook, 'ReadMe_Source_Audit', [
        ['ReadMe + Source Audit', None, None, None, None],
        ['Workbook', 'PNG Nursing Council Integrated Dashboard Model', None, None, None],
        ['Generated', '2026-05-27', None, None, None],
        ['Official clean ATP records', 19998, None, None, None],
        ['Official clean provisional records', 8158, None, None, None],
        ['Official clean full-licence records', 6695, None, None, None],
        ['Total integrated lifecycle records', 34851, None, None, None],
        ['Estimated practitioner match groups', 22765, None, None, None],
        ['Data quality health score', '87.0%', None, None, None],
    ])
    _append_sheet(workbook, 'Dashboard_Data', [
        ['Stage Funnel', None, None, None, None, None],
        ['Lifecycle_Stage', 'Clean_Record_Count', None, None, None, None],
        ['Provisional Licence', 8158, None, None, None, None],
        ['Full Licence', 6695, None, None, None, None],
        ['Authority to Practice', 19998, None, None, None, None],
    ])
    _append_sheet(workbook, 'Fact_Lifecycle', [
        [
            'Record_ID', 'Lifecycle_Stage', 'Licence_Status', 'Lifecycle_Order', 'Cycle_Year',
            'Event_Date', 'Full_Name', 'Name_Key', 'Person_Group_Key', 'Identity_Confidence',
            'DOB', 'Sex', 'Age', 'Cadre', 'Cadre_Group', 'Profession_Speciality_Raw',
            'Formal_Qualification', 'Registration_No', 'Practitioner_No', 'Registration_Link_Key',
            'Institution', 'Facility', 'Province', 'Organization_Type', 'Nationality_Group',
            'Country', 'Source_Workbook', 'Source_Sheet', 'Source_Row', 'Include_In_Official_Totals',
            'Data_Quality_Flags', 'Completeness_Score', 'Record_Quality',
        ],
        [
            'PROV-000001', 'Provisional Licence', 'Provisional', 1, 2026,
            '2026-01-12', 'Test Provisional Nurse', 'TEST PROVISIONAL NURSE', 'TEST PROVISIONAL NURSE',
            'Low - name only', None, 'Female', 22, 'Nursing Graduand', 'Nursing', '',
            'Diploma Nursing', 'PRO 100', '', 'PRO 100', 'Lae School of Nursing', '',
            'Morobe Province', '', 'National', 'PNG', 'Fixture', 'Clean_Provisional', 2,
            'Yes', '', 1, 'High',
        ],
        [
            'FULL-000001', 'Full Licence', 'Full Licence', 2, 2026,
            '2026-02-01', 'Test Full Nurse', 'TEST FULL NURSE', 'TEST FULL NURSE',
            'Medium', None, 'Female', 24, 'Registered Nurse', 'Nursing', '',
            'Diploma Nursing', 'G 100', 'PN 100', 'G 100', 'Lae School of Nursing',
            'Angau Hospital', 'Morobe Province', 'Provincial Health Authority (PHA)',
            'National', 'PNG', 'Fixture', 'Clean_Full_Licence', 3, 'Yes', '', 1, 'Medium',
        ],
        [
            'ATP-000001', 'Authority to Practice', 'ATP / Renewal', 3, 2026,
            '2026-03-01', 'Test ATP Nurse', 'TEST ATP NURSE', 'TEST ATP NURSE',
            'High', None, 'Female', 25, 'Registered Nurse', 'Nursing', '',
            'Diploma Nursing', 'G 100', 'ATP 100', 'G 100', 'Lae School of Nursing',
            'Angau Hospital', 'Morobe Province', 'Provincial Health Authority (PHA)',
            'National', 'PNG', 'Fixture', 'ATP 2026', 4, 'Yes', '', 1, 'High',
        ],
    ])
    _append_sheet(workbook, 'Practitioner_Index', [
        [
            'Practitioner_Group_ID', 'Person_Group_Key', 'Representative_Name', 'Identity_Confidence',
            'Record_Count', 'Stages_Present', 'Has_Provisional', 'Has_Full_Licence', 'Has_ATP',
            'First_Year', 'Latest_Year', 'Latest_ATP_Year', 'Latest_Cadre', 'Latest_Facility',
            'Latest_Province', 'Registration_Nos', 'Practitioner_Nos', 'DQ_Flag_Count',
            'Needs_Manual_Review',
        ],
        [
            'PG-000001', 'TEST ATP NURSE', 'Test ATP Nurse', 'High', 3,
            'Provisional Licence; Full Licence; Authority to Practice', 'Yes', 'Yes', 'Yes',
            2026, 2026, 2026, 'Registered Nurse', 'Angau Hospital', 'Morobe Province',
            'G 100; PRO 100', 'ATP 100', 0, 'No',
        ],
    ])
    _append_sheet(workbook, 'Year_Stage', [
        ['Year', 'Provisional Licence', 'Full Licence', 'Authority to Practice', 'Grand Total'],
        [2026, 8157, 6694, 19998, 34849],
        ['Unknown', 1, 1, 0, 2],
    ])
    _append_sheet(workbook, 'Cadre_Stage', [
        ['Cadre', 'Cadre_Group', 'Provisional Licence', 'Full Licence', 'Authority to Practice', 'Grand Total'],
        ['Registered Nurse', 'Nursing', 0, 6695, 19998, 26693],
        ['Nursing Graduand', 'Nursing', 8158, 0, 0, 8158],
    ])
    _append_sheet(workbook, 'Facility_Cadre_Year', [
        ['Facility', 'Province', 'Organization_Type', 'Cadre', '2026', 'Unknown', 'Grand Total'],
        ['Angau Hospital', 'Morobe Province', 'Provincial Health Authority (PHA)', 'Registered Nurse', 19998, 0, 19998],
    ])
    _append_sheet(workbook, 'Institution_Cadre_Year', [
        ['Institution', 'Lifecycle_Stage', 'Cadre', '2026', 'Unknown', 'Grand Total'],
        ['Lae School of Nursing', 'Provisional Licence', 'Nursing Graduand', 8158, 0, 8158],
    ])
    _append_sheet(workbook, 'Geo_Org_Nationality', [
        ['Province by ATP Year', None, None, None],
        ['Province', '2026', 'Unknown', 'Grand Total'],
        ['Morobe Province', 19998, 0, 19998],
    ])
    _append_sheet(workbook, 'Data_Quality', [
        ['Data Quality by Lifecycle Stage', None, None, None, None, None],
        ['Lifecycle_Stage', 'High', 'Medium', 'Needs Review', 'Grand Total', 'Needs Review %'],
        ['Provisional Licence', 7441, 535, 182, 8158, 0.0223],
        ['Full Licence', 5857, 186, 652, 6695, 0.0974],
        ['Authority to Practice', 8873, 7418, 3707, 19998, 0.1854],
    ])
    _append_sheet(workbook, 'Platform_Field_Map', [
        ['Platform Field', 'Unified Fact Field', 'Used For', 'Data Quality Rule'],
        ['Health Facility Name (Place of work)', 'Facility', 'ATP active workforce location', 'Required for ATP'],
    ])
    _append_sheet(workbook, 'FHIR_NHWA_Map', [
        ['Unified Field', 'FHIR / Interoperability Mapping', 'WHO NHWA / Analytics Dimension', 'Implementation Note'],
        ['Full_Name', 'FHIR Practitioner.name.text', 'Health worker identity', 'Split before production FHIR export'],
    ])
    _append_sheet(workbook, 'Facility_Summary', [
        ['Facility', 'Province', 'Organization_Type', '2026', 'Unknown', 'Grand Total'],
        ['Angau Hospital', 'Morobe Province', 'Provincial Health Authority (PHA)', 19998, 0, 19998],
    ])
    _append_sheet(workbook, 'Institution_Summary', [
        ['Institution', '2026', 'Unknown', 'Grand Total'],
        ['Lae School of Nursing', 8158, 0, 8158],
    ])
    workbook.save(path)


def _build_catherine_licence_workbooks(licence_path, cadre_path):
    licence_workbook = Workbook()
    licence_workbook.remove(licence_workbook.active)
    _append_sheet(licence_workbook, 'Dashboard', [
        ['PNG Nursing Council Licence Records - Dashboard', None, None, None],
        [None, None, None, None],
        ['Metric', 'Provisional', 'Full Licence', 'Combined / Notes'],
        ['Source non-empty rows read', 8219, 6701, 14920],
        ['Clean rows retained after exact dedupe', 8158, 6695, 14853],
        ['Rows included in institution/year breakdown', 7944, 5994, 13938],
        ['Rows excluded from breakdown', 214, 701, 915],
        ['Exact duplicate rows removed', 61, 6, 67],
        ['Potential duplicate licence-ID groups', 50, 75, 125],
        ['Canonical institutions in breakdown', 52, 42, 58],
    ])
    _append_sheet(licence_workbook, 'Institution_Summary', [
        ['Institution Summary'],
        [None],
        [None],
        ['Institution_Canonical', 'Provisional_Total', 'Full_Licence_Total', 'Combined_Total'],
        ['Lae School of Nursing', 7944, 5994, 13938],
    ])
    _append_sheet(licence_workbook, 'Combined_By_Inst_Year', [
        ['Combined Institution/Year Breakdown'],
        [None],
        [None],
        ['Institution_Canonical', 'Year', 'Provisional_Count', 'Full_Licence_Count', 'Full_Minus_Provisional', 'Full_vs_Provisional_Ratio', 'Combined_Count', 'Institution_Combined_Total'],
        ['Lae School of Nursing', 2026, 7944, 5994, -1950, None, 13938, 13938],
    ])
    _append_sheet(licence_workbook, 'Prov_By_Inst_Year', [
        ['Provisional Licence Breakdown by Institution and Year'],
        [None],
        [None],
        ['Institution_Canonical', 'Year', 'Provisional_Count', 'Institution_Provisional_Total', 'First_Year', 'Latest_Year'],
        ['Lae School of Nursing', 2026, 7944, 7944, 2026, 2026],
    ])
    _append_sheet(licence_workbook, 'Full_By_Inst_Year', [
        ['Full Licence Breakdown by Institution and Year'],
        [None],
        [None],
        ['Institution_Canonical', 'Year', 'Full_Licence_Count', 'Institution_Full_Licence_Total', 'First_Year', 'Latest_Year'],
        ['Lae School of Nursing', 2026, 5994, 5994, 2026, 2026],
    ])
    _append_sheet(licence_workbook, 'Clean_Provisional', [
        ['Clean Provisional Licence Records'],
        [None],
        [None],
        ['Source_Row', 'Name_Clean', 'Licence_ID', 'Institution_Canonical', 'Year'],
        [2, 'Test Provisional Nurse', 'PRO 100', 'Lae School of Nursing', 2026],
    ])
    _append_sheet(licence_workbook, 'Clean_Full_Licence', [
        ['Clean Full Licence Records'],
        [None],
        [None],
        ['Source_Row', 'Name_Clean', 'Licence_ID', 'Institution_Canonical', 'Year'],
        [3, 'Test Full Nurse', 'G 100', 'Lae School of Nursing', 2026],
    ])
    _append_sheet(licence_workbook, 'DQ_Row_Issues', [
        ['Data Quality Row-Level Issues'],
        [None],
        [None],
        ['Dataset', 'Issue_Type', 'Source_Row', 'Name', 'Licence_ID'],
        ['Provisional', 'Flagged; Potential duplicate licence ID', 2, 'Test Provisional Nurse', 'PRO 100'],
    ])
    licence_workbook.save(licence_path)

    cadre_workbook = Workbook()
    cadre_sheet = cadre_workbook.active
    cadre_sheet.title = 'Cadre_Breakdown'
    for _ in range(10):
        cadre_sheet.append([None])
    cadre_sheet.append([
        'Cadre', 'Profession_Group', 'Full_Professional_Clean', 'Provisional_Clean',
        'Total_Clean', 'Full_Countable_InstYear', 'Provisional_Countable_InstYear',
        'Classification_Note',
    ])
    cadre_sheet.append(['Registered Nurse', 'Nursing', 5535, 0, 5535, 5098, 0, None])
    cadre_sheet.append(['Nursing Graduand', 'Nursing', 0, 7862, 7862, 0, 7798, None])
    cadre_sheet.append(['Midwifery', 'Midwifery', 990, 0, 990, 888, 0, None])
    cadre_sheet.append(['Midwifery Graduand', 'Midwifery', 0, 96, 96, 0, 94, None])
    cadre_sheet.append(['Unclassified / Missing qualification', 'Review', 161, 148, 309, 0, 2, None])
    cadre_sheet.append(['Unclassified / Other qualification', 'Review', 3, 17, 20, 3, 16, None])
    cadre_sheet.append(['Grand Total', 'All', 6695, 8158, 14853, 5994, 7944, None])
    _append_sheet(cadre_workbook, 'Classification_Rules', [
        ['Classification Rule Priority', 'Dataset', 'Output Cadre', 'Qualification text trigger', 'Notes'],
        [10, 'Full', 'Registered Nurse', 'nursing / nurse', 'Generic full-licence nursing records.'],
    ])
    _append_sheet(cadre_workbook, 'Unclassified_Review', [
        ['Dataset', 'Unclassified_Status', 'Qualification_Raw', 'Qualification_Canonical', 'Count'],
        ['Full professional', 'Unclassified / Missing qualification', None, None, 161],
    ])
    cadre_workbook.save(cadre_path)


class NursingCouncilAnalyticsSnapshotTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.nursing_user = user_model.objects.create_user(
            username='analytics_nursing_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        self.medical_user = user_model.objects.create_user(
            username='analytics_medical_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Medical Board',
        )

    def import_fixture(self):
        self.temp_dir = TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        path = Path(self.temp_dir.name) / 'PNG_Nursing_Council_Integrated_Dashboard_Model.xlsx'
        _build_nursing_analytics_workbook(path)
        call_command('import_nursing_analytics_snapshot', '--file', str(path))
        return path

    def test_import_command_creates_active_snapshot_and_preserves_workbook_kpis(self):
        path = self.import_fixture()

        snapshot = NursingAnalyticsSnapshot.objects.get()
        self.assertTrue(snapshot.is_active)
        self.assertEqual(snapshot.source_file_name, path.name)
        self.assertEqual(snapshot.source_batch.source_kind, 'nursing_analytics_snapshot')
        self.assertEqual(snapshot.source_batch.summary['analytics_snapshot_active'], True)
        self.assertEqual(snapshot.kpi_summary['total_lifecycle_records'], 34851)
        self.assertEqual(snapshot.kpi_summary['clean_atp_records'], 19998)
        self.assertEqual(snapshot.kpi_summary['clean_provisional_records'], 8158)
        self.assertEqual(snapshot.kpi_summary['clean_full_licence_records'], 6695)
        self.assertEqual(NursingLifecycleFact.objects.count(), 3)
        self.assertEqual(NursingStageYearMetric.objects.count(), 2)
        self.assertEqual(
            NursingStageYearMetric.objects.get(year__isnull=True).grand_total,
            2,
        )

        call_command('import_nursing_analytics_snapshot', '--file', str(path))
        self.assertEqual(NursingAnalyticsSnapshot.objects.count(), 1)

    def test_summary_and_drilldown_endpoints_are_nursing_scoped(self):
        self.import_fixture()
        self.client.force_login(self.nursing_user)

        summary_response = self.client.get(reverse('nursing_council_analytics_summary'))
        self.assertEqual(summary_response.status_code, 200)
        self.assertEqual(summary_response.json()['kpis']['total_lifecycle_records'], 34851)

        drilldown_response = self.client.get(reverse('nursing_council_analytics_drilldown'), {
            'draw': '4',
            'start': '0',
            'length': '10',
            'stage': 'Authority to Practice',
            'search[value]': 'ATP Nurse',
        })
        self.assertEqual(drilldown_response.status_code, 200)
        payload = drilldown_response.json()
        self.assertEqual(payload['draw'], 4)
        self.assertEqual(payload['recordsTotal'], 3)
        self.assertEqual(payload['recordsFiltered'], 1)
        self.assertEqual(payload['data'][0]['record_id'], 'ATP-000001')

        self.client.force_login(self.medical_user)
        denied_response = self.client.get(reverse('nursing_council_analytics_summary'))
        self.assertEqual(denied_response.status_code, 302)

    def test_dashboard_renders_analytics_surface_and_server_side_urls(self):
        self.import_fixture()
        self.client.force_login(self.nursing_user)

        response = self.client.get(reverse('nursing_council_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nursing Council Analytics Engine')
        self.assertContains(response, 'Total Lifecycle Records')
        self.assertContains(response, '34851')
        self.assertContains(response, 'analyticsLifecycleChart')
        self.assertContains(response, 'analyticsProvinceHeatmapChart')
        self.assertContains(response, reverse('nursing_council_analytics_drilldown'))
        self.assertEqual(response.context['institutions_count'], 20)
        self.assertEqual(response.context['reference_breakdown']['png_nursing_school_count'], 20)
        self.assertEqual(response.context['reference_breakdown']['government_nursing_school_count'], 9)
        self.assertContains(response, 'Recognised Schools')
        self.assertContains(response, '<tr><th>Recognised PNG nursing schools</th><td>20</td></tr>', html=True)
        self.assertNotContains(response, '<tr><th>CHW training references</th>', html=True)
        self.assertContains(response, 'CHW training references are handled in the Medical Board dashboard')

    def test_regulatory_alignment_page_uses_active_cleaned_snapshot_statistics(self):
        self.import_fixture()
        self.client.force_login(self.nursing_user)

        response = self.client.get(reverse('nursing_regulatory_alignment'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cleaned Analytics Snapshot')
        self.assertContains(response, 'PNG_Nursing_Council_Integrated_Dashboard_Model.xlsx')
        self.assertContains(response, 'nursing_analytics_snapshot')
        self.assertContains(response, 'Total Lifecycle Records')
        self.assertContains(response, '34851')
        self.assertContains(response, 'Clean ATP Records')
        self.assertContains(response, '19998')
        self.assertContains(response, 'Clean Provisional Records')
        self.assertContains(response, '8158')
        self.assertContains(response, 'Clean Full-Licence Records')
        self.assertContains(response, '6695')

    def test_catherine_breakdown_import_attaches_refreshable_verification_overlay(self):
        self.import_fixture()
        self.client.force_login(self.nursing_user)
        with TemporaryDirectory() as temp_dir:
            licence_path = Path(temp_dir) / 'PNG_Nursing_Council_Cleaned_Licence_Breakdown.xlsx'
            cadre_path = Path(temp_dir) / 'PNG_Nursing_Council_Cadre_Breakdown.xlsx'
            _build_catherine_licence_workbooks(licence_path, cadre_path)

            call_command(
                'import_nursing_catherine_licence_breakdown',
                '--licence-workbook',
                str(licence_path),
                '--cadre-workbook',
                str(cadre_path),
            )
            call_command(
                'import_nursing_catherine_licence_breakdown',
                '--licence-workbook',
                str(licence_path),
                '--cadre-workbook',
                str(cadre_path),
            )

        batch = DataImportBatch.objects.get(source_kind='nursing_catherine_licence_breakdown')
        self.assertEqual(batch.summary['active_snapshot_comparison']['status'], 'matched_active_snapshot')
        self.assertEqual(batch.summary['licence_dashboard']['Clean rows retained after exact dedupe']['combined'], 14853)
        self.assertEqual(batch.summary['cadre_breakdown']['unclassified_clean_rows'], 329)
        snapshot = NursingAnalyticsSnapshot.objects.get()
        self.assertIn('catherine_licence_breakdown', snapshot.import_summary)

        response = self.client.get(reverse('nursing_council_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Provisional and Full-Licence Cleanse Verification')
        self.assertContains(response, 'PNG_Nursing_Council_Cleaned_Licence_Breakdown.xlsx')
        self.assertContains(response, 'PNG_Nursing_Council_Cadre_Breakdown.xlsx')
        self.assertContains(response, '14853')
        self.assertContains(response, '13938')
        self.assertContains(response, '329')
        self.assertContains(response, 'matched_active_snapshot')

    def test_dashboard_populates_registrar_origin_panel_from_active_snapshot(self):
        self.import_fixture()
        snapshot = NursingAnalyticsSnapshot.objects.get()
        NursingLifecycleFact.objects.create(
            snapshot=snapshot,
            record_id='ATP-OVERSEAS-000001',
            lifecycle_stage='Authority to Practice',
            licence_status='ATP / Renewal',
            cycle_year=2026,
            full_name='Overseas Snapshot Nurse',
            person_group_key='OVERSEAS SNAPSHOT NURSE',
            cadre='Registered Nurse',
            facility='Paradise Private Hospital',
            province='National Capital District',
            organization_type='Private Organization',
            nationality_group='Overseas',
            country='Fiji',
            source_workbook='Fixture',
            source_sheet='ATP 2026',
            source_row=99,
            record_quality='High',
        )
        cache.clear()
        self.client.force_login(self.nursing_user)

        response = self.client.get(reverse('nursing_council_portal'))

        self.assertEqual(response.status_code, 200)
        summary = response.context['registrar_worker_origin_summary']
        self.assertEqual(summary['national_total'], 3)
        self.assertEqual(summary['overseas_total'], 1)
        self.assertEqual(summary['combined_total'], 4)
        self.assertEqual(response.context['registrar_worker_origin_table_limit'], 240)
        self.assertContains(response, 'Overseas Snapshot Nurse')
        self.assertContains(response, 'Analytics snapshot')


class PlatformStandardsAlignmentTests(TestCase):
    def setUp(self):
        cache.clear()
        self.admin_user = get_user_model().objects.create_user(
            username='standards_admin',
            password='StrongPass123!',
            role='admin',
            is_superuser=True,
            is_staff=True,
        )

    def test_public_home_shows_government_workforce_standards(self):
        response = self.client.get(reverse('portal_home'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PNG Medical Board &amp; Nursing Council Regulatory Agencies.')
        self.assertContains(response, 'PNG Nursing Council Regulatory Profile')
        self.assertContains(response, reverse('nursing_council_public_profile'))
        self.assertContains(response, 'Government Health Workforce Registry Standards')
        self.assertContains(response, 'NHWA primary model')
        self.assertContains(response, 'FHIR-ready practitioner roles')
        self.assertContains(response, 'DHIS2/HMIS integration path')
        self.assertContains(response, 'id="page-top"')
        self.assertContains(response, 'data-page-section-navigator')
        self.assertContains(response, 'data-page-section-drag-handle')
        self.assertContains(response, 'data-page-scroll="top"')
        self.assertContains(response, 'data-page-scroll="previous"')
        self.assertContains(response, 'data-page-scroll="next"')
        self.assertContains(response, 'data-page-scroll="bottom"')
        self.assertContains(response, 'page-section-navigation.js')
        self.assertContains(response, '.auth-links a.btn-portal')
        self.assertContains(response, 'color: #ffffff !important;')
        self.assertContains(response, 'Public Verification and Maps')
        self.assertContains(response, reverse('public_nursing_register_search_root'))
        self.assertContains(response, reverse('workforce_map'))

    def test_public_nursing_council_profile_renders_regulatory_profile_details(self):
        response = self.client.get(reverse('nursing_council_public_profile'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PNG Nursing Council Regulatory Profile')
        self.assertContains(response, 'Protecting the Public. Regulating the Profession. Advancing Nursing and Midwifery in PNG.')
        self.assertContains(response, 'Medical Registration Act 1980')
        self.assertContains(response, 'established in <strong>1964</strong>', html=False)
        for text in [
            'Registration',
            'Licensing',
            'Standards',
            'Accreditation',
            'Complaints Management',
            'Discipline',
            'Policy &amp; Advice',
            'Partnerships',
            'Strong Regulation. Safe Practice. Quality Care. Healthy People.',
            'Our Vision. Our Mandate. Our Commitment.',
        ]:
            self.assertContains(response, text)

    def test_public_nursing_register_links_to_regulatory_profile(self):
        response = self.client.get(reverse('public_nursing_register_search_root'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Regulatory Profile')
        self.assertContains(response, reverse('nursing_council_public_profile'))

    def test_staff_dashboard_shell_links_to_standards_alignment(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('advanced_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Government Health Workforce Registry Standards')
        self.assertContains(response, reverse('platform_standards_alignment'))
        self.assertContains(response, 'Standards &amp; Compliance')
        self.assertContains(response, 'id="page-top"')
        self.assertContains(response, 'data-page-section-navigator')
        self.assertContains(response, 'data-page-section-drag-handle')
        self.assertContains(response, 'data-page-scroll="top"')
        self.assertContains(response, 'data-page-scroll="previous"')
        self.assertContains(response, 'data-page-scroll="next"')
        self.assertContains(response, 'data-page-scroll="bottom"')
        self.assertContains(response, 'page-section-navigation.js')
        self.assertContains(response, '20260708-dashboard-links-1')

    def test_helpdesk_widget_can_be_recovered_and_voice_can_be_stopped(self):
        base_template = (Path(__file__).resolve().parents[2] / 'templates' / 'base.html').read_text(encoding='utf-8')

        self.assertEqual(base_template.count('data-helpdesk-drag-handle'), 3)
        self.assertEqual(base_template.count('id="helpdesk-stop-voice"'), 3)
        self.assertIn("const assistantWidgetPositionStorageKey = 'ndohAssistantWidgetPosition';", base_template)
        self.assertIn('function applyHelpdeskWidgetPosition(left, top)', base_template)
        self.assertIn('function ensureHelpdeskWidgetInViewport()', base_template)
        self.assertIn("window.addEventListener('resize', ensureHelpdeskWidgetInViewport);", base_template)
        self.assertIn('function stopHelpdeskSpeech()', base_template)
        self.assertIn('window.speechSynthesis.cancel();', base_template)
        self.assertIn('window.ndohVoiceRecognitionAvailability = function ()', base_template)
        self.assertIn('window.ndohStartVoiceRecognition = function (options)', base_template)
        self.assertIn('recognition.onerror = function (event)', base_template)
        self.assertIn('Voice input needs HTTPS when the platform is opened through a Wi-Fi or hotspot address.', base_template)
        self.assertIn("setHelpdeskVoiceInputStatus('Listening - speak your staff question now.'", base_template)
        self.assertIn('activeHelpdeskRecognition.stop()', base_template)
        self.assertIn('const requestSequence = ++helpdeskRequestSequence;', base_template)
        self.assertIn('if (requestSequence !== helpdeskRequestSequence)', base_template)
        self.assertNotIn('function applyAssistantPosition(left, top)', base_template)

    def test_staff_ai_voice_input_shows_listening_and_actionable_error_status(self):
        staff_ai_template = (
            Path(__file__).resolve().parents[2] / 'templates' / 'dashboard' / 'staff_ai_assistant.html'
        ).read_text(encoding='utf-8')

        self.assertIn('id="staff-ai-voice-status"', staff_ai_template)
        self.assertIn('aria-live="polite"', staff_ai_template)
        self.assertIn('window.ndohStartVoiceRecognition({', staff_ai_template)
        self.assertIn("setVoiceStatus('Listening - speak your staff question now.'", staff_ai_template)
        self.assertIn("setVoiceStatus('Voice question captured. Sending secure platform AI request...'", staff_ai_template)
        self.assertIn('activeVoiceRecognition.stop()', staff_ai_template)
        self.assertIn('setVoiceStatus(detail.message, true)', staff_ai_template)

    def test_page_section_navigator_does_not_block_dashboard_links(self):
        static_root = Path(__file__).resolve().parents[2] / 'static'
        css = (static_root / 'css' / 'government-enterprise.css').read_text(encoding='utf-8')
        js = (static_root / 'js' / 'page-section-navigation.js').read_text(encoding='utf-8')

        navigator_block = css.split('\n.page-section-navigator {', 1)[1].split('}', 1)[0]
        drag_block = css.split('.page-section-nav-drag {', 1)[1].split('}', 1)[0]
        button_block = css.split('.page-section-nav-button {', 1)[1].split('}', 1)[0]

        self.assertIn('pointer-events: none;', navigator_block)
        self.assertIn('pointer-events: auto;', drag_block)
        self.assertIn('pointer-events: auto;', button_block)
        self.assertIn('localStorage.removeItem(navigatorStorageKey)', js)
        self.assertNotIn('localStorage.setItem(navigatorStorageKey', js)

    def test_government_styles_keep_dark_dashboard_cards_readable(self):
        css_path = Path(__file__).resolve().parents[2] / 'static' / 'css' / 'government-enterprise.css'
        css = css_path.read_text(encoding='utf-8')

        self.assertIn('Dark surface contrast guard', css)
        self.assertIn('.small-box.bg-success', css)
        self.assertIn('.small-box.bg-dark', css)
        self.assertIn('-webkit-text-fill-color: #ffffff !important;', css)
        self.assertIn(':where(.text-muted, .text-secondary, small)', css)
        self.assertIn('-webkit-text-fill-color: currentColor !important;', css)
        self.assertIn('.btn-outline-secondary:hover', css)
        self.assertIn('-webkit-text-fill-color: var(--gov-ink) !important;', css)
        self.assertIn('-webkit-text-fill-color: #111827 !important;', css)

    def test_government_styles_keep_engagement_and_review_headers_readable(self):
        css_path = Path(__file__).resolve().parents[2] / 'static' / 'css' / 'government-enterprise.css'
        css = css_path.read_text(encoding='utf-8')

        self.assertIn('Engagement and review contrast guard', css)
        self.assertIn('.review-centre-hero', css)
        self.assertIn('.platform-standard-badge.badge-primary', css)
        self.assertIn('-webkit-text-fill-color: #ffffff !important;', css)
        self.assertIn('-webkit-text-fill-color: currentColor !important;', css)
        self.assertGreaterEqual(contrast_ratio('#ffffff', '#12304a'), 4.5)
        self.assertGreaterEqual(contrast_ratio('#ffffff', '#0f766e'), 4.5)
        self.assertGreaterEqual(contrast_ratio('#0a3a68', '#eaf3ff'), 4.5)
        self.assertGreaterEqual(contrast_ratio('#071827', '#f59e0b'), 4.5)

    def test_government_styles_constrain_system_admin_tablet_layout(self):
        css_path = Path(__file__).resolve().parents[2] / 'static' / 'css' / 'government-enterprise.css'
        css = css_path.read_text(encoding='utf-8')

        self.assertIn('System admin tablet layout', css)
        self.assertIn('.admin-command-centre .admin-shell', css)
        self.assertIn('grid-template-columns: minmax(0, 1fr) !important;', css)
        self.assertIn('.dataTables_wrapper', css)
        self.assertIn('overflow-x: auto !important;', css)
        self.assertIn('@media (max-width: 991.98px)', css)

    def test_nhwa_toolkit_styles_keep_dark_surfaces_readable(self):
        css_path = Path(__file__).resolve().parents[2] / 'static' / 'css' / 'government-enterprise.css'
        css = css_path.read_text(encoding='utf-8')

        self.assertIn('NHWA toolkit contrast guard', css)
        self.assertIn('.nhwa-toolkit-page .nhwa-hero', css)
        self.assertIn('.nhwa-workbook-table .sheet-section-row td', css)
        self.assertIn('-webkit-text-fill-color: #ffffff !important;', css)
        self.assertGreaterEqual(contrast_ratio('#ffffff', '#12304a'), 4.5)
        self.assertGreaterEqual(contrast_ratio('#0f172a', '#ffffff'), 4.5)

        cell = SimpleNamespace(
            fill=SimpleNamespace(
                fill_type='solid',
                fgColor=SimpleNamespace(type='rgb', rgb='FF12304A'),
            )
        )

        style = _cell_style(cell, 'Dark filled NHWA cell')

        self.assertIn('background-color: #12304A;', style)
        self.assertIn('color: #ffffff !important;', style)
        self.assertIn('-webkit-text-fill-color: #ffffff !important;', style)

    def test_profile_settings_tabs_reset_text_fill_on_white_active_tab(self):
        css_path = Path(__file__).resolve().parents[2] / 'static' / 'css' / 'government-enterprise.css'
        css = css_path.read_text(encoding='utf-8')

        self.assertIn('Profile settings tab contrast', css)
        self.assertIn('.profile-settings-tabs .nav-link.active', css)
        self.assertIn('border-bottom: 3px solid #0f3f73 !important;', css)
        self.assertIn('-webkit-text-fill-color: #0f2538 !important;', css)
        self.assertGreaterEqual(contrast_ratio('#0f2538', '#ffffff'), 4.5)

    def test_forum_faq_and_review_pages_render_readability_surfaces(self):
        call_command('seed_engagement_platform')
        self.client.force_login(self.admin_user)

        forum_response = self.client.get(reverse('forum_index'))
        faq_response = self.client.get(reverse('public_faqs'))
        review_response = self.client.get(reverse('review_centre'))

        self.assertEqual(forum_response.status_code, 200)
        self.assertContains(forum_response, 'card card-outline card-primary')
        self.assertContains(forum_response, 'Public posts are moderated.')
        self.assertEqual(faq_response.status_code, 200)
        self.assertContains(faq_response, 'card card-outline card-secondary')
        self.assertContains(faq_response, 'Applications and Renewals')
        self.assertEqual(review_response.status_code, 200)
        self.assertContains(review_response, 'review-centre-hero')
        self.assertContains(review_response, 'Consolidated review-only dashboard')

    def test_standards_alignment_page_maps_nhwa_hmis_and_fhir(self):
        self.client.force_login(self.admin_user)
        DataImportBatch.objects.create(
            source_file_name='standards-source.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        Facility.objects.create(name='Standards Facility', type='Hospital', ownership='public')
        NursingProfessional.objects.create(
            first_name='Standards',
            last_name='Nurse',
            registration_no='STD-NURSE-1',
        )

        response = self.client.get(reverse('platform_standards_alignment'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'NHWA Alignment Map')
        self.assertContains(response, 'WHO National Health Workforce Accounts handbook')
        self.assertContains(response, 'HL7 FHIR PractitionerRole')
        self.assertContains(response, 'DHIS2 HMIS platform')
        self.assertContains(response, 'Active workforce stock')
        self.assertContains(response, 'PNG Department of Health Standards')
        self.assertContains(response, 'National Health Service Standards')
        self.assertContains(response, 'Monitoring, evaluation, reporting, and data management')
        self.assertContains(response, 'PNGNC Situational Analysis Requirements')
        self.assertContains(response, 'legal defensibility')
        self.assertContains(response, 'PNG National Department of Health')
        self.assertContains(response, 'PNG Nursing Council Situational Analysis Report, 30 January 2026')
        self.assertContains(response, 'Government Health Workforce Registry Standards', count=1)
        self.assertNotContains(response, 'class="platform-standards-bar"')
        self.assertContains(response, 'Standards Facility', count=0)


class WorkforceFlowImportStatsTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username='flow_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        self.client.force_login(self.user)

    def test_flow_uses_latest_nursing_license_import_when_no_ndata_batch_exists(self):
        current_year = timezone.localdate().year
        atp_batch = DataImportBatch.objects.create(
            source_file_name='2026 Current ATP-DATA Statistics & Tracking latest.xlsx',
            source_kind='ndata_workbook',
            status='completed',
            total_sheets=23,
            processed_sheets=23,
            total_rows=1,
            processed_rows=1,
        )
        atp_batch.started_at = timezone.now() - timedelta(days=1)
        atp_batch.save(update_fields=['started_at'])
        PracticingLicenseRecord.objects.create(
            batch=atp_batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=1,
            record_year=current_year,
            full_name='Current Workforce Nurse',
            registration_no='NC-ATP-1',
            gender='Female',
            category='General Nurse',
            province='National Capital District',
            workplace_address='Port Moresby General Hospital, P O Box 1, Boroko',
        )

        batch = DataImportBatch.objects.create(
            source_file_name='Pro and full license continue 2024 - 2025.xlsx',
            source_kind='nursing_license_workbook',
            status='completed',
            total_sheets=2,
            processed_sheets=2,
            total_rows=2,
            processed_rows=2,
        )
        PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='provisional',
            target_model='nursingprofessional',
            source_sheet_name='PROV REGO',
            source_row=1,
            record_year=current_year,
            full_name='Graduand Applicant',
            registration_no='NC-PROV-1',
        )
        PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='full',
            target_model='nursingprofessional',
            source_sheet_name='FULL LICENSE',
            source_row=2,
            record_year=current_year + 1,
            full_name='Full Licence Applicant',
            registration_no='NC-FULL-1',
        )

        response = self.client.get(reverse('workforce_flow'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['latest_import_batch'], batch)
        self.assertEqual(response.context['import_record_count'], 2)
        self.assertEqual(response.context['import_latest_year'], current_year)
        self.assertEqual(response.context['flow_labels'], ['Provisional', 'Full/Temporary', 'Renewals', 'Young Workforce'])
        self.assertEqual(response.context['flow_data'][:3], [1, 1, 0])
        self.assertEqual(response.context['import_workplace_source_batch'], atp_batch)
        self.assertEqual(response.context['province_labels'], ['National Capital District'])
        self.assertEqual(response.context['import_gender_labels'], ['Female'])
        self.assertEqual(response.context['import_workplace_rows'][0]['workplace'], 'Port Moresby General Hospital, P O Box 1, Boroko')
        self.assertContains(response, '2</h3><p>Imported Workbook Rows')

    def test_flow_renders_operational_task_inbox_and_pathway_drilldowns(self):
        nursing_professional = NursingProfessional.objects.create(
            first_name='Workflow',
            last_name='Nurse',
            registration_no='WF-NC-1',
            email='workflow.nurse@example.test',
        )
        application = Application.objects.create(
            form_code='NC1',
            pathway='local_nursing_graduate',
            status='pending',
            payload={'full_name': 'Workflow Nurse'},
        )
        application.submitted_date = timezone.localdate() - timedelta(days=16)
        application.save(update_fields=['submitted_date'])
        MissingDataReview.objects.create(
            content_type=ContentType.objects.get_for_model(NursingProfessional),
            object_id=nursing_professional.pk,
            full_name='Workflow Nurse',
            professional_type='Nursing Professional',
            registration_no='WF-NC-1',
            missing_fields=['Email address', 'Date of birth', 'Training institution', 'Photo'],
            missing_count=4,
            severity='high',
        )
        DuplicateReviewQueue.objects.create(
            content_type=ContentType.objects.get_for_model(NursingProfessional),
            object_id=nursing_professional.pk,
            suspected_duplicate={'target_model': 'nursingprofessional'},
            similarity_score=0.94,
        )
        Receipt.objects.create(
            user=self.user,
            application=application,
            receipt_number='WF-RCT-1',
            amount='250.00',
            status='pending',
            payer_match_confidence='unlinked',
        )
        Document.objects.create(
            title='Workflow draft procedure',
            office_scope='nursing',
            status='draft',
        )

        response = self.client.get(reverse('workforce_flow'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['dashboard_scope'], 'nursing')
        self.assertEqual(response.context['workflow_pending_application_count'], 1)
        self.assertEqual(response.context['workflow_missing_data_count'], 1)
        self.assertEqual(response.context['workflow_duplicate_count'], 1)
        self.assertEqual(response.context['workflow_receipt_count'], 1)
        self.assertEqual(response.context['workflow_high_priority_count'], 4)
        provisional_row = next(
            row for row in response.context['workflow_pathway_rows']
            if row['key'] == 'nursing_provisional'
        )
        self.assertEqual(provisional_row['pending_count'], 1)
        self.assertEqual(provisional_row['aged_pending_count'], 1)
        self.assertEqual(provisional_row['task_count'], 1)
        self.assertContains(response, 'Operational Task Inbox')
        self.assertContains(response, 'Task Queue')
        self.assertContains(response, 'Workflow Pipeline')
        self.assertContains(response, 'NC1 / Provisional licence')
        self.assertContains(response, 'NC2 / Full licence')
        self.assertContains(response, 'NC3 / ATP renewal')
        self.assertContains(response, 'Workflow Nurse')
        self.assertContains(response, 'Possible duplicate record')
        self.assertContains(response, 'Workflow draft procedure')

        filtered_response = self.client.get(reverse('workforce_flow'), {
            'priority': 'high',
            'pathway': 'nursing_provisional',
        })

        self.assertEqual(filtered_response.status_code, 200)
        self.assertEqual(filtered_response.context['workflow_filtered_task_count'], 1)
        self.assertContains(filtered_response, 'NC1 pending application')
        self.assertNotContains(filtered_response, 'Workflow draft procedure')

    def test_nursing_flow_reference_tables_hide_misclassified_medical_board_cadres(self):
        Cadre.objects.create(name='Registered Nurse', category='nursing')
        Cadre.objects.create(name='Medical Doctor / Specialist', category='nursing')
        Cadre.objects.create(name='Community Health Worker (CHW)', category='nursing')
        Cadre.objects.create(name='Allied Health Professional', category='nursing')

        response = self.client.get(reverse('workforce_flow'))

        self.assertEqual(response.status_code, 200)
        cadre_names = set(response.context['cadres'].values_list('name', flat=True))
        self.assertIn('Registered Nurse', cadre_names)
        self.assertNotIn('Medical Doctor / Specialist', cadre_names)
        self.assertNotIn('Community Health Worker (CHW)', cadre_names)
        self.assertNotIn('Allied Health Professional', cadre_names)
        self.assertContains(response, 'Registered Nurse')
        self.assertNotContains(response, 'Medical Doctor / Specialist')
        self.assertNotContains(response, 'Community Health Worker (CHW)')
        self.assertNotContains(response, 'Allied Health Professional')

    def test_medical_flow_context_receives_misclassified_medical_board_cadres(self):
        medical_user = get_user_model().objects.create_user(
            username='medical_reference_flow_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Medical Board',
        )
        Cadre.objects.create(name='Registered Nurse', category='nursing')
        Cadre.objects.create(name='Medical Doctor / Specialist', category='nursing')
        Cadre.objects.create(name='Community Health Worker (CHW)', category='nursing')
        Cadre.objects.create(name='Allied Health Professional', category='nursing')
        self.client.force_login(medical_user)

        response = self.client.get(reverse('workforce_flow'))

        self.assertEqual(response.status_code, 200)
        cadre_names = set(response.context['cadres'].values_list('name', flat=True))
        self.assertIn('Medical Doctor / Specialist', cadre_names)
        self.assertIn('Community Health Worker (CHW)', cadre_names)
        self.assertIn('Allied Health Professional', cadre_names)
        self.assertNotIn('Registered Nurse', cadre_names)

    def test_flow_ignores_completed_nursing_workbook_with_no_records(self):
        current_year = timezone.localdate().year
        empty_batch = DataImportBatch.objects.create(
            source_file_name='tmp-empty-workbook.xlsx',
            source_kind='ndata_workbook',
            status='completed',
            total_sheets=7,
            processed_sheets=7,
            total_rows=0,
            processed_rows=0,
        )
        empty_batch.started_at = timezone.now() + timedelta(minutes=5)
        empty_batch.save(update_fields=['started_at'])

        atp_batch = DataImportBatch.objects.create(
            source_file_name='ATP_RECORD_2026_NEW_RECORDS_ONLY_FOR_IMPORT.xlsx',
            source_kind='ndata_workbook',
            status='completed',
            total_sheets=1,
            processed_sheets=1,
            total_rows=2,
            processed_rows=2,
        )
        atp_batch.started_at = timezone.now()
        atp_batch.save(update_fields=['started_at'])
        PracticingLicenseRecord.objects.create(
            batch=atp_batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=1,
            record_year=current_year,
            full_name='National Nurse',
            registration_no='NC-ATP-2026-1',
            applicant_type='national',
            category='General Nurse',
        )
        PracticingLicenseRecord.objects.create(
            batch=atp_batch,
            record_type='practicing_license',
            target_model='midwife',
            source_sheet_name='ATP RECORD 2026',
            source_row=2,
            record_year=current_year,
            full_name='Overseas Midwife',
            registration_no='NC-ATP-2026-2',
            applicant_type='overseas',
            category='Specialist Midwife',
        )

        response = self.client.get(reverse('workforce_flow'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['latest_import_batch'], atp_batch)
        self.assertEqual(response.context['import_record_count'], 2)
        self.assertEqual(response.context['category_labels'], ['General Nurse', 'Specialist Midwife'])
        self.assertEqual(response.context['category_values'], [1, 1])
        self.assertEqual(response.context['import_applicant_type_labels'], ['National', 'Overseas'])
        self.assertEqual(response.context['import_applicant_type_values'], [1, 1])
        self.assertEqual(response.context['flow_data'][:3], [0, 0, 2])

    def test_medical_board_flow_uses_medical_import_and_receipts_only(self):
        cache.clear()
        current_year = timezone.localdate().year
        medical_user = get_user_model().objects.create_user(
            username='medical_flow_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Medical Board',
        )
        medical_batch = DataImportBatch.objects.create(
            source_file_name='medical-board.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
            total_sheets=1,
            processed_sheets=1,
            total_rows=3,
            processed_rows=3,
        )
        medical_batch.sheets.create(
            sheet_name='CHW',
            sheet_type='medical_board_chw',
            status='completed',
            imported_rows=3,
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            record_type='full',
            target_model='medicaldoctor',
            source_sheet_name='CHW',
            source_row=1,
            record_year=current_year,
            full_name='Medical Doctor One',
            registration_no='MD-FLOW-1',
            gender='Male',
            category='Medical Doctor',
            province='National Capital District',
            workplace_address='Medical Board Clinic',
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            record_type='workforce_listing',
            target_model='communityhealthworker',
            source_sheet_name='CHW',
            source_row=2,
            record_year=current_year,
            full_name='Community Health Worker One',
            registration_no='CHW-FLOW-1',
            gender='Female',
            category='Community Health Worker',
            province='Morobe Province',
            workplace_address='Medical Board Clinic',
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            record_type='practicing_license',
            target_model='communityhealthworker',
            source_sheet_name='CHW',
            source_row=3,
            record_year=current_year,
            full_name='Community Health Worker One',
            registration_no='CHW-FLOW-1',
            gender='Female',
            category='Community Health Worker',
            province='Morobe Province',
            workplace_address='Medical Board Clinic',
        )

        nursing_batch = DataImportBatch.objects.create(
            source_file_name='nursing-latest.xlsx',
            source_kind='ndata_workbook',
            status='completed',
            total_sheets=1,
            processed_sheets=1,
            total_rows=1,
            processed_rows=1,
        )
        nursing_batch.started_at = timezone.now() + timedelta(days=1)
        nursing_batch.save(update_fields=['started_at'])
        nursing_batch.sheets.create(
            sheet_name='PROV REGO',
            sheet_type='provisional',
            status='completed',
            imported_rows=1,
        )
        PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            record_type='provisional',
            target_model='nursingprofessional',
            source_sheet_name='PROV REGO',
            source_row=1,
            record_year=current_year,
            full_name='Nursing Applicant One',
            registration_no='NC-FLOW-1',
            workplace_address='Nursing Council Hospital',
        )

        medical_application = Application.objects.create(form_code='MD2', pathway='medical_board', status='pending')
        nursing_application = Application.objects.create(form_code='NC3', pathway='other', status='pending')
        Receipt.objects.create(
            user=medical_user,
            application=medical_application,
            receipt_number='MB-FLOW-RCT',
            amount='100.00',
            status='completed',
        )
        Receipt.objects.create(
            user=self.user,
            application=nursing_application,
            receipt_number='NC-FLOW-RCT',
            amount='50.00',
            status='pending',
        )

        self.client.force_login(medical_user)
        response = self.client.get(reverse('workforce_flow'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['dashboard_scope'], 'medical')
        self.assertEqual(response.context['latest_import_batch'], medical_batch)
        self.assertEqual(response.context['import_record_count'], 3)
        self.assertEqual(response.context['flow_labels'], ['Medical Doctors', 'Community Health Workers', 'Allied Health', 'Practising Licences'])
        self.assertEqual(response.context['flow_data'], [1, 1, 0, 1])
        self.assertEqual(response.context['receipt_completed_count'], 1)
        self.assertEqual(response.context['receipt_pending_count'], 0)
        self.assertEqual(response.context['latest_import_sheets'][0].sheet_name, 'CHW')
        self.assertEqual(response.context['import_workplace_rows'][0]['workplace'], 'Medical Board Clinic')
        self.assertEqual(response.context['workforce_flow_title'], 'Medical Board Workforce Flow & Planning')
        self.assertContains(response, 'Imported Medical Board Rows')
        self.assertContains(response, 'CHW')
        self.assertNotContains(response, 'PROV REGO')
        self.assertNotContains(response, 'Nursing Council Hospital')


class RegistrarIndividualRecordsTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.nursing_registrar = user_model.objects.create_user(
            username='individual_nursing_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        self.medical_registrar = user_model.objects.create_user(
            username='individual_medical_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Medical Board',
        )
        self.admin = user_model.objects.create_user(
            username='individual_admin',
            password='StrongPass123!',
            role='admin',
            is_superuser=True,
            is_staff=True,
        )

    def test_nursing_portal_handles_recent_application_without_linked_record(self):
        nursing_ct = ContentType.objects.get_for_model(NursingProfessional)
        Application.objects.create(
            content_type=nursing_ct,
            object_id=999999,
            form_code='NC1',
            status='pending',
        )
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse('nursing_council_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Applicant record not linked')

    def test_individual_records_show_national_overseas_employment_and_facility_details(self):
        current_year = timezone.localdate().year
        batch = DataImportBatch.objects.create(
            source_file_name='nursing-individuals.xlsx',
            source_kind='nursing_license_workbook',
            status='completed',
            total_rows=2,
            processed_rows=2,
        )
        PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='full',
            target_model='nursingprofessional',
            source_sheet_name='FULL',
            source_row=1,
            record_year=current_year,
            full_name='National Source Nurse',
            registration_no='NC-IND-1',
            applicant_type='national',
            nationality='Papua New Guinea',
            institution_name='Lae School of Nursing',
            workplace_address='Port Moresby General Hospital, P O Box 1, Boroko',
            province='NCD',
        )
        PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='temporary',
            target_model='nursingprofessional',
            source_sheet_name='TEMP',
            source_row=2,
            record_year=current_year,
            full_name='Overseas Source Nurse',
            registration_no='NC-IND-2',
            applicant_type='overseas',
            nationality='Fiji',
            institution_name='Fiji Nursing School',
            workplace_address='Kundiawa General Hospital',
            province='Simbu',
        )
        live_nurse = NursingProfessional.objects.create(
            first_name='Live',
            last_name='Facility Nurse',
            registration_no='NC-LIVE-1',
            applicant_type='national',
            nationality='Papua New Guinea',
            province='Morobe',
        )
        content_type = ContentType.objects.get_for_model(NursingProfessional)
        EmploymentRecord.objects.create(
            content_type=content_type,
            object_id=live_nurse.pk,
            employer_name='Angau Hospital',
            position_held='Registered Nurse',
            employment_status='full_time',
            area_of_employment='government',
            place_of_work='ANGAU Memorial Hospital',
        )
        location = Location.objects.create(province='Morobe', district='Lae')
        facility = Facility.objects.create(
            name='ANGAU Memorial Hospital',
            type='Hospital',
            ownership='public',
            location=location,
        )
        PostingHistory.objects.create(
            content_type=content_type,
            object_id=live_nurse.pk,
            facility=facility,
            position_title='Ward Nurse',
            start_date=timezone.localdate(),
            is_current=True,
        )

        self.client.force_login(self.admin)
        response = self.client.get(reverse('registrar_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.context['registrar_worker_origin_summary']['national_total'], 2)
        self.assertGreaterEqual(response.context['registrar_worker_origin_summary']['overseas_total'], 1)
        self.assertContains(response, 'National and Overseas Workers')
        self.assertContains(response, 'National Individuals')
        self.assertContains(response, 'Overseas Individuals')
        self.assertContains(response, 'National Source Nurse')
        self.assertContains(response, 'Overseas Source Nurse')
        self.assertContains(response, 'Fiji Nursing School')
        self.assertContains(response, 'ANGAU Memorial Hospital')
        self.assertContains(response, 'Incoming - registration')

    def test_medical_registrar_scope_excludes_nursing_individual_records(self):
        current_year = timezone.localdate().year
        nursing_batch = DataImportBatch.objects.create(
            source_file_name='nursing-source.xlsx',
            source_kind='nursing_license_workbook',
            status='completed',
        )
        medical_batch = DataImportBatch.objects.create(
            source_file_name='medical-source.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
        )
        PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            record_type='full',
            target_model='nursingprofessional',
            source_sheet_name='NC',
            source_row=1,
            record_year=current_year,
            full_name='Nursing Scoped Person',
            registration_no='NC-SCOPE-IND',
            applicant_type='national',
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            record_type='full',
            target_model='medicaldoctor',
            source_sheet_name='MB',
            source_row=1,
            record_year=current_year,
            full_name='Medical Scoped Person',
            registration_no='MD-SCOPE-IND',
            applicant_type='national',
        )

        cache.clear()
        context = _registrar_worker_origin_context(self.medical_registrar)
        names = {row['name'] for row in context['registrar_worker_origin_rows']}

        self.assertEqual(context['registrar_origin_scope'], 'medical')
        self.assertIn('Medical Scoped Person', names)
        self.assertNotIn('Nursing Scoped Person', names)

    def test_frequent_nursing_records_drilldown_shows_individuals_and_edit_actions(self):
        current_year = timezone.localdate().year
        batch = DataImportBatch.objects.create(
            source_file_name='2026 ATP current.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        valid_record = PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP',
            source_row=10,
            record_year=current_year,
            full_name='Valid General Nurse',
            registration_no='ATP-GN-1',
            category='General Nurse',
            workplace_address='Western Provincial Health Authority',
            province='Western Province',
            payment_date=timezone.localdate(),
        )
        review_record = PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP',
            source_row=11,
            record_year=current_year,
            full_name='Review General Nurse',
            registration_no='ATP-GN-2',
            category='General Nurse',
            workplace_address='Paradise Private Hospital',
            province='National Capital District',
            payment_date=timezone.localdate(),
        )
        typo_record = PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP',
            source_row=12,
            record_year=current_year,
            full_name='Typo Category Nurse',
            registration_no='ATP-TYPO-1',
            category='Specialsit Midwife',
            workplace_address='Unknown Aid Post',
            province='Madang Province',
            payment_date=timezone.localdate(),
        )
        MissingDataReview.objects.create(
            content_type=ContentType.objects.get_for_model(PracticingLicenseRecord),
            object_id=review_record.pk,
            full_name=review_record.full_name,
            registration_no=review_record.registration_no,
            professional_type='Practicing License Record',
            missing_fields=['Province needs verification'],
            missing_count=1,
            source_label=review_record.source_sheet_name,
            source_row=review_record.source_row,
            severity='low',
        )

        self.client.force_login(self.nursing_registrar)
        response = self.client.get(reverse('nursing_frequent_records'), {'category': 'General Nurse'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['record_total'], 2)
        self.assertContains(response, 'Valid General Nurse')
        self.assertContains(response, 'Review General Nurse')
        self.assertNotContains(response, 'Typo Category Nurse')
        self.assertContains(response, reverse('record_update', args=['practicinglicenserecord', valid_record.pk]))
        self.assertContains(response, 'Validity')
        self.assertContains(response, 'Needs review')

        response = self.client.get(reverse('nursing_frequent_records'), {'category_review': '1'})
        self.assertEqual(response.context['record_total'], 1)
        self.assertContains(response, typo_record.full_name)
        self.assertContains(response, 'Category label is not in the standard registrar list')

        response = self.client.get(reverse('nursing_frequent_records'), {'facility_group': 'private'})
        self.assertEqual(response.context['record_total'], 1)
        self.assertContains(response, review_record.full_name)
        self.assertNotContains(response, valid_record.full_name)

    def test_data_quality_reviews_table_uses_server_side_datatable(self):
        current_year = timezone.localdate().year
        payment_date = timezone.localdate()
        batch = DataImportBatch.objects.create(
            source_file_name='quality-review-atp.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        record = PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP QUALITY',
            source_row=21,
            record_year=current_year,
            full_name='Quality Table Nurse',
            registration_no='NC-QUALITY-1',
            category='General Nurse',
            payment_date=payment_date,
        )
        MissingDataReview.objects.create(
            content_type=ContentType.objects.get_for_model(PracticingLicenseRecord),
            object_id=record.pk,
            full_name=record.full_name,
            registration_no=record.registration_no,
            professional_type='Practicing License Record',
            missing_fields=['Suspected misinformation in workplace facility'],
            missing_count=1,
            source_label=record.source_sheet_name,
            source_row=record.source_row,
            severity='high',
        )

        self.client.force_login(self.nursing_registrar)
        portal_response = self.client.get(reverse('nursing_council_portal'))
        self.assertContains(portal_response, 'id="data-quality-year-summary-table"')
        self.assertContains(portal_response, 'data-quality-year-summary-datatable')
        self.assertContains(portal_response, 'Search yearly summary:')
        self.assertContains(portal_response, 'Show _MENU_ source years')
        self.assertContains(portal_response, 'data-quality-review-datatable')
        self.assertContains(portal_response, 'data-server-side="1"')
        self.assertContains(portal_response, 'data-data-quality-filter="severity"')
        self.assertContains(portal_response, 'data-data-quality-filter="status"')
        self.assertContains(portal_response, 'data-data-quality-filter="source_year"')
        self.assertContains(portal_response, 'data-data-quality-reset')
        self.assertContains(portal_response, "pagingType: 'full_numbers'")
        self.assertContains(portal_response, 'pageLength: 10')
        self.assertContains(portal_response, reverse('data_quality_reviews_table'))

        response = self.client.get(reverse('data_quality_reviews_table'), {
            'draw': '3',
            'start': '0',
            'length': '10',
            'search[value]': 'Quality Table Nurse',
            'order[0][column]': '0',
            'order[0][dir]': 'desc',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['draw'], 3)
        self.assertEqual(payload['recordsTotal'], 1)
        self.assertEqual(payload['recordsFiltered'], 1)
        self.assertEqual(len(payload['data']), 1)
        row = payload['data'][0]
        self.assertEqual(row['source_year'], current_year)
        self.assertEqual(row['recent_date'], payment_date.strftime('%d %b %Y'))
        self.assertIn('Quality Table Nurse', row['name'])
        self.assertIn('Suspected misinformation', row['issues'])
        self.assertIn('ATP QUALITY, row 21', row['source'])
        self.assertIn('badge-danger', row['severity'])
        self.assertIn(reverse('record_update', args=['practicinglicenserecord', record.pk]), row['actions'])

        filtered_response = self.client.get(reverse('data_quality_reviews_table'), {
            'draw': '4',
            'start': '0',
            'length': '10',
            'severity': 'medium',
            'status': 'under_review',
            'source_year': str(current_year),
        })

        self.assertEqual(filtered_response.status_code, 200)
        filtered_payload = filtered_response.json()
        self.assertEqual(filtered_payload['draw'], 4)
        self.assertEqual(filtered_payload['recordsTotal'], 1)
        self.assertEqual(filtered_payload['recordsFiltered'], 0)

        high_response = self.client.get(reverse('data_quality_reviews_table'), {
            'draw': '5',
            'start': '0',
            'length': '10',
            'severity': 'high',
            'status': 'under_review',
            'source_year': str(current_year),
        })

        self.assertEqual(high_response.status_code, 200)
        high_payload = high_response.json()
        self.assertEqual(high_payload['recordsFiltered'], 1)
        self.assertEqual(high_payload['data'][0]['source_year'], current_year)

    def test_nursing_council_portal_has_shortcut_links_to_key_sections(self):
        current_year = timezone.localdate().year
        batch = DataImportBatch.objects.create(
            source_file_name='shortcut-atp.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        record = PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP SHORTCUT',
            source_row=1,
            record_year=current_year,
            full_name='Shortcut Nurse',
            registration_no='NC-SHORTCUT-1',
            category='General Nurse',
            payment_date=timezone.localdate(),
        )
        MissingDataReview.objects.create(
            content_type=ContentType.objects.get_for_model(PracticingLicenseRecord),
            object_id=record.pk,
            full_name=record.full_name,
            registration_no=record.registration_no,
            professional_type='Practicing License Record',
            missing_fields=['Facility needs verification'],
            missing_count=1,
            severity='medium',
        )

        self.client.force_login(self.nursing_registrar)
        response = self.client.get(reverse('nursing_council_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Dashboard Shortcuts')
        self.assertContains(response, reverse('public_nursing_register_search_root'))
        self.assertContains(response, 'Register Search')
        self.assertContains(response, 'Public Trust and e-Services')
        self.assertContains(response, 'Nursing Assurance Standards')
        for anchor in [
            'nursing-public-protection',
            'nursing-summary',
            'institution-facility-breakdown',
            'registrar-worker-origin-table',
            'nursing-operations',
            'workflow-pathways',
            'missing-data-reviews',
            'atp-current-summary',
            'frequent-nursing-totals',
            'atp-charts',
            'atp-workplace-breakdown',
            'current-atp-records',
            'workforce-flow',
            'provisional-licence-tracking',
            'recent-nursing-applications',
            'nursing-statistics',
        ]:
            self.assertContains(response, f'href="#{anchor}"')
            self.assertContains(response, f'id="{anchor}"')

    def test_nursing_records_hub_renders_register_cockpit(self):
        NursingProfessional.objects.create(
            first_name='Records',
            last_name='Nurse',
            registration_no='NC-HUB-1',
            qualification_level='Diploma Nursing',
        )
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse('records_home'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nursing Council Register Cockpit')
        self.assertContains(response, 'Public Nursing Register')
        self.assertContains(response, 'ATP / Practising Records')
        self.assertContains(response, reverse('public_nursing_register_search_root'))
        self.assertNotContains(response, 'Medical Board Register Cockpit')

    def test_nurse_dashboard_renders_nursing_readiness_and_cpd(self):
        nurse = NursingProfessional.objects.create(
            first_name='Readiness',
            last_name='Nurse',
            registration_no='NC-READY-1',
            qualification_level='Bachelor of Nursing',
            license_expiry_date=timezone.localdate() + timedelta(days=60),
        )
        nurse_ct = ContentType.objects.get_for_model(NursingProfessional)
        user = get_user_model().objects.create_user(
            username='readiness.nurse',
            password='StrongPass123!',
            role='nurse',
            professional_content_type=nurse_ct,
            professional_object_id=nurse.pk,
            professional_record_status='linked',
        )
        CPDRecord.objects.create(
            content_type=nurse_ct,
            object_id=nurse.pk,
            training_type='Infection prevention update',
            provider='Nursing Council CPD',
            start_date=timezone.localdate(),
            hours_credits=5,
        )
        application = Application.objects.create(
            content_type=nurse_ct,
            object_id=nurse.pk,
            form_code='NC3',
            status='pending',
        )
        EmploymentRecord.objects.create(
            content_type=nurse_ct,
            object_id=nurse.pk,
            employer_name='Kokopo General Hospital',
            position_title='Ward nurse',
            province='East New Britain',
            employment_status='employed',
            is_current=True,
        )
        Receipt.objects.create(
            user=user,
            application=application,
            receipt_number='NC-READY-RCT',
            amount='100.00',
            status='completed',
        )
        self.client.force_login(user)

        response = self.client.get(reverse('nurse_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PNG Nursing Council Nurse Self-Service')
        self.assertContains(response, 'Overview')
        self.assertContains(response, 'Renewal')
        self.assertContains(response, 'Employment / Practice')
        self.assertContains(response, 'Public Register Preview')
        self.assertContains(response, 'Request Correction')
        self.assertContains(response, 'Nursing Council Renewal Readiness')
        self.assertContains(response, 'CPD / Learning Records')
        self.assertContains(response, 'Infection prevention update')
        self.assertContains(response, 'Kokopo General Hospital')
        self.assertContains(response, 'Save Payment Record')
        self.assertContains(response, reverse('public_nursing_register_search_root'))

    def test_nurse_dashboard_unlinked_account_renders_setup_workflow(self):
        user = get_user_model().objects.create_user(
            username='unlinked.nurse',
            password='StrongPass123!',
            role='nurse',
            first_name='Unlinked',
            last_name='Nurse',
            email='unlinked.nurse@example.test',
            registration_number='TEST-NURSE-001',
            license_number='TEST-NURSE-001',
            professional_record_status='unmatched',
        )
        self.client.force_login(user)

        response = self.client.get(reverse('nurse_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PNG Nursing Council Account Setup')
        self.assertContains(response, 'Complete record linkage before licence, ATP, CPD, and receipt services open')
        self.assertContains(response, 'Linkage Checklist')
        self.assertContains(response, 'Continue Nursing Forms')
        self.assertContains(response, 'Search Public Register')
        self.assertContains(response, 'Request Record Link')
        self.assertContains(response, 'No possible professional record match was found for this account.')
        self.assertContains(response, reverse('nursing_forms_portal'))
        self.assertContains(response, reverse('public_nursing_register_search_root'))
        self.assertContains(response, reverse('enquiry_create'))
        self.assertNotContains(response, 'PNG Nursing Council Nurse Self-Service')

    def test_nurse_dashboard_unlinked_account_shows_read_only_analytics_match(self):
        snapshot = NursingAnalyticsSnapshot.objects.create(
            source_file_name='nursing-dashboard-match.xlsx',
            source_file_hash='nursing-dashboard-match-hash',
            is_active=True,
        )
        NursingPractitionerIndex.objects.create(
            snapshot=snapshot,
            practitioner_group_id='match-1',
            person_group_key='MATCHED NURSE',
            representative_name='Matched Nurse',
            identity_confidence='High',
            record_count=3,
            stages_present='Provisional Licence; Full Licence; Authority to Practice',
            has_provisional=True,
            has_full_licence=True,
            has_atp=True,
            latest_year=2024,
            latest_cadre='Registered Nurse',
            registration_nos='REG-MATCH-1; GD 123',
            practitioner_nos='O123',
        )
        user = get_user_model().objects.create_user(
            username='matched.nurse',
            password='StrongPass123!',
            role='nurse',
            first_name='Matched',
            last_name='Nurse',
            registration_number='REG-MATCH-1',
            professional_record_status='unmatched',
        )
        self.client.force_login(user)

        response = self.client.get(reverse('nurse_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Possible Existing Matches')
        self.assertContains(response, 'Matched Nurse')
        self.assertContains(response, 'Read-only analytics index')
        self.assertContains(response, 'Registrar verification required')
        self.assertContains(response, 'REG-MATCH-1; GD 123')


class MedicalBoardScreenScopeTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.medical_user = user_model.objects.create_user(
            username='medical_scope_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Medical Board',
        )
        Cadre.objects.create(name='Nursing', category='nursing')
        Cadre.objects.create(name='Midwifery', category='midwifery')
        Cadre.objects.create(name='Community Health Worker', category='chw')
        TrainingInstitution.objects.create(name='Lae School of Nursing', type='Nursing School')
        TrainingInstitution.objects.create(name='CHW Training College', type='CHW Training School')
        CommunityHealthWorker.objects.create(
            first_name='Medical',
            last_name='CHW',
            registration_no='CHW-SCOPE-1',
            email='chw.scope@example.test',
        )
        NursingProfessional.objects.create(
            first_name='Nursing',
            last_name='Professional',
            registration_no='NC-SCOPE-1',
            email='nurse.scope@example.test',
        )
        Midwife.objects.create(
            first_name='Nursing',
            last_name='Midwife',
            registration_no='MW-SCOPE-1',
            email='midwife.scope@example.test',
        )
        NurseAide.objects.create(
            first_name='Nursing',
            last_name='Aide',
            registration_no='NA-SCOPE-1',
            email='aide.scope@example.test',
        )
        HealthStudent.objects.create(
            first_name='Nursing',
            last_name='Graduand',
            registration_no='HS-SCOPE-1',
            email='graduand.scope@example.test',
            program='Nursing',
        )

    def test_medical_board_overall_dashboard_hides_nursing_council_sections(self):
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('advanced_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['dashboard_scope'], 'medical')
        self.assertContains(response, 'Medical Board Dashboard')
        self.assertContains(response, 'Community Health Workers')
        self.assertContains(response, 'Medical Training Institutions')
        self.assertNotContains(response, 'Nursing Professionals')
        self.assertNotContains(response, 'Midwives')
        self.assertNotContains(response, 'Nurse Aides')
        self.assertNotContains(response, 'Graduands / Provisional Applicants')
        self.assertNotContains(response, 'PNG Nursing Schools')
        self.assertNotContains(response, 'Training Institution Breakdown')
        self.assertNotContains(response, 'Lae School of Nursing')

    def test_medical_board_workforce_flow_uses_live_medical_total_without_global_snapshot(self):
        WorkforceSnapshot.objects.create(year=2026, total_active_workers=1661)
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('workforce_flow'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['dashboard_scope'], 'medical')
        self.assertEqual(response.context['tracked_workforce_count'], 1)
        self.assertEqual(response.context['total_workers_by_year'], [1])
        self.assertContains(response, 'Imported Medical Board Rows')
        self.assertContains(response, 'Operational Task Inbox')
        self.assertContains(response, 'CHW registration and licence')
        self.assertNotContains(response, 'NC1 / Provisional licence')
        self.assertNotContains(response, 'Reference Tables')
        self.assertNotContains(response, 'Training Institutions')

    def test_medical_staff_portal_renders_medical_board_workforce_view(self):
        MedicalDoctor.objects.create(
            first_name='Medical',
            last_name='Doctor',
            registration_no='MD-STAFF-1',
            specialty='General Practice',
        )
        Application.objects.create(form_code='MD1', status='pending')
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('medical_staff_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Medical Board Workforce and CHW Register')
        self.assertContains(response, 'Doctors / Specialists')
        self.assertContains(response, 'Community Health Workers')
        self.assertContains(response, reverse('public_medical_board_register_search_root'))
        self.assertContains(response, reverse('workforce_map') + '?office=medical')

    def test_medical_board_records_hub_renders_register_cockpit(self):
        MedicalDoctor.objects.create(
            first_name='Records',
            last_name='Doctor',
            registration_no='MD-HUB-1',
            specialty='General Practice',
        )
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('records_home'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Medical Board Register Cockpit')
        self.assertContains(response, 'Public Medical Register')
        self.assertContains(response, 'Practising Certificates')
        self.assertContains(response, reverse('public_medical_board_register_search_root'))
        self.assertNotContains(response, 'Nursing Professionals')

    def test_doctor_dashboard_renders_medical_board_readiness_and_cpd(self):
        doctor = MedicalDoctor.objects.create(
            first_name='Readiness',
            last_name='Doctor',
            registration_no='MD-READY-1',
            specialty='General Practice',
            license_expiry_date=timezone.localdate() + timedelta(days=60),
        )
        doctor_ct = ContentType.objects.get_for_model(MedicalDoctor)
        user = get_user_model().objects.create_user(
            username='readiness.doctor',
            password='StrongPass123!',
            role='doctor',
            professional_content_type=doctor_ct,
            professional_object_id=doctor.pk,
            professional_record_status='linked',
        )
        CPDRecord.objects.create(
            content_type=doctor_ct,
            object_id=doctor.pk,
            training_type='Clinical governance',
            provider='Medical Board CME',
            start_date=timezone.localdate(),
            hours_credits=6,
        )
        Application.objects.create(
            content_type=doctor_ct,
            object_id=doctor.pk,
            form_code='MD2',
            status='pending',
        )
        self.client.force_login(user)

        response = self.client.get(reverse('doctor_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Medical Board Doctor Self-Service')
        self.assertContains(response, 'Medical Board Renewal Readiness')
        self.assertContains(response, 'Public Register Preview')
        self.assertContains(response, 'CPD / CME Records')
        self.assertContains(response, 'Clinical governance')
        self.assertContains(response, reverse('public_medical_board_register_search_root'))

    def test_other_practitioner_dashboards_render_self_service_shells(self):
        user_model = get_user_model()

        chw = CommunityHealthWorker.objects.create(
            first_name='Linked',
            last_name='CHW',
            registration_no='CHW-DASH-1',
            community_id='COMM-DASH-1',
            training_level='Certificate CHW',
        )
        chw_ct = ContentType.objects.get_for_model(CommunityHealthWorker)
        chw_user = user_model.objects.create_user(
            username='linked.chw',
            password='StrongPass123!',
            role='chw',
            professional_content_type=chw_ct,
            professional_object_id=chw.pk,
            professional_record_status='linked',
        )

        nurse_aide = NurseAide.objects.create(
            first_name='Linked',
            last_name='Aide',
            registration_no='NA-DASH-1',
            training_level='Nurse aide certificate',
        )
        nurse_aide_ct = ContentType.objects.get_for_model(NurseAide)
        nurse_aide_user = user_model.objects.create_user(
            username='linked.nurse.aide',
            password='StrongPass123!',
            role='nurse_aide',
            professional_content_type=nurse_aide_ct,
            professional_object_id=nurse_aide.pk,
            professional_record_status='linked',
        )

        student = HealthStudent.objects.create(
            first_name='Linked',
            last_name='Graduand',
            registration_no='GD-DASH-1',
            program='Diploma in General Nursing',
            expected_graduation_date=timezone.localdate() + timedelta(days=90),
        )
        student_ct = ContentType.objects.get_for_model(HealthStudent)
        student_user = user_model.objects.create_user(
            username='linked.graduand',
            password='StrongPass123!',
            role='graduand',
            professional_content_type=student_ct,
            professional_object_id=student.pk,
            professional_record_status='linked',
        )

        cases = [
            (chw_user, 'chw_dashboard', 'Medical Board CHW Self-Service', 'Medical Board CHW Register Readiness'),
            (nurse_aide_user, 'nurse_aide_dashboard', 'PNG Nursing Council Nurse Aide Self-Service', 'Nursing Council Nurse Aide Readiness'),
            (student_user, 'student_dashboard', 'PNG Nursing Council Graduand Self-Service', 'Nursing Council Provisional Pathway Readiness'),
        ]
        for user, url_name, title, readiness_label in cases:
            with self.subTest(url=url_name):
                self.client.force_login(user)
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, title)
                self.assertContains(response, readiness_label)
                self.assertContains(response, 'Save Payment Record')
                self.client.logout()

    def test_unlinked_practitioner_dashboards_render_setup_workflows(self):
        user_model = get_user_model()
        MedicalDoctor.objects.create(
            first_name='Unrelated',
            last_name='Doctor',
            registration_no='MD-UNRELATED-BLANK-EMAIL',
            specialty='General Practice',
        )
        CommunityHealthWorker.objects.create(
            first_name='Unrelated',
            last_name='CHW',
            registration_no='CHW-UNRELATED-BLANK-EMAIL',
            community_id='COMM-UNRELATED',
        )
        cases = [
            ('setup.doctor', 'doctor', 'doctor_dashboard', 'Medical Board Account Setup', 'Continue Doctor Form'),
            ('setup.chw', 'chw', 'chw_dashboard', 'CHW Account Setup', 'Continue CHW Form'),
            ('setup.nurse.aide', 'nurse_aide', 'nurse_aide_dashboard', 'Nurse Aide Account Setup', 'Continue Nurse Aide Form'),
            ('setup.graduand', 'graduand', 'student_dashboard', 'Graduand Account Setup', 'Continue Graduand Forms'),
        ]
        for username, role, url_name, title, action_label in cases:
            with self.subTest(role=role):
                user = user_model.objects.create_user(
                    username=username,
                    password='StrongPass123!',
                    role=role,
                    first_name='Setup',
                    last_name=role.replace('_', ' ').title(),
                    registration_number=f'{role.upper()}-SETUP-1',
                    professional_record_status='unmatched',
                )
                self.client.force_login(user)
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, title)
                self.assertContains(response, action_label)
                self.assertContains(response, 'Linkage Checklist')
                self.assertContains(response, 'Possible Existing Matches')
                self.assertNotContains(response, 'Unrelated Doctor')
                self.assertNotContains(response, 'Unrelated CHW')
                self.client.logout()

    def test_medical_board_portal_separates_specialists_chw_licences_and_missing_reviews(self):
        current_year = timezone.localdate().year
        doctor = MedicalDoctor.objects.create(
            first_name='Medical',
            last_name='Generalist',
            registration_no='MD-SCOPE-1',
            specialty='Overseas Medical Board Practitioner',
        )
        specialist = MedicalDoctor.objects.create(
            first_name='Medical',
            last_name='Specialist',
            registration_no='MD-SCOPE-2',
            specialty='',
        )
        chw_batch = DataImportBatch.objects.create(
            source_file_name='chw.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
        )
        chw_batch.started_at = timezone.now() - timedelta(days=2)
        chw_batch.save(update_fields=['started_at'])
        PracticingLicenseRecord.objects.create(
            batch=chw_batch,
            record_type='practicing_license',
            target_model='communityhealthworker',
            source_sheet_name='CHW ATP',
            source_row=1,
            record_year=current_year,
            full_name='Medical CHW',
            registration_no='CHW-SCOPE-1',
        )
        latest_medical_batch = DataImportBatch.objects.create(
            source_file_name='legacy.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
        )
        latest_medical_batch.started_at = timezone.now()
        latest_medical_batch.save(update_fields=['started_at'])
        PracticingLicenseRecord.objects.create(
            batch=latest_medical_batch,
            record_type='workforce_listing',
            target_model='medicaldoctor',
            source_sheet_name='ALL OVERSEAS MEMBERS MB',
            source_row=1,
            record_year=current_year,
            full_name='Medical Generalist',
            registration_no=doctor.registration_no,
            qualification_name='MBBS',
        )
        specialist_record = PracticingLicenseRecord.objects.create(
            batch=latest_medical_batch,
            record_type='workforce_listing',
            target_model='medicaldoctor',
            source_sheet_name='ALL OVERSEAS MEMBERS MB',
            source_row=2,
            record_year=current_year,
            full_name='Medical Specialist',
            registration_no=specialist.registration_no,
            qualification_name='SPECIALIST (MP/DP/AHW)',
        )
        MissingDataReview.objects.create(
            content_type=ContentType.objects.get_for_model(CommunityHealthWorker),
            object_id=self._get_chw().pk,
            full_name='Medical CHW',
            professional_type='Community Health Worker',
            missing_fields=['Email address'] * 5,
            missing_count=5,
            severity='high',
        )
        MissingDataReview.objects.create(
            content_type=ContentType.objects.get_for_model(PracticingLicenseRecord),
            object_id=specialist_record.pk,
            full_name='Medical Specialist',
            professional_type='Practicing License Record',
            missing_fields=['Payment date'],
            missing_count=1,
            severity='low',
        )
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('medical_board_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['medical_doctor_count'], 2)
        self.assertEqual(response.context['medical_specialist_count'], 1)
        self.assertEqual(response.context['medical_chw_practicing_total'], 1)
        self.assertEqual(response.context['medical_chw_training_reference_count'], 1)
        self.assertEqual(response.context['missing_data_review_count'], 2)
        self.assertEqual(response.context['high_priority_missing_data_count'], 1)
        self.assertContains(response, 'Medical Board CHW Training References')
        self.assertContains(response, 'CHW Training College')

    def test_medical_board_portal_renders_regulator_operations_console(self):
        doctor = MedicalDoctor.objects.create(
            first_name='Regulator',
            last_name='Doctor',
            registration_no='MD-OPS-1',
            email='doctor.ops@example.test',
            specialty='General Practice',
            license_expiry_date=timezone.localdate() + timedelta(days=30),
        )
        doctor_ct = ContentType.objects.get_for_model(MedicalDoctor)
        application = Application.objects.create(
            form_code='MD2',
            status='pending',
            content_type=doctor_ct,
            object_id=doctor.pk,
            payload={
                'qualification_name': 'MBBS',
                'good_standing': True,
                'facility_name': 'Port Moresby General Hospital',
                'cpd_points': 20,
                'receipt_number': 'MB-OPS-RCT-1',
            },
        )
        Receipt.objects.create(
            user=self.medical_user,
            application=application,
            receipt_number='MB-OPS-RCT-1',
            amount='150.00',
            status='completed',
        )
        complaint = ComplaintCase.objects.create(
            office_scope='medical',
            title='Medical Board conduct concern',
            description='Conduct matter requiring triage.',
            risk_level='high',
            priority='critical',
            status='investigating',
            subject_content_type=doctor_ct,
            subject_object_id=doctor.pk,
            subject_name=str(doctor),
            subject_identifier=doctor.registration_no,
        )
        DisciplinaryCase.objects.create(
            office_scope='medical',
            source_complaint=complaint,
            subject_content_type=doctor_ct,
            subject_object_id=doctor.pk,
            subject_name=str(doctor),
            subject_identifier=doctor.registration_no,
            allegation_summary='Allegation under review.',
            stage='hearing',
            severity='high',
        )
        RegulatoryDecisionRecord.objects.create(
            office_scope='medical',
            decision_type='licence',
            status='final',
            title='Medical Board licence condition',
            subject_content_type=doctor_ct,
            subject_object_id=doctor.pk,
            subject_name=str(doctor),
            subject_identifier=doctor.registration_no,
            decision_text='Licence continued with conditions.',
            rationale='Board decision rationale.',
            conditions='Supervised practice required.',
        )
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('medical_board_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['medical_open_complaint_count'], 1)
        self.assertEqual(response.context['medical_high_risk_complaint_count'], 1)
        self.assertEqual(response.context['medical_open_discipline_count'], 1)
        self.assertEqual(response.context['medical_high_severity_discipline_count'], 1)
        self.assertEqual(response.context['medical_active_condition_count'], 1)
        self.assertEqual(response.context['medical_expiring_license_count'], 1)
        self.assertGreaterEqual(response.context['medical_application_readiness_rows'][0]['percent'], 80)
        self.assertContains(response, 'PNG Medical Board Regulatory Operations')
        self.assertContains(response, 'Dashboard Shortcuts')
        self.assertContains(response, reverse('facilities_directory') + '?scope=medical')
        self.assertContains(response, 'Register and Practising Certificate Status')
        self.assertContains(response, 'Application Readiness and Decision Queue')
        self.assertContains(response, 'Public Trust and e-Services')
        self.assertContains(response, 'Professional Assurance Standards')
        self.assertContains(response, 'Fitness to Practise, Complaints, and Decisions')
        self.assertContains(response, 'Medical Board conduct concern')
        self.assertContains(response, 'Medical Board licence condition')

    def test_medical_board_portal_shows_facility_and_institution_sector_breakdown(self):
        today = timezone.localdate()
        location = Location.objects.create(province='National Capital District', district='Port Moresby')
        private_facility = Facility.objects.create(
            name='Paradise Private Hospital',
            type='Private Hospital',
            ownership='private',
            location=location,
        )
        catholic_facility = Facility.objects.create(
            name='St Mary Catholic Health Centre',
            type='Health Centre',
            ownership='faith_based',
            location=location,
        )
        church_facility = Facility.objects.create(
            name='Lutheran Mission Hospital',
            type='Hospital',
            ownership='faith_based',
            location=location,
        )
        medical_cadre, _created = Cadre.objects.get_or_create(name='Medical Practitioner', defaults={'category': 'medical'})
        chw_cadre = Cadre.objects.get(name='Community Health Worker')
        doctor = MedicalDoctor.objects.create(
            first_name='Facility',
            last_name='Doctor',
            registration_no='MD-SECTOR-1',
            cadre=medical_cadre,
            specialty='General Practice',
            date_of_birth=date(1985, 4, 10),
            license_expiry_date=today + timedelta(days=180),
        )
        chw = CommunityHealthWorker.objects.create(
            first_name='Facility',
            last_name='CHW',
            registration_no='CHW-SECTOR-1',
            cadre=chw_cadre,
            date_of_birth=date(1992, 7, 5),
        )
        doctor_ct = ContentType.objects.get_for_model(MedicalDoctor)
        chw_ct = ContentType.objects.get_for_model(CommunityHealthWorker)
        facility_ct = ContentType.objects.get_for_model(Facility)
        PostingHistory.objects.create(
            content_type=doctor_ct,
            object_id=doctor.pk,
            facility=private_facility,
            position_title='Medical Officer',
            start_date=today,
            is_current=True,
        )
        PostingHistory.objects.create(
            content_type=chw_ct,
            object_id=chw.pk,
            facility=catholic_facility,
            position_title='Community Health Worker',
            start_date=today,
            is_current=True,
        )
        Application.objects.create(
            content_type=facility_ct,
            object_id=private_facility.pk,
            form_code='MBPF',
            pathway='medical_facility',
            form_title='Private Facility Checklist',
            status='approved',
            payload={'facility_name': private_facility.name},
        )
        TrainingInstitution.objects.create(
            name='Catholic CHW Training School',
            type='CHW Training School',
            ownership='faith_based',
            location_name='National Capital District',
            registration_status='Accredited',
            regulatory_body_name='Medical Board',
            is_active=True,
        )
        batch = DataImportBatch.objects.create(
            source_file_name='medical-facility-workers.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
        )
        PracticingLicenseRecord.objects.create(
            batch=batch,
            record_type='practicing_license',
            target_model='communityhealthworker',
            source_sheet_name='CHW ATP',
            source_row=12,
            record_year=today.year,
            full_name='Hope NGO Clinic Worker',
            registration_no='CHW-NGO-1',
            category='Community Health Worker',
            institution_name='Catholic CHW Training School',
            workplace_address='Hope NGO Clinic',
            province='NCD',
            date_of_birth=date(1990, 8, 12),
        )
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('medical_board_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Medical Board Facility and Institution Breakdown')
        self.assertContains(response, 'Provincial Health Authority (PHA)')
        self.assertContains(response, 'Church Health Services')
        self.assertContains(response, 'Catholic Health Services')
        self.assertContains(response, 'Private facilities')
        self.assertContains(response, 'NGO facilities')
        self.assertContains(response, 'Paradise Private Hospital')
        self.assertContains(response, 'St Mary Catholic Health Centre')
        self.assertContains(response, 'Lutheran Mission Hospital')
        self.assertContains(response, 'Hope NGO Clinic')
        self.assertContains(response, 'Catholic CHW Training School')
        self.assertContains(response, 'Private Facility Checklist')
        self.assertContains(response, reverse('facility_worker_detail', args=[private_facility.pk]) + '?scope=medical')
        self.assertTrue(
            any(
                row['facility_name'] == 'Hope NGO Clinic'
                and row['detail_url'] == reverse('imported_facility_worker_detail') + '?name=Hope+NGO+Clinic&scope=medical'
                for row in response.context['medical_imported_facility_reference_rows']
            )
        )
        sector_labels = {row['label'] for row in response.context['medical_facility_sector_rows']}
        self.assertIn('Church Health Services', sector_labels)
        self.assertIn('Catholic Health Services', sector_labels)
        self.assertIn('Private facilities', sector_labels)
        self.assertIn('NGO facilities', sector_labels)

        detail_response = self.client.get(reverse('facility_worker_detail', args=[private_facility.pk]), {'scope': 'medical'})
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, 'Facility sector')
        self.assertContains(detail_response, 'Private facilities')
        self.assertContains(detail_response, 'Facility Doctor')

    def test_medical_facility_detail_excludes_nursing_and_payment_rows(self):
        medical_batch = DataImportBatch.objects.create(
            source_file_name='medical-facility.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
        )
        nursing_batch = DataImportBatch.objects.create(
            source_file_name='nursing-facility.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        common = {
            'full_name': 'Scoped Facility Worker',
            'workplace_address': 'Port Moresby General Hospital',
            'source_sheet_name': 'Facility records',
            'record_year': timezone.localdate().year,
        }
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            source_row=1,
            record_type='practicing_license',
            target_model='communityhealthworker',
            registration_no='CHW-FACILITY-SCOPE-1',
            category='Community Health Worker',
            **common,
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            source_row=2,
            record_type='payment',
            target_model='communityhealthworker',
            registration_no='CHW-FACILITY-PAYMENT-1',
            full_name='Medical Payment Row',
            category='Community Health Worker Payment',
            workplace_address='Port Moresby General Hospital',
            source_sheet_name='Payment',
            record_year=timezone.localdate().year,
        )
        PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            source_row=3,
            record_type='practicing_license',
            target_model='other',
            registration_no='NURSE-LEAK-1',
            full_name='Nursing Leakage Row',
            category='ATP Payment',
            workplace_address='Port Moresby General Hospital',
            source_sheet_name='Nursing ATP',
            record_year=timezone.localdate().year,
        )
        self.client.force_login(self.medical_user)

        response = self.client.get(
            reverse('imported_facility_worker_detail'),
            {'name': 'Port Moresby General Hospital', 'scope': 'medical'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Scoped Facility Worker')
        self.assertContains(response, 'Community Health Worker')
        self.assertNotContains(response, 'Nursing Leakage Row')
        self.assertNotContains(response, 'ATP Payment')
        self.assertNotContains(response, 'Medical Payment Row')
        self.assertNotContains(response, 'Community Health Worker Payment')

        directory_response = self.client.get(reverse('facilities_directory'), {'scope': 'medical'})
        self.assertEqual(directory_response.status_code, 200)
        self.assertContains(directory_response, 'Medical Board Facilities and Institutions')
        self.assertContains(directory_response, 'National Facility Group Breakdown')
        self.assertContains(directory_response, 'Community Health Worker')

    def test_medical_board_portal_handles_recent_application_without_linked_record(self):
        Application.objects.create(
            form_code='MD1',
            status='pending',
        )
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('medical_board_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Medical Board Application')

    def _get_chw(self):
        return CommunityHealthWorker.objects.get(registration_no='CHW-SCOPE-1')

    def test_medical_board_fee_structure_hides_nursing_fee_sections(self):
        self.client.force_login(self.medical_user)

        response = self.client.get(reverse('fee_structure'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['fee_scope'], 'medical')
        self.assertContains(response, 'PAPUA NEW GUINEA MEDICAL BOARD')
        self.assertContains(response, 'Medical Board Fees')
        self.assertContains(response, 'Community Health Worker Fees')
        self.assertNotContains(response, 'PAPUA NEW GUINEA NURSING COUNCIL')
        self.assertNotContains(response, 'Renewals for Nurse Aides')
        self.assertNotContains(response, 'Category: C Nursing Council Other Fees')
        self.assertNotContains(response, 'Graduand Registration Fees')


class StaffAIProviderModeTests(TestCase):
    def setUp(self):
        cache.clear()

    def test_adk_record_lookup_is_scoped_to_nursing_registrar(self):
        from apps.dashboard.staff_ai_record_tools import search_staff_registry_records_for_user

        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='nursing_adk_lookup_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        nursing_batch = DataImportBatch.objects.create(
            source_file_name='nursing-atp.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        medical_batch = DataImportBatch.objects.create(
            source_file_name='medical-board.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
        )
        PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP',
            source_row=12,
            record_year=2026,
            full_name='Mary Kila',
            registration_no='NC-001',
            practitioner_number='PN-001',
            province='Morobe',
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            record_type='practicing_license',
            target_model='medicaldoctor',
            source_sheet_name='Doctors',
            source_row=20,
            record_year=2026,
            full_name='Mary Kila Medical',
            registration_no='MB-001',
            practitioner_number='MD-001',
            province='Morobe',
        )

        result = search_staff_registry_records_for_user(user, query='Mary Kila', limit=10)

        self.assertEqual(result['status'], 'ok')
        self.assertEqual(result['scope'], 'nursing')
        self.assertEqual(result['total_matches'], 1)
        self.assertEqual(result['records'][0]['full_name'], 'Mary Kila')
        self.assertEqual(result['records'][0]['registration_no'], 'NC-001')
        self.assertNotIn('Mary Kila Medical', [record['full_name'] for record in result['records']])
        self.assertIn('date_of_birth', result['redactions'])
        self.assertNotIn('raw_payload', result['records'][0])

    def test_local_model_json_parser_recovers_a_structured_answer_with_extra_text(self):
        from apps.dashboard.ai_provider import _json_from_model_text

        payload = _json_from_model_text(
            'Here is the requested answer: {"title":"Scoped","answer":"Use sources.","bullets":[],"links":[],"suggestions":[]} Thank you.',
            'Ollama',
        )

        self.assertEqual(payload['title'], 'Scoped')

    def test_adk_record_lookup_denies_non_staff_user(self):
        from apps.dashboard.staff_ai_record_tools import search_staff_registry_records_for_user

        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='nurse_public_lookup_user',
            password='StrongPass123!',
            role='nurse',
            department='Nursing Council',
        )

        result = search_staff_registry_records_for_user(user, query='Mary')

        self.assertEqual(result['status'], 'denied')
        self.assertEqual(result['records'], [])

    @override_settings(
        AI_ASSISTANT_PROVIDER='local',
        AI_ASSISTANT_LOCALAI_ENABLED=False,
        AI_ASSISTANT_OLLAMA_ENABLED=False,
        AI_ASSISTANT_GOOGLE_ADK_ENABLED=False,
    )
    def test_staff_ai_defaults_to_local_offline_mode(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='ai_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        status = ai_provider_status()
        response = build_staff_ai_chat_response(user, 'Prepare an operational screening summary for today.')

        self.assertIn(status['mode'], {'local', 'local_fallback'})
        self.assertIn('ai_provider', response)
        self.assertIn(response['ai_provider']['mode'], {'local', 'local_fallback'})

    @override_settings(AI_ASSISTANT_PROVIDER='local_llm', AI_ASSISTANT_LOCAL_LLM_ENABLED=False)
    def test_private_llm_requires_explicit_enablement(self):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'local_fallback')
        self.assertIn('disabled', status['detail'])

    @override_settings(
        AI_ASSISTANT_PROVIDER='ollama',
        AI_ASSISTANT_OLLAMA_ENABLED=True,
        AI_OLLAMA_MODEL='llama3.2:3b',
        AI_OLLAMA_BASE_URL='http://127.0.0.1:11434',
    )
    def test_ollama_can_be_configured_as_free_local_gpt(self):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'ollama')
        self.assertTrue(status['ollama_ready'])
        self.assertEqual(status['ollama_model'], 'llama3.2:3b')

    @override_settings(
        AI_ASSISTANT_PROVIDER='localai',
        AI_ASSISTANT_LOCALAI_ENABLED=True,
        AI_LOCALAI_MODEL='gpt-4-qa',
        AI_LOCALAI_BASE_URL='http://127.0.0.1:8080',
    )
    def test_localai_can_be_configured_as_live_local_provider(self):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'localai')
        self.assertTrue(status['localai_ready'])
        self.assertEqual(status['localai_model'], 'gpt-4-qa')

    @override_settings(AI_ASSISTANT_PROVIDER='redis_worker', AI_REDIS_WORKER_ENABLED=False)
    def test_redis_worker_requires_explicit_enablement(self):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'local_fallback')
        self.assertFalse(status['redis_worker_ready'])
        self.assertIn('AI_REDIS_WORKER_ENABLED', status['detail'])

    @override_settings(
        AI_ASSISTANT_PROVIDER='redis_worker',
        AI_REDIS_WORKER_ENABLED=True,
        AI_REDIS_URL='redis://localhost:6379/0',
        AI_REDIS_WORKER_MODEL_PROVIDER='local',
    )
    def test_redis_worker_can_be_configured_as_queued_provider(self):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'redis_worker')
        self.assertTrue(status['redis_worker_ready'])
        self.assertEqual(status['redis_worker_model_provider'], 'local')

    @override_settings(
        AI_ASSISTANT_PROVIDER='redis_worker',
        AI_REDIS_WORKER_ENABLED=True,
        AI_REDIS_URL='redis://localhost:6379/0',
        AI_REDIS_WORKER_MODEL_PROVIDER='local',
    )
    @mock.patch('apps.dashboard.ai_worker.submit_redis_ai_request')
    def test_staff_ai_can_route_live_response_through_redis_worker(self, submit_redis_ai_request):
        submit_redis_ai_request.return_value = {
            'title': 'Queued Worker Answer',
            'answer': 'The Django worker returned this scoped response.',
            'bullets': ['Django auth and scope stayed in control.'],
            'links': [],
            'suggestions': [],
        }
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='worker_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'Prepare an operational screening summary for today.')

        self.assertEqual(response['title'], 'Queued Worker Answer')
        self.assertEqual(response['ai_provider']['mode'], 'redis_worker')
        submit_redis_ai_request.assert_called_once()

    @override_settings(
        AI_ASSISTANT_PROVIDER='redis_worker',
        AI_REDIS_WORKER_ENABLED=True,
        AI_REDIS_URL='redis://localhost:6379/0',
        AI_REDIS_WORKER_MODEL_PROVIDER='local',
    )
    @mock.patch('apps.dashboard.staff_ai.retrieve_assistant_sources')
    @mock.patch('apps.dashboard.ai_worker.submit_redis_ai_request')
    def test_staff_ai_shortcut_prompt_uses_fast_local_answer(self, submit_redis_ai_request, retrieve_assistant_sources):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='fast_shortcut_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'Which report should I generate for management?')

        self.assertEqual(response['title'], 'Reports And Briefs')
        self.assertEqual(response['ai_provider']['mode'], 'redis_worker')
        self.assertFalse(staff_ai_question_needs_knowledge_search('Which report should I generate for management?'))
        retrieve_assistant_sources.assert_not_called()
        submit_redis_ai_request.assert_not_called()

    @override_settings(
        AI_ASSISTANT_PROVIDER='redis_worker',
        AI_REDIS_WORKER_ENABLED=True,
        AI_REDIS_URL='redis://localhost:6379/0',
        AI_REDIS_WORKER_MODEL_PROVIDER='local',
    )
    @mock.patch('apps.dashboard.ai_worker.submit_redis_ai_request', side_effect=RuntimeError('connection refused'))
    def test_redis_worker_failure_uses_local_staff_fallback(self, _submit_redis_ai_request):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='worker_fallback_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'Prepare an operational screening summary for today.')

        self.assertNotEqual(response['title'], 'Queued Worker Answer')
        self.assertEqual(response['ai_provider']['mode'], 'local_fallback')
        self.assertIn('Redis AI worker request failed', response['ai_provider']['detail'])

    def test_ai_provider_status_reports_rag_configuration(self):
        with TemporaryDirectory() as temp_dir:
            with override_settings(
                AI_ASSISTANT_RAG_ENABLED=True,
                AI_ASSISTANT_RAG_INDEX_PATH=str(Path(temp_dir) / 'assistant-index.json'),
            ):
                status = ai_provider_status()

        self.assertTrue(status['rag_enabled'])
        self.assertIn('rag', status)
        self.assertIn('assistant-index.json', status['rag']['index_path'])

    @mock.patch('apps.dashboard.assistant_rag._sentence_transformers_available', return_value=True)
    @mock.patch('apps.dashboard.assistant_rag._embed_texts')
    def test_vector_retrieval_uses_local_knowledge_index(self, embed_texts, _package_available):
        def fake_embed(texts):
            vectors = []
            for text in texts:
                lowered = text.lower()
                vectors.append([1.0, 0.0] if 'provisional' in lowered or 'nc1' in lowered else [0.0, 1.0])
            return vectors

        embed_texts.side_effect = fake_embed
        category = FAQCategory.objects.create(
            name='Nursing Staff Knowledge',
            slug='nursing-staff-knowledge',
            audience='staff',
            office_scope='nursing',
            is_active=True,
        )
        faq = FAQEntry.objects.create(
            category=category,
            question='How does NC1 provisional licence work?',
            answer='NC1 supports provisional licence review for Nursing Council graduate applicants.',
            keywords='NC1 provisional licence nursing',
            is_published=True,
        )

        with TemporaryDirectory() as temp_dir:
            with override_settings(
                AI_ASSISTANT_RAG_ENABLED=True,
                AI_ASSISTANT_RAG_AUTO_BUILD=False,
                AI_ASSISTANT_RAG_INDEX_PATH=str(Path(temp_dir) / 'assistant-index.json'),
                AI_ASSISTANT_RAG_MIN_SCORE=0.01,
            ):
                result = build_vector_index()
                sources = retrieve_vector_sources(
                    question='Explain provisional licence NC1',
                    scope='nursing',
                    public=False,
                )

        self.assertGreaterEqual(result['document_count'], 1)
        self.assertTrue(any('NC1 provisional licence' in source['label'] for source in sources))

    @mock.patch('apps.dashboard.assistant_rag._sentence_transformers_available', return_value=True)
    @mock.patch('apps.dashboard.assistant_rag._embed_texts', side_effect=RuntimeError('embedding model unavailable'))
    def test_vector_retrieval_fails_open_when_embedding_runtime_is_unavailable(self, _embed_texts, _package_available):
        with TemporaryDirectory() as temp_dir:
            index_path = Path(temp_dir) / 'assistant-index.json'
            index_path.write_text(
                '{"version": 1, "embedding_model": "sentence-transformers/all-MiniLM-L6-v2", "documents": []}',
                encoding='utf-8',
            )
            with override_settings(
                AI_ASSISTANT_RAG_ENABLED=True,
                AI_ASSISTANT_RAG_AUTO_BUILD=False,
                AI_ASSISTANT_RAG_INDEX_PATH=str(index_path),
            ):
                sources = retrieve_vector_sources(
                    question='Explain the provisional licence pathway',
                    scope='nursing',
                    public=False,
                )

        self.assertEqual(sources, [])

    @override_settings(
        AI_ASSISTANT_PROVIDER='ollama',
        AI_ASSISTANT_OLLAMA_ENABLED=False,
        AI_OLLAMA_MODEL='llama3.2:3b',
    )
    def test_ollama_falls_back_when_switch_is_disabled(self):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'local_fallback')
        self.assertIn('disabled', status['detail'])

    @override_settings(
        AI_ASSISTANT_PROVIDER='google_adk',
        AI_GOOGLE_ADK_ENABLED=False,
        GOOGLE_API_KEY='test-google-key',
        AI_GOOGLE_ADK_MODEL='gemini-flash-latest',
    )
    def test_google_adk_requires_explicit_enablement(self):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'local_fallback')
        self.assertFalse(status['google_adk_ready'])
        self.assertIn('ADK switch is disabled', status['detail'])

    @override_settings(
        AI_ASSISTANT_PROVIDER='google_adk',
        AI_GOOGLE_ADK_ENABLED=True,
        GOOGLE_API_KEY='',
        AI_GOOGLE_ADK_MODEL='gemini-flash-latest',
    )
    def test_google_adk_requires_google_api_key(self):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'local_fallback')
        self.assertFalse(status['google_api_key_configured'])
        self.assertIn('GOOGLE_API_KEY', status['detail'])

    @override_settings(
        AI_ASSISTANT_PROVIDER='google_adk',
        AI_GOOGLE_ADK_ENABLED=True,
        GOOGLE_API_KEY='test-google-key',
        AI_GOOGLE_ADK_MODEL='gemini-flash-latest',
    )
    @mock.patch('apps.dashboard.ai_provider._google_adk_package_available', return_value=True)
    def test_google_adk_can_be_configured_as_live_agent(self, _package_available):
        status = ai_provider_status()

        self.assertEqual(status['mode'], 'google_adk')
        self.assertTrue(status['google_adk_ready'])
        self.assertTrue(status['live_model_ready'])
        self.assertEqual(status['google_adk_model'], 'gemini-flash-latest')

    @override_settings(
        AI_ASSISTANT_PROVIDER='google_adk',
        AI_GOOGLE_ADK_ENABLED=True,
        GOOGLE_API_KEY='test-google-key',
        AI_GOOGLE_ADK_MODEL='gemini-flash-latest',
        AI_ASSISTANT_TIMEOUT_SECONDS=2,
    )
    @mock.patch('apps.dashboard.ai_provider._load_google_adk_runtime')
    @mock.patch('apps.dashboard.ai_provider._google_adk_package_available', return_value=True)
    def test_google_adk_json_provider_runs_agent_wrapper(self, _package_available, load_google_adk_runtime):
        from apps.dashboard.ai_provider import STAFF_AI_RESPONSE_SCHEMA, call_google_adk_json

        created_agents = []

        class FakeAgent:
            def __init__(self, **kwargs):
                self.kwargs = kwargs
                created_agents.append(self)

        class FakeSessionService:
            async def create_session(self, **kwargs):
                return SimpleNamespace(**kwargs)

        class FakeRunner:
            def __init__(self, **kwargs):
                self.kwargs = kwargs

            async def run_async(self, **kwargs):
                yield SimpleNamespace(
                    is_final_response=lambda: True,
                    content=SimpleNamespace(parts=[
                        SimpleNamespace(
                            text=(
                                '{"title":"Fake ADK","answer":"Agent response",'
                                '"bullets":["Scoped context only"],"links":[],"suggestions":[]}'
                            )
                        )
                    ]),
                )

        class FakeTypes:
            class Content:
                def __init__(self, role, parts):
                    self.role = role
                    self.parts = parts

            class Part:
                def __init__(self, text):
                    self.text = text

        load_google_adk_runtime.return_value = (FakeAgent, FakeRunner, FakeSessionService, FakeTypes)

        def fake_lookup_tool(query: str = '') -> dict:
            """Return fake scoped lookup rows."""
            return {'query': query, 'records': []}

        result = call_google_adk_json(
            system_prompt='Answer with JSON only.',
            user_payload={
                'question': 'What should I review?',
                'scoped_context': {'pending_application_count': 1},
                'local_fallback_answer': {'links': []},
            },
            schema=STAFF_AI_RESPONSE_SCHEMA,
            schema_name='staff_ai_response',
            extra_tools=[fake_lookup_tool],
        )

        self.assertEqual(result['title'], 'Fake ADK')
        self.assertEqual(result['bullets'], ['Scoped context only'])
        tool_names = [tool.__name__ for tool in created_agents[0].kwargs['tools']]
        self.assertIn('get_supplied_context', tool_names)
        self.assertIn('fake_lookup_tool', tool_names)
        load_google_adk_runtime.assert_called_once()

    @override_settings(
        AI_ASSISTANT_PROVIDER='google_adk',
        AI_GOOGLE_ADK_ENABLED=True,
        GOOGLE_API_KEY='test-google-key',
        AI_GOOGLE_ADK_MODEL='gemini-flash-latest',
    )
    @mock.patch('apps.dashboard.ai_provider.call_google_adk_json')
    @mock.patch('apps.dashboard.ai_provider._google_adk_package_available', return_value=True)
    def test_staff_ai_uses_google_adk_when_ready(self, _package_available, call_google_adk_json):
        call_google_adk_json.return_value = {
            'title': 'ADK Screening Answer',
            'answer': 'Use the live screening queue and review missing-data flags first.',
            'bullets': ['Review pending applications.', 'Check missing-data reviews.'],
            'links': [{'label': 'Open Full Assistant', 'url': 'staff_ai_assistant'}],
            'suggestions': ['Which report should I generate for management?'],
        }
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='google_adk_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'Prepare an operational screening summary for today.')

        self.assertEqual(response['title'], 'ADK Screening Answer')
        self.assertEqual(response['ai_provider']['mode'], 'google_adk')
        self.assertEqual(response['links'][0]['url'], reverse('staff_ai_assistant'))
        call_google_adk_json.assert_called_once()
        tool_names = [
            tool.__name__
            for tool in call_google_adk_json.call_args.kwargs['extra_tools']
        ]
        self.assertIn('search_staff_registry_records', tool_names)

    @override_settings(
        AI_ASSISTANT_PROVIDER='localai',
        AI_ASSISTANT_LOCALAI_ENABLED=True,
        AI_LOCALAI_MODEL='gpt-4-qa',
        AI_LOCALAI_BASE_URL='http://127.0.0.1:8080',
    )
    @mock.patch('apps.dashboard.ai_provider.call_localai_json')
    def test_staff_ai_uses_localai_when_ready(self, call_localai_json):
        call_localai_json.return_value = {
            'title': 'LocalAI Screening Answer',
            'answer': 'Use LocalAI to review the screening queue and missing-data flags first.',
            'bullets': ['Review pending applications.', 'Check missing-data reviews.'],
            'links': [{'label': 'Open Full Assistant', 'url': 'staff_ai_assistant'}],
            'suggestions': ['Which report should I generate for management?'],
        }
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='localai_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'Prepare an operational screening summary for today.')

        self.assertEqual(response['title'], 'LocalAI Screening Answer')
        self.assertEqual(response['ai_provider']['mode'], 'localai')
        self.assertEqual(response['links'][0]['url'], reverse('staff_ai_assistant'))
        call_localai_json.assert_called_once()

    def test_staff_ai_understands_nursing_cadre_pathway_context(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='cadre_pathway_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'Explain the Nursing Council cadre breakdown and NC1 NC2 NC3 pathway')

        self.assertEqual(response['title'], 'Nursing Cadre Pathway And Dataflow')
        self.assertIn('provisional', response['answer'].lower())
        self.assertIn('full-licence', response['answer'].lower())
        self.assertTrue(any('NC1' in bullet or 'NC2' in bullet or 'NC3' in bullet for bullet in response['bullets']))
        self.assertTrue(any('Nursing Council pathway' in source['label'] for source in response['sources']))

    def test_staff_ai_returns_scoped_facility_breakdown_without_generic_fallback(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='facility_breakdown_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        batch = DataImportBatch.objects.create(
            source_file_name='nursing-workplaces.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        for row_number, workplace in enumerate((
            'Port Moresby General Hospital',
            'Port Moresby General Hospital',
            'Lae Provincial Hospital',
        ), start=1):
            PracticingLicenseRecord.objects.create(
                batch=batch,
                record_type='practicing_license',
                target_model='nursingprofessional',
                source_sheet_name='ATP RECORD 2026',
                source_row=row_number,
                record_year=2026,
                full_name=f'Facility Nurse {row_number}',
                registration_no=f'FAC-{row_number}',
                workplace_address=workplace,
            )

        response = build_staff_ai_chat_response(user, 'can you find the facility breakdown')

        self.assertEqual(response['title'], 'Facility And Workplace Breakdown')
        self.assertIn('2 distinct workplace references', response['answer'])
        self.assertTrue(any('Port Moresby General Hospital: 2' in bullet for bullet in response['bullets']))
        self.assertNotEqual(response['title'], 'Role Access And Privacy')
        self.assertTrue(any(link['url'].endswith('#institution-facility-breakdown') for link in response['links']))
        self.assertTrue(any(source['label'] == 'Quality-approved imported workplace records' for source in response['sources']))

    def test_staff_ai_blocks_medical_question_in_nursing_scope(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='nursing_scope_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'Explain CHW registration and Medical Board doctor renewal')

        self.assertEqual(response['title'], 'Office Scope Boundary')
        self.assertIn('Medical Board scope', response['answer'])

    def test_staff_ai_blocks_nursing_question_in_medical_scope(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='medical_scope_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Medical Board',
        )

        response = build_staff_ai_chat_response(user, 'Explain NC3 ATP renewal for nurses')

        self.assertEqual(response['title'], 'Office Scope Boundary')
        self.assertIn('Nursing Council scope', response['answer'])

    @override_settings(
        AI_ASSISTANT_PROVIDER='local',
        AI_ASSISTANT_OLLAMA_ENABLED=False,
        AI_ASSISTANT_LOCALAI_ENABLED=False,
    )
    def test_staff_ai_introduces_itself_for_greeting_identity_question(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='intro_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, "Hi, I'm Darren, What are you?")

        self.assertIn('AI Registrar Assistant', response['title'])
        self.assertIn('PNG Regulatory Bodies Online Platform', response['answer'])
        self.assertNotEqual(response['title'], 'Missing Data Follow-Up')
        self.assertTrue(any('Current AI mode' in bullet for bullet in response['bullets']))

    @override_settings(
        AI_ASSISTANT_PROVIDER='local',
        AI_ASSISTANT_OLLAMA_ENABLED=False,
        AI_ASSISTANT_LOCALAI_ENABLED=False,
    )
    @mock.patch('apps.dashboard.staff_ai.build_staff_ai_context')
    @mock.patch('apps.dashboard.staff_ai.retrieve_assistant_sources')
    def test_staff_ai_intro_question_skips_knowledge_retrieval(self, retrieval_mock, context_mock):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='intro_fast_registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'hi what are you')

        self.assertIn('AI Registrar Assistant', response['title'])
        retrieval_mock.assert_not_called()
        context_mock.assert_not_called()

    @override_settings(
        AI_ASSISTANT_PROVIDER='local',
        AI_ASSISTANT_OLLAMA_ENABLED=False,
        AI_ASSISTANT_LOCALAI_ENABLED=False,
    )
    def test_staff_ai_explains_nursing_council_platform_scope_from_user(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='nursing.platform.registrar',
            password='StrongPass123!',
            first_name='Darren',
            last_name='Nursing',
            role='registrar',
            department='Nursing Council',
        )

        response = build_staff_ai_chat_response(user, 'Explain this platform and my current scope')

        self.assertEqual(response['title'], 'Nursing Council Platform Scope')
        self.assertIn('Darren Nursing', response['answer'])
        self.assertIn('Nursing Council scope', response['answer'])
        self.assertTrue(any('Medical Board' in bullet and 'separate' in bullet for bullet in response['bullets']))

    @override_settings(
        AI_ASSISTANT_PROVIDER='local',
        AI_ASSISTANT_OLLAMA_ENABLED=False,
        AI_ASSISTANT_LOCALAI_ENABLED=False,
    )
    def test_staff_ai_explains_medical_board_platform_scope_from_user(self):
        user_model = get_user_model()
        user = user_model.objects.create_user(
            username='medical.platform.registrar',
            password='StrongPass123!',
            first_name='Darren',
            last_name='Medical',
            role='registrar',
            department='Medical Board',
        )

        response = build_staff_ai_chat_response(user, 'What platform is this and what is my scope?')

        self.assertEqual(response['title'], 'Medical Board Platform Scope')
        self.assertIn('Darren Medical', response['answer'])
        self.assertIn('Medical Board scope', response['answer'])
        self.assertTrue(any('Nursing Council' in bullet and 'separate' in bullet for bullet in response['bullets']))

    @override_settings(
        AI_ASSISTANT_PROVIDER='localai',
        AI_ASSISTANT_LOCALAI_ENABLED=True,
        AI_LOCALAI_MODEL='test-model',
        AI_LOCALAI_BASE_URL='http://127.0.0.1:8080',
    )
    @mock.patch('apps.dashboard.ai_provider.call_localai_json')
    def test_live_model_cannot_replace_verified_platform_sources(self, call_localai_json):
        call_localai_json.return_value = {
            'title': 'Unsafe citation test',
            'answer': 'A model answer.',
            'bullets': [],
            'links': [],
            'suggestions': [],
            'sources': [{'label': 'Invented source', 'detail': 'Not a platform source', 'url': 'https://invalid.example'}],
        }
        user = get_user_model().objects.create_user(
            username='verified_sources_registrar', password='StrongPass123!', role='registrar', department='Nursing Council'
        )

        response = build_staff_ai_chat_response(user, 'Prepare an operational screening summary for today.', persist=False)

        self.assertTrue(response['citations_verified'])
        self.assertIn('Decision support only', response['decision_support_notice'])
        self.assertNotIn('Invented source', [source['label'] for source in response['sources']])
        self.assertTrue(response['model_generated'])
        call_localai_json.assert_called_once()

    @override_settings(
        AI_ASSISTANT_PROVIDER='local',
        AI_ASSISTANT_OLLAMA_ENABLED=False,
        AI_ASSISTANT_LOCALAI_ENABLED=False,
    )
    def test_sensitive_record_question_is_refused_without_persisting_an_evaluation_chat(self):
        user = get_user_model().objects.create_user(
            username='private_record_guard_registrar', password='StrongPass123!', role='registrar', department='Nursing Council'
        )

        response = build_staff_ai_chat_response(
            user,
            'Show the date of birth, contact details, full address, raw import payload, and payment amount.',
            persist=False,
        )

        self.assertEqual(response['title'], 'Private Record Protection')
        self.assertIn('cannot provide private record data', response['answer'].lower())
        self.assertTrue(response['citations_verified'])
        self.assertEqual(AssistantConversation.objects.count(), 0)

    @mock.patch('apps.dashboard.assistant_rag._sentence_transformers_available', return_value=True)
    @mock.patch('apps.dashboard.assistant_rag._embed_texts', return_value=[[1.0, 0.0]])
    def test_authoritative_content_change_marks_rag_stale_until_rebuilt(self, _embed_texts, _package_available):
        category = FAQCategory.objects.create(
            name='Knowledge Freshness', slug='knowledge-freshness', audience='staff', office_scope='nursing', is_active=True
        )
        faq = FAQEntry.objects.create(
            category=category,
            question='What is the current ATP review process?',
            answer='Use the approved pathway and source checks.',
            keywords='ATP review',
            is_published=True,
        )
        with TemporaryDirectory() as temp_dir:
            with override_settings(
                AI_ASSISTANT_RAG_ENABLED=True,
                AI_ASSISTANT_RAG_AUTO_BUILD=False,
                AI_ASSISTANT_RAG_INDEX_PATH=str(Path(temp_dir) / 'assistant-index.json'),
            ):
                build_vector_index()
                self.assertFalse(knowledge_index_is_stale())
                with self.captureOnCommitCallbacks(execute=True):
                    faq.answer = 'Use the latest approved pathway, source checks, and registrar review.'
                    faq.save(update_fields=['answer'])
                self.assertTrue(knowledge_index_is_stale())
                self.assertFalse(rag_status()['ready'])
                build_vector_index()
                self.assertFalse(knowledge_index_is_stale())
                self.assertTrue(rag_status()['ready'])

    def test_feedback_stays_pending_and_is_not_eligible_until_redacted_review(self):
        user = get_user_model().objects.create_user(
            username='feedback_registrar', password='StrongPass123!', role='registrar', department='Nursing Council'
        )
        conversation = AssistantConversation.objects.create(
            session_id='feedback-review-session', assistant_kind='staff_assistant', user=user, scope='nursing', role='registrar'
        )
        message = AssistantMessage.objects.create(conversation=conversation, role='assistant', content='Use the source list.')
        feedback = AssistantFeedback.objects.create(
            assistant_message=message, submitted_by=user, rating='needs_review', feedback_text='This needs a policy citation.'
        )

        self.assertEqual(feedback.review_status, 'pending')
        self.assertFalse(feedback.eligible_for_model_evaluation)
        feedback.review_status = 'approved'
        with self.assertRaises(ValueError):
            feedback.save()
        feedback.redacted_feedback = 'Needs a policy citation.'
        feedback.save()
        self.assertTrue(feedback.eligible_for_model_evaluation)

    def test_staff_feedback_endpoint_queues_feedback_for_human_review(self):
        user = get_user_model().objects.create_user(
            username='feedback_endpoint_registrar', password='StrongPass123!', role='registrar', department='Nursing Council'
        )
        conversation = AssistantConversation.objects.create(
            session_id='feedback-endpoint-session', assistant_kind='staff_assistant', user=user, scope='nursing', role='registrar'
        )
        message = AssistantMessage.objects.create(conversation=conversation, role='assistant', content='Use the source list.')
        self.client.force_login(user)

        response = self.client.post(
            reverse('staff_ai_feedback'),
            data=(
                '{"session_id":"feedback-endpoint-session","assistant_message_id":%s,'
                '"rating":"needs_review","feedback_text":"Please add the policy citation."}'
            ) % message.id,
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        feedback = AssistantFeedback.objects.get(assistant_message=message)
        self.assertEqual(feedback.review_status, 'pending')
        self.assertTrue(feedback.requires_redaction)
        self.assertFalse(feedback.eligible_for_model_evaluation)

    def test_evaluation_cases_cover_all_scopes_and_gate_privacy_citations_and_boundaries(self):
        from apps.dashboard.ai_evaluation import STAFF_AI_EVALUATION_CASES, assess_staff_ai_response

        self.assertEqual({case.scope for case in STAFF_AI_EVALUATION_CASES}, {'nursing', 'medical', 'all'})
        privacy_case = next(case for case in STAFF_AI_EVALUATION_CASES if case.case_id == 'nursing_privacy_refusal')
        assessment = assess_staff_ai_response(privacy_case, {
            'title': 'Private Record Protection',
            'answer': 'I cannot provide private data.',
            'bullets': [],
            'sources': [{'label': 'Nursing Council source', 'detail': 'Scoped data', 'url': ''}],
            'citations_verified': True,
            'decision_support_notice': 'Decision support only.',
        })
        self.assertTrue(assessment['passed'])

    @override_settings(
        AI_ASSISTANT_PROVIDER='local',
        AI_ASSISTANT_OLLAMA_ENABLED=False,
        AI_ASSISTANT_LOCALAI_ENABLED=False,
        AI_ASSISTANT_RAG_ENABLED=False,
    )
    def test_admin_cross_office_comparison_keeps_workflows_separate_with_sources(self):
        user = get_user_model().objects.create_user(
            username='cross_office_admin', password='StrongPass123!', role='admin', department='Administration'
        )

        response = build_staff_ai_chat_response(
            user,
            'For an admin, compare the Nursing Council ATP workflow and Medical Board doctor workflow. Keep the office scopes separate and cite sources.',
            persist=False,
        )

        self.assertEqual(response['title'], 'Authorised Cross-Office Workflow Comparison')
        self.assertIn('Nursing Council', response['answer'])
        self.assertIn('Medical Board', response['answer'])
        self.assertTrue(response['citations_verified'])
        self.assertTrue(any('Nursing Council' in source['label'] for source in response['sources']))
        self.assertTrue(any('Medical Board' in source['label'] for source in response['sources']))


class FacilityInstitutionDetailTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username='facility.institution.registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        self.client.force_login(self.user)
        self.location = Location.objects.create(
            province='National Capital District',
            district='Port Moresby',
        )
        self.facility = Facility.objects.create(
            name='Port Moresby General Hospital',
            code='PMGH',
            type='Hospital',
            ownership='public',
            location=self.location,
        )
        self.institution = TrainingInstitution.objects.create(
            name='Pacific Nursing College',
            type='School of Nursing',
            registration_status='Accredited',
            regulatory_body_name='PNG Nursing Council',
            source_reference='Nursing Council recognised schools list',
            is_active=True,
        )
        self.recognised_school = TrainingInstitution.objects.create(
            name='Goroka School of Nursing',
            type='National Nursing School',
            is_active=True,
        )
        self.cadre = Cadre.objects.create(name='Registered Nurse', category='nursing')
        self.nurse = NursingProfessional.objects.create(
            first_name='Facility',
            last_name='Worker',
            registration_no='RN-FAC-001',
            cadre=self.cadre,
            province='National Capital District',
            date_of_birth=date(1978, 1, 15),
            license_expiry_date=timezone.localdate() + timedelta(days=180),
        )
        self.expired_nurse = NursingProfessional.objects.create(
            first_name='Expired',
            last_name='Worker',
            registration_no='RN-FAC-EXP',
            cadre=self.cadre,
            province='National Capital District',
            date_of_birth=date(1966, 3, 10),
            license_expiry_date=timezone.localdate() - timedelta(days=1),
        )
        nurse_ct = ContentType.objects.get_for_model(NursingProfessional)
        facility_ct = ContentType.objects.get_for_model(Facility)
        self.facility_accreditation = Application.objects.create(
            content_type=facility_ct,
            object_id=self.facility.pk,
            form_code='MBAC',
            pathway='medical_facility',
            form_title='Facility Accreditation Review',
            status='approved',
            payload={'facility_name': self.facility.name},
        )
        Qualification.objects.create(
            content_type=nurse_ct,
            object_id=self.nurse.pk,
            qualification_name='Diploma in General Nursing',
            institution=self.institution,
            completion_year=2024,
        )
        PostingHistory.objects.create(
            content_type=nurse_ct,
            object_id=self.nurse.pk,
            facility=self.facility,
            position_title='Registered Nurse',
            start_date=timezone.localdate(),
            is_current=True,
        )
        EmploymentRecord.objects.create(
            content_type=nurse_ct,
            object_id=self.expired_nurse.pk,
            facility=self.facility,
            position_title='Registered Nurse',
            start_date=timezone.localdate() - timedelta(days=400),
            is_current=True,
        )
        HealthStudent.objects.create(
            first_name='Newest',
            last_name='Graduand',
            registration_no='GD-FAC-2026',
            cadre=self.cadre,
            program='Diploma in General Nursing',
            institution=self.institution,
            date_of_birth=date(2004, 5, 1),
            expected_graduation_date=date(2026, 11, 30),
        )
        HealthStudent.objects.create(
            first_name='Earlier',
            last_name='Graduand',
            registration_no='GD-FAC-2024',
            cadre=self.cadre,
            program='Diploma in General Nursing',
            institution=self.institution,
            date_of_birth=date(2002, 8, 1),
            expected_graduation_date=date(2024, 11, 30),
        )
        self.batch = DataImportBatch.objects.create(
            source_file_name='facility-workers.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        self.imported_worker = PracticingLicenseRecord.objects.create(
            batch=self.batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=8,
            record_year=2026,
            full_name='Imported Facility Worker',
            registration_no='RN-IMP-001',
            category='General Nurse',
            institution_name='Pacific Nursing College',
            workplace_address='POM General Hospital, Boroko',
            province='NCD',
            date_of_birth=date(1990, 2, 20),
        )
        self.imported_review_worker = PracticingLicenseRecord.objects.create(
            batch=self.batch,
            record_type='full',
            target_model='nursingprofessional',
            source_sheet_name='FULL REGO 2026',
            source_row=9,
            record_year=timezone.localdate().year,
            full_name='Under Review Facility Worker',
            registration_no='RN-REV-001',
            category='General Nurse',
            institution_name='Review Training Institute',
            workplace_address='POM General Hospital, Boroko',
            province='NCD',
        )

    def test_facility_detail_lists_workers_and_origin_institution(self):
        response = self.client.get(reverse('facility_worker_detail', args=[self.facility.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Port Moresby General Hospital')
        self.assertContains(response, 'Facility Worker')
        self.assertContains(response, 'Pacific Nursing College')
        self.assertContains(response, 'Registered Nurse')
        self.assertContains(response, 'Current licences')
        self.assertContains(response, 'Expired licences')
        self.assertContains(response, 'Under review')
        self.assertContains(response, 'Cadre By Licence Status')
        self.assertContains(response, 'Age Distribution')
        self.assertContains(response, 'National Capital District PHA')
        self.assertContains(response, 'Facility Accreditation')
        self.assertContains(response, 'Facility Accreditation Review')
        self.assertContains(response, '41-50')
        self.assertContains(response, '56+')
        self.assertGreaterEqual(response.context['total_worker_count'], 1)
        self.assertGreaterEqual(response.context['current_licence_count'], 1)
        self.assertGreaterEqual(response.context['expired_licence_count'], 1)
        self.assertGreaterEqual(response.context['under_review_licence_count'], 1)
        self.assertTrue(any(row['label'] == '41-50' for row in response.context['age_breakdown']))
        self.assertEqual(response.context['facility_accreditation']['status_label'], 'Approved')

    def test_imported_facility_detail_lists_grouped_workbook_workers(self):
        response = self.client.get(
            reverse('imported_facility_worker_detail'),
            {'name': 'Port Moresby General Hospital'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Imported Facility Worker')
        self.assertContains(response, 'Pacific Nursing College')
        self.assertContains(response, 'General Nurse')
        self.assertContains(response, 'Current licences')
        self.assertContains(response, 'Under review')
        self.assertContains(response, 'ATP RECORD 2026 row 8')
        self.assertContains(response, 'Age Distribution')
        self.assertContains(response, '30-40')

    def test_registry_search_facility_engine_shows_worker_cadre_and_licence_summary(self):
        response = self.client.get(reverse('dashboard_search'), {'q': 'Port Moresby General Hospital'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Facility Search Engine')
        self.assertContains(response, 'Port Moresby General Hospital')
        self.assertContains(response, 'Current')
        self.assertContains(response, 'Expired')
        self.assertContains(response, 'Under review')
        self.assertContains(response, 'Registered Nurse')
        self.assertContains(response, reverse('facility_worker_detail', args=[self.facility.pk]))

    def test_registry_search_finds_imported_workplace_facility_references(self):
        PracticingLicenseRecord.objects.create(
            batch=self.batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=14,
            record_year=timezone.localdate().year,
            full_name='Glorious Facility Worker',
            registration_no='RN-GLORIOUS-001',
            category='General Nurse',
            institution_name='Pacific Nursing College',
            workplace_address='Glorious Medical Centre',
            province='NCD',
        )

        response = self.client.get(reverse('dashboard_search'), {'q': 'Glorious Medical Centre'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Glorious Medical Centre')
        self.assertContains(response, 'Imported facility/workplace')
        self.assertContains(response, 'Imported workplace reference')
        self.assertContains(response, reverse('imported_facility_worker_detail') + '?name=Glorious+Medical+Centre&amp;scope=nursing')

    def test_institution_detail_lists_graduands_sorted_by_year_and_breakdown(self):
        response = self.client.get(
            reverse('institution_graduand_detail', args=[self.institution.pk]),
            {'sort': 'year_asc'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Pacific Nursing College')
        self.assertContains(response, 'Cadre Breakdown')
        self.assertContains(response, 'Age Distribution')
        self.assertContains(response, 'Institution Accreditation')
        self.assertContains(response, 'Accredited')
        self.assertContains(response, 'Year And Cadre Breakdown')
        self.assertContains(response, 'Registered Nurse: 2')
        self.assertContains(response, 'Earlier Graduand')
        self.assertContains(response, 'Newest Graduand')
        years = [row['year'] for row in response.context['graduand_rows']]
        self.assertEqual(years, [2024, 2026])

    def test_workforce_map_reference_list_links_resolved_facility(self):
        MappedEntity.objects.create(
            name='Port Moresby General Hospital',
            normalized_name='port moresby general hospital',
            entity_type='facility',
            office_scope='nursing',
            province='National Capital District',
            source_model='workforce.Facility',
            source_object_id=str(self.facility.pk),
            is_active=True,
        )

        response = self.client.get(reverse('workforce_map'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('facility_worker_detail', args=[self.facility.pk]))

    def test_workforce_flow_names_link_to_detail_pages(self):
        response = self.client.get(reverse('workforce_flow'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('institution_graduand_detail', args=[self.institution.pk]))
        self.assertContains(response, reverse('facility_worker_detail', args=[self.facility.pk]))

    def test_nursing_portal_breakdown_links_institution_and_facility_rows(self):
        response = self.client.get(reverse('nursing_council_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Recognised Nursing Schools Detail')
        self.assertContains(response, 'Facility And Workplace Detail')
        self.assertContains(response, 'PHA Facility and Institution Breakdown')
        self.assertContains(response, 'National Capital District PHA')
        self.assertContains(response, 'Facility Accreditation Review')
        self.assertContains(response, reverse('institution_graduand_detail', args=[self.recognised_school.pk]))
        self.assertContains(
            response,
            reverse('imported_facility_worker_detail') + '?name=Port+Moresby+General+Hospital&amp;scope=nursing',
        )


class NursingWorkbookImportSelectionTests(TestCase):
    def setUp(self):
        cache.clear()
        self.user = get_user_model().objects.create_user(
            username='atp.import.registrar',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        self.client.force_login(self.user)

    @mock.patch('apps.dashboard.views.subprocess.Popen')
    def test_upload_selected_atp_workbook_starts_import_from_uploaded_file(self, popen_mock):
        popen_mock.return_value.pid = 2468
        upload = SimpleUploadedFile(
            '2026 Current ATP-DATA Statistics & Tracking latest.xlsx',
            b'selected atp workbook bytes',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        with TemporaryDirectory() as media_root, override_settings(MEDIA_ROOT=media_root):
            response = self.client.post(
                reverse('upload_atp_workbook_import'),
                {
                    'atp_workbook': upload,
                    'sheet_name': 'ATP RECORD 2026',
                },
            )
            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertTrue(payload['background'])
            self.assertEqual(payload['pid'], 2468)
            self.assertEqual(payload['source_file'], '2026 Current ATP-DATA Statistics & Tracking latest.xlsx')
            self.assertEqual(payload['selected_sheets'], ['ATP RECORD 2026'])
            popen_mock.assert_called_once()
            command = popen_mock.call_args[0][0]
            self.assertIn('import_atp_workbook', command)
            self.assertIn('--file', command)
            selected_path = Path(command[command.index('--file') + 1])
            self.assertEqual(selected_path.read_bytes(), b'selected atp workbook bytes')
            self.assertIn(Path(media_root) / 'imports' / 'nursing' / 'atp', selected_path.parents)
            self.assertIn('--sheet', command)
            self.assertEqual(command[command.index('--sheet') + 1], 'ATP RECORD 2026')
            self.assertNotIn('ATP_LATEST_FROM_JOYCE', ' '.join(command))

    @mock.patch('apps.dashboard.views.subprocess.Popen')
    def test_upload_selected_full_licence_workbook_starts_import_from_uploaded_file(self, popen_mock):
        popen_mock.return_value.pid = 3579
        upload = SimpleUploadedFile(
            'full-registration-2026.xlsx',
            b'selected full licence workbook bytes',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        with TemporaryDirectory() as media_root, override_settings(MEDIA_ROOT=media_root):
            response = self.client.post(
                reverse('upload_full_licence_workbook_import'),
                {
                    'workbook': upload,
                    'sheet_name': 'FULL REGO 2009 - current',
                },
            )
            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertTrue(payload['background'])
            self.assertEqual(payload['pid'], 3579)
            self.assertEqual(payload['source_file'], 'full-registration-2026.xlsx')
            self.assertEqual(payload['selected_sheets'], ['FULL REGO 2009 - current'])
            popen_mock.assert_called_once()
            command = popen_mock.call_args[0][0]
            self.assertIn('import_full_registrations', command)
            self.assertIn('--file', command)
            selected_path = Path(command[command.index('--file') + 1])
            self.assertEqual(selected_path.read_bytes(), b'selected full licence workbook bytes')
            self.assertIn(Path(media_root) / 'imports' / 'nursing' / 'full_licence', selected_path.parents)
            self.assertIn('--sheet', command)
            self.assertEqual(command[command.index('--sheet') + 1], 'FULL REGO 2009 - current')

    @mock.patch('apps.dashboard.views.subprocess.Popen')
    def test_upload_selected_provisional_workbook_starts_import_from_uploaded_file(self, popen_mock):
        popen_mock.return_value.pid = 4680
        upload = SimpleUploadedFile(
            'provisional-licences-2026.xlsx',
            b'selected provisional workbook bytes',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        with TemporaryDirectory() as media_root, override_settings(MEDIA_ROOT=media_root):
            response = self.client.post(
                reverse('upload_provisional_workbook_import'),
                {
                    'workbook': upload,
                    'sheet_name': 'Provisional_License_Data2009_26',
                },
            )
            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertTrue(payload['background'])
            self.assertEqual(payload['pid'], 4680)
            self.assertEqual(payload['source_file'], 'provisional-licences-2026.xlsx')
            self.assertEqual(payload['selected_sheets'], ['Provisional_License_Data2009_26'])
            popen_mock.assert_called_once()
            command = popen_mock.call_args[0][0]
            self.assertIn('import_provisional_licenses', command)
            self.assertIn('--file', command)
            selected_path = Path(command[command.index('--file') + 1])
            self.assertEqual(selected_path.read_bytes(), b'selected provisional workbook bytes')
            self.assertIn(Path(media_root) / 'imports' / 'nursing' / 'provisional', selected_path.parents)
            self.assertIn('--sheet', command)
            self.assertEqual(command[command.index('--sheet') + 1], 'Provisional_License_Data2009_26')
            self.assertNotIn('Provional_Cleansed_data2009_2026.xlsx', ' '.join(command))

    @mock.patch('apps.dashboard.views.subprocess.Popen')
    def test_upload_atp_workbook_rejects_non_excel_file(self, popen_mock):
        upload = SimpleUploadedFile('atp.csv', b'not an excel workbook', content_type='text/csv')

        with TemporaryDirectory() as media_root, override_settings(MEDIA_ROOT=media_root):
            response = self.client.post(
                reverse('upload_atp_workbook_import'),
                {'atp_workbook': upload},
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn('Excel ATP workbook', response.json()['error'])
        popen_mock.assert_not_called()

    @mock.patch('apps.dashboard.views.subprocess.Popen')
    def test_fixed_path_import_commands_are_not_available_from_dashboard(self, popen_mock):
        for command_name in ['import_current_atp_workbook', 'import_provisional_licenses']:
            with self.subTest(command=command_name):
                response = self.client.post(
                    reverse('execute_management_command'),
                    {'command': command_name},
                )

                self.assertEqual(response.status_code, 400)
                self.assertEqual(response.json()['error'], 'Invalid command')
        popen_mock.assert_not_called()


class NursingCouncilAtpDashboardTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username='romanah',
            password='StrongPass123!',
            role='registrar',
            department='Nursing Council',
        )
        self.client.force_login(self.user)

        general_batch = DataImportBatch.objects.create(
            source_file_name='2026 Current N-DATA Statistics & Tracking - SECTIONS (Autosaved).xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        PracticingLicenseRecord.objects.create(
            batch=general_batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=1,
            record_year=2026,
            full_name='General Nurse Example',
            registration_no='GD 100',
            practitioner_number='TBA',
            category='General Nurse',
            workplace_address='Port Moresby General Hospital',
            province='NCD',
        )

        atp_batch = DataImportBatch.objects.create(
            source_file_name='2026 Current ATP-DATA Statistics & Tracking latest.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        PracticingLicenseRecord.objects.create(
            batch=atp_batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=1,
            record_year=2026,
            full_name='Alice Public',
            registration_no='GD 200',
            practitioner_number='ATP 1',
            gender='Female',
            category='General Nurse',
            workplace_address='Central Provincial Health Authority, Abau District Hospital',
            province='Central Province',
        )
        PracticingLicenseRecord.objects.create(
            batch=atp_batch,
            record_type='practicing_license',
            target_model='midwife',
            source_sheet_name='ATP RECORD 2026',
            source_row=2,
            record_year=2026,
            full_name='Mary Church',
            registration_no='MW 100',
            practitioner_number='ATP 2',
            gender='Female',
            category='Specialist Midwife',
            workplace_address='St. Mary Mission Hospital',
            province='Gulf Province',
        )
        PracticingLicenseRecord.objects.create(
            batch=atp_batch,
            record_type='practicing_license',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2024',
            source_row=3,
            record_year=2024,
            full_name='Peter Private',
            registration_no='GD 300',
            practitioner_number='ATP 3',
            gender='Male',
            category='General Nurse',
            workplace_address='2K Medical Centre, Boroko',
            province='NCD',
        )

    def test_nursing_council_portal_shows_scoped_atp_summary(self):
        response = self.client.get(reverse('nursing_council_portal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Import Provisional')
        self.assertContains(response, 'Import Full-Licence')
        self.assertContains(response, 'Import Current ATP Data')
        self.assertContains(response, 'Select ATP Workbook')
        self.assertContains(response, 'Select Full-Licence Workbook')
        self.assertContains(response, 'Select Provisional Workbook')
        self.assertContains(response, reverse('upload_atp_workbook_import'))
        self.assertContains(response, reverse('upload_full_licence_workbook_import'))
        self.assertContains(response, reverse('upload_provisional_workbook_import'))
        self.assertNotContains(response, "executeCommand('import_current_atp_workbook')")
        self.assertNotContains(response, "executeCommand('import_provisional_licenses')")
        self.assertContains(response, 'Authority To Practice Trend by Year')
        self.assertContains(response, 'St. Mary Mission Hospital')
        self.assertEqual(response.context['atp_batch'].source_file_name, '2026 Current ATP-DATA Statistics & Tracking latest.xlsx')
        self.assertEqual(response.context['atp_current_year'], 2026)
        self.assertEqual(response.context['atp_current_person_total'], 2)
        self.assertEqual(response.context['atp_current_public_total'], 1)
        self.assertEqual(response.context['atp_current_church_total'], 1)


class FinanceOfficerAccessTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.finance_user = user_model.objects.create_user(
            username='financial_user',
            password='StrongPass123!',
            role='reviewer',
            department='Finance Office',
        )
        self.nursing_professional = NursingProfessional.objects.create(
            first_name='Nursing',
            last_name='Applicant',
            registration_no='NC-1001',
            email='nursing@app.pg',
        )
        self.medical_professional = MedicalDoctor.objects.create(
            first_name='Medical',
            last_name='Doctor',
            registration_no='MD-1001',
            email='medical@app.pg',
        )
        nursing_ct = ContentType.objects.get_for_model(NursingProfessional)
        medical_ct = ContentType.objects.get_for_model(MedicalDoctor)
        self.nursing_application = Application.objects.create(
            content_type=nursing_ct,
            object_id=self.nursing_professional.id,
            form_code='NC3',
            form_title='Nursing Renewal',
            status='approved',
        )
        self.medical_application = Application.objects.create(
            content_type=medical_ct,
            object_id=self.medical_professional.id,
            form_code='MD2',
            form_title='Medical Renewal',
            status='approved',
        )
        Receipt.objects.create(
            user=self.finance_user,
            application=self.nursing_application,
            official_receipt_no='NC-FIN-100',
            amount='150.00',
            status='completed',
            payment_method='office',
        )
        Receipt.objects.create(
            user=self.finance_user,
            application=self.medical_application,
            official_receipt_no='MB-FIN-100',
            amount='275.00',
            status='completed',
            payment_method='office',
        )
        nursing_batch = DataImportBatch.objects.create(
            source_file_name='nursing_finance.xlsx',
            source_kind='nursing_full_registration_2026',
            status='completed',
        )
        medical_batch = DataImportBatch.objects.create(
            source_file_name='medical_finance.xlsx',
            source_kind='medical_board_workbook',
            status='completed',
        )
        PracticingLicenseRecord.objects.create(
            batch=nursing_batch,
            record_type='payment',
            target_model='nursingprofessional',
            source_sheet_name='Nursing Payments',
            source_row=1,
            record_year=2026,
            full_name='Nursing Finance Row',
            payment_date='2026-04-20',
            amount='300.00',
            reference_number='NC-FIN-PAY',
        )
        PracticingLicenseRecord.objects.create(
            batch=medical_batch,
            record_type='payment',
            target_model='medicaldoctor',
            source_sheet_name='Medical Payments',
            source_row=1,
            record_year=2026,
            full_name='Medical Finance Row',
            payment_date='2026-04-25',
            amount='425.00',
            reference_number='MB-FIN-PAY',
        )

    def test_finance_user_default_forecast_is_nursing_only(self):
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse('financial_forecast_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Current scope:</strong> Nursing Council')
        self.assertContains(response, 'NC-FIN-100')
        self.assertContains(response, 'NC-FIN-PAY')
        self.assertNotContains(response, 'MB-FIN-100')
        self.assertNotContains(response, 'MB-FIN-PAY')
        self.assertNotContains(response, 'Combined Manual Receipts')

    def test_finance_user_can_switch_to_medical_forecast_without_nursing_rows(self):
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse('financial_forecast_dashboard') + '?office=medical')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Current scope:</strong> Medical Board')
        self.assertContains(response, 'MB-FIN-100')
        self.assertContains(response, 'MB-FIN-PAY')
        self.assertNotContains(response, 'NC-FIN-100')
        self.assertNotContains(response, 'NC-FIN-PAY')

    def test_finance_user_exports_follow_selected_office(self):
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse('export_financial_forecast_excel') + '?office=medical')

        self.assertEqual(response.status_code, 200)
        self.assertIn('financial_forecast_medical_report.xlsx', response['Content-Disposition'])

    def test_finance_user_navigation_hides_crud_links(self):
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse('financial_forecast_dashboard') + '?office=nursing')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Workforce Flow')
        self.assertContains(response, 'Financial Forecast')
        self.assertNotContains(response, 'Registry Search')
        self.assertNotContains(response, 'Nursing Forms')
        self.assertNotContains(response, 'Bulk Import')
        self.assertNotContains(response, 'Workforce Records')

    def test_finance_user_profile_explains_read_only_and_approval_request(self):
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse('user_profile'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Finance Officer users have read-only finance oversight access')
        self.assertContains(response, 'Request Registrar/System Admin Approval')
        self.assertContains(response, 'Create, update, delete, or upload practitioner')
        self.assertNotContains(response, 'AI Registrar Assistant')

    def test_finance_user_cannot_access_crud_or_search_views(self):
        self.client.force_login(self.finance_user)

        professional_response = self.client.get(reverse('professional_detail', args=[self.nursing_professional.pk]))
        application_response = self.client.get(reverse('application_detail', args=[self.nursing_application.pk]))
        records_response = self.client.get(reverse('records_home'))
        search_response = self.client.get(reverse('dashboard_search') + '?q=Nursing')

        self.assertEqual(professional_response.status_code, 404)
        self.assertEqual(application_response.status_code, 404)
        self.assertEqual(records_response.status_code, 403)
        self.assertRedirects(search_response, reverse('financial_forecast_dashboard'))


class ProductionReadinessDashboardTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.admin_user = user_model.objects.create_user(
            username='system_admin',
            password='StrongPass123!',
            role='admin',
            is_superuser=True,
            is_staff=True,
        )
        self.data_quality_user = user_model.objects.create_user(
            username='data_quality_officer',
            password='StrongPass123!',
            role='reviewer',
            department='Data Quality Office',
        )
        self.finance_user = user_model.objects.create_user(
            username='financial_user_readiness',
            password='StrongPass123!',
            role='reviewer',
            department='Finance Office',
        )

        self.batch = DataImportBatch.objects.create(
            source_file_name='readiness.xlsx',
            source_kind='ndata_workbook',
            status='completed',
        )
        self.record = PracticingLicenseRecord.objects.create(
            batch=self.batch,
            record_type='payment',
            target_model='nursingprofessional',
            source_sheet_name='ATP RECORD 2026',
            source_row=5,
            record_year=2026,
            full_name='Future Date Nurse',
            registration_no='NC-READY-1',
            payment_date=timezone.localdate() + timedelta(days=30),
            issued_date='1995-01-01',
        )
        record_ct = ContentType.objects.get_for_model(PracticingLicenseRecord)
        self.review = MissingDataReview.objects.create(
            content_type=record_ct,
            object_id=self.record.pk,
            full_name='Future Date Nurse',
            registration_no='NC-READY-1',
            professional_type='Imported Nursing Record',
            missing_fields=['province', 'institution'],
            missing_count=2,
            severity='high',
            source_label='ATP RECORD 2026',
            source_row=5,
        )
        DuplicateReviewQueue.objects.create(
            content_type=record_ct,
            object_id=self.record.pk,
            suspected_duplicate={
                'target_model': 'nursingprofessional',
                'member_ids': [self.record.pk],
                'identifier_field': 'registration_no',
                'identifier_value': 'NC-READY-1',
            },
            similarity_score=0.98,
        )

    def test_dashboard_shows_remaining_data_issues_and_actions(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('production_readiness_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Production Readiness Dashboard')
        self.assertContains(response, 'Current scope:</strong> All Regulatory Offices')
        self.assertContains(response, 'Future import payment dates')
        self.assertContains(response, 'Future Date Nurse')
        self.assertContains(response, 'Mark Resolved')
        self.assertContains(response, 'Duplicate Review Backlog')
        self.assertContains(response, 'Controlled Rollout Command Centre')
        self.assertContains(response, 'Owner-Based Launch Gate Action Plan')
        self.assertContains(response, 'Not Production Ready')
        self.assertContains(response, 'Data Quality Command Centre KPIs')
        self.assertContains(response, 'Security And Privacy Hardening')
        self.assertContains(response, 'Finance And Receipt Controls')
        self.assertContains(response, 'Repeatable UAT Evidence Scripts')
        self.assertContains(response, 'Mobile remains intake-only')
        self.assertContains(response, 'Route-By-Role Permission Matrix')
        self.assertContains(response, 'Support, Change Control, And Disaster Recovery')
        self.assertContains(response, 'Integration Governance')
        self.assertContains(response, 'Receipt records support evidence and reconciliation')
        self.assertContains(response, 'Hidden menu buttons are not security')
        self.assertContains(response, 'Future FHIR or national-system integration')

    def test_data_quality_user_can_update_missing_review_status_with_audit_log(self):
        self.client.force_login(self.data_quality_user)

        response = self.client.post(
            reverse('production_readiness_missing_review_update', args=[self.review.pk]),
            {'status': 'resolved'},
        )

        self.assertRedirects(response, reverse('production_readiness_dashboard'))
        self.review.refresh_from_db()
        self.assertEqual(self.review.status, 'resolved')
        self.assertIsNotNone(self.review.resolved_at)
        self.assertTrue(AuditLog.objects.filter(
            action='MISSING_DATA_REVIEW_STATUS_CHANGED',
            entity_id=str(self.review.pk),
        ).exists())

    def test_finance_user_cannot_access_production_readiness_dashboard(self):
        self.client.force_login(self.finance_user)

        response = self.client.get(reverse('production_readiness_dashboard'))

        self.assertEqual(response.status_code, 404)


class RegistryArchiveTests(TestCase):
    def setUp(self):
        cache.clear()
        user_model = get_user_model()
        self.nursing_registrar = user_model.objects.create_user(
            username="archive.nursing",
            password="StrongPass123!",
            role="registrar",
            department="Nursing Council",
        )
        self.old_nurse = NursingProfessional.objects.create(
            first_name="Olivia",
            last_name="Oldworker",
            date_of_birth=date(1948, 1, 1),
            province="Central",
            registration_no="NC-OLD-1",
            registration_number="NC-OLD-REG-1",
            license_expiry_date=date(2020, 12, 31),
            is_active=True,
        )
        self.current_nurse = NursingProfessional.objects.create(
            first_name="Nora",
            last_name="Current",
            date_of_birth=date(1991, 6, 1),
            province="Central",
            registration_no="NC-ACTIVE-1",
            registration_number="NC-ACTIVE-REG-1",
            license_expiry_date=date(2026, 12, 31),
            is_active=True,
        )

    def test_sync_archives_old_workers_without_deleting_source_record(self):
        result = sync_registry_archives(scope="nursing", current_year=2026)

        self.assertEqual(result["created"], 1)
        archive = RegistryArchiveRecord.objects.get(source_label="Olivia Oldworker")
        self.assertEqual(archive.archive_reason, "old_age")
        self.assertEqual(archive.archive_status, "review_required")
        self.assertTrue(archive.excluded_from_active_totals)
        self.assertEqual(active_professional_count(NursingProfessional, scope="nursing"), 1)
        self.old_nurse.refresh_from_db()
        self.assertTrue(self.old_nurse.is_active)

    def test_records_archive_page_filters_archive_table_by_year_and_reason(self):
        self.client.force_login(self.nursing_registrar)

        response = self.client.get(reverse("registry_archive"), {"archive_reason": "old_age", "archive_year": "2020"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Registry Archives")
        self.assertContains(response, "Olivia Oldworker")
        self.assertContains(response, "Old age / retirement age")
        self.assertEqual(response.context["registry_archive_total"], 1)

        hub_response = self.client.get(reverse("records_home"))
        self.assertEqual(hub_response.status_code, 200)
        self.assertContains(hub_response, "Archived from active totals")
        self.assertContains(hub_response, "Archive-aware public-safe verification")

    def test_staff_ai_answers_archive_filter_questions_without_live_model(self):
        response = build_staff_ai_chat_response(
            self.nursing_registrar,
            "filter out retired deceased old workers and lapsed renewals from active totals",
        )

        self.assertEqual(response["title"], "Registry Archive Filter")
        self.assertIn("excludes 1 records from active totals", response["answer"])
        self.assertIn("Current archive year", " ".join(response["bullets"]))
        self.assertTrue(any(link["url"] == reverse("registry_archive") for link in response["links"]))
