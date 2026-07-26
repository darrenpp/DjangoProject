from collections import Counter
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
import hashlib
import re

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.dashboard.models import (
    NursingAnalyticsSnapshot,
    NursingCadreStageMetric,
    NursingDataQualityMetric,
    NursingFacilityAlias,
    NursingFacilityCadreYearMetric,
    NursingInstitutionAlias,
    NursingInstitutionCadreYearMetric,
    NursingLifecycleFact,
    NursingPractitionerIndex,
    NursingProvinceYearMetric,
    NursingStageYearMetric,
    NursingStandardsFieldMap,
)
from apps.workforce.models import DataImportBatch


DEFAULT_NURSING_ANALYTICS_WORKBOOK = Path(
    r"C:\Users\darre\Documents\ProjectApps\databasedocuments\spreadsheets\Nursing_Council_Cleansed_data_current\PNG_Nursing_Council_Integrated_Dashboard_Model.xlsx"
)

ANALYTICS_SOURCE_KIND = "nursing_analytics_snapshot"
BULK_CREATE_SIZE = 1000
QUALITY_LIFECYCLE_STAGES = {
    "Provisional Licence",
    "Full Licence",
    "Authority to Practice",
}


def file_sha256(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_header(value):
    text = normalize_text(value).lower().replace("%", " percent")
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"-?\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def normalize_name_key(value):
    text = normalize_text(value)
    text = re.sub(r"[^A-Za-z0-9]+", " ", text).strip()
    return re.sub(r"\s+", " ", text).title()


def parse_int(value, default=0):
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return int(value)
    try:
        return int(Decimal(str(value).replace(",", "").strip()))
    except (InvalidOperation, ValueError):
        return default


def parse_optional_int(value):
    parsed = parse_int(value, default=None)
    return parsed if parsed is not None else None


def parse_decimal(value):
    if value in (None, ""):
        return None
    text = str(value).replace("%", "").replace(",", "").strip()
    try:
        decimal_value = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if "%" in str(value):
        decimal_value = decimal_value / Decimal("100")
    return decimal_value


def parse_bool(value):
    return normalize_text(value).lower() in {"1", "true", "yes", "y"}


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        serial_value = float(value)
        if 20000 <= serial_value <= 60000:
            return date(1899, 12, 30) + timedelta(days=int(serial_value))
    text = normalize_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def sheet_rows(workbook, sheet_name, required_first_header=None):
    if sheet_name not in workbook.sheetnames:
        return [], []
    rows = workbook[sheet_name].iter_rows(values_only=True)
    header = []
    for row in rows:
        values = list(row)
        normalized = [normalize_header(value) for value in values]
        if required_first_header:
            if normalized and normalized[0] == required_first_header:
                header = normalized
                break
        elif any(normalized):
            header = normalized
            break
    if not header:
        return [], []

    dict_rows = []
    for row in rows:
        values = list(row)
        if not any(value not in (None, "") for value in values):
            continue
        item = {}
        for index, key in enumerate(header):
            if not key:
                continue
            item[key] = values[index] if index < len(values) else None
        dict_rows.append(item)
    return header, dict_rows


def data_quality_lifecycle_rows(workbook):
    if "Data_Quality" not in workbook.sheetnames:
        return []
    rows = workbook["Data_Quality"].iter_rows(values_only=True)
    header = []
    for row in rows:
        normalized = [normalize_header(value) for value in row]
        if normalized[:4] == ["lifecycle_stage", "high", "medium", "needs_review"]:
            header = normalized
            break
    if not header:
        return []

    dict_rows = []
    for row in rows:
        if not any(value not in (None, "") for value in row):
            break
        item = {}
        for index, key in enumerate(header):
            if not key:
                continue
            item[key] = row[index] if index < len(row) else None
        if normalize_text(item.get("lifecycle_stage")) in QUALITY_LIFECYCLE_STAGES:
            dict_rows.append(item)
    return dict_rows


def row_payload(row):
    return {
        key: value.isoformat() if isinstance(value, (date, datetime)) else value
        for key, value in row.items()
        if value not in (None, "")
    }


def workbook_audit_metadata(workbook):
    metadata = {
        "workbook_title": "",
        "workbook_generated_on": None,
        "official_clean_atp_records": None,
        "official_clean_provisional_records": None,
        "official_clean_full_licence_records": None,
        "total_integrated_lifecycle_records": None,
        "estimated_practitioner_match_groups": None,
        "data_quality_health_score": None,
    }
    if "ReadMe_Source_Audit" not in workbook.sheetnames:
        return metadata
    for row in workbook["ReadMe_Source_Audit"].iter_rows(values_only=True):
        key = normalize_text(row[0] if row else "")
        value = row[1] if len(row) > 1 else None
        key_lookup = normalize_header(key)
        if key_lookup == "workbook":
            metadata["workbook_title"] = normalize_text(value)
        elif key_lookup == "generated":
            metadata["workbook_generated_on"] = parse_date(value)
        elif key_lookup == "official_clean_atp_records":
            metadata["official_clean_atp_records"] = parse_int(value)
        elif key_lookup == "official_clean_provisional_records":
            metadata["official_clean_provisional_records"] = parse_int(value)
        elif key_lookup == "official_clean_full_licence_records":
            metadata["official_clean_full_licence_records"] = parse_int(value)
        elif key_lookup == "total_integrated_lifecycle_records":
            metadata["total_integrated_lifecycle_records"] = parse_int(value)
        elif key_lookup == "estimated_practitioner_match_groups":
            metadata["estimated_practitioner_match_groups"] = parse_int(value)
        elif key_lookup == "data_quality_health_score":
            metadata["data_quality_health_score"] = float(parse_decimal(value) or 0) * 100
    return metadata


def dashboard_stage_counts(workbook):
    counts = {}
    if "Dashboard_Data" not in workbook.sheetnames:
        return counts
    rows = list(workbook["Dashboard_Data"].iter_rows(values_only=True))
    for index, row in enumerate(rows):
        normalized = [normalize_header(value) for value in row]
        if normalized[:2] == ["lifecycle_stage", "clean_record_count"]:
            for stage_row in rows[index + 1:]:
                stage = normalize_text(stage_row[0] if stage_row else "")
                if not stage:
                    break
                counts[stage] = parse_int(stage_row[1] if len(stage_row) > 1 else None)
            break
    return counts


def build_kpi_summary(workbook):
    metadata = workbook_audit_metadata(workbook)
    stage_counts = dashboard_stage_counts(workbook)
    provisional_count = metadata["official_clean_provisional_records"] or stage_counts.get("Provisional Licence", 0)
    full_count = metadata["official_clean_full_licence_records"] or stage_counts.get("Full Licence", 0)
    atp_count = metadata["official_clean_atp_records"] or stage_counts.get("Authority to Practice", 0)
    total = metadata["total_integrated_lifecycle_records"] or sum(stage_counts.values())
    high = medium = needs_review = quality_total = 0
    quality_rows = data_quality_lifecycle_rows(workbook)
    for row in quality_rows:
        if normalize_text(row.get("lifecycle_stage")) not in QUALITY_LIFECYCLE_STAGES:
            continue
        high += parse_int(row.get("high"))
        medium += parse_int(row.get("medium"))
        needs_review += parse_int(row.get("needs_review"))
        quality_total += parse_int(row.get("grand_total"))
    health_score = metadata["data_quality_health_score"]
    if not health_score and quality_total:
        health_score = round(((quality_total - needs_review) / quality_total) * 100, 1)
    return {
        "total_lifecycle_records": total,
        "clean_atp_records": atp_count,
        "clean_provisional_records": provisional_count,
        "clean_full_licence_records": full_count,
        "data_quality_health_score": round(float(health_score or 0), 1),
        "data_quality_high_count": high,
        "data_quality_medium_count": medium,
        "data_quality_needs_review_count": needs_review,
        "estimated_practitioner_match_groups": metadata["estimated_practitioner_match_groups"] or 0,
    }, metadata


def bulk_create(model, rows):
    if rows:
        model.objects.bulk_create(rows, batch_size=BULK_CREATE_SIZE)


def import_lifecycle_facts(workbook, snapshot):
    _header, rows = sheet_rows(workbook, "Fact_Lifecycle")
    objects = []
    for index, row in enumerate(rows, start=1):
        record_id = normalize_text(row.get("record_id")) or f"FACT-{index:06d}"
        source_sheet = normalize_text(row.get("source_sheet"))
        source_row = parse_optional_int(row.get("source_row"))
        objects.append(NursingLifecycleFact(
            snapshot=snapshot,
            record_id=record_id[:80],
            lifecycle_stage=normalize_text(row.get("lifecycle_stage"))[:80],
            licence_status=normalize_text(row.get("licence_status"))[:100],
            lifecycle_order=parse_optional_int(row.get("lifecycle_order")),
            cycle_year=parse_optional_int(row.get("cycle_year")),
            event_date=parse_date(row.get("event_date")),
            full_name=normalize_text(row.get("full_name"))[:255],
            name_key=normalize_text(row.get("name_key"))[:255],
            person_group_key=normalize_text(row.get("person_group_key"))[:255],
            identity_confidence=normalize_text(row.get("identity_confidence"))[:120],
            dob=parse_date(row.get("dob")),
            sex=normalize_text(row.get("sex"))[:40],
            age=parse_optional_int(row.get("age")),
            cadre=normalize_text(row.get("cadre"))[:150],
            cadre_group=normalize_text(row.get("cadre_group"))[:120],
            profession_speciality_raw=normalize_text(row.get("profession_speciality_raw"))[:255],
            formal_qualification=normalize_text(row.get("formal_qualification"))[:255],
            registration_no=normalize_text(row.get("registration_no"))[:100],
            practitioner_no=normalize_text(row.get("practitioner_no"))[:100],
            registration_link_key=normalize_text(row.get("registration_link_key"))[:120],
            institution=normalize_text(row.get("institution"))[:255],
            facility=normalize_text(row.get("facility"))[:255],
            province=normalize_text(row.get("province"))[:120],
            organization_type=normalize_text(row.get("organization_type"))[:160],
            nationality_group=normalize_text(row.get("nationality_group"))[:100],
            country=normalize_text(row.get("country"))[:100],
            include_in_official_totals=parse_bool(row.get("include_in_official_totals")),
            data_quality_flags=normalize_text(row.get("data_quality_flags")),
            completeness_score=parse_decimal(row.get("completeness_score")),
            record_quality=normalize_text(row.get("record_quality"))[:80],
            source_workbook=normalize_text(row.get("source_workbook"))[:255],
            source_sheet=source_sheet[:255],
            source_row=source_row,
            source_lineage=f"{source_sheet} row {source_row}" if source_sheet and source_row else source_sheet[:255],
            raw_payload=row_payload(row),
        ))
        if len(objects) >= BULK_CREATE_SIZE:
            bulk_create(NursingLifecycleFact, objects)
            objects = []
    bulk_create(NursingLifecycleFact, objects)
    return len(rows)


def import_practitioner_index(workbook, snapshot):
    _header, rows = sheet_rows(workbook, "Practitioner_Index")
    objects = []
    for index, row in enumerate(rows, start=1):
        group_id = normalize_text(row.get("practitioner_group_id")) or f"PG-{index:06d}"
        objects.append(NursingPractitionerIndex(
            snapshot=snapshot,
            practitioner_group_id=group_id[:80],
            person_group_key=normalize_text(row.get("person_group_key"))[:255],
            representative_name=normalize_text(row.get("representative_name"))[:255],
            identity_confidence=normalize_text(row.get("identity_confidence"))[:120],
            record_count=parse_int(row.get("record_count")),
            stages_present=normalize_text(row.get("stages_present"))[:255],
            has_provisional=parse_bool(row.get("has_provisional")),
            has_full_licence=parse_bool(row.get("has_full_licence")),
            has_atp=parse_bool(row.get("has_atp")),
            first_year=parse_optional_int(row.get("first_year")),
            latest_year=parse_optional_int(row.get("latest_year")),
            latest_atp_year=parse_optional_int(row.get("latest_atp_year")),
            latest_cadre=normalize_text(row.get("latest_cadre"))[:150],
            latest_facility=normalize_text(row.get("latest_facility"))[:255],
            latest_province=normalize_text(row.get("latest_province"))[:120],
            registration_nos=normalize_text(row.get("registration_nos")),
            practitioner_nos=normalize_text(row.get("practitioner_nos")),
            dq_flag_count=parse_int(row.get("dq_flag_count")),
            needs_manual_review=parse_bool(row.get("needs_manual_review")),
            raw_payload=row_payload(row),
        ))
        if len(objects) >= BULK_CREATE_SIZE:
            bulk_create(NursingPractitionerIndex, objects)
            objects = []
    bulk_create(NursingPractitionerIndex, objects)
    return len(rows)


def import_stage_year_metrics(workbook, snapshot):
    _header, rows = sheet_rows(workbook, "Year_Stage")
    objects = []
    for row in rows:
        year_label = normalize_text(row.get("year"))
        year = parse_optional_int(year_label)
        if not year and not year_label:
            continue
        objects.append(NursingStageYearMetric(
            snapshot=snapshot,
            year=year,
            year_label=year_label[:40],
            provisional_licence_count=parse_int(row.get("provisional_licence")),
            full_licence_count=parse_int(row.get("full_licence")),
            authority_to_practice_count=parse_int(row.get("authority_to_practice")),
            grand_total=parse_int(row.get("grand_total")),
            raw_payload=row_payload(row),
        ))
    bulk_create(NursingStageYearMetric, objects)
    return len(objects)


def import_cadre_stage_metrics(workbook, snapshot):
    _header, rows = sheet_rows(workbook, "Cadre_Stage")
    objects = [
        NursingCadreStageMetric(
            snapshot=snapshot,
            cadre=normalize_text(row.get("cadre"))[:150],
            cadre_group=normalize_text(row.get("cadre_group"))[:120],
            provisional_licence_count=parse_int(row.get("provisional_licence")),
            full_licence_count=parse_int(row.get("full_licence")),
            authority_to_practice_count=parse_int(row.get("authority_to_practice")),
            grand_total=parse_int(row.get("grand_total")),
            raw_payload=row_payload(row),
        )
        for row in rows
        if normalize_text(row.get("cadre"))
    ]
    bulk_create(NursingCadreStageMetric, objects)
    return len(objects)


def year_columns(header, excluded):
    return [
        key for key in header
        if key not in excluded and key != "grand_total" and (key == "unknown" or re.fullmatch(r"\d{4}", key))
    ]


def import_facility_cadre_year_metrics(workbook, snapshot):
    header, rows = sheet_rows(workbook, "Facility_Cadre_Year")
    year_keys = year_columns(header, {"facility", "province", "organization_type", "cadre"})
    objects = []
    for row in rows:
        facility = normalize_text(row.get("facility"))
        if not facility:
            continue
        for key in year_keys:
            count = parse_int(row.get(key))
            if not count:
                continue
            year = parse_optional_int(key)
            objects.append(NursingFacilityCadreYearMetric(
                snapshot=snapshot,
                facility=facility[:255],
                province=normalize_text(row.get("province"))[:120],
                organization_type=normalize_text(row.get("organization_type"))[:160],
                cadre=normalize_text(row.get("cadre"))[:150],
                year=year,
                year_label=key.title(),
                count=count,
                raw_payload=row_payload(row),
            ))
            if len(objects) >= BULK_CREATE_SIZE:
                bulk_create(NursingFacilityCadreYearMetric, objects)
                objects = []
    bulk_create(NursingFacilityCadreYearMetric, objects)
    return sum(1 for row in rows if normalize_text(row.get("facility")))


def import_institution_cadre_year_metrics(workbook, snapshot):
    header, rows = sheet_rows(workbook, "Institution_Cadre_Year")
    year_keys = year_columns(header, {"institution", "lifecycle_stage", "cadre"})
    objects = []
    for row in rows:
        institution = normalize_text(row.get("institution"))
        if not institution:
            continue
        for key in year_keys:
            count = parse_int(row.get(key))
            if not count:
                continue
            year = parse_optional_int(key)
            objects.append(NursingInstitutionCadreYearMetric(
                snapshot=snapshot,
                institution=institution[:255],
                lifecycle_stage=normalize_text(row.get("lifecycle_stage"))[:80],
                cadre=normalize_text(row.get("cadre"))[:150],
                year=year,
                year_label=key.title(),
                count=count,
                raw_payload=row_payload(row),
            ))
            if len(objects) >= BULK_CREATE_SIZE:
                bulk_create(NursingInstitutionCadreYearMetric, objects)
                objects = []
    bulk_create(NursingInstitutionCadreYearMetric, objects)
    return sum(1 for row in rows if normalize_text(row.get("institution")))


def import_province_year_metrics(workbook, snapshot):
    header, rows = sheet_rows(workbook, "Geo_Org_Nationality", "province")
    year_keys = year_columns(header, {"province"})
    objects = []
    for row in rows:
        province = normalize_text(row.get("province"))
        if not province:
            continue
        for key in year_keys:
            count = parse_int(row.get(key))
            if not count:
                continue
            year = parse_optional_int(key)
            objects.append(NursingProvinceYearMetric(
                snapshot=snapshot,
                province=province[:120],
                year=year,
                year_label=key.title(),
                count=count,
                raw_payload=row_payload(row),
            ))
    bulk_create(NursingProvinceYearMetric, objects)
    return len(rows)


def import_quality_metrics(workbook, snapshot):
    rows = data_quality_lifecycle_rows(workbook)
    grouped = {}
    for row in rows:
        stage = normalize_text(row.get("lifecycle_stage"))[:80]
        if stage not in QUALITY_LIFECYCLE_STAGES:
            continue
        metric = grouped.setdefault(stage, {
            "high_count": 0,
            "medium_count": 0,
            "needs_review_count": 0,
            "grand_total": 0,
            "raw_rows": [],
        })
        metric["high_count"] += parse_int(row.get("high"))
        metric["medium_count"] += parse_int(row.get("medium"))
        metric["needs_review_count"] += parse_int(row.get("needs_review"))
        metric["grand_total"] += parse_int(row.get("grand_total"))
        metric["raw_rows"].append(row_payload(row))

    objects = []
    for stage, metric in grouped.items():
        needs_review_percent = None
        if metric["grand_total"]:
            needs_review_percent = Decimal(metric["needs_review_count"]) / Decimal(metric["grand_total"])
        objects.append(NursingDataQualityMetric(
            snapshot=snapshot,
            lifecycle_stage=stage,
            high_count=metric["high_count"],
            medium_count=metric["medium_count"],
            needs_review_count=metric["needs_review_count"],
            grand_total=metric["grand_total"],
            needs_review_percent=needs_review_percent,
            raw_payload={"rows": metric["raw_rows"]},
        ))
    bulk_create(NursingDataQualityMetric, objects)
    return len(objects)


def import_standards_maps(workbook, snapshot):
    imported = 0
    _header, platform_rows = sheet_rows(workbook, "Platform_Field_Map")
    platform_objects = [
        NursingStandardsFieldMap(
            snapshot=snapshot,
            map_type="platform",
            platform_field=normalize_text(row.get("platform_field"))[:255],
            unified_field=normalize_text(row.get("unified_fact_field"))[:255],
            used_for=normalize_text(row.get("used_for")),
            data_quality_rule=normalize_text(row.get("data_quality_rule")),
            raw_payload=row_payload(row),
        )
        for row in platform_rows
        if normalize_text(row.get("unified_fact_field"))
    ]
    bulk_create(NursingStandardsFieldMap, platform_objects)
    imported += len(platform_objects)

    _header, fhir_rows = sheet_rows(workbook, "FHIR_NHWA_Map")
    fhir_objects = [
        NursingStandardsFieldMap(
            snapshot=snapshot,
            map_type="fhir_nhwa",
            unified_field=normalize_text(row.get("unified_field"))[:255],
            fhir_mapping=normalize_text(row.get("fhir_interoperability_mapping")),
            nhwa_dimension=normalize_text(row.get("who_nhwa_analytics_dimension")),
            implementation_note=normalize_text(row.get("implementation_note")),
            raw_payload=row_payload(row),
        )
        for row in fhir_rows
        if normalize_text(row.get("unified_field"))
    ]
    bulk_create(NursingStandardsFieldMap, fhir_objects)
    imported += len(fhir_objects)
    return imported


def import_aliases(workbook, snapshot):
    institution_names = set()
    for sheet_name, first_header in (("Institution_Summary", None), ("Institution_Cadre_Year", None)):
        _header, rows = sheet_rows(workbook, sheet_name, first_header)
        for row in rows:
            name = normalize_text(row.get("institution"))
            if name:
                institution_names.add(name)
    institution_objects = [
        NursingInstitutionAlias(
            snapshot=snapshot,
            raw_name=name[:255],
            normalized_name=normalize_name_key(name)[:255],
            raw_payload={"source": "analytics_workbook"},
        )
        for name in sorted(institution_names)
    ]
    bulk_create(NursingInstitutionAlias, institution_objects)

    facility_keys = {}
    for sheet_name in ("Facility_Summary", "Facility_Cadre_Year"):
        _header, rows = sheet_rows(workbook, sheet_name)
        for row in rows:
            name = normalize_text(row.get("facility"))
            if not name:
                continue
            key = (
                name,
                normalize_text(row.get("province")),
                normalize_text(row.get("organization_type")),
            )
            facility_keys[key] = row_payload(row)
    facility_objects = [
        NursingFacilityAlias(
            snapshot=snapshot,
            raw_name=name[:255],
            normalized_name=normalize_name_key(name)[:255],
            province=province[:120],
            organization_type=organization_type[:160],
            raw_payload=payload,
        )
        for (name, province, organization_type), payload in sorted(facility_keys.items())
    ]
    bulk_create(NursingFacilityAlias, facility_objects)
    return len(institution_objects) + len(facility_objects)


def sheet_row_counts(workbook):
    counts = {}
    for name in workbook.sheetnames:
        sheet = workbook[name]
        counts[name] = max((getattr(sheet, "max_row", 0) or 0) - 1, 0)
    return counts


def build_filter_options(snapshot):
    return {
        "stages": list(snapshot.lifecycle_facts.order_by("lifecycle_stage").values_list("lifecycle_stage", flat=True).distinct()),
        "years": list(
            snapshot.stage_year_metrics
            .exclude(year__isnull=True)
            .order_by("year")
            .values_list("year", flat=True)
        ),
        "cadres": list(snapshot.cadre_stage_metrics.order_by("cadre").values_list("cadre", flat=True)),
        "provinces": list(snapshot.province_year_metrics.order_by("province").values_list("province", flat=True).distinct()),
    }


class NursingAnalyticsSnapshotImporter:
    def __init__(self, workbook_path, activate=True, force=False, initiated_by=None):
        self.workbook_path = Path(workbook_path)
        self.activate = activate
        self.force = force
        self.initiated_by = initiated_by

    def import_workbook(self):
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        digest = file_sha256(self.workbook_path)
        existing = NursingAnalyticsSnapshot.objects.filter(source_file_hash=digest).order_by("-created_at").first()
        if existing and not self.force:
            if self.activate and not existing.is_active:
                self.activate_snapshot(existing)
            return existing, False

        if existing and self.force:
            existing.delete()

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        sheet_counts = sheet_row_counts(workbook)
        kpi_summary, metadata = build_kpi_summary(workbook)
        batch = DataImportBatch.objects.create(
            source_file_name=self.workbook_path.name,
            source_file_path=str(self.workbook_path),
            source_kind=ANALYTICS_SOURCE_KIND,
            status="running",
            total_sheets=len(workbook.sheetnames),
            total_rows=sum(sheet_counts.values()),
            summary={
                "source_file_hash": digest,
                "analytics_snapshot_active": False,
                "kpi_summary": kpi_summary,
            },
            initiated_by=self.initiated_by,
        )

        try:
            with transaction.atomic():
                snapshot = NursingAnalyticsSnapshot.objects.create(
                    source_batch=batch,
                    source_file_name=self.workbook_path.name,
                    source_file_path=str(self.workbook_path),
                    source_file_hash=digest,
                    workbook_generated_on=metadata.get("workbook_generated_on"),
                    workbook_title=metadata.get("workbook_title", ""),
                    is_active=False,
                    total_sheets=len(workbook.sheetnames),
                    sheet_row_counts=sheet_counts,
                    kpi_summary=kpi_summary,
                    import_summary={"source_kind": ANALYTICS_SOURCE_KIND},
                )
                imported = Counter()
                imported["Fact_Lifecycle"] = import_lifecycle_facts(workbook, snapshot)
                imported["Practitioner_Index"] = import_practitioner_index(workbook, snapshot)
                imported["Year_Stage"] = import_stage_year_metrics(workbook, snapshot)
                imported["Cadre_Stage"] = import_cadre_stage_metrics(workbook, snapshot)
                imported["Facility_Cadre_Year"] = import_facility_cadre_year_metrics(workbook, snapshot)
                imported["Institution_Cadre_Year"] = import_institution_cadre_year_metrics(workbook, snapshot)
                imported["Geo_Org_Nationality"] = import_province_year_metrics(workbook, snapshot)
                imported["Data_Quality"] = import_quality_metrics(workbook, snapshot)
                imported["Standards_Field_Map"] = import_standards_maps(workbook, snapshot)
                imported["Reference_Aliases"] = import_aliases(workbook, snapshot)

                snapshot.imported_rows = sum(imported.values())
                snapshot.processed_sheets = len([name for name, count in imported.items() if count >= 0])
                snapshot.filter_options = build_filter_options(snapshot)
                snapshot.import_summary = {
                    "source_kind": ANALYTICS_SOURCE_KIND,
                    "imported_counts": dict(imported),
                    "analytics_snapshot_active": self.activate,
                }
                snapshot.save(update_fields=[
                    "imported_rows",
                    "processed_sheets",
                    "filter_options",
                    "import_summary",
                ])

                if self.activate:
                    self.activate_snapshot(snapshot)

            batch.status = "completed"
            batch.processed_sheets = len(workbook.sheetnames)
            batch.processed_rows = snapshot.imported_rows
            batch.completed_at = timezone.now()
            batch.summary = {
                **batch.summary,
                "analytics_snapshot_id": str(snapshot.snapshot_id),
                "analytics_snapshot_pk": snapshot.pk,
                "analytics_snapshot_active": snapshot.is_active,
                "sheet_row_counts": sheet_counts,
                "imported_rows": snapshot.imported_rows,
            }
            batch.save(update_fields=["status", "processed_sheets", "processed_rows", "completed_at", "summary"])
        except Exception as exc:
            batch.status = "failed"
            batch.completed_at = timezone.now()
            batch.summary = {**batch.summary, "error": str(exc)}
            batch.save(update_fields=["status", "completed_at", "summary"])
            raise
        finally:
            workbook.close()

        return snapshot, True

    def activate_snapshot(self, snapshot):
        NursingAnalyticsSnapshot.objects.exclude(pk=snapshot.pk).filter(is_active=True).update(is_active=False)
        snapshot.is_active = True
        snapshot.activated_at = timezone.now()
        snapshot.save(update_fields=["is_active", "activated_at"])
        if snapshot.source_batch_id:
            summary = dict(snapshot.source_batch.summary or {})
            summary["analytics_snapshot_active"] = True
            snapshot.source_batch.summary = summary
            snapshot.source_batch.save(update_fields=["summary"])
