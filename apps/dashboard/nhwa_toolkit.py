import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from django.conf import settings
from django.core.cache import cache

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - surfaced on the page if dependency is missing.
    load_workbook = None
    get_column_letter = None


SOURCE_DIR = Path(settings.BASE_DIR) / "docs" / "nhwa_toolkit" / "source"

NHWA_SOURCE_DOCUMENTS = {
    "workbook": {
        "title": "PNG NHWA Data Collection Toolkit v2",
        "filename": "PNG_NHWA_Data_Collection_Toolkit_v2_May2026.xlsx",
        "type": "Excel workbook",
    },
    "presentation": {
        "title": "PNG NHWA Output 2 Presentation",
        "filename": "PNG_NHWA_Output2_Presentation_May2026.pptx",
        "type": "PowerPoint presentation",
    },
}

NHWA_PLATFORM_ALIGNMENT = [
    {
        "title": "PHA Establishment",
        "module": "NHWA Modules 1 and 2",
        "platform": "Facility, workforce stock, province, district, establishment, in-post, vacancy, sex, age-band, and funding-source evidence.",
        "action": "Use Workforce Flow and Records Hub imports as staging sources before official reporting.",
    },
    {
        "title": "Training Schools and MEU",
        "module": "NHWA Module 3",
        "platform": "Training institutions, nursing schools, CHW schools, medical school programmes, graduates, enrolment, accreditation, and faculty staffing.",
        "action": "Keep Nursing schools and CHW schools separated; report graduate output and enrolment by programme.",
    },
    {
        "title": "Professional Councils",
        "module": "NHWA Module 1",
        "platform": "Nursing Council, Medical Board, Pharmacy Board, and Dental Board registered versus practising totals.",
        "action": "Separate total registered records from current practising certificate holders before publishing counts.",
    },
    {
        "title": "Health Workforce Finance",
        "module": "NHWA Module 4",
        "platform": "Finance dashboards and evidence uploads support salary, allowance, training, CPD, and partner-disbursement review.",
        "action": "Keep finance office-scoped and mark expenditure as payment rows, not live people counts.",
    },
    {
        "title": "Cadre Mapping",
        "module": "Cross-module terminology",
        "platform": "Maps local PHA/Alesco titles into NHWA categories for nurses, midwives, CHWs, HEOs, doctors, allied health, and support staff.",
        "action": "Apply most-specific category first; classify midwifery separately from general nursing where title/function indicates it.",
    },
    {
        "title": "Data Quality Checklist",
        "module": "Pre-submission control",
        "platform": "Duplicate review, missing-data review, audit trail, import provenance, and correction workflow.",
        "action": "Do not submit official NHWA figures until checklist items, source provenance, and sign-off are complete.",
    },
]

NHWA_REPORTING_GATES = [
    "Blank is not zero: blanks mean unknown or unavailable; zero means confirmed zero.",
    "Practising means payroll-confirmed staff or current practising certificate holders, not only registered persons.",
    "Nursing and midwifery rows must be separated before reporting.",
    "Council reports must distinguish registered totals from active and practising certificate holders.",
    "Age-band data is a core NHWA requirement and must be sourced from councils if payroll age is incomplete.",
    "Finance entries are expenditure or disbursement evidence, not budget allocations unless specifically labelled.",
    "All toolkit sheets must retain province/institution, reporting period, data source, completed by, and date fields.",
    "Formula or auto-calculated cells must not be manually edited.",
]

NHWA_DISCLOSURE_CONTROLS = [
    {
        "title": "No Source File Downloads",
        "detail": "The workbook and presentation remain server-side reference material. The platform renders controlled summaries, template structures, and reporting rules only.",
    },
    {
        "title": "Role-Restricted Access",
        "detail": "The NHWA section is available only to authorised regulatory, data-quality, and finance users. It is not exposed through public register pages.",
    },
    {
        "title": "Safe Reporting Boundary",
        "detail": "Published outputs must use aggregate counts, approved definitions, and limitation notes. Person-level payroll, receipt, or identity evidence stays in protected review workflows.",
    },
    {
        "title": "Source Provenance Retained",
        "detail": "The platform preserves source document status and NHWA mapping logic without redistributing the original documents from the web interface.",
    },
]


def get_nhwa_document_path(document_key):
    document = NHWA_SOURCE_DOCUMENTS.get(document_key)
    if not document:
        return None
    path = SOURCE_DIR / document["filename"]
    if path.exists():
        return path
    return None


def _document_status(document_key, document):
    path = SOURCE_DIR / document["filename"]
    return {
        "key": document_key,
        "title": document["title"],
        "type": document["type"],
        "filename": document["filename"],
        "exists": path.exists(),
        "size_kb": round(path.stat().st_size / 1024, 1) if path.exists() else None,
        "access_policy": "Internal source only; not downloadable from the platform.",
    }


def _clean_cell_value(value):
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return text


def _display_workbook_text(raw_text):
    if not raw_text:
        return ""
    if raw_text.startswith("="):
        return "Calculated by platform"
    text = re.sub(
        r"=\s*(SUM|IF|COUNTIF|COUNT|AVERAGE|MIN|MAX)\([^)]*\)",
        "by platform calculation",
        raw_text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"\b(SUM|IF|COUNTIF|COUNT|AVERAGE|MIN|MAX)\([^)]*\)",
        "platform calculation",
        text,
        flags=re.IGNORECASE,
    )
    return text


def _rgb_from_fill(cell):
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type not in {"solid", "gray125"}:
        return ""
    color = fill.fgColor
    if not color:
        return ""
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:]
        if rgb.upper() not in {"000000", "FFFFFF"}:
            return rgb.upper()
    return ""


def _is_dark_hex(rgb):
    if not rgb or len(rgb) != 6:
        return False
    red = int(rgb[0:2], 16)
    green = int(rgb[2:4], 16)
    blue = int(rgb[4:6], 16)
    return (red * 0.299 + green * 0.587 + blue * 0.114) < 135


def _cell_style(cell, text):
    styles = []
    rgb = _rgb_from_fill(cell)
    if rgb:
        styles.append(f"background-color: #{rgb};")
        text_color = "#ffffff" if _is_dark_hex(rgb) else "#0f172a"
        styles.append(f"color: {text_color} !important;")
        styles.append(f"-webkit-text-fill-color: {text_color} !important;")
    if text.startswith("="):
        styles.append("font-family: Consolas, monospace;")
    return " ".join(styles)


def _cell_class(cell, text):
    classes = []
    if text.startswith("="):
        classes.append("formula-cell")
    if cell.row <= 3:
        classes.append("sheet-heading-cell")
    if cell.font and cell.font.bold:
        classes.append("bold-cell")
    return " ".join(classes)


def _row_class(row_index, cells):
    nonempty = [cell["text"] for cell in cells if cell["text"]]
    if row_index <= 3:
        return "sheet-heading-row"
    if len(nonempty) == 1 and nonempty[0].upper() == nonempty[0]:
        return "sheet-section-row"
    return ""


def _read_workbook():
    workbook_path = get_nhwa_document_path("workbook")
    if not workbook_path:
        return {"available": False, "error": "NHWA workbook source file is not available.", "sheets": []}
    if load_workbook is None:
        return {"available": False, "error": "openpyxl is not installed for workbook rendering.", "sheets": []}

    workbook = load_workbook(workbook_path, data_only=False)
    sheets = []
    for worksheet in workbook.worksheets:
        max_column = min(worksheet.max_column or 1, 24)
        rows = []
        non_empty_count = 0
        for excel_row in worksheet.iter_rows(max_col=max_column):
            cells = []
            row_has_value = False
            for cell in excel_row:
                raw_text = _clean_cell_value(cell.value)
                text = _display_workbook_text(raw_text)
                if raw_text:
                    row_has_value = True
                    non_empty_count += 1
                cells.append(
                    {
                        "text": text,
                        "style": _cell_style(cell, raw_text),
                        "class_name": _cell_class(cell, raw_text),
                    }
                )
            if row_has_value:
                rows.append(
                    {
                        "index": excel_row[0].row,
                        "class_name": _row_class(excel_row[0].row, cells),
                        "cells": cells,
                    }
                )

        sheets.append(
            {
                "name": worksheet.title,
                "display_name": worksheet.title.replace("_", " ").title(),
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "non_empty_count": non_empty_count,
                "columns": [get_column_letter(index) for index in range(1, max_column + 1)],
                "rows": rows,
            }
        )
    return {"available": True, "error": "", "sheets": sheets}


def _read_presentation():
    presentation_path = get_nhwa_document_path("presentation")
    if not presentation_path:
        return {"available": False, "error": "NHWA presentation source file is not available.", "slides": []}

    namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    slide_matcher = re.compile(r"ppt/slides/slide(\d+)\.xml$")
    slides = []
    with zipfile.ZipFile(presentation_path) as archive:
        slide_names = sorted(
            [name for name in archive.namelist() if slide_matcher.match(name)],
            key=lambda name: int(slide_matcher.match(name).group(1)),
        )
        for slide_number, slide_name in enumerate(slide_names, start=1):
            root = ET.fromstring(archive.read(slide_name))
            text_parts = []
            for text_node in root.findall(".//a:t", namespace):
                text = _clean_cell_value(text_node.text)
                if text:
                    text_parts.append(text)
            title = text_parts[0] if text_parts else f"Slide {slide_number}"
            subtitle = text_parts[1] if len(text_parts) > 1 else ""
            slide_points = [
                part
                for part in (text_parts[2:] if len(text_parts) > 2 else [])
                if not re.fullmatch(r"0?\d", part)
            ]
            slides.append(
                {
                    "number": slide_number,
                    "title": title,
                    "subtitle": subtitle,
                    "text_parts": slide_points,
                    "metrics": _slide_metrics(slide_points),
                    "full_text": " / ".join(text_parts),
                }
            )
    return {"available": True, "error": "", "slides": slides}


def _slide_metrics(parts):
    metrics = []
    seen = set()
    metric_pattern = re.compile(
        r"\b\d{1,3}(?:,\d{3})+\b|\b\d+(?:\.\d+)?%\b|\b\d+\b(?=\s+(?:files|positions|provinces|cadres|slides|minutes|rows|records|datasets))",
        flags=re.IGNORECASE,
    )
    for part in parts:
        for match in metric_pattern.finditer(part):
            value = match.group(0)
            key = (value, part[:60])
            if key in seen:
                continue
            seen.add(key)
            width = 58
            numeric = value.replace(",", "").replace("%", "")
            try:
                number = float(numeric)
                width = max(18, min(100, int(number if value.endswith("%") else number / 300)))
            except ValueError:
                pass
            metrics.append(
                {
                    "value": value,
                    "label": part[:90],
                    "bar_width": width,
                }
            )
            if len(metrics) >= 4:
                return metrics
    return metrics


def build_nhwa_toolkit_context():
    cache_key = "dashboard:nhwa-toolkit:v4"
    cached = cache.get(cache_key)
    if cached:
        return cached

    workbook = _read_workbook()
    presentation = _read_presentation()
    context = {
        "source_documents": [
            _document_status(key, document)
            for key, document in NHWA_SOURCE_DOCUMENTS.items()
        ],
        "workbook": workbook,
        "presentation": presentation,
        "alignment_cards": NHWA_PLATFORM_ALIGNMENT,
        "reporting_gates": NHWA_REPORTING_GATES,
        "disclosure_controls": NHWA_DISCLOSURE_CONTROLS,
    }
    cache.set(cache_key, context, 300)
    return context
