import hashlib
from pathlib import Path

from django.db import transaction
from django.db.models import Sum
from django.utils import timezone
from openpyxl import load_workbook

from apps.dashboard.models import NursingLifecycleFact
from apps.dashboard.nursing_analytics import active_nursing_analytics_snapshot
from apps.dashboard.nursing_analytics_import import file_sha256
from apps.workforce.models import DataImportBatch


CATHERINE_BREAKDOWN_SOURCE_KIND = "nursing_catherine_licence_breakdown"
DEFAULT_CATHERINE_SOURCE_DIR = (
    Path.home()
    / "Documents"
    / "ProjectApps"
    / "databasedocuments"
    / "spreadsheets"
    / "Nursing_Council_Cleansed_data_current"
)
DEFAULT_CLEANED_LICENCE_WORKBOOK = DEFAULT_CATHERINE_SOURCE_DIR / "PNG_Nursing_Council_Cleaned_Licence_Breakdown.xlsx"
DEFAULT_CADRE_BREAKDOWN_WORKBOOK = DEFAULT_CATHERINE_SOURCE_DIR / "PNG_Nursing_Council_Cadre_Breakdown.xlsx"


def _clean_key(value):
    return str(value or "").strip()


def _normalise_identity(value):
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _safe_int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _sheet_dicts(workbook, sheet_name, header_row):
    sheet = workbook[sheet_name]
    headers = [
        str(value).strip() if value is not None else ""
        for value in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    ]
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None for value in values):
            continue
        yield dict(zip(headers, values))


def _workbook_sheet_row_counts(workbook):
    counts = {}
    for sheet in workbook.worksheets:
        max_row = getattr(sheet, "max_row", None)
        if max_row:
            counts[sheet.title] = _safe_int(max_row)
        else:
            counts[sheet.title] = sum(1 for _ in sheet.iter_rows(values_only=True))
    return counts


def _dashboard_metrics(workbook):
    sheet = workbook["Dashboard"]
    metrics = {}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        label = _clean_key(row[0])
        if not label:
            continue
        metrics[label] = {
            "provisional": _safe_int(row[1]),
            "full_licence": _safe_int(row[2]),
            "combined": _safe_int(row[3]),
        }
    return metrics


def _institution_totals(workbook):
    rows = list(_sheet_dicts(workbook, "Institution_Summary", 4))
    return {
        "canonical_institution_rows": len(rows),
        "provisional_total": sum(_safe_int(row.get("Provisional_Total")) for row in rows),
        "full_licence_total": sum(_safe_int(row.get("Full_Licence_Total")) for row in rows),
        "combined_total": sum(_safe_int(row.get("Combined_Total")) for row in rows),
    }


def _year_breakdown_totals(workbook):
    combined_rows = list(_sheet_dicts(workbook, "Combined_By_Inst_Year", 4))
    provisional_rows = list(_sheet_dicts(workbook, "Prov_By_Inst_Year", 4))
    full_rows = list(_sheet_dicts(workbook, "Full_By_Inst_Year", 4))
    return {
        "combined_rows": len(combined_rows),
        "provisional_rows": len(provisional_rows),
        "full_licence_rows": len(full_rows),
        "provisional_total": sum(_safe_int(row.get("Provisional_Count")) for row in combined_rows),
        "full_licence_total": sum(_safe_int(row.get("Full_Licence_Count")) for row in combined_rows),
        "combined_total": sum(_safe_int(row.get("Combined_Count")) for row in combined_rows),
        "provisional_sheet_total": sum(_safe_int(row.get("Provisional_Count")) for row in provisional_rows),
        "full_licence_sheet_total": sum(_safe_int(row.get("Full_Licence_Count")) for row in full_rows),
    }


def _data_quality_summary(workbook):
    issue_rows = list(_sheet_dicts(workbook, "DQ_Row_Issues", 4))
    by_issue = {}
    by_dataset = {}
    for row in issue_rows:
        issue = _clean_key(row.get("Issue_Type")) or "Unspecified"
        dataset = _clean_key(row.get("Dataset")) or "Unspecified"
        by_issue[issue] = by_issue.get(issue, 0) + 1
        by_dataset[dataset] = by_dataset.get(dataset, 0) + 1
    return {
        "row_issue_count": len(issue_rows),
        "issues_by_type": by_issue,
        "issues_by_dataset": by_dataset,
    }


def _licence_identity_sets(workbook):
    provisional_ids = {
        _clean_key(row.get("Licence_ID")).upper()
        for row in _sheet_dicts(workbook, "Clean_Provisional", 4)
        if _clean_key(row.get("Licence_ID"))
    }
    full_ids = {
        _clean_key(row.get("Licence_ID")).upper()
        for row in _sheet_dicts(workbook, "Clean_Full_Licence", 4)
        if _clean_key(row.get("Licence_ID"))
    }
    return provisional_ids, full_ids


def _cadre_breakdown(workbook):
    sheet = workbook["Cadre_Breakdown"]
    headers = [
        str(value).strip() if value is not None else ""
        for value in next(sheet.iter_rows(min_row=11, max_row=11, values_only=True))
    ]
    rows = []
    grand_total = {}
    for values in sheet.iter_rows(min_row=12, values_only=True):
        if not any(value is not None for value in values):
            continue
        row = dict(zip(headers, values))
        payload = {
            "cadre": _clean_key(row.get("Cadre")),
            "profession_group": _clean_key(row.get("Profession_Group")),
            "full_professional_clean": _safe_int(row.get("Full_Professional_Clean")),
            "provisional_clean": _safe_int(row.get("Provisional_Clean")),
            "total_clean": _safe_int(row.get("Total_Clean")),
            "full_countable_inst_year": _safe_int(row.get("Full_Countable_InstYear")),
            "provisional_countable_inst_year": _safe_int(row.get("Provisional_Countable_InstYear")),
            "classification_note": _clean_key(row.get("Classification_Note")),
        }
        if payload["cadre"] == "Grand Total":
            grand_total = payload
        elif payload["total_clean"] or payload["full_countable_inst_year"] or payload["provisional_countable_inst_year"]:
            rows.append(payload)
    return {
        "rows": rows,
        "grand_total": grand_total,
        "unclassified_clean_rows": sum(row["total_clean"] for row in rows if row["profession_group"] == "Review"),
        "mapped_clean_rows": sum(row["total_clean"] for row in rows if row["profession_group"] != "Review"),
    }


def _classification_rules(workbook):
    rules = []
    for row in _sheet_dicts(workbook, "Classification_Rules", 1):
        rules.append({
            "priority": _safe_int(row.get("Classification Rule Priority")),
            "dataset": _clean_key(row.get("Dataset")),
            "output_cadre": _clean_key(row.get("Output Cadre")),
            "trigger": _clean_key(row.get("Qualification text trigger")),
            "notes": _clean_key(row.get("Notes")),
        })
    return rules


def _unclassified_review(workbook):
    rows = []
    for row in _sheet_dicts(workbook, "Unclassified_Review", 1):
        rows.append({
            "dataset": _clean_key(row.get("Dataset")),
            "status": _clean_key(row.get("Unclassified_Status")),
            "qualification_raw": _clean_key(row.get("Qualification_Raw")),
            "qualification_canonical": _clean_key(row.get("Qualification_Canonical")),
            "count": _safe_int(row.get("Count")),
        })
    return rows


def _active_snapshot_comparison(snapshot, provisional_ids, full_ids, dashboard_metrics, cadre_breakdown):
    if not snapshot:
        return {
            "has_active_snapshot": False,
            "status": "no_active_snapshot",
            "database_action": "verification_batch_created_only",
        }

    kpis = snapshot.kpi_summary or {}
    active_provisional_ids = {
        _clean_key(value).upper()
        for value in NursingLifecycleFact.objects.filter(
            snapshot=snapshot,
            lifecycle_stage="Provisional Licence",
        ).values_list("registration_no", flat=True)
        if _clean_key(value)
    }
    active_full_ids = {
        _clean_key(value).upper()
        for value in NursingLifecycleFact.objects.filter(
            snapshot=snapshot,
            lifecycle_stage="Full Licence",
        ).values_list("registration_no", flat=True)
        if _clean_key(value)
    }
    active_cadre_totals = snapshot.cadre_stage_metrics.aggregate(
        provisional=Sum("provisional_licence_count"),
        full_licence=Sum("full_licence_count"),
    )
    clean_rows = dashboard_metrics.get("Clean rows retained after exact dedupe", {})
    normalised_full_workbook = {_normalise_identity(value) for value in full_ids}
    normalised_full_active = {_normalise_identity(value) for value in active_full_ids}

    stage_matches = {
        "provisional": _safe_int(kpis.get("clean_provisional_records")) == _safe_int(clean_rows.get("provisional")),
        "full_licence": _safe_int(kpis.get("clean_full_licence_records")) == _safe_int(clean_rows.get("full_licence")),
    }
    identity_matches = {
        "provisional_exact": len(provisional_ids - active_provisional_ids) == 0 and len(active_provisional_ids - provisional_ids) == 0,
        "full_exact": len(full_ids - active_full_ids) == 0 and len(active_full_ids - full_ids) == 0,
        "full_normalised": normalised_full_workbook == normalised_full_active,
    }
    cadre_matches = {
        "provisional": _safe_int(active_cadre_totals["provisional"]) == _safe_int(cadre_breakdown["grand_total"].get("provisional_clean")),
        "full_licence": _safe_int(active_cadre_totals["full_licence"]) == _safe_int(cadre_breakdown["grand_total"].get("full_professional_clean")),
    }
    all_match = all(stage_matches.values()) and identity_matches["provisional_exact"] and identity_matches["full_normalised"] and all(cadre_matches.values())

    return {
        "has_active_snapshot": True,
        "snapshot_pk": snapshot.pk,
        "snapshot_id": str(snapshot.snapshot_id),
        "snapshot_source_file_name": snapshot.source_file_name,
        "status": "matched_active_snapshot" if all_match else "review_required",
        "database_action": "verification_overlay_only_no_lifecycle_facts_created",
        "stage_matches": stage_matches,
        "identity_matches": identity_matches,
        "cadre_matches": cadre_matches,
        "provisional_ids": {
            "workbook_distinct": len(provisional_ids),
            "active_distinct": len(active_provisional_ids),
            "workbook_only": len(provisional_ids - active_provisional_ids),
            "active_only": len(active_provisional_ids - provisional_ids),
        },
        "full_licence_ids": {
            "workbook_distinct": len(full_ids),
            "active_distinct": len(active_full_ids),
            "workbook_only_exact": len(full_ids - active_full_ids),
            "active_only_exact": len(active_full_ids - full_ids),
            "workbook_only_normalised": len(normalised_full_workbook - normalised_full_active),
            "active_only_normalised": len(normalised_full_active - normalised_full_workbook),
        },
        "normalisation_note": (
            "Full-licence exact ID differences are punctuation-only when full_normalised is true, "
            "for example decimal separators converted to spaces during analytics import."
        ),
    }


def build_catherine_breakdown_summary(licence_workbook_path, cadre_workbook_path):
    licence_path = Path(licence_workbook_path)
    cadre_path = Path(cadre_workbook_path)
    if not licence_path.exists():
        raise FileNotFoundError(f"Cleaned licence workbook not found: {licence_path}")
    if not cadre_path.exists():
        raise FileNotFoundError(f"Cadre breakdown workbook not found: {cadre_path}")

    licence_hash = file_sha256(licence_path)
    cadre_hash = file_sha256(cadre_path)
    combined_hash = hashlib.sha256(f"{licence_hash}:{cadre_hash}".encode("utf-8")).hexdigest()

    licence_workbook = load_workbook(licence_path, read_only=True, data_only=True)
    cadre_workbook = load_workbook(cadre_path, read_only=True, data_only=True)
    try:
        dashboard_metrics = _dashboard_metrics(licence_workbook)
        provisional_ids, full_ids = _licence_identity_sets(licence_workbook)
        institution_totals = _institution_totals(licence_workbook)
        year_breakdown = _year_breakdown_totals(licence_workbook)
        data_quality = _data_quality_summary(licence_workbook)
        cadre_breakdown = _cadre_breakdown(cadre_workbook)
        classification_rules = _classification_rules(cadre_workbook)
        unclassified_review = _unclassified_review(cadre_workbook)
        active_snapshot = active_nursing_analytics_snapshot()
        active_comparison = _active_snapshot_comparison(
            active_snapshot,
            provisional_ids,
            full_ids,
            dashboard_metrics,
            cadre_breakdown,
        )
        licence_sheet_rows = _workbook_sheet_row_counts(licence_workbook)
        cadre_sheet_rows = _workbook_sheet_row_counts(cadre_workbook)
    finally:
        licence_workbook.close()
        cadre_workbook.close()

    return {
        "source_kind": CATHERINE_BREAKDOWN_SOURCE_KIND,
        "combined_source_hash": combined_hash,
        "source_files": {
            "cleaned_licence": {
                "file_name": licence_path.name,
                "file_path": str(licence_path),
                "sha256": licence_hash,
            },
            "cadre_breakdown": {
                "file_name": cadre_path.name,
                "file_path": str(cadre_path),
                "sha256": cadre_hash,
            },
        },
        "licence_dashboard": dashboard_metrics,
        "institution_totals": institution_totals,
        "year_breakdown_totals": year_breakdown,
        "data_quality": data_quality,
        "cadre_breakdown": cadre_breakdown,
        "classification_rules": classification_rules,
        "unclassified_review": unclassified_review,
        "active_snapshot_comparison": active_comparison,
        "sheet_row_counts": {
            licence_path.name: licence_sheet_rows,
            cadre_path.name: cadre_sheet_rows,
        },
        "refreshed_at": timezone.now().isoformat(),
        "logic": (
            "These workbooks verify the Provisional and Full-Licence sections already present in the active Nursing analytics snapshot. "
            "They are not inserted into operational legal registry tables and are not added a second time to lifecycle facts."
        ),
    }


def _attach_summary_to_active_snapshot(summary):
    snapshot = active_nursing_analytics_snapshot()
    if not snapshot:
        return None
    import_summary = dict(snapshot.import_summary or {})
    import_summary["catherine_licence_breakdown"] = summary
    snapshot.import_summary = import_summary
    snapshot.save(update_fields=["import_summary"])
    return snapshot


def import_catherine_breakdown(licence_workbook_path, cadre_workbook_path, force=False, initiated_by=None):
    summary = build_catherine_breakdown_summary(licence_workbook_path, cadre_workbook_path)
    existing = (
        DataImportBatch.objects
        .filter(
            source_kind=CATHERINE_BREAKDOWN_SOURCE_KIND,
            summary__combined_source_hash=summary["combined_source_hash"],
            status="completed",
        )
        .order_by("-completed_at", "-started_at")
        .first()
    )
    if existing and not force:
        snapshot = _attach_summary_to_active_snapshot(existing.summary)
        return existing, snapshot, False

    with transaction.atomic():
        batch = DataImportBatch.objects.create(
            source_file_name=" + ".join([
                summary["source_files"]["cleaned_licence"]["file_name"],
                summary["source_files"]["cadre_breakdown"]["file_name"],
            ]),
            source_file_path="; ".join([
                summary["source_files"]["cleaned_licence"]["file_path"],
                summary["source_files"]["cadre_breakdown"]["file_path"],
            ]),
            source_kind=CATHERINE_BREAKDOWN_SOURCE_KIND,
            status="completed",
            total_sheets=sum(len(rows) for rows in summary["sheet_row_counts"].values()),
            processed_sheets=sum(len(rows) for rows in summary["sheet_row_counts"].values()),
            total_rows=sum(sum(rows.values()) for rows in summary["sheet_row_counts"].values()),
            processed_rows=summary["licence_dashboard"].get("Clean rows retained after exact dedupe", {}).get("combined", 0),
            completed_at=timezone.now(),
            summary=summary,
            initiated_by=initiated_by,
        )
        snapshot = _attach_summary_to_active_snapshot(summary)
    return batch, snapshot, True


def catherine_breakdown_overlay_payload(snapshot):
    if not snapshot:
        return {}
    summary = (snapshot.import_summary or {}).get("catherine_licence_breakdown") or {}
    if not summary:
        return {}
    metrics = summary.get("licence_dashboard") or {}
    display_metric_names = [
        "Clean rows retained after exact dedupe",
        "Rows included in institution/year breakdown",
        "Rows excluded from breakdown",
        "Exact duplicate rows removed",
        "Potential duplicate licence-ID groups",
        "Canonical institutions in breakdown",
    ]
    metric_rows = [
        {
            "label": name,
            "provisional": metrics.get(name, {}).get("provisional", 0),
            "full_licence": metrics.get(name, {}).get("full_licence", 0),
            "combined": metrics.get(name, {}).get("combined", 0),
        }
        for name in display_metric_names
        if name in metrics
    ]
    cadre_rows = [
        row for row in (summary.get("cadre_breakdown") or {}).get("rows", [])
        if row.get("total_clean")
    ]
    return {
        "source_files": summary.get("source_files", {}),
        "refreshed_at": summary.get("refreshed_at", ""),
        "logic": summary.get("logic", ""),
        "metric_rows": metric_rows,
        "cadre_rows": cadre_rows,
        "cadre_grand_total": (summary.get("cadre_breakdown") or {}).get("grand_total", {}),
        "unclassified_clean_rows": (summary.get("cadre_breakdown") or {}).get("unclassified_clean_rows", 0),
        "mapped_clean_rows": (summary.get("cadre_breakdown") or {}).get("mapped_clean_rows", 0),
        "data_quality": summary.get("data_quality", {}),
        "comparison": summary.get("active_snapshot_comparison", {}),
    }
